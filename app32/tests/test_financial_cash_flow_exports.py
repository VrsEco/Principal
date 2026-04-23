import io
import os
import sys

from openpyxl import load_workbook
from pypdf import PdfReader

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from services.financial_report_service import FinancialReportService


def _cash_flow_report_payload():
    return {
        "report_type": "cash_flow",
        "report_slug": "fluxo-caixa",
        "title": "Fluxo de Caixa",
        "subtitle": "Fluxo diário com saldo inicial, realizado, projeções abertas e saldo acumulado.",
        "company_name": "Versus Gestao Corporativa",
        "generated_at": "23/04/2026 20:00",
        "filters": [
            {"label": "Janela analisada", "value": "01/05/2026 até 31/05/2026"},
            {"label": "Contas correntes", "value": "Inter Versus, Efi Banco - Conta 01"},
            {"label": "Periodicidade", "value": "Semanal"},
        ],
        "summary_cards": [
            {"label": "Saldo inicial do período", "value": "517.487,61", "tone": "primary"},
            {"label": "Entradas no fluxo", "value": "1.250,00", "tone": "positive"},
            {"label": "Saídas no fluxo", "value": "- 1.050,00", "tone": "negative"},
            {"label": "Saldo final do período", "value": "517.687,61", "tone": "primary"},
        ],
        "bank_balance_reference_label": "23/04/2026",
        "bank_account_summary_rows": [
            {
                "description": "Inter Versus",
                "limit": "R$ 50.000,00",
                "limit_value": 50000.0,
                "balance": "R$ 529.917,61",
                "balance_value": 529917.61,
                "available_total": "R$ 579.917,61",
                "available_total_value": 579917.61,
            },
            {
                "description": "Efi Banco - Conta 01",
                "limit": "R$ 0,00",
                "limit_value": 0.0,
                "balance": "R$ -3.230,00",
                "balance_value": -3230.0,
                "available_total": "R$ -3.230,00",
                "available_total_value": -3230.0,
            },
        ],
        "bank_account_summary_totals": {
            "limit": "R$ 50.000,00",
            "limit_value": 50000.0,
            "balance": "R$ 526.687,61",
            "balance_value": 526687.61,
            "available_total": "R$ 576.687,61",
            "available_total_value": 576687.61,
        },
        "columns": [
            {"key": "periodo", "label": "Período"},
            {"key": "data_inicial", "label": "Data Inicial"},
            {"key": "data_final", "label": "Data Final"},
            {"key": "saldo_inicial", "label": "Saldo Inicial"},
            {"key": "entrada", "label": "Entrada"},
            {"key": "saida", "label": "Saída"},
            {"key": "saldo_final", "label": "Saldo Final"},
            {"key": "limite", "label": "Limite"},
            {"key": "disponivel_total_final", "label": "Disp. Total Final"},
        ],
        "rows": [
            {
                "periodo": "Semana 1",
                "data_inicial": "01/05/2026",
                "data_final": "07/05/2026",
                "saldo_inicial": "R$ 517.487,61",
                "entrada": "R$ 1.250,00",
                "saida": "R$ -1.050,00",
                "saldo_final": "R$ 517.687,61",
                "limite": "R$ 50.000,00",
                "disponivel_total_final": "R$ 567.687,61",
            }
        ],
        "projected_amount_label": "Saldo do Principal Corrigido",
        "selected_receivables": [
            {
                "id": 44,
                "type_code": "RCB",
                "title_amount": "R$ 1.250,00",
                "title_amount_value": 1250.0,
                "open_amount": "R$ 1.250,00",
                "open_amount_value": 1250.0,
                "counterparty": "001 - Cliente Teste 01",
                "due_date": "02/05/2026",
            }
        ],
        "selected_receivables_totals": {
            "count": 1,
            "title_amount": "R$ 1.250,00",
            "title_amount_value": 1250.0,
            "open_amount": "R$ 1.250,00",
            "open_amount_value": 1250.0,
        },
        "selected_payables": [
            {
                "id": 45,
                "type_code": "PGT",
                "title_amount": "R$ -1.050,00",
                "title_amount_value": -1050.0,
                "open_amount": "R$ -1.050,00",
                "open_amount_value": -1050.0,
                "counterparty": "002 - Fornecedor Teste 02",
                "due_date": "02/05/2026",
            }
        ],
        "selected_payables_totals": {
            "count": 1,
            "title_amount": "R$ -1.050,00",
            "title_amount_value": -1050.0,
            "open_amount": "R$ -1.050,00",
            "open_amount_value": -1050.0,
        },
    }


def test_cash_flow_export_xlsx_creates_executive_workbook():
    content = FinancialReportService.export_xlsx(_cash_flow_report_payload())

    workbook = load_workbook(io.BytesIO(content), data_only=True)

    assert workbook.sheetnames == [
        "Resumo Executivo",
        "Fluxo Consolidado",
        "Titulos Receber",
        "Titulos Pagar",
    ]
    assert workbook["Resumo Executivo"]["A1"].value == "Fluxo de Caixa"
    assert workbook["Fluxo Consolidado"]["A5"].value == "Contas correntes"
    assert workbook["Fluxo Consolidado"]["A11"].value == "Fluxo do período"
    assert workbook["Fluxo Consolidado"]["E13"].value == 1250
    assert workbook["Fluxo Consolidado"]["F13"].value == -1050
    assert workbook["Titulos Receber"]["A1"].value == "Contas a Receber Selecionadas"
    assert workbook["Titulos Receber"]["C5"].value == 1250
    assert workbook["Titulos Pagar"]["D5"].value == -1050


def test_cash_flow_export_pdf_contains_executive_sections():
    content = FinancialReportService.export_pdf(_cash_flow_report_payload())

    assert content.startswith(b"%PDF")

    reader = PdfReader(io.BytesIO(content))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)

    assert "Fluxo de Caixa" in text
    assert "Versus Gestao Corporativa" in text
    assert "Contas correntes" in text
    assert "Fluxo do período" in text
    assert "Contas a receber selecionadas" in text
    assert "Contas a pagar selecionadas" in text
    assert "Cliente Teste 01" in text
    assert "Fornecedor Teste 02" in text
