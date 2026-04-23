import io
import os
import sys

from openpyxl import load_workbook
from pypdf import PdfReader

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from services.financial_report_service import FinancialReportService


def _working_capital_report_payload():
    return {
        "report_type": "working_capital",
        "report_slug": "capital-circulante-liquido",
        "title": "Capital Circulante Líquido",
        "subtitle": "Composição do capital circulante líquido com disponibilidades, recebíveis, exigibilidades e folga financeira.",
        "company_name": "Versus Gestao Corporativa",
        "generated_at": "23/04/2026 20:00",
        "filters": [
            {"label": "Posição base", "value": "23/04/2026"},
        ],
        "general_info": [
            {"label": "Contas programadas", "value": "6"},
        ],
        "summary_cards": [
            {"label": "Ativo circulante", "value": "20.514,93"},
            {"label": "Passivo circulante", "value": "4.586,00"},
            {"label": "CCL", "value": "15.928,93"},
        ],
        "columns": [
            {"key": "descricao", "label": "Descrição"},
            {"key": "tipo", "label": "Tipo"},
            {"key": "classe", "label": "Classe"},
            {"key": "valor", "label": "Valor"},
        ],
        "rows": [
            {"descricao": "1.1.1 - Disponibilidades", "tipo": "Ativo", "classe": "Circulante", "valor": 1776.27},
            {"descricao": "1.3.2 - Investimentos Contratados a Receber - Vencidos", "tipo": "Ativo", "classe": "Circulante", "valor": 18738.66},
            {"descricao": "2.1.2 - Contas a Pagar / Fornecedores - Vencidas", "tipo": "Passivo", "classe": "Circulante", "valor": -4586.00},
        ],
        "balance_sheet": {
            "asset": {
                "title": "Ativo",
                "current": {
                    "title": "Circulante",
                    "amount": "20.514,93",
                    "groups": [
                        {
                            "code": "1.1",
                            "label": "Disponibilidades",
                            "amount": "1.776,27",
                            "items": [{"code": "1.1.1", "label": "Disponibilidades", "amount": "1.776,27"}],
                        },
                        {
                            "code": "1.3",
                            "label": "Investimentos",
                            "amount": "18.738,66",
                            "items": [{"code": "1.3.2", "label": "Investimentos Contratados a Receber - Vencidos", "amount": "18.738,66"}],
                        },
                    ],
                },
                "non_current": {"title": "Ativo Não Circulante", "amount": "0,00"},
            },
            "liability": {
                "title": "Passivo",
                "current": {
                    "title": "Circulante",
                    "amount": "4.586,00",
                    "groups": [
                        {
                            "code": "2.1",
                            "label": "Fornecedores",
                            "amount": "4.586,00",
                            "items": [{"code": "2.1.2", "label": "Contas a Pagar / Fornecedores - Vencidas", "amount": "4.586,00"}],
                        }
                    ],
                },
                "non_current": {"title": "Passivo Não Circulante", "amount": "0,00"},
                "equity": {"title": "Patrimônio Líquido", "amount": "15.928,93"},
            },
            "working_capital": {"title": "Capital Circulante Líquido", "amount": "15.928,93"},
            "patrimonial_status": {"title": "Situação Patrimonial", "amount": "15.928,93"},
        },
    }


def test_working_capital_export_xlsx_creates_balance_workbook():
    content = FinancialReportService.export_xlsx(_working_capital_report_payload())

    workbook = load_workbook(io.BytesIO(content), data_only=True)

    assert workbook.sheetnames == ["Balanço CCL", "Filtros e resumo", "Base analítica"]
    assert workbook["Balanço CCL"]["A1"].value == "Capital Circulante Líquido"
    assert workbook["Balanço CCL"]["A6"].value == "ATIVO"
    assert workbook["Balanço CCL"]["E6"].value == "PASSIVO"
    assert workbook["Balanço CCL"]["A8"].value == "1.1 - Disponibilidades"
    labels = [row[0] for row in workbook["Balanço CCL"].iter_rows(min_col=1, max_col=1, values_only=True) if row and row[0]]
    assert "Situação Patrimonial" in labels


def test_working_capital_export_pdf_contains_balance_sections():
    content = FinancialReportService.export_pdf(_working_capital_report_payload())

    assert content.startswith(b"%PDF")

    reader = PdfReader(io.BytesIO(content))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)

    assert "Capital Circulante Líquido" in text
    assert "Versus Gestao Corporativa" in text
    assert "Ativo" in text
    assert "Passivo" in text
    assert "Patrimônio Líquido" in text
    assert "Situação Patrimonial" in text
