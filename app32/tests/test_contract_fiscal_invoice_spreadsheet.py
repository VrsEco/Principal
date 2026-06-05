from io import BytesIO

from openpyxl import load_workbook

from services.contract_fiscal_invoice_spreadsheet import (
    NFSE_INTEGRATION_HEADERS,
    NFSE_REFERENCE_FIELDS,
    build_nfse_integration_workbook,
)


def test_nfse_integration_workbook_matches_save_water_template_columns():
    workbook_bytes = build_nfse_integration_workbook(
        [
            {
                "CPF_CNPJ": "22424335000170",
                "Nome": "CONDOMINIO DO NORTH WAY SHOPPING",
                "Email": "flavio@aclf.com.br",
                "Valor": 9250,
                "Codigo_Servico": "1401001",
                "Endereco_Pais": "BRA",
                "Endereco_Cep": "53401445",
                "Descricao": "Tratamento e monitoramento do sistema de agua e refrigeracao.",
                "IBSCBS_Indicador_Operacao": "050101",
                "IBSCBS_Codigo_Classificacao": "000001",
                "IBSCBS_Tipo_Operacao": "1401101",
                "NBS": "120011000",
                "CNAE": "4322302",
                "Aliquota_ISS": "5",
                "Valor_ISS": "462,50",
                "Retencao_OUTROS": "462,50",
            }
        ]
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    sheet = workbook["Sheet1"]
    reference_sheet = workbook["Sheet2"]

    assert sheet.max_column == 206
    assert [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)] == NFSE_INTEGRATION_HEADERS
    assert reference_sheet.max_row == 210
    assert [reference_sheet.cell(row=index, column=1).value for index in range(1, reference_sheet.max_row + 1)] == NFSE_REFERENCE_FIELDS
    assert sheet["A2"].value == "22424335000170"
    assert sheet["D2"].value == 9250
    assert sheet["AB2"].value == "462,50"
    assert sheet["AI2"].value == "462,50"
