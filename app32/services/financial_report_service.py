from __future__ import annotations

import io
import os
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Tuple

from flask import current_app, has_app_context
from financial_domain import build_title_operational_state_metadata
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import or_

from models import db
from models.financial import (
    FinancialAccountCategory,
    FinancialAssetAccount,
    FinancialBankAccount,
    FinancialBordero,
    FinancialBorderoItem,
    FinancialChartAccount,
    FinancialCostCenter,
    FinancialCounterparty,
    FinancialEntry,
    FinancialEntryAllocation,
    FinancialSchedule,
    FinancialSettlement,
)
from models.process import Process
from models.project import Project
from schemas.financial_reports import FinancialManagementReportFiltersInput
from services.financial_dashboard_analytics import FinancialDashboardAnalytics
from services.financial_service import FinancialService
from services.financial_title_amount_service import FinancialTitleAmountService
from services.financial_title_balance_service import FinancialTitleBalanceService


class FinancialReportService:
    """Relatórios gerenciais do módulo financeiro com saída HTML/PDF/XLSX."""

    WORKING_CAPITAL_PROGRAMMED_ACCOUNTS: tuple[Dict[str, Any], ...] = (
        {"id": 1, "description": "1.1.1 - Disponibilidades", "type": "Ativo", "class_name": "Circulante", "category": "Saldo Conta Corrente", "rule": "bank_balance"},
        {"id": 2, "description": "1.2.1 - Recebíveis a Vencer (até 06 meses)", "type": "Ativo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "receivable_due_180"},
        {"id": 3, "description": "1.2.2 - Recebíveis em Atraso", "type": "Ativo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "receivable_overdue"},
        {"id": 4, "description": "2.1.1 - Contas a Pagar / Fornecedores - A Vencer", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_supplier_due"},
        {"id": 5, "description": "2.1.2 - Contas a Pagar / Fornecedores - Vencidas", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_supplier_overdue"},
        {"id": 6, "description": "2.2.1 - Funcionários e Colaboradores - A Vencer", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_people_due"},
        {"id": 7, "description": "2.2.2 - Funcionários e Colaboradores - Vencidas", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_people_overdue"},
        {"id": 8, "description": "2.3.1 - Impostos a Vencer", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_tax_due"},
        {"id": 9, "description": "2.3.2 - Impostos Vencidos", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_tax_overdue"},
        {"id": 10, "description": "2.4.1 - Empréstimos e Financiamentos a Vencer", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_financing_due"},
        {"id": 11, "description": "2.4.2 - Empréstimos e Financiamentos Vencidos", "type": "Passivo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "payable_financing_overdue"},
        {"id": 12, "description": "1.3.1 - Investimentos Contratados a Receber - A Vencer", "type": "Ativo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "receivable_investment_due"},
        {"id": 13, "description": "1.3.2 - Investimentos Contratados a Receber - Vencidos", "type": "Ativo", "class_name": "Circulante", "category": "Data de Vencimento", "rule": "receivable_investment_overdue"},
    )

    REPORT_TYPES: tuple[str, ...] = (
        "schedule_report",
        "bank_statement",
        "bank_statement_dossier",
        "income_statement",
        "income_statement_2",
        "cash_flow",
        "ledger",
        "working_capital",
    )

    REPORT_DEFINITIONS: Dict[str, Dict[str, Any]] = {
        "schedule_report": {
            "code": "schedule_report",
            "slug": "agendamento",
            "label": "Relatório de Títulos Financeiros",
            "description": "Mapa operacional dos títulos financeiros com saldos, baixas e vínculo com borderô.",
            "filters": ("schedule_report_config",),
        },
        "bank_statement": {
            "code": "bank_statement",
            "slug": "extrato-bancario",
            "label": "Extrato Bancário",
            "description": "Extrato gerencial das baixas por conta bancária com saldo inicial, entradas, saídas, saldo final e composição consolidada.",
            "filters": ("period", "bank_account", "advanced_bank_statement"),
        },
        "bank_statement_dossier": {
            "code": "bank_statement_dossier",
            "slug": "dossie-extrato-bancario",
            "label": "Dossiê do Extrato Bancário",
            "description": "Dossiê documental com extrato bancário, DRE por liquidação e comprovantes anexados.",
            "filters": ("period", "bank_account", "advanced_bank_statement"),
        },
        "income_statement": {
            "code": "income_statement",
            "slug": "demonstrativo-resultados",
            "label": "Demonstração de Resultados 01",
            "description": "DRE contábil com período único e apuração independente por competência, vencimento e baixa.",
            "filters": (
                "period",
                "chart_account_multi",
                "cost_center_multi",
                "project_multi",
                "income_statement_config",
            ),
        },
        "income_statement_2": {
            "code": "income_statement_2",
            "slug": "demonstrativo-resultados-02",
            "label": "Demonstração de Resultados 02",
            "description": "DRE contábil hierárquica nas visões de competência, vencimento e baixa por conta contábil.",
            "filters": (
                "competence_period",
                "due_period",
                "settlement_period",
                "chart_account_multi",
                "cost_center_multi",
                "project_multi",
                "income_statement_config",
            ),
        },
        "cash_flow": {
            "code": "cash_flow",
            "slug": "fluxo-caixa",
            "label": "Fluxo de Caixa",
            "description": "Fluxo diário com saldo inicial, realizado, projeções abertas e saldo acumulado do período.",
            "filters": ("cash_flow_config",),
        },
        "ledger": {
            "code": "ledger",
            "slug": "razao",
            "label": "Razão",
            "description": "Razão gerencial por conta contábil, com histórico dos lançamentos, baixa vinculada e saldo acumulado.",
            "filters": ("ledger_config",),
        },
        "working_capital": {
            "code": "working_capital",
            "slug": "capital-circulante-liquido",
            "label": "Capital Circulante Líquido",
            "description": "Composição do capital circulante líquido com disponibilidades, recebíveis, exigibilidades e folga financeira.",
            "filters": ("working_capital_config",),
        },
    }

    @staticmethod
    def _definition_by_slug(report_slug: str) -> Optional[Dict[str, Any]]:
        normalized = str(report_slug or "").strip().lower()
        for definition in FinancialReportService.REPORT_DEFINITIONS.values():
            if definition["slug"] == normalized:
                return definition
        return None

    @staticmethod
    def get_report_definition(identifier: str) -> Optional[Dict[str, Any]]:
        normalized = str(identifier or "").strip().lower()
        if normalized in FinancialReportService.REPORT_DEFINITIONS:
            return FinancialReportService.REPORT_DEFINITIONS[normalized]
        return FinancialReportService._definition_by_slug(normalized)

    @staticmethod
    def get_report_definition_or_error(identifier: str) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        definition = FinancialReportService.get_report_definition(identifier)
        if not definition:
            return None, "Tipo de relatório financeiro inválido."
        return definition, None

    @staticmethod
    def default_period() -> Tuple[date, date]:
        today = date.today()
        return today.replace(day=1), today

    @staticmethod
    def _normalize_filters(report_type: str, raw_filters: Optional[Dict[str, Any]]) -> Tuple[Optional[FinancialManagementReportFiltersInput], Optional[str]]:
        payload = dict(raw_filters or {})
        if report_type == "cash_flow":
            payload.setdefault("frequency", "daily")
            payload.setdefault("include_projected", "true")
            payload.setdefault("projected_values_mode", "with_financial_correction")
            if not payload.get("excluded_projected_refs"):
                legacy_ids = payload.get("excluded_entry_ids") or []
                if legacy_ids:
                    payload["excluded_projected_refs"] = [
                        ref
                        for ref in (
                            FinancialReportService._cash_flow_projection_ref("entry", item)
                            for item in legacy_ids
                        )
                        if ref
                    ]
        if report_type == "working_capital":
            payload.setdefault("include_overdraft", "true")
        payload["report_type"] = report_type
        for key, value in list(payload.items()):
            if value in ("", None):
                payload.pop(key, None)
        try:
            data = FinancialManagementReportFiltersInput(**payload)
        except Exception as exc:
            return None, f"Filtros inválidos para o relatório: {exc}"

        period_start, period_end = data.period_start, data.period_end
        if not period_start or not period_end:
            default_start, default_end = FinancialReportService.default_period()
            period_start = period_start or data.competence_start or default_start
            period_end = period_end or data.competence_end or default_end
            data = data.model_copy(update={"period_start": period_start, "period_end": period_end})

        if data.report_type == "income_statement_2":
            updates = {}
            if not data.competence_start:
                updates["competence_start"] = data.period_start
            if not data.competence_end:
                updates["competence_end"] = data.period_end
            if updates:
                data = data.model_copy(update=updates)

        for start_attr, end_attr, label in [
            ("period_start", "period_end", "Período"),
            ("competence_start", "competence_end", "Competência"),
            ("due_start", "due_end", "Vencimento"),
            ("settlement_start", "settlement_end", "Baixa"),
        ]:
            start_value = getattr(data, start_attr)
            end_value = getattr(data, end_attr)
            if (start_value and not end_value) or (end_value and not start_value):
                return None, f"Faixa de {label.lower()} incompleta para relatório."
            if start_value and end_value and start_value > end_value:
                return None, f"Faixa de {label.lower()} inválida para relatório."

        if data.report_type in {"income_statement", "income_statement_2"}:
            data = data.model_copy(update={
                "show_code": True,
                "show_description": True,
                "order_by": "code",
                "order_direction": "asc",
                "include_budget_vs_actual": False,
            })
            if not any([data.include_open, data.include_settled]):
                return None, "Selecione ao menos um status para o DRE."
            if not any([data.include_receivable, data.include_payable]):
                return None, "Selecione ao menos um tipo para o DRE."
            if not any([data.show_competence_column, data.show_due_column, data.show_liquidation_column]):
                return None, "Selecione ao menos uma coluna principal para o DRE."
        if data.report_type == "bank_statement_dossier" and "orientation" not in payload:
            data = data.model_copy(update={"orientation": "portrait"})
        if data.report_type in {"bank_statement", "bank_statement_dossier"}:
            updates = {}
            if "order_by" not in payload:
                updates["order_by"] = "settlement_date"
            if "show_competence_date" not in payload:
                updates["show_competence_date"] = False
            if "show_due_date" not in payload:
                updates["show_due_date"] = False
            if updates:
                data = data.model_copy(update=updates)
            if not any([data.include_settled, data.include_partial, data.include_open]):
                return None, "Selecione ao menos um status para o extrato bancário."
            if not any([data.include_receivable, data.include_payable]):
                return None, "Selecione ao menos um tipo para o extrato bancário."
            if not any([
                data.show_settlement_date,
                data.show_code,
                data.show_title_number,
                data.show_description,
                data.show_counterparty,
                data.show_competence_date,
                data.show_due_date,
                data.show_title_amount,
                data.show_balance_amount,
            ]):
                return None, "Selecione ao menos uma coluna para exibir no extrato bancário."

        if data.report_type == "schedule_report":
            updates = {"orientation": "portrait"}
            if not data.competence_start:
                updates["competence_start"] = data.period_start
            if not data.competence_end:
                updates["competence_end"] = data.period_end
            if updates:
                data = data.model_copy(update=updates)
            if not any([data.include_settled, data.include_partial, data.include_open, data.include_bordero]):
                return None, "Selecione ao menos um status para o relatório de títulos financeiros."
            if not any([data.include_receivable, data.include_payable]):
                return None, "Selecione ao menos um tipo para o relatório de títulos financeiros."
            if not any([
                data.show_title_number,
                data.show_installment,
                data.show_history,
                data.show_counterparty,
                data.show_title_amount,
                data.show_correction_amount,
                data.show_balance_amount,
                data.show_competence_date,
                data.show_due_date,
                data.show_settlement_date,
            ]):
                return None, "Selecione ao menos uma coluna para exibir no relatório de títulos financeiros."
        if data.report_type == "working_capital" and data.reference_date:
            data = data.model_copy(update={"period_end": data.reference_date, "period_start": data.reference_date})
        return data, None

    @staticmethod
    def _serialize_money(value: Decimal | float | int) -> float:
        return float(Decimal(value or 0).quantize(Decimal("0.01")))

    @staticmethod
    def _format_currency(value: Decimal | float | int) -> str:
        amount = FinancialReportService._serialize_money(value)
        inteiro, decimal = f"{amount:,.2f}".split(".")
        return f"R$ {inteiro.replace(',', '.')},{decimal}"

    @staticmethod
    def _format_signed_currency(value: Decimal | float | int, *, positive_sign: bool = False) -> str:
        amount = Decimal(value or 0)
        signal = ""
        if amount < 0:
            signal = "- "
        elif positive_sign and amount > 0:
            signal = "+ "
        return f"{signal}{FinancialReportService._format_currency(abs(amount))}"

    @staticmethod
    def _format_decimal_br(value: Decimal | float | int) -> str:
        amount = FinancialReportService._serialize_money(value)
        inteiro, decimal = f"{abs(amount):,.2f}".split(".")
        label = f"{inteiro.replace(',', '.')},{decimal}"
        return f"- {label}" if amount < 0 else label

    @staticmethod
    def _reconciliation_status_label(value: Optional[str]) -> str:
        return {
            "pending": "Pendente",
            "suggested": "Sugerida",
            "matched": "Casada",
            "reconciled": "Conciliada",
            "rejected": "Rejeitada",
        }.get(str(value or "").strip().lower(), "Não informada")

    @staticmethod
    def _reconciliation_status_tone(value: Optional[str]) -> str:
        return {
            "pending": "neutral",
            "suggested": "primary",
            "matched": "primary",
            "reconciled": "positive",
            "rejected": "negative",
        }.get(str(value or "").strip().lower(), "neutral")

    @staticmethod
    def _report_card(label: str, value: Any, tone: str = "neutral") -> Dict[str, Any]:
        return {"label": label, "value": value, "tone": tone}

    @staticmethod
    def _report_info(label: str, value: Any) -> Dict[str, Any]:
        return {"label": label, "value": value}

    @staticmethod
    def _report_payload(
        definition: Dict[str, Any],
        *,
        summary_cards: List[Dict[str, Any]],
        general_info: List[Dict[str, Any]],
        columns: List[Dict[str, Any]],
        rows: List[Dict[str, Any]],
        totals: Dict[str, Any],
        extra: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        payload = {
            "title": definition["label"],
            "subtitle": definition["description"],
            "summary_cards": summary_cards,
            "general_info": general_info,
            "columns": columns,
            "rows": rows,
            "totals": totals,
        }
        if extra:
            payload.update(extra)
        return payload

    @staticmethod
    def _name_map(model, company_id: int) -> Dict[int, str]:
        query = model.query.filter(model.company_id == company_id)
        if hasattr(model, "deleted_at"):
            query = query.filter(model.deleted_at.is_(None))
        items = query.all()
        result: Dict[int, str] = {}
        for item in items:
            code = getattr(item, "code", None)
            prefix = f"{code} - " if code else ""
            result[item.id] = f"{prefix}{getattr(item, 'name', str(item.id))}"
        return result

    @staticmethod
    def _cash_flow_header_dimension_line(
        label: str,
        selected_ids: Sequence[int],
        names_by_id: Dict[int, str],
    ) -> Optional[Dict[str, str]]:
        ids = FinancialReportService._selected_ids(None, selected_ids)
        if not ids:
            return None
        return {
            "label": label,
            "value": ", ".join(names_by_id.get(item, str(item)) for item in ids),
        }

    @staticmethod
    def _working_capital_type_label(value: Optional[str]) -> str:
        return {
            "asset": "Ativo",
            "liability": "Passivo",
            "equity": "Patrimônio Líquido",
        }.get(str(value or "").strip().lower(), "Ativo")

    @staticmethod
    def _working_capital_class_label(value: Optional[str]) -> str:
        return {
            "current": "Circulante",
            "non_current": "Não Circulante",
        }.get(str(value or "").strip().lower(), "Circulante")

    @staticmethod
    def _working_capital_value_label(config_mode: str, metadata: Dict[str, Any]) -> str:
        if config_mode == "bank_balances":
            return "Saldo Conta Corrente"
        if config_mode == "manual_value":
            return "Informado na emissão"
        due_scope = str(metadata.get("due_scope") or "overdue").strip().lower()
        if due_scope == "all_future":
            return "Todas a vencer"
        if due_scope == "due_in_days":
            days = metadata.get("due_in_days")
            return f"A vencer em {days} dias" if days not in (None, "") else "A vencer em dias"
        return "Vencidas"

    @staticmethod
    def _get_working_capital_accounts(company_id: int) -> List[Dict[str, Any]]:
        category_names = FinancialReportService._name_map(FinancialAccountCategory, company_id)
        records = FinancialAssetAccount.query.filter(
            FinancialAssetAccount.company_id == company_id,
            FinancialAssetAccount.deleted_at.is_(None),
            FinancialAssetAccount.is_active.is_(True),
        ).order_by(FinancialAssetAccount.code.asc(), FinancialAssetAccount.name.asc()).all()

        if records:
            options: List[Dict[str, Any]] = []
            for item in records:
                metadata = dict(item.metadata_json or {})
                config_mode = str(metadata.get("config_mode") or "due_dates").strip().lower()
                category_id = metadata.get("category_id")
                category_label = category_names.get(category_id) if category_id else None
                options.append(
                    {
                        "id": item.id,
                        "description": f"{item.code} - {item.name}" if item.code else item.name,
                        "type": FinancialReportService._working_capital_type_label(metadata.get("patrimonial_type")),
                        "class_name": FinancialReportService._working_capital_class_label(metadata.get("account_class")),
                        "category": category_label or FinancialReportService._working_capital_type_label(metadata.get("patrimonial_type")),
                        "rule": config_mode,
                        "config_mode": config_mode,
                        "value_label": FinancialReportService._working_capital_value_label(config_mode, metadata),
                        "requires_manual_value": config_mode == "manual_value",
                        "metadata_json": metadata,
                        "code": item.code,
                        "name": item.name,
                    }
                )
            return options

        return [dict(item) for item in FinancialReportService.WORKING_CAPITAL_PROGRAMMED_ACCOUNTS]

    @staticmethod
    def get_filter_options(*, company_id: int, allowed_company_ids: Optional[Sequence[int]] = None) -> Tuple[Optional[Dict[str, List[Dict[str, Any]]]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error
        from services.financial_schedule_service import FinancialScheduleService

        def _natural_sort_parts(value: str) -> Tuple[Any, ...]:
            return tuple(int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value or ""))

        def _code_sort_key(item):
            code = str(getattr(item, "code", "") or "")
            name = str(getattr(item, "name", "") or "")
            return (_natural_sort_parts(code), _natural_sort_parts(name), int(getattr(item, "id", 0) or 0))

        def _base_records(model):
            query = model.query.filter(model.company_id == company_id)
            if hasattr(model, "deleted_at"):
                query = query.filter(model.deleted_at.is_(None))
            if hasattr(model, "is_active"):
                query = query.filter(model.is_active.is_(True))
            return sorted(query.all(), key=_code_sort_key)

        def _label(item, *, level: int = 0) -> str:
            code = str(getattr(item, "code", "") or "")
            name = str(getattr(item, "name", "") or "")
            prefix = "— " * max(level, 0)
            return f"{prefix}{code} - {name}" if code else f"{prefix}{name}"

        def _flat_list(model):
            return [
                {
                    "id": item.id,
                    "label": _label(item),
                    "code": str(getattr(item, "code", "") or ""),
                    "selectable": True,
                    "level": 0,
                }
                for item in _base_records(model)
            ]

        def _hierarchical_list(model):
            records = _base_records(model)
            by_id = {item.id: item for item in records}
            parent_ids = {item.parent_id for item in records if getattr(item, "parent_id", None)}
            level_cache: Dict[int, int] = {}

            def _level(item) -> int:
                if item.id in level_cache:
                    return level_cache[item.id]
                seen = {item.id}
                current = item
                level = 0
                while getattr(current, "parent_id", None) and current.parent_id in by_id and current.parent_id not in seen:
                    level += 1
                    seen.add(current.parent_id)
                    current = by_id[current.parent_id]
                level_cache[item.id] = level
                return level

            return [
                {
                    "id": item.id,
                    "label": _label(item, level=_level(item)),
                    "code": str(getattr(item, "code", "") or ""),
                    "selectable": bool(getattr(item, "accepts_posting", True)) and item.id not in parent_ids,
                    "level": _level(item),
                }
                for item in records
            ]

        enabled_domains, enabled_error = FinancialScheduleService.list_enabled_domains(
            company_id=company_id,
            allowed_company_ids=allowed_company_ids,
        )
        if enabled_error:
            return None, enabled_error
        enabled_domains = enabled_domains or []
        enabled_project_ids = {
            int(item.get("source_id"))
            for item in enabled_domains
            if item.get("domain_type") == "project" and str(item.get("source_kind") or "routine").strip().lower() == "routine"
        }
        enabled_process_ids = {
            int(item.get("source_id"))
            for item in enabled_domains
            if item.get("domain_type") == "process" and str(item.get("source_kind") or "routine").strip().lower() == "routine"
        }

        def _flat_list_from_enabled(model, enabled_ids: set[int]):
            return [
                {
                    "id": item.id,
                    "label": _label(item),
                    "code": str(getattr(item, "code", "") or ""),
                    "selectable": True,
                    "level": 0,
                }
                for item in _base_records(model)
                if int(item.id) in enabled_ids
            ]

        return {
            "bank_accounts": _flat_list(FinancialBankAccount),
            "chart_accounts": _hierarchical_list(FinancialChartAccount),
            "cost_centers": _hierarchical_list(FinancialCostCenter),
            "projects": _flat_list_from_enabled(Project, enabled_project_ids),
            "processes": _flat_list_from_enabled(Process, enabled_process_ids),
            "working_capital_accounts": FinancialReportService._get_working_capital_accounts(company_id),
            "counterparties": _flat_list(FinancialCounterparty),
            "movement_natures": [{"id": "credit", "label": "Entrada"}, {"id": "debit", "label": "Saída"}],
            "schedule_statuses": [
                {"id": "draft", "label": "Rascunho"},
                {"id": "active", "label": "Ativo"},
                {"id": "paused", "label": "Pausado"},
                {"id": "completed", "label": "Concluído"},
                {"id": "cancelled", "label": "Cancelado"},
            ],
            "frequencies": [
                {"id": "one_time", "label": "Uma vez"},
                {"id": "weekly", "label": "Semanal"},
                {"id": "monthly", "label": "Mensal"},
                {"id": "yearly", "label": "Anual"},
            ],
        }, None

    @staticmethod
    def list_report_types(*, company_id: int, allowed_company_ids: Optional[Sequence[int]] = None) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        return [
            {"code": code, "slug": definition["slug"], "label": definition["label"], "description": definition["description"]}
            for code, definition in FinancialReportService.REPORT_DEFINITIONS.items()
        ], None

    @staticmethod
    def _selected_ids(
        single_id: Optional[int],
        multiple_ids: Optional[Sequence[int]],
        *,
        preserve_empty_marker: bool = False,
    ) -> List[int]:
        values: List[int] = []
        has_empty_marker = False
        for current in list(multiple_ids or []) + ([single_id] if single_id else []):
            try:
                parsed = int(current)
            except (TypeError, ValueError):
                continue
            if parsed == -1 and preserve_empty_marker:
                has_empty_marker = True
                continue
            if parsed > 0 and parsed not in values:
                values.append(parsed)
        if preserve_empty_marker and has_empty_marker and not values:
            return [-1]
        return values

    @staticmethod
    def _parse_positive_int(value: Any) -> Optional[int]:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    @staticmethod
    def _cash_flow_projection_ref(kind: str, identifier: Any) -> Optional[str]:
        normalized_kind = str(kind or "").strip().lower()
        parsed_id = FinancialReportService._parse_positive_int(identifier)
        if normalized_kind not in {"entry", "schedule"} or not parsed_id:
            return None
        return f"{normalized_kind}:{parsed_id}"

    @staticmethod
    def _cash_flow_selected_projection_refs(filters: Any) -> set[str]:
        refs = {
            str(item).strip()
            for item in list(getattr(filters, "excluded_projected_refs", None) or [])
            if str(item).strip()
        }
        legacy_entry_ids = FinancialReportService._selected_ids(
            None,
            getattr(filters, "excluded_entry_ids", None),
        )
        refs.update(
            ref
            for ref in (
                FinancialReportService._cash_flow_projection_ref("entry", entry_id)
                for entry_id in legacy_entry_ids
            )
            if ref
        )
        return refs

    @staticmethod
    def _entry_settlement_totals(
        company_id: int,
        *,
        entry_ids: Optional[Sequence[int]] = None,
    ) -> Dict[int, Decimal]:
        query = db.session.query(
            FinancialSettlement.financial_entry_id,
            db.func.coalesce(db.func.sum(FinancialSettlement.principal_amount), 0),
        ).filter(
            FinancialSettlement.company_id == company_id,
            FinancialSettlement.deleted_at.is_(None),
            FinancialSettlement.settlement_status != "cancelled",
            FinancialSettlement.financial_entry_id.isnot(None),
        )
        positive_entry_ids = [int(item) for item in (entry_ids or []) if FinancialReportService._parse_positive_int(item)]
        if entry_ids is not None and not positive_entry_ids:
            return {}
        if positive_entry_ids:
            query = query.filter(FinancialSettlement.financial_entry_id.in_(positive_entry_ids))
        rows = query.group_by(FinancialSettlement.financial_entry_id).all()
        return {
            int(entry_id): Decimal(str(total or 0))
            for entry_id, total in rows
            if entry_id is not None
        }

    @staticmethod
    def _entry_outstanding_amount(entry: FinancialEntry, settled_amount: Decimal | float | int = Decimal("0")) -> Decimal:
        original_amount = Decimal(str(getattr(entry, "original_amount", None) or 0))
        settled_total = Decimal(str(settled_amount or 0))
        return max(original_amount - settled_total, Decimal("0"))

    @staticmethod
    def _projected_values_mode_label(value: Optional[str]) -> str:
        return {
            "with_financial_correction": "Com correção financeira",
            "without_financial_correction": "Sem correção financeira",
        }.get(str(value or "").strip().lower(), "Com correção financeira")

    @staticmethod
    def _entry_schedule_id(entry: FinancialEntry) -> Optional[int]:
        raw_schedule_id = getattr(entry, "financial_schedule_id", None)
        if raw_schedule_id not in (None, ""):
            parsed = FinancialReportService._parse_positive_int(raw_schedule_id)
            if parsed:
                return parsed
        external_reference = str(getattr(entry, "external_reference", "") or "").strip()
        prefix = "financial_schedule:"
        if external_reference.startswith(prefix):
            parsed = FinancialReportService._parse_positive_int(external_reference[len(prefix):])
            if parsed:
                return parsed
        return None

    @staticmethod
    def _schedule_projected_balance_snapshot(
        schedule: FinancialSchedule,
        *,
        reference_date: Optional[date] = None,
    ) -> Dict[str, Decimal]:
        from services.financial_title_adjustment_service import FinancialTitleAdjustmentService

        balance = FinancialTitleBalanceService.calculate_for_schedule(
            schedule=schedule,
            reference_date=reference_date,
        )
        principal_amount = Decimal(str(balance.get("principal_amount") or 0))
        principal_open = Decimal(str(balance.get("principal_open") or 0))
        principal_settled = Decimal(str(balance.get("principal_settled") or 0))
        settlement_total_amount = Decimal(str(balance.get("settlement_total_amount") or 0))
        suggested_financial_correction = Decimal("0")
        suggested_discount = Decimal("0")
        try:
            adjustment_simulation = FinancialTitleAdjustmentService.simulate_for_schedule(
                schedule=schedule,
                reference_date=reference_date or date.today(),
                base_amount=principal_open,
            )
            totals = adjustment_simulation.get("totals") or {}
            suggested_financial_correction = Decimal(str(totals.get("positive_adjustments") or 0))
            suggested_discount = Decimal(str(totals.get("discount") or 0))
        except Exception:
            suggested_financial_correction = Decimal("0")
            suggested_discount = Decimal("0")

        principal_corrected_open = max(
            principal_open + suggested_financial_correction - suggested_discount,
            Decimal("0"),
        )
        return {
            "principal_amount": principal_amount,
            "principal_open": principal_open,
            "principal_settled": principal_settled,
            "settlement_total_amount": settlement_total_amount,
            "financial_correction_open": suggested_financial_correction,
            "discount_open": suggested_discount,
            "principal_corrected_open": principal_corrected_open,
        }

    @staticmethod
    def _entry_projected_open_amount(
        entry: FinancialEntry,
        *,
        settled_amount: Decimal | float | int = Decimal("0"),
        projected_values_mode: Optional[str] = None,
        schedule_cache: Optional[Dict[int, FinancialSchedule]] = None,
        schedule_projection_cache: Optional[Dict[int, Dict[str, Decimal]]] = None,
    ) -> Decimal:
        principal_outstanding = FinancialReportService._entry_outstanding_amount(entry, settled_amount)
        if principal_outstanding <= Decimal("0"):
            return Decimal("0")
        if str(projected_values_mode or "with_financial_correction").strip().lower() != "with_financial_correction":
            return principal_outstanding

        schedule_id = FinancialReportService._entry_schedule_id(entry)
        if not schedule_id:
            return principal_outstanding

        snapshot = (schedule_projection_cache or {}).get(schedule_id)
        if snapshot is None:
            schedule = (schedule_cache or {}).get(schedule_id)
            if schedule is None:
                return principal_outstanding
            snapshot = FinancialReportService._schedule_projected_balance_snapshot(schedule)
            if schedule_projection_cache is not None:
                schedule_projection_cache[schedule_id] = snapshot

        schedule_principal_open = Decimal(str(snapshot.get("principal_open") or 0))
        schedule_corrected_open = Decimal(str(snapshot.get("principal_corrected_open") or 0))
        if schedule_principal_open <= Decimal("0"):
            return principal_outstanding
        if schedule_corrected_open <= Decimal("0"):
            return Decimal("0")
        ratio = principal_outstanding / schedule_principal_open
        corrected_amount = (schedule_corrected_open * ratio).quantize(Decimal("0.01"))
        return max(corrected_amount, Decimal("0"))

    @staticmethod
    def _entry_installment_label(entry: FinancialEntry) -> str:
        metadata = dict(getattr(entry, "metadata_json", None) or {})
        value = (
            metadata.get("installment_number")
            or metadata.get("parcela")
            or metadata.get("installment_label")
            or metadata.get("parcel_label")
            or metadata.get("repeat_label")
        )
        return str(value).strip() if value not in (None, "") else "-"

    @staticmethod
    def _entry_history_label(entry: FinancialEntry) -> str:
        metadata = dict(getattr(entry, "metadata_json", None) or {})
        history = (
            metadata.get("history")
            or metadata.get("historico")
            or metadata.get("description")
            or getattr(entry, "description", None)
            or getattr(entry, "memo", None)
            or getattr(entry, "notes", None)
        )
        return str(history or "Sem histórico").strip()

    @staticmethod
    def _entry_number_installment_label(entry: FinancialEntry) -> str:
        number = getattr(entry, "document_number", None) or getattr(entry, "entry_code", None) or str(getattr(entry, "id", "-"))
        installment = FinancialReportService._entry_installment_label(entry)
        return f"{number} / {installment}" if installment != "-" else str(number)

    @staticmethod
    def _format_date_br(value: Any) -> str:
        if value in (None, "", "-"):
            return "-"
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        try:
            return date.fromisoformat(str(value)).strftime("%d/%m/%Y")
        except (TypeError, ValueError):
            return str(value)

    @staticmethod
    def _cash_flow_frequency_label(value: Optional[str]) -> str:
        return {
            "daily": "Diário",
            "weekly": "Semanal",
            "monthly": "Mensal",
        }.get(str(value or "daily").strip().lower(), str(value or "Diário"))

    @staticmethod
    def _cash_flow_period_buckets(
        period_start: date,
        period_end: date,
        frequency: Optional[str],
    ) -> List[Dict[str, Any]]:
        bucket_mode = str(frequency or "daily").strip().lower()
        buckets: List[Dict[str, Any]] = []

        if bucket_mode == "weekly":
            cursor = period_start
            index = 1
            while cursor <= period_end:
                bucket_end = min(cursor + timedelta(days=6), period_end)
                buckets.append(
                    {
                        "key": f"week-{index}",
                        "label": f"Semana {index}",
                        "start": cursor,
                        "end": bucket_end,
                    }
                )
                cursor = bucket_end + timedelta(days=1)
                index += 1
            return buckets

        if bucket_mode == "monthly":
            cursor = period_start
            while cursor <= period_end:
                month_end = date(cursor.year, cursor.month, monthrange(cursor.year, cursor.month)[1])
                bucket_end = min(month_end, period_end)
                buckets.append(
                    {
                        "key": cursor.strftime("%Y-%m"),
                        "label": cursor.strftime("%m/%Y"),
                        "start": cursor,
                        "end": bucket_end,
                    }
                )
                cursor = bucket_end + timedelta(days=1)
            return buckets

        cursor = period_start
        while cursor <= period_end:
            buckets.append(
                {
                    "key": cursor.isoformat(),
                    "label": FinancialReportService._format_date_br(cursor),
                    "start": cursor,
                    "end": cursor,
                }
            )
            cursor += timedelta(days=1)
        return buckets

    @staticmethod
    def _cash_flow_entry_type_code(entry: FinancialEntry) -> str:
        entry_type = str(getattr(entry, "entry_type", None) or "").strip().lower()
        movement_nature = str(getattr(entry, "movement_nature", None) or "").strip().lower()
        if entry_type == "receivable" or movement_nature == "credit":
            return "RCB"
        return "PGT"

    @staticmethod
    def _cash_flow_bank_account_limit(account: FinancialBankAccount) -> Decimal:
        return FinancialDashboardAnalytics.extract_decimal_from_mapping(
            dict(getattr(account, "metadata_json", None) or {}),
            (
                "overdraft_limit",
                "cheque_especial_limit",
                "special_limit",
                "credit_limit",
                "limite_cheque_especial",
                "limite",
            ),
        )

    @staticmethod
    def _entry_project_ids(entry: FinancialEntry) -> List[int]:
        metadata = getattr(entry, "metadata_json", None) or {}
        candidates = [metadata.get("project_id"), metadata.get("app_project_id"), metadata.get("grv_project_id")]
        values = metadata.get("project_ids") or []
        if isinstance(values, (list, tuple, set)):
            candidates.extend(values)
        project_ids: List[int] = []
        for value in candidates:
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                continue
            if parsed > 0 and parsed not in project_ids:
                project_ids.append(parsed)
        return project_ids

    @staticmethod
    def _entry_process_ids(entry: FinancialEntry) -> List[int]:
        metadata = dict(getattr(entry, "metadata_json", None) or {})
        candidates = [
            metadata.get("process_id"),
            metadata.get("app_process_id"),
            metadata.get("grv_process_id"),
        ]
        values = metadata.get("process_ids") or []
        if isinstance(values, (list, tuple, set)):
            candidates.extend(values)

        process_instance = getattr(entry, "process_instance", None)
        if process_instance is not None:
            candidates.append(getattr(process_instance, "process_id", None))

        activity = getattr(entry, "activity", None)
        if activity is not None:
            candidates.append(getattr(activity, "process_id", None))

        process_ids: List[int] = []
        for value in candidates:
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                continue
            if parsed > 0 and parsed not in process_ids:
                process_ids.append(parsed)
        return process_ids

    @staticmethod
    def _entry_matches_projects(entry: FinancialEntry, project_ids: Sequence[int]) -> bool:
        selected = {int(item) for item in project_ids if item}
        if not selected:
            return True
        return bool(selected.intersection(FinancialReportService._entry_project_ids(entry)))

    @staticmethod
    def _entry_matches_processes(entry: FinancialEntry, process_ids: Sequence[int]) -> bool:
        selected = {int(item) for item in process_ids if item}
        if not selected:
            return True
        return bool(selected.intersection(FinancialReportService._entry_process_ids(entry)))

    @staticmethod
    def _sort_income_statement_account_ids(
        account_ids: Sequence[int],
        hierarchy_nodes: Dict[int, Dict[str, Any]],
        *,
        order_by: str = "code",
        reverse: bool = False,
    ) -> List[int]:
        sorted_ids = [account_id for account_id in account_ids if account_id in hierarchy_nodes]
        sorted_ids.sort(
            key=lambda account_id: (
                (hierarchy_nodes.get(account_id) or {}).get("codigo")
                if order_by == "code"
                else (hierarchy_nodes.get(account_id) or {}).get("descricao"),
                (hierarchy_nodes.get(account_id) or {}).get("descricao"),
            ),
            reverse=reverse,
        )
        return sorted_ids

    @staticmethod
    def _resolve_income_statement_root_ids(
        hierarchy_nodes: Dict[int, Dict[str, Any]],
        *,
        order_by: str = "code",
        reverse: bool = False,
    ) -> List[int]:
        root_ids: List[int] = []
        for account_id, node in hierarchy_nodes.items():
            parent_id = node.get("parent_id")
            if not parent_id or parent_id not in hierarchy_nodes:
                root_ids.append(account_id)
        if not root_ids:
            root_ids = list(hierarchy_nodes.keys())
        return FinancialReportService._sort_income_statement_account_ids(
            root_ids,
            hierarchy_nodes,
            order_by=order_by,
            reverse=reverse,
        )

    @staticmethod
    def _schedule_project_ids(schedule: FinancialSchedule) -> List[int]:
        metadata = dict(getattr(schedule, "metadata_json", None) or {})
        candidates = [
            metadata.get("project_id"),
            metadata.get("app_project_id"),
            metadata.get("grv_project_id"),
        ]
        values = metadata.get("project_ids") or []
        if isinstance(values, (list, tuple, set)):
            candidates.extend(values)

        allocations = metadata.get("allocations") or []
        if isinstance(allocations, list):
            for allocation in allocations:
                if not isinstance(allocation, dict):
                    continue
                candidates.extend([
                    allocation.get("project_id"),
                    allocation.get("app_project_id"),
                    allocation.get("grv_project_id"),
                ])
                allocation_values = allocation.get("project_ids") or []
                if isinstance(allocation_values, (list, tuple, set)):
                    candidates.extend(allocation_values)

        project_ids: List[int] = []
        for value in candidates:
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                continue
            if parsed > 0 and parsed not in project_ids:
                project_ids.append(parsed)
        return project_ids

    @staticmethod
    def _schedule_matches_projects(schedule: FinancialSchedule, project_ids: Sequence[int]) -> bool:
        selected = {int(item) for item in project_ids if item}
        if not selected:
            return True
        return bool(selected.intersection(FinancialReportService._schedule_project_ids(schedule)))

    @staticmethod
    def _schedule_process_ids(schedule: FinancialSchedule) -> List[int]:
        metadata = dict(getattr(schedule, "metadata_json", None) or {})
        candidates = [
            metadata.get("process_id"),
            metadata.get("app_process_id"),
            metadata.get("grv_process_id"),
        ]
        values = metadata.get("process_ids") or []
        if isinstance(values, (list, tuple, set)):
            candidates.extend(values)

        process_instance = getattr(schedule, "process_instance", None)
        if process_instance is not None:
            candidates.append(getattr(process_instance, "process_id", None))

        activity = getattr(schedule, "activity", None)
        if activity is not None:
            candidates.append(getattr(activity, "process_id", None))

        process_ids: List[int] = []
        for value in candidates:
            try:
                parsed = int(value)
            except (TypeError, ValueError):
                continue
            if parsed > 0 and parsed not in process_ids:
                process_ids.append(parsed)
        return process_ids

    @staticmethod
    def _schedule_matches_processes(schedule: FinancialSchedule, process_ids: Sequence[int]) -> bool:
        selected = {int(item) for item in process_ids if item}
        if not selected:
            return True
        return bool(selected.intersection(FinancialReportService._schedule_process_ids(schedule)))

    @staticmethod
    def _entry_query(company_id: int, filters: FinancialManagementReportFiltersInput):
        query = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
            FinancialEntry.competence_date >= filters.period_start,
            FinancialEntry.competence_date <= filters.period_end,
        )
        bank_account_ids = FinancialReportService._selected_ids(filters.bank_account_id, filters.bank_account_ids)
        if bank_account_ids:
            query = query.filter(FinancialEntry.bank_account_id.in_(bank_account_ids))
        chart_account_ids = FinancialReportService._selected_ids(filters.chart_account_id, filters.chart_account_ids)
        if chart_account_ids:
            query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(filters.cost_center_id, filters.cost_center_ids)
        if cost_center_ids:
            query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
        counterparty_ids = FinancialReportService._selected_ids(filters.counterparty_id, filters.counterparty_ids)
        if counterparty_ids:
            query = query.filter(FinancialEntry.counterparty_id.in_(counterparty_ids))
        if filters.movement_nature:
            query = query.filter(FinancialEntry.movement_nature == filters.movement_nature)
        return query

    @staticmethod
    def _open_entries_until(company_id: int, filters: FinancialManagementReportFiltersInput, until: date) -> List[FinancialEntry]:
        query = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
            FinancialEntry.status.in_(["draft", "pending_review", "scheduled", "posted", "partially_settled"]),
            FinancialEntry.due_date.isnot(None),
            FinancialEntry.due_date <= until,
        )
        bank_account_ids = FinancialReportService._selected_ids(filters.bank_account_id, filters.bank_account_ids)
        if bank_account_ids:
            query = query.filter(FinancialEntry.bank_account_id.in_(bank_account_ids))
        if filters.chart_account_id:
            query = query.filter(FinancialEntry.chart_account_id == filters.chart_account_id)
        if filters.cost_center_id:
            query = query.filter(FinancialEntry.cost_center_id == filters.cost_center_id)
        return query.order_by(FinancialEntry.due_date.asc(), FinancialEntry.id.asc()).all()

    @staticmethod
    def _cash_flow_dimension_filtered_entries(company_id: int, filters: FinancialManagementReportFiltersInput) -> List[FinancialEntry]:
        query = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
        )
        chart_account_ids = FinancialReportService._selected_ids(
            getattr(filters, "chart_account_id", None),
            getattr(filters, "chart_account_ids", None),
        )
        if chart_account_ids:
            query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(
            getattr(filters, "cost_center_id", None),
            getattr(filters, "cost_center_ids", None),
        )
        if cost_center_ids:
            query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
        entries = query.all()
        project_ids = list(getattr(filters, "project_ids", None) or [])
        if project_ids:
            entries = [entry for entry in entries if FinancialReportService._entry_matches_projects(entry, project_ids)]
        process_ids = list(getattr(filters, "process_ids", None) or [])
        if process_ids:
            entries = [entry for entry in entries if FinancialReportService._entry_matches_processes(entry, process_ids)]
        return entries

    @staticmethod
    def _settlement_query(company_id: int, filters: FinancialManagementReportFiltersInput):
        query = FinancialSettlement.query.filter(
            FinancialSettlement.company_id == company_id,
            FinancialSettlement.deleted_at.is_(None),
            FinancialSettlement.settlement_status != "cancelled",
            FinancialSettlement.settlement_date >= filters.period_start,
            FinancialSettlement.settlement_date <= filters.period_end,
        )
        bank_account_ids = FinancialReportService._selected_ids(
            filters.bank_account_id,
            filters.bank_account_ids,
            preserve_empty_marker=filters.report_type == "cash_flow",
        )
        if bank_account_ids:
            query = query.filter(FinancialSettlement.bank_account_id.in_(bank_account_ids))
        if filters.include_reconciled_only:
            query = query.filter(FinancialSettlement.reconciliation_status == "reconciled")
        return query

    @staticmethod
    def _normalize_cash_flow_title_selection_filters(raw_filters: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        payload = dict(raw_filters or {})
        movement_nature = str(payload.get("movement_nature") or "").strip().lower()
        search = str(payload.get("search") or "").strip()
        return {
            "movement_nature": movement_nature if movement_nature in {"credit", "debit"} else None,
            "counterparty_id": FinancialReportService._parse_positive_int(payload.get("counterparty_id")),
            "chart_account_id": FinancialReportService._parse_positive_int(payload.get("chart_account_id")),
            "cost_center_id": FinancialReportService._parse_positive_int(payload.get("cost_center_id")),
            "search": search or None,
        }

    @staticmethod
    def _cash_flow_projected_entry_query(
        company_id: int,
        filters: FinancialManagementReportFiltersInput,
        *,
        entry_ids: Optional[Sequence[int]] = None,
        selection_filters: Optional[Dict[str, Any]] = None,
    ):
        query = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
            FinancialEntry.status.in_(["draft", "pending_review", "scheduled", "posted", "partially_settled"]),
            FinancialEntry.due_date.isnot(None),
            FinancialEntry.due_date >= filters.period_start,
            FinancialEntry.due_date <= filters.period_end,
        )
        bank_account_ids = FinancialReportService._selected_ids(
            filters.bank_account_id,
            filters.bank_account_ids,
            preserve_empty_marker=True,
        )
        if bank_account_ids == [-1]:
            query = query.filter(FinancialEntry.bank_account_id.is_(None))
        elif bank_account_ids:
            query = query.filter(
                or_(
                    FinancialEntry.bank_account_id.in_(bank_account_ids),
                    FinancialEntry.bank_account_id.is_(None),
                )
            )

        chart_account_ids = FinancialReportService._selected_ids(
            getattr(filters, "chart_account_id", None),
            getattr(filters, "chart_account_ids", None),
        )
        if chart_account_ids:
            query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(
            getattr(filters, "cost_center_id", None),
            getattr(filters, "cost_center_ids", None),
        )
        if cost_center_ids:
            query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))

        positive_entry_ids = FinancialReportService._selected_ids(None, entry_ids or [])
        if positive_entry_ids:
            query = query.filter(FinancialEntry.id.in_(positive_entry_ids))

        normalized_selection_filters = FinancialReportService._normalize_cash_flow_title_selection_filters(selection_filters)
        movement_nature = normalized_selection_filters.get("movement_nature")
        if movement_nature:
            query = query.filter(FinancialEntry.movement_nature == movement_nature)
        counterparty_id = normalized_selection_filters.get("counterparty_id")
        if counterparty_id:
            query = query.filter(FinancialEntry.counterparty_id == counterparty_id)
        chart_account_id = normalized_selection_filters.get("chart_account_id")
        if chart_account_id:
            query = query.filter(FinancialEntry.chart_account_id == chart_account_id)
        cost_center_id = normalized_selection_filters.get("cost_center_id")
        if cost_center_id:
            query = query.filter(FinancialEntry.cost_center_id == cost_center_id)

        search = normalized_selection_filters.get("search")
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                or_(
                    FinancialEntry.entry_code.ilike(pattern),
                    FinancialEntry.document_number.ilike(pattern),
                    FinancialEntry.description.ilike(pattern),
                    FinancialEntry.memo.ilike(pattern),
                    FinancialEntry.external_reference.ilike(pattern),
                    FinancialEntry.origin_reference.ilike(pattern),
                )
        )
        return query.order_by(FinancialEntry.due_date.asc(), FinancialEntry.id.asc())

    @staticmethod
    def _cash_flow_projected_schedule_query(
        company_id: int,
        filters: FinancialManagementReportFiltersInput,
        *,
        selection_filters: Optional[Dict[str, Any]] = None,
    ):
        query = FinancialSchedule.query.filter(
            FinancialSchedule.company_id == company_id,
            FinancialSchedule.deleted_at.is_(None),
            FinancialSchedule.next_due_date.isnot(None),
            FinancialSchedule.next_due_date >= filters.period_start,
            FinancialSchedule.next_due_date <= filters.period_end,
        )
        bank_account_ids = FinancialReportService._selected_ids(
            filters.bank_account_id,
            filters.bank_account_ids,
            preserve_empty_marker=True,
        )
        if bank_account_ids == [-1]:
            query = query.filter(FinancialSchedule.bank_account_id.is_(None))
        elif bank_account_ids:
            query = query.filter(
                or_(
                    FinancialSchedule.bank_account_id.in_(bank_account_ids),
                    FinancialSchedule.bank_account_id.is_(None),
                )
            )

        chart_account_ids = FinancialReportService._selected_ids(
            getattr(filters, "chart_account_id", None),
            getattr(filters, "chart_account_ids", None),
        )
        if chart_account_ids:
            query = query.filter(FinancialSchedule.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(
            getattr(filters, "cost_center_id", None),
            getattr(filters, "cost_center_ids", None),
        )
        if cost_center_ids:
            query = query.filter(FinancialSchedule.cost_center_id.in_(cost_center_ids))

        normalized_selection_filters = FinancialReportService._normalize_cash_flow_title_selection_filters(selection_filters)
        movement_nature = normalized_selection_filters.get("movement_nature")
        if movement_nature:
            query = query.filter(FinancialSchedule.movement_nature == movement_nature)
        counterparty_id = normalized_selection_filters.get("counterparty_id")
        if counterparty_id:
            query = query.filter(FinancialSchedule.counterparty_id == counterparty_id)
        chart_account_id = normalized_selection_filters.get("chart_account_id")
        if chart_account_id:
            query = query.filter(FinancialSchedule.chart_account_id == chart_account_id)
        cost_center_id = normalized_selection_filters.get("cost_center_id")
        if cost_center_id:
            query = query.filter(FinancialSchedule.cost_center_id == cost_center_id)

        search = normalized_selection_filters.get("search")
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                or_(
                    FinancialSchedule.schedule_code.ilike(pattern),
                    FinancialSchedule.document_number_prefix.ilike(pattern),
                    FinancialSchedule.name.ilike(pattern),
                    FinancialSchedule.description.ilike(pattern),
                    FinancialSchedule.memo.ilike(pattern),
                )
            )
        return query.order_by(FinancialSchedule.next_due_date.asc(), FinancialSchedule.id.asc())

    @staticmethod
    def _serialize_cash_flow_excluded_title(
        entry: FinancialEntry,
        *,
        settled_amount: Decimal,
        counterparty_names: Dict[int, str],
        open_amount: Optional[Decimal] = None,
    ) -> Dict[str, Any]:
        outstanding = open_amount if open_amount is not None else FinancialReportService._entry_outstanding_amount(entry, settled_amount)
        counterparty_label = (
            counterparty_names.get(getattr(entry, "counterparty_id", None))
            or (dict(getattr(entry, "metadata_json", None) or {}).get("counterparty_name"))
            or "Não informado"
        )
        title_amount = Decimal(str(getattr(entry, "original_amount", None) or 0))
        return {
            "id": entry.id,
            "projection_ref": FinancialReportService._cash_flow_projection_ref("entry", entry.id),
            "entry_code": getattr(entry, "entry_code", None) or str(entry.id),
            "history": FinancialReportService._entry_history_label(entry),
            "type": "Entrada" if getattr(entry, "movement_nature", None) == "credit" else "Saída",
            "title_amount": FinancialReportService._format_currency(title_amount),
            "title_amount_value": FinancialReportService._serialize_money(title_amount),
            "open_amount": FinancialReportService._format_currency(outstanding),
            "open_amount_value": FinancialReportService._serialize_money(outstanding),
            "counterparty": counterparty_label,
            "number_installment": FinancialReportService._entry_number_installment_label(entry),
            "competence_date": entry.competence_date.isoformat() if getattr(entry, "competence_date", None) else "-",
            "due_date": entry.due_date.isoformat() if getattr(entry, "due_date", None) else "-",
            "status_label": "Retirado do fluxo",
        }

    @staticmethod
    def _schedule_installment_label(schedule: FinancialSchedule) -> str:
        metadata = dict(getattr(schedule, "metadata_json", None) or {})
        value = (
            metadata.get("installment_number")
            or metadata.get("parcela")
            or metadata.get("installment_label")
            or metadata.get("parcel_label")
            or metadata.get("repeat_label")
        )
        return str(value).strip() if value not in (None, "") else "-"

    @staticmethod
    def _schedule_history_label(schedule: FinancialSchedule) -> str:
        metadata = dict(getattr(schedule, "metadata_json", None) or {})
        history = (
            metadata.get("history")
            or metadata.get("historico")
            or getattr(schedule, "name", None)
            or getattr(schedule, "description", None)
            or getattr(schedule, "memo", None)
        )
        return str(history or "Sem histórico").strip()

    @staticmethod
    def _cash_flow_schedule_open_amounts(
        schedule: FinancialSchedule,
        filters: FinancialManagementReportFiltersInput,
    ) -> Optional[Tuple[Decimal, Decimal, date]]:
        normalized_status = str(getattr(schedule, "status", None) or "").strip().lower()
        if normalized_status in {"draft", "cancelled", "completed"}:
            return None
        if normalized_status == "forecast" and not bool(filters.include_budget_vs_actual):
            return None
        if schedule.entry_type == "receivable" and not filters.include_receivable:
            return None
        if schedule.entry_type == "payable" and not filters.include_payable:
            return None

        due_date = getattr(schedule, "next_due_date", None) or getattr(schedule, "first_due_date", None) or getattr(schedule, "start_date", None)
        if not due_date:
            return None

        projection_snapshot = FinancialReportService._schedule_projected_balance_snapshot(schedule)
        title_amount = Decimal(str(projection_snapshot.get("principal_amount") or getattr(schedule, "template_amount", None) or 0))
        if title_amount <= Decimal("0"):
            title_amount = Decimal(str(getattr(schedule, "template_amount", None) or 0))
        open_amount = Decimal(
            str(
                projection_snapshot.get(
                    "principal_corrected_open"
                    if filters.projected_values_mode == "with_financial_correction"
                    else "principal_open"
                )
                or 0
            )
        )
        if open_amount <= Decimal("0") and title_amount > Decimal("0"):
            open_amount = title_amount
        if open_amount <= Decimal("0"):
            return None
        return title_amount, open_amount, due_date

    @staticmethod
    def _serialize_cash_flow_projected_schedule(
        schedule: FinancialSchedule,
        *,
        title_amount: Decimal,
        open_amount: Decimal,
        counterparty_label: str,
        due_date: Optional[date] = None,
    ) -> Dict[str, Any]:
        title_due_date = due_date or getattr(schedule, "next_due_date", None) or getattr(schedule, "first_due_date", None) or getattr(schedule, "start_date", None)
        return {
            "id": schedule.id,
            "projection_ref": FinancialReportService._cash_flow_projection_ref("schedule", schedule.id),
            "entry_code": getattr(schedule, "schedule_code", None) or str(schedule.id),
            "history": FinancialReportService._schedule_history_label(schedule),
            "type": "Entrada" if getattr(schedule, "movement_nature", None) == "credit" else "Saída",
            "type_code": FinancialReportService._cash_flow_entry_type_code(schedule),
            "title_amount": FinancialReportService._format_currency(title_amount),
            "title_amount_value": FinancialReportService._serialize_money(title_amount),
            "open_amount": FinancialReportService._format_currency(open_amount),
            "open_amount_value": FinancialReportService._serialize_money(open_amount),
            "counterparty": counterparty_label,
            "number_installment": FinancialReportService._schedule_installment_label(schedule),
            "competence_date": getattr(schedule, "competence_date", None).isoformat() if getattr(schedule, "competence_date", None) else "-",
            "due_date": title_due_date.isoformat() if title_due_date else "-",
            "status_label": "Retirado do fluxo",
        }

    @staticmethod
    def build_cash_flow_title_preview(
        *,
        company_id: int,
        filters: Optional[Dict[str, Any]] = None,
        selection_filters: Optional[Dict[str, Any]] = None,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        normalized_filters, error = FinancialReportService._normalize_filters("cash_flow", filters)
        if error:
            return None, error

        selected_projection_refs = FinancialReportService._cash_flow_selected_projection_refs(normalized_filters)
        entries = FinancialReportService._cash_flow_projected_entry_query(
            company_id,
            normalized_filters,
            selection_filters=selection_filters,
        ).all()
        if normalized_filters.project_ids:
            entries = [
                entry
                for entry in entries
                if FinancialReportService._entry_matches_projects(entry, normalized_filters.project_ids)
            ]
        if normalized_filters.process_ids:
            entries = [
                entry
                for entry in entries
                if FinancialReportService._entry_matches_processes(entry, normalized_filters.process_ids)
            ]
        settlement_totals = FinancialReportService._entry_settlement_totals(
            company_id,
            entry_ids=[entry.id for entry in entries],
        )
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
        schedule_projection_cache: Dict[int, Dict[str, Decimal]] = {}
        schedule_ids = {FinancialReportService._entry_schedule_id(entry) for entry in entries}
        schedule_ids = {item for item in schedule_ids if item}
        schedule_cache: Dict[int, FinancialSchedule] = {}
        if schedule_ids:
            for schedule in FinancialSchedule.query.filter(
                FinancialSchedule.company_id == company_id,
                FinancialSchedule.id.in_(list(schedule_ids)),
                FinancialSchedule.deleted_at.is_(None),
            ).all():
                schedule_cache[int(schedule.id)] = schedule

        titles: List[Dict[str, Any]] = []
        total_open_amount = Decimal("0")
        for entry in entries:
            settled_amount = settlement_totals.get(entry.id, Decimal("0"))
            outstanding = FinancialReportService._entry_projected_open_amount(
                entry,
                settled_amount=settled_amount,
                projected_values_mode=normalized_filters.projected_values_mode,
                schedule_cache=schedule_cache,
                schedule_projection_cache=schedule_projection_cache,
            )
            if outstanding <= Decimal("0"):
                continue
            item = FinancialReportService._serialize_cash_flow_excluded_title(
                entry,
                settled_amount=settled_amount,
                counterparty_names=counterparty_names,
                open_amount=outstanding,
            )
            item["selected"] = item.get("projection_ref") in selected_projection_refs
            titles.append(item)
            total_open_amount += outstanding

        projected_schedules = FinancialReportService._cash_flow_projected_schedule_query(
            company_id,
            normalized_filters,
            selection_filters=selection_filters,
        ).all()
        if normalized_filters.process_ids:
            projected_schedules = [
                item
                for item in projected_schedules
                if FinancialReportService._schedule_matches_processes(item, normalized_filters.process_ids)
            ]
        if normalized_filters.project_ids:
            projected_schedules = [
                item
                for item in projected_schedules
                if FinancialReportService._schedule_matches_projects(item, normalized_filters.project_ids)
            ]

        projected_schedule_ids = [item.id for item in projected_schedules]
        projected_schedule_refs = {f"financial_schedule:{item.id}": item.id for item in projected_schedules}
        linked_entries_by_schedule: Dict[int, List[FinancialEntry]] = {item.id: [] for item in projected_schedules}
        if projected_schedule_ids:
            linked_entries = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
                or_(
                    FinancialEntry.financial_schedule_id.in_(projected_schedule_ids),
                    FinancialEntry.external_reference.in_(list(projected_schedule_refs.keys())),
                ),
            ).all()
            for linked_entry in linked_entries:
                schedule_id = getattr(linked_entry, "financial_schedule_id", None) or projected_schedule_refs.get(linked_entry.external_reference)
                if schedule_id in linked_entries_by_schedule:
                    linked_entries_by_schedule[schedule_id].append(linked_entry)

        for schedule in projected_schedules:
            if linked_entries_by_schedule.get(schedule.id):
                continue

            open_payload = FinancialReportService._cash_flow_schedule_open_amounts(schedule, normalized_filters)
            if open_payload is None:
                continue
            title_amount, open_amount, due_date = open_payload
            metadata = dict(getattr(schedule, "metadata_json", None) or {})
            counterparty_label = (
                counterparty_names.get(getattr(schedule, "counterparty_id", None))
                or metadata.get("counterparty_name")
                or "Não informado"
            )
            item = FinancialReportService._serialize_cash_flow_projected_schedule(
                schedule,
                title_amount=title_amount,
                open_amount=open_amount,
                counterparty_label=counterparty_label,
                due_date=due_date,
            )
            item["selected"] = item.get("projection_ref") in selected_projection_refs
            titles.append(item)
            total_open_amount += open_amount

        titles.sort(
            key=lambda item: (
                str(item.get("due_date") or ""),
                str(item.get("history") or "").lower(),
                str(item.get("projection_ref") or item.get("id") or ""),
            )
        )

        return {
            "titles": titles,
            "summary": {
                "count": len(titles),
                "selected_count": len(selected_projection_refs),
                "total_open_amount": FinancialReportService._serialize_money(total_open_amount),
                "total_open_amount_label": FinancialReportService._format_currency(total_open_amount),
                "period_label": f"{normalized_filters.period_start.isoformat()} até {normalized_filters.period_end.isoformat()}",
                "projected_values_mode": normalized_filters.projected_values_mode,
                "projected_values_mode_label": FinancialReportService._projected_values_mode_label(normalized_filters.projected_values_mode),
            },
        }, None

    @staticmethod
    def _build_schedule_report(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
        query = FinancialSchedule.query.filter(
            FinancialSchedule.company_id == company_id,
            FinancialSchedule.deleted_at.is_(None),
        )
        if filters.competence_start and filters.competence_end:
            query = query.filter(
                FinancialSchedule.competence_date >= filters.competence_start,
                FinancialSchedule.competence_date <= filters.competence_end,
            )
        if filters.due_start and filters.due_end:
            query = query.filter(
                FinancialSchedule.next_due_date >= filters.due_start,
                FinancialSchedule.next_due_date <= filters.due_end,
            )
        if filters.bank_account_id:
            query = query.filter(FinancialSchedule.bank_account_id == filters.bank_account_id)
        counterparty_ids = FinancialReportService._selected_ids(filters.counterparty_id, filters.counterparty_ids)
        if counterparty_ids:
            query = query.filter(FinancialSchedule.counterparty_id.in_(counterparty_ids))
        chart_account_ids = FinancialReportService._selected_ids(filters.chart_account_id, filters.chart_account_ids)
        if chart_account_ids:
            query = query.filter(FinancialSchedule.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(filters.cost_center_id, filters.cost_center_ids)
        if cost_center_ids:
            query = query.filter(FinancialSchedule.cost_center_id.in_(cost_center_ids))
        allowed_entry_types = []
        if filters.include_payable:
            allowed_entry_types.append("payable")
        if filters.include_receivable:
            allowed_entry_types.append("receivable")
        if allowed_entry_types:
            query = query.filter(FinancialSchedule.entry_type.in_(allowed_entry_types))

        schedules = query.order_by(FinancialSchedule.next_due_date.asc(), FinancialSchedule.id.asc()).all()
        if filters.process_ids:
            schedules = [item for item in schedules if FinancialReportService._schedule_matches_processes(item, filters.process_ids)]
        if filters.project_ids:
            schedules = [item for item in schedules if FinancialReportService._schedule_matches_projects(item, filters.project_ids)]

        schedule_map = {item.id: item for item in schedules}
        schedule_ids = list(schedule_map.keys())
        entry_refs = {f"financial_schedule:{item.id}": item.id for item in schedules}
        entries = []
        if schedule_ids:
            entries = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
                or_(
                    FinancialEntry.financial_schedule_id.in_(schedule_ids),
                    FinancialEntry.external_reference.in_(list(entry_refs.keys())),
                ),
            ).all()

        entries_by_schedule: Dict[int, List[FinancialEntry]] = {item.id: [] for item in schedules}
        for entry in entries:
            schedule_id = getattr(entry, "financial_schedule_id", None) or entry_refs.get(entry.external_reference)
            if schedule_id and schedule_id in entries_by_schedule:
                entries_by_schedule[schedule_id].append(entry)

        settlements_by_entry: Dict[int, List[FinancialSettlement]] = {}
        if entries:
            entry_ids = [item.id for item in entries]
            settlements = FinancialSettlement.query.filter(
                FinancialSettlement.company_id == company_id,
                FinancialSettlement.financial_entry_id.in_(entry_ids),
                FinancialSettlement.deleted_at.is_(None),
                FinancialSettlement.settlement_status != "cancelled",
            ).order_by(
                FinancialSettlement.financial_entry_id.asc(),
                FinancialSettlement.settlement_date.asc(),
                FinancialSettlement.id.asc(),
            ).all()
            for settlement in settlements:
                settlements_by_entry.setdefault(settlement.financial_entry_id, []).append(settlement)

        active_borderos_by_schedule: Dict[int, FinancialBordero] = {}
        if schedules:
            bordero_rows = db.session.query(FinancialBorderoItem, FinancialBordero).join(
                FinancialBordero,
                FinancialBorderoItem.bordero_id == FinancialBordero.id,
            ).filter(
                FinancialBorderoItem.company_id == company_id,
                FinancialBorderoItem.financial_schedule_id.in_(list(schedule_map.keys())),
                FinancialBorderoItem.deleted_at.is_(None),
                FinancialBordero.company_id == company_id,
                FinancialBordero.deleted_at.is_(None),
                FinancialBordero.status.in_(("draft", "open", "partially_settled")),
            ).order_by(FinancialBordero.id.desc()).all()
            for bordero_item, bordero in bordero_rows:
                active_borderos_by_schedule.setdefault(bordero_item.financial_schedule_id, bordero)

        rows: List[Dict[str, Any]] = []
        receivable_title_total = Decimal("0")
        receivable_correction_total = Decimal("0")
        receivable_open_total = Decimal("0")
        payable_title_total = Decimal("0")
        payable_correction_total = Decimal("0")
        payable_open_total = Decimal("0")
        open_count = 0
        bordero_count = 0

        for schedule in schedules:
            metadata = dict(schedule.metadata_json or {})
            schedule_entries = entries_by_schedule.get(schedule.id, [])
            projection_snapshot = FinancialReportService._schedule_projected_balance_snapshot(schedule)
            original_total = Decimal("0")
            settled_total = Decimal("0")
            open_total = Decimal("0")
            settled_entries = 0
            partial_entries = 0
            open_entries = 0
            latest_settlement_date = None

            if not schedule_entries:
                template_amount = Decimal(str(schedule.template_amount or 0))
                original_total = template_amount
                if schedule.status not in {"completed", "cancelled"} and template_amount > 0:
                    open_total = template_amount
                    open_entries = 1

            for entry in schedule_entries:
                original_amount = Decimal(str(entry.original_amount or 0))
                entry_settlements = settlements_by_entry.get(entry.id, [])
                settlements_total = sum(Decimal(str(item.principal_amount or 0)) for item in entry_settlements)
                outstanding = max(original_amount - settlements_total, Decimal("0"))
                original_total += original_amount
                settled_total += settlements_total
                open_total += outstanding
                if entry_settlements:
                    latest_entry_settlement = entry_settlements[-1].settlement_date
                    if latest_entry_settlement and (latest_settlement_date is None or latest_entry_settlement > latest_settlement_date):
                        latest_settlement_date = latest_entry_settlement
                if outstanding == 0 and original_amount > 0:
                    settled_entries += 1
                elif settlements_total > 0:
                    partial_entries += 1
                else:
                    open_entries += 1

            settlement_state = "open"
            if schedule_entries:
                if open_entries == 0 and partial_entries == 0:
                    settlement_state = "settled"
                elif partial_entries > 0 or settled_entries > 0:
                    settlement_state = "partial"

            active_bordero = active_borderos_by_schedule.get(schedule.id)
            operational_state = build_title_operational_state_metadata(
                schedule_status=schedule.status,
                settlement_state=settlement_state,
                entry_type=schedule.entry_type,
                metadata_json=schedule.metadata_json,
            )
            report_state = "bordero" if active_bordero else operational_state["code"]
            if report_state == "settled" and not filters.include_settled:
                continue
            if report_state == "partial" and not filters.include_partial:
                continue
            if report_state == "open" and not filters.include_open:
                continue
            if report_state in {"draft", "cancelled"}:
                continue
            if report_state == "forecast" and not bool(filters.include_budget_vs_actual):
                continue
            if report_state == "bordero" and not filters.include_bordero:
                continue

            if filters.settlement_start and filters.settlement_end:
                if not latest_settlement_date:
                    continue
                if latest_settlement_date < filters.settlement_start or latest_settlement_date > filters.settlement_end:
                    continue

            correction_amount = Decimal(str(projection_snapshot.get("financial_correction_open") or 0))
            corrected_open_total = Decimal(str(projection_snapshot.get("principal_corrected_open") or open_total or 0))
            signed_title_amount = Decimal(str(FinancialService.get_signed_amount(original_total, schedule.movement_nature)))
            signed_correction_amount = Decimal(str(FinancialService.get_signed_amount(correction_amount, schedule.movement_nature)))
            signed_open_amount = Decimal(str(FinancialService.get_signed_amount(corrected_open_total, schedule.movement_nature)))
            if schedule.entry_type == "receivable":
                receivable_title_total += original_total
                receivable_correction_total += correction_amount
                receivable_open_total += signed_open_amount
            else:
                payable_title_total += original_total
                payable_correction_total += correction_amount
                payable_open_total += signed_open_amount
            if report_state not in {"settled", "cancelled", "draft"}:
                open_count += 1
            if report_state == "bordero":
                bordero_count += 1

            installment_value = (
                metadata.get("installment_number")
                or metadata.get("parcela")
                or metadata.get("installment_label")
                or metadata.get("parcel_label")
                or metadata.get("repeat_label")
                or "-"
            )
            title_number_value = schedule.document_number_prefix or schedule.schedule_code or str(schedule.id)
            history_value = schedule.name or schedule.description or "-"
            counterparty_value = counterparty_names.get(schedule.counterparty_id) or metadata.get("counterparty_name") or "Não informado"
            competence_date_value = schedule.competence_date.isoformat() if getattr(schedule, "competence_date", None) else "-"
            due_date_value = (schedule.next_due_date or schedule.first_due_date).isoformat() if (schedule.next_due_date or schedule.first_due_date) else "-"
            settlement_date_value = latest_settlement_date.isoformat() if latest_settlement_date else "-"
            status_label = {
                "draft": "Rascunho",
                "forecast": "Projetado",
                "settled": "Baixado",
                "partial": "Baixado Parcial",
                "open": "Aberto",
                "cancelled": "Cancelado",
                "bordero": "Borderô",
            }[report_state]
            type_label = "Recebimento" if schedule.entry_type == "receivable" else "Pagamento"

            rows.append(
                {
                    "title_number": title_number_value,
                    "installment": str(installment_value),
                    "history": history_value,
                    "counterparty": counterparty_value,
                    "title_amount": FinancialReportService._format_currency(signed_title_amount),
                    "correction_amount": FinancialReportService._format_currency(signed_correction_amount),
                    "balance_amount": FinancialReportService._format_currency(signed_open_amount),
                    "competence_date": competence_date_value,
                    "due_date": due_date_value,
                    "settlement_date": settlement_date_value,
                    "status": status_label,
                    "type": type_label,
                    "_title_number_sort": str(title_number_value or "").lower(),
                    "_installment_sort": str(installment_value or "").lower(),
                    "_history_sort": str(history_value or "").lower(),
                    "_counterparty_sort": str(counterparty_value or "").lower(),
                    "_title_amount_sort": float(signed_title_amount),
                    "_correction_amount_sort": float(signed_correction_amount),
                    "_balance_amount_sort": float(signed_open_amount),
                    "_competence_date_sort": competence_date_value if competence_date_value != "-" else "",
                    "_due_date_sort": due_date_value if due_date_value != "-" else "",
                    "_settlement_date_sort": settlement_date_value if settlement_date_value != "-" else "",
                }
            )

        sort_key_map = {
            "title_number": "_title_number_sort",
            "installment": "_installment_sort",
            "history": "_history_sort",
            "counterparty": "_counterparty_sort",
            "title_amount": "_title_amount_sort",
            "correction_amount": "_correction_amount_sort",
            "balance_amount": "_balance_amount_sort",
            "competence_date": "_competence_date_sort",
            "due_date": "_due_date_sort",
            "settlement_date": "_settlement_date_sort",
        }
        reverse = filters.order_direction == "desc"
        sort_key = sort_key_map.get(filters.order_by, "_title_number_sort")
        rows.sort(key=lambda item: item.get(sort_key) or "", reverse=reverse)

        columns = []
        for key, label, enabled in [
            ("title_number", "Nº Título", filters.show_title_number),
            ("counterparty", "Favorecido", filters.show_counterparty),
            ("title_amount", "Valor Original", filters.show_title_amount),
            ("correction_amount", "Valor da Correção", filters.show_correction_amount),
            ("balance_amount", "Saldo do Principal Corrigido", filters.show_balance_amount),
            ("competence_date", "Competência", filters.show_competence_date),
            ("due_date", "Vencimento", filters.show_due_date),
            ("settlement_date", "Dt. da Última Liquid.", filters.show_settlement_date),
        ]:
            if enabled:
                columns.append({"key": key, "label": label})

        total_general = receivable_title_total - payable_title_total
        total_general_net = receivable_open_total + payable_open_total

        return FinancialReportService._report_payload(
            definition,
            summary_cards=[
                FinancialReportService._report_card("Quantidade de registros", len(rows)),
                FinancialReportService._report_card("Total a receber", FinancialReportService._format_currency(receivable_title_total), "positive"),
                FinancialReportService._report_card("Correções a receber", FinancialReportService._format_currency(receivable_correction_total), "positive"),
                FinancialReportService._report_card("Total líquido a receber", FinancialReportService._format_currency(receivable_open_total), "positive"),
                FinancialReportService._report_card("Total a pagar", FinancialReportService._format_currency(payable_title_total), "negative"),
                FinancialReportService._report_card("Correções a pagar", FinancialReportService._format_currency(payable_correction_total), "negative"),
                FinancialReportService._report_card("Total líquido a pagar", FinancialReportService._format_currency(abs(payable_open_total)), "negative"),
                FinancialReportService._report_card("Total geral", FinancialReportService._format_currency(total_general), "primary"),
                FinancialReportService._report_card("Total geral líquido", FinancialReportService._format_currency(total_general_net), "primary"),
            ],
            general_info=[
                FinancialReportService._report_info("Competência base", f"{filters.competence_start.isoformat()} até {filters.competence_end.isoformat()}"),
                FinancialReportService._report_info("Critério principal", "Título financeiro operacional por saldo e baixa"),
                FinancialReportService._report_info("Em aberto / borderô", f"{open_count} / {bordero_count}"),
            ],
            columns=columns,
            rows=rows,
            totals={
                "count": len(rows),
                "receivable_title_total": FinancialReportService._serialize_money(receivable_title_total),
                "receivable_correction_total": FinancialReportService._serialize_money(receivable_correction_total),
                "receivable_open_total": FinancialReportService._serialize_money(receivable_open_total),
                "payable_title_total": FinancialReportService._serialize_money(payable_title_total),
                "payable_correction_total": FinancialReportService._serialize_money(payable_correction_total),
                "payable_open_total": FinancialReportService._serialize_money(abs(payable_open_total)),
                "total_general": FinancialReportService._serialize_money(total_general),
                "total_general_net": FinancialReportService._serialize_money(total_general_net),
            },
        )

    @staticmethod
    def _iso_date_or_none(value: Any) -> Optional[str]:
        return value.isoformat() if hasattr(value, "isoformat") else None

    @staticmethod
    def _bank_statement_dimension_labels(
        *,
        entry: Any,
        allocation_breakdown: Dict[str, Any],
        allocations_by_entry: Dict[int, List[Any]],
        dimension_key: str,
        names_by_id: Dict[int, str],
    ) -> List[str]:
        identifiers: List[int] = []

        def _add(raw_value: Any) -> None:
            try:
                parsed = int(raw_value)
            except (TypeError, ValueError):
                return
            if parsed > 0 and parsed not in identifiers:
                identifiers.append(parsed)

        entry_id = getattr(entry, "id", None)
        if entry_id is not None:
            for allocation in allocations_by_entry.get(int(entry_id), []):
                _add(getattr(allocation, dimension_key, None))

        for component_payload in dict(allocation_breakdown or {}).values():
            component = dict(component_payload or {})
            for item in component.get("items") or []:
                _add(dict(item or {}).get(dimension_key))

        _add(getattr(entry, dimension_key, None))
        return [names_by_id.get(item_id, str(item_id)) for item_id in identifiers]

    @staticmethod
    def _metadata_attachments(entity: Any) -> List[Dict[str, Any]]:
        metadata = dict(getattr(entity, "metadata_json", {}) or {})
        attachments = metadata.get("attachments") or []
        if not isinstance(attachments, list):
            return []
        return [dict(item or {}) for item in attachments if isinstance(item, dict)]

    @staticmethod
    def _bank_statement_dossier_attachment_payloads(
        *,
        row_context: Dict[str, Any],
        source_type: str,
        source_label: str,
        attachments: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        documents: List[Dict[str, Any]] = []
        for attachment in attachments or []:
            name = str(attachment.get("name") or attachment.get("file_name") or "Anexo").strip() or "Anexo"
            documents.append(
                {
                    **row_context,
                    "source_type": source_type,
                    "source_label": source_label,
                    "document_name": name,
                    "document_url": attachment.get("url"),
                    "content_type": attachment.get("content_type") or attachment.get("mime_type"),
                    "attachment": dict(attachment),
                }
            )
        return documents

    @staticmethod
    def _bank_statement_columns(filters: FinancialManagementReportFiltersInput) -> List[Dict[str, Any]]:
        configured_columns = [
            (bool(getattr(filters, "show_settlement_date", True)), "data", "Data"),
            (bool(getattr(filters, "show_code", True)), "codigo", "Liquidação"),
            (True, "conta_bancaria", "Conta bancária"),
            (bool(getattr(filters, "show_title_number", True)), "lancamento", "Lançamento"),
            (bool(getattr(filters, "show_description", True)), "descricao", "Descrição"),
            (bool(getattr(filters, "show_counterparty", True)), "favorecido", "Favorecido"),
            (bool(getattr(filters, "show_competence_date", False)), "competencia", "Competência"),
            (bool(getattr(filters, "show_due_date", False)), "vencimento", "Vencimento"),
            (bool(getattr(filters, "show_title_amount", True)), "valor", "Valor"),
            (True, "conciliacao", "Conciliação"),
            (bool(getattr(filters, "show_balance_amount", True)), "saldo", "Saldo"),
        ]
        return [{"key": key, "label": label} for enabled, key, label in configured_columns if enabled]

    @staticmethod
    def _bank_statement_entry_status_bucket(entry: Any) -> str:
        status = str(getattr(entry, "status", "") or "").lower()
        if status in {"partially_settled", "partial", "settled_partial"}:
            return "partial"
        return "settled"

    @staticmethod
    def _bank_statement_accepts_entry(entry: Any, filters: FinancialManagementReportFiltersInput) -> bool:
        movement_nature = str(getattr(entry, "movement_nature", "") or "").lower()
        if movement_nature == "credit" and not bool(getattr(filters, "include_receivable", True)):
            return False
        if movement_nature == "debit" and not bool(getattr(filters, "include_payable", True)):
            return False
        status_bucket = FinancialReportService._bank_statement_entry_status_bucket(entry)
        if status_bucket == "partial":
            return bool(getattr(filters, "include_partial", True))
        return bool(getattr(filters, "include_settled", True))

    @staticmethod
    def _bank_statement_sort_key(
        settlement: Any,
        entry: Any,
        *,
        counterparty_names: Dict[int, str],
        filters: FinancialManagementReportFiltersInput,
    ):
        order_by = getattr(filters, "order_by", "settlement_date") or "settlement_date"
        counterparty_label = counterparty_names.get(getattr(entry, "counterparty_id", None), "")
        key_map = {
            "settlement_date": getattr(settlement, "settlement_date", None),
            "due_date": getattr(entry, "due_date", None),
            "competence_date": getattr(entry, "competence_date", None),
            "code": getattr(settlement, "settlement_code", "") or "",
            "title_number": getattr(entry, "entry_code", "") or "",
            "description": getattr(entry, "description", "") or "",
            "history": getattr(entry, "description", "") or "",
            "counterparty": counterparty_label,
            "movement_nature": getattr(entry, "movement_nature", "") or "",
            "title_amount": Decimal(str(getattr(settlement, "net_amount", None) or 0)),
            "correction_amount": Decimal("0"),
            "balance_amount": Decimal(str(getattr(settlement, "net_amount", None) or 0)),
            "installment": getattr(entry, "entry_code", "") or "",
            "project": "",
        }
        value = key_map.get(order_by, key_map["settlement_date"])
        return (value is None, value, getattr(settlement, "id", 0) or 0)

    @staticmethod
    def _bank_statement_opening_balance(
        company_id: int,
        filters: FinancialManagementReportFiltersInput,
    ) -> Decimal:
        history_query = FinancialSettlement.query.filter(
            FinancialSettlement.company_id == company_id,
            FinancialSettlement.deleted_at.is_(None),
            FinancialSettlement.settlement_status != "cancelled",
            FinancialSettlement.settlement_date < filters.period_start,
        )
        bank_account_ids = FinancialReportService._selected_ids(
            getattr(filters, "bank_account_id", None),
            getattr(filters, "bank_account_ids", None),
        )
        if bank_account_ids:
            history_query = history_query.filter(FinancialSettlement.bank_account_id.in_(bank_account_ids))
        if filters.include_reconciled_only:
            history_query = history_query.filter(FinancialSettlement.reconciliation_status == "reconciled")

        history_settlements = history_query.all()
        entry_ids = [item.financial_entry_id for item in history_settlements if item.financial_entry_id]
        entries_by_id: Dict[int, Any] = {}
        if entry_ids:
            entries_by_id = {
                entry.id: entry
                for entry in FinancialEntry.query.filter(
                    FinancialEntry.company_id == company_id,
                    FinancialEntry.id.in_(entry_ids),
                    FinancialEntry.deleted_at.is_(None),
                ).all()
            }

        balance = Decimal("0")
        for settlement in history_settlements:
            entry = entries_by_id.get(settlement.financial_entry_id)
            if not entry or not FinancialReportService._bank_statement_accepts_entry(entry, filters):
                continue
            amount = Decimal(settlement.net_amount or 0)
            balance += amount if entry.movement_nature == "credit" else -amount
        return balance

    @staticmethod
    def _build_bank_statement(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        bank_names = FinancialReportService._name_map(FinancialBankAccount, company_id)
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
        chart_names = FinancialReportService._name_map(FinancialChartAccount, company_id)
        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        settlements = FinancialReportService._settlement_query(company_id, filters).order_by(FinancialSettlement.settlement_date.asc(), FinancialSettlement.id.asc()).all()
        entry_ids = [item.financial_entry_id for item in settlements]
        entries = {}
        if entry_ids:
            for entry in FinancialEntry.query.filter(FinancialEntry.company_id == company_id, FinancialEntry.id.in_(entry_ids), FinancialEntry.deleted_at.is_(None)).all():
                entries[entry.id] = entry
        settlements = [
            settlement
            for settlement in settlements
            if entries.get(settlement.financial_entry_id)
            and FinancialReportService._bank_statement_accepts_entry(entries[settlement.financial_entry_id], filters)
        ]
        reverse_order = (getattr(filters, "order_direction", "asc") or "asc") == "desc"
        settlements.sort(
            key=lambda item: FinancialReportService._bank_statement_sort_key(
                item,
                entries[item.financial_entry_id],
                counterparty_names=counterparty_names,
                filters=filters,
            ),
            reverse=reverse_order,
        )
        entry_ids = [item.financial_entry_id for item in settlements]
        is_dossier = filters.report_type == "bank_statement_dossier"
        schedules_by_id: Dict[int, Any] = {}
        allocations_by_entry: Dict[int, List[Any]] = defaultdict(list)
        if is_dossier and entries:
            schedule_ids = {
                int(schedule_id)
                for schedule_id in (getattr(entry, "financial_schedule_id", None) for entry in entries.values())
                if schedule_id
            }
            if schedule_ids:
                for schedule in FinancialSchedule.query.filter(
                    FinancialSchedule.company_id == company_id,
                    FinancialSchedule.id.in_(schedule_ids),
                    FinancialSchedule.deleted_at.is_(None),
                ).all():
                    schedules_by_id[int(schedule.id)] = schedule
            try:
                for allocation in FinancialEntryAllocation.query.filter(
                    FinancialEntryAllocation.company_id == company_id,
                    FinancialEntryAllocation.financial_entry_id.in_(entry_ids),
                    FinancialEntryAllocation.deleted_at.is_(None),
                ).all():
                    allocations_by_entry[int(allocation.financial_entry_id)].append(allocation)
            except Exception:
                allocations_by_entry = defaultdict(list)

        balance_base = FinancialReportService._bank_statement_opening_balance(company_id, filters)

        running = Decimal(balance_base)
        inflow = Decimal("0")
        outflow = Decimal("0")
        rows: List[Dict[str, Any]] = []
        dossier_documents: List[Dict[str, Any]] = []
        for settlement in settlements:
            entry = entries.get(settlement.financial_entry_id)
            if not entry:
                continue
            settlement_payload = FinancialService.serialize_settlement(settlement, entry=entry, include_components=True)
            component_summary = dict(settlement_payload.get("settlement_component_summary") or {})
            allocation_breakdown = dict(settlement_payload.get("settlement_allocation_breakdown") or {})
            chart_labels = FinancialReportService._bank_statement_dimension_labels(
                entry=entry,
                allocation_breakdown=allocation_breakdown,
                allocations_by_entry=allocations_by_entry,
                dimension_key="chart_account_id",
                names_by_id=chart_names,
            )
            center_labels = FinancialReportService._bank_statement_dimension_labels(
                entry=entry,
                allocation_breakdown=allocation_breakdown,
                allocations_by_entry=allocations_by_entry,
                dimension_key="cost_center_id",
                names_by_id=center_names,
            )
            schedule = schedules_by_id.get(int(getattr(entry, "financial_schedule_id", 0) or 0))
            counterparty_label = counterparty_names.get(getattr(entry, "counterparty_id", None), "Não informado")
            amount = Decimal(settlement.net_amount or 0)
            movement_tone = "positive" if entry.movement_nature == "credit" else "negative"
            signed_amount = amount if entry.movement_nature == "credit" else -amount
            if entry.movement_nature == "credit":
                inflow += amount
                running += amount
            else:
                outflow += amount
                running -= amount
            row_payload = {
                    "settlement_id": getattr(settlement, "id", None),
                    "financial_entry_id": getattr(entry, "id", None),
                    "financial_schedule_id": getattr(entry, "financial_schedule_id", None),
                    "data": settlement.settlement_date.isoformat(),
                    "codigo": settlement.settlement_code,
                    "conta_bancaria": bank_names.get(settlement.bank_account_id, "Não informada"),
                    "lancamento": entry.entry_code,
                    "descricao": entry.description,
                    "favorecido": counterparty_label,
                    "competencia": FinancialReportService._iso_date_or_none(getattr(entry, "competence_date", None)),
                    "vencimento": FinancialReportService._iso_date_or_none(getattr(entry, "due_date", None)),
                    "liquidacao": FinancialReportService._iso_date_or_none(getattr(settlement, "settlement_date", None)),
                    "plano_contas": ", ".join(chart_labels) if chart_labels else "Não informado",
                    "centro_resultados": ", ".join(center_labels) if center_labels else "Não informado",
                    "titulo": getattr(schedule, "schedule_code", None) or getattr(entry, "financial_schedule_id", None) or "Não informado",
                    "titulo_nome": getattr(schedule, "name", None),
                    "movimento": "Entrada" if entry.movement_nature == "credit" else "Saída",
                    "movimento_tone": movement_tone,
                    "valor": FinancialReportService._serialize_money(signed_amount),
                    "valor_label": FinancialReportService._format_decimal_br(signed_amount),
                    "valor_tone": "negative" if signed_amount < 0 else ("positive" if signed_amount > 0 else "neutral"),
                    "valor_principal": FinancialReportService._serialize_money(component_summary.get("principal") or settlement.principal_amount or 0),
                    "valor_correcao": FinancialReportService._serialize_money(component_summary.get("financial_correction") or 0),
                    "valor_desconto": FinancialReportService._serialize_money(component_summary.get("discount") or 0),
                    "rateio_principal_itens": len(dict(allocation_breakdown.get("principal") or {}).get("items") or []),
                    "rateio_correcao_itens": len(dict(allocation_breakdown.get("financial_correction") or {}).get("items") or []),
                    "rateio_desconto_itens": len(dict(allocation_breakdown.get("discount") or {}).get("items") or []),
                    "conciliacao": settlement.reconciliation_status,
                    "conciliacao_label": FinancialReportService._reconciliation_status_label(settlement.reconciliation_status),
                    "conciliacao_tone": FinancialReportService._reconciliation_status_tone(settlement.reconciliation_status),
                    "saldo": FinancialReportService._serialize_money(running),
                    "saldo_label": FinancialReportService._format_decimal_br(running),
                    "saldo_tone": "negative" if running < 0 else ("positive" if running > 0 else "neutral"),
                }
            rows.append(row_payload)
            if is_dossier:
                row_context = {
                    "settlement_id": row_payload["settlement_id"],
                    "financial_entry_id": row_payload["financial_entry_id"],
                    "financial_schedule_id": row_payload["financial_schedule_id"],
                    "settlement_code": row_payload["codigo"],
                    "entry_code": row_payload["lancamento"],
                    "title_code": row_payload["titulo"],
                    "description": row_payload["descricao"],
                    "competence_date": row_payload["competencia"],
                    "due_date": row_payload["vencimento"],
                    "settlement_date": row_payload["liquidacao"],
                    "chart_account": row_payload["plano_contas"],
                    "cost_center": row_payload["centro_resultados"],
                    "counterparty": row_payload["favorecido"],
                    "bank_account": row_payload["conta_bancaria"],
                    "movement": row_payload["movimento"],
                    "amount": row_payload["valor_label"],
                }
                dossier_documents.extend(
                    FinancialReportService._bank_statement_dossier_attachment_payloads(
                        row_context=row_context,
                        source_type="title",
                        source_label="Título financeiro",
                        attachments=FinancialReportService._metadata_attachments(schedule) if schedule else [],
                    )
                )
                dossier_documents.extend(
                    FinancialReportService._bank_statement_dossier_attachment_payloads(
                        row_context=row_context,
                        source_type="entry",
                        source_label="Lançamento",
                        attachments=FinancialReportService._metadata_attachments(entry),
                    )
                )
                dossier_documents.extend(
                    FinancialReportService._bank_statement_dossier_attachment_payloads(
                        row_context=row_context,
                        source_type="settlement",
                        source_label="Baixa",
                        attachments=FinancialReportService._metadata_attachments(settlement),
                    )
                )
        totals = {
            "opening_balance": FinancialReportService._serialize_money(balance_base),
            "inflow": FinancialReportService._serialize_money(inflow),
            "outflow": FinancialReportService._serialize_money(outflow),
            "closing_balance": FinancialReportService._serialize_money(running),
        }
        extra: Dict[str, Any] = {}
        if is_dossier:
            income_statement_payload: Dict[str, Any]
            income_updates = {
                "report_type": "income_statement",
                "orientation": "portrait",
                "show_competence_column": False,
                "show_due_column": False,
                "show_liquidation_column": True,
                "include_open": False,
                "include_settled": True,
                "period_start": filters.period_start,
                "period_end": filters.period_end,
            }
            try:
                if hasattr(filters, "model_copy"):
                    income_statement_filters = filters.model_copy(update=income_updates)
                else:
                    income_statement_filters = FinancialManagementReportFiltersInput(**income_updates)
                income_statement_payload = FinancialReportService._build_income_statement(company_id, income_statement_filters)
            except Exception:
                income_statement_payload = {
                    "title": "Demonstração de Resultados 01",
                    "filters": [
                        FinancialReportService._report_info(
                            "Período",
                            f"{filters.period_start.isoformat()} até {filters.period_end.isoformat()}",
                        )
                    ],
                    "hierarchy_rows": [],
                }
            totals["dossier_document_count"] = len(dossier_documents)
            extra.update(
                {
                    "statement_title": "Extrato Bancário",
                    "statement_subtitle": FinancialReportService.REPORT_DEFINITIONS["bank_statement"]["description"],
                    "dossier_income_statement": income_statement_payload,
                    "dossier_documents": dossier_documents,
                    "dossier_document_count": len(dossier_documents),
                    "dossier_modes": [
                        {"key": "complete", "label": "Extrato, DRE e comprovantes"},
                        {"key": "simple", "label": "Extrato, DRE e comprovantes"},
                    ],
                }
            )
        return FinancialReportService._report_payload(
            definition,
            summary_cards=[
                FinancialReportService._report_card("Saldo inicial", FinancialReportService._format_currency(balance_base)),
                FinancialReportService._report_card("Entradas", FinancialReportService._format_currency(inflow), "positive"),
                FinancialReportService._report_card("Saídas", FinancialReportService._format_currency(outflow), "negative"),
                FinancialReportService._report_card("Saldo final", FinancialReportService._format_currency(running), "primary" if running >= 0 else "negative"),
            ],
            general_info=[
                FinancialReportService._report_info("Janela analisada", f"{filters.period_start.isoformat()} até {filters.period_end.isoformat()}"),
                FinancialReportService._report_info("Recorte", bank_names.get(getattr(filters, "bank_account_id", None), "Todas as contas bancárias")),
                FinancialReportService._report_info("Movimentos", str(len(rows))),
                FinancialReportService._report_info("Somente conciliados", "Sim" if filters.include_reconciled_only else "Não"),
                FinancialReportService._report_info("Projetar abertos", "Sim" if getattr(filters, "include_projected", False) else "Não"),
                FinancialReportService._report_info("Considerar limites", "Sim" if getattr(filters, "include_overdraft", True) else "Não"),
                FinancialReportService._report_info("Ordenação", f"{getattr(filters, 'order_by', 'settlement_date')} / {getattr(filters, 'order_direction', 'asc')}"),
            ],
            columns=FinancialReportService._bank_statement_columns(filters),
            rows=rows,
            totals=totals,
            extra=extra,
        )

    @staticmethod
    def _build_income_statement(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        return FinancialReportService._build_income_statement_base(company_id, filters, consolidated_by_period=True)

    @staticmethod
    def _build_income_statement_2(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        return FinancialReportService._build_income_statement_base(company_id, filters, consolidated_by_period=False)

    @staticmethod
    def _build_income_statement_base(
        company_id: int,
        filters: FinancialManagementReportFiltersInput,
        *,
        consolidated_by_period: bool,
    ) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        chart_accounts = FinancialChartAccount.query.filter(
            FinancialChartAccount.company_id == company_id,
            FinancialChartAccount.deleted_at.is_(None),
        ).order_by(FinancialChartAccount.code.asc(), FinancialChartAccount.name.asc()).all()
        chart_map = {
            item.id: {
                "id": item.id,
                "parent_id": item.parent_id,
                "code": getattr(item, "code", None) or f"CTA-{item.id}",
                "name": item.name,
                "accepts_posting": bool(getattr(item, "accepts_posting", False)),
            }
            for item in chart_accounts
        }
        chart_children_map: Dict[Optional[int], List[int]] = {}
        for item in chart_accounts:
            chart_children_map.setdefault(item.parent_id, []).append(item.id)
        for child_ids in chart_children_map.values():
            child_ids.sort(key=lambda account_id: ((chart_map.get(account_id) or {}).get("code") or "", (chart_map.get(account_id) or {}).get("name") or ""))
        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        project_names = FinancialReportService._name_map(Project, company_id)

        period_start = filters.period_start
        period_end = filters.period_end
        competence_start = filters.competence_start or filters.period_start
        competence_end = filters.competence_end or filters.period_end
        due_start = filters.due_start
        due_end = filters.due_end
        settlement_start = filters.settlement_start
        settlement_end = filters.settlement_end

        chart_account_ids = FinancialReportService._selected_ids(filters.chart_account_id, filters.chart_account_ids)
        cost_center_ids = FinancialReportService._selected_ids(filters.cost_center_id, filters.cost_center_ids)
        allowed_entry_types = []
        if filters.include_receivable:
            allowed_entry_types.append("receivable")
        if filters.include_payable:
            allowed_entry_types.append("payable")
        if filters.include_budget_vs_actual or filters.show_budget_column:
            allowed_entry_types.append("forecast")

        grouped: Dict[int, Dict[str, Any]] = {}

        def _account_slot(chart_account_id: Optional[int]) -> Dict[str, Any]:
            account_id = chart_account_id or 0
            account = chart_map.get(account_id, {"id": account_id, "parent_id": None, "code": f"CTA-{account_id}", "name": "Sem conta contábil", "accepts_posting": True})
            return grouped.setdefault(
                account_id,
                {
                    "id": account["id"],
                    "parent_id": account["parent_id"],
                    "codigo": account["code"],
                    "descricao": account["name"],
                    "orcamento": Decimal("0"),
                    "competencia": Decimal("0"),
                    "vencimento": Decimal("0"),
                    "liquidacao": Decimal("0"),
                    "aberto": Decimal("0"),
                    "baixado": Decimal("0"),
                    "centros": set(),
                    "projetos": set(),
                    "tipos": set(),
                    "accepts_posting": account["accepts_posting"],
                },
            )

        def _passes_financial_status(total_amount: Decimal, settled_amount: Decimal, explicit_status: Optional[str] = None) -> Tuple[bool, bool]:
            normalized_status = str(explicit_status or "").strip().lower()
            if normalized_status in {"draft", "cancelled"}:
                return False, False
            if normalized_status == "forecast":
                return bool(filters.include_budget_vs_actual or filters.show_budget_column), True
            is_settled = normalized_status in {"settled", "completed"} or (total_amount > Decimal("0") and settled_amount >= total_amount)
            is_partial = not is_settled and settled_amount > Decimal("0")
            is_open = not is_settled
            if is_settled:
                return bool(filters.include_settled), is_open
            if is_partial:
                return bool(filters.include_partial or filters.include_open), is_open
            return bool(filters.include_open), is_open

        def _add_type(slot: Dict[str, Any], entry_type: Optional[str]) -> None:
            if entry_type == "forecast":
                slot["tipos"].add("Orçado")
            elif entry_type == "receivable":
                slot["tipos"].add("Recebimento")
            elif entry_type == "payable":
                slot["tipos"].add("Pagamento")
            elif entry_type:
                slot["tipos"].add(entry_type)

        def _decimal_from_item(item: Dict[str, Any], *keys: str) -> Decimal:
            for key in keys:
                value = item.get(key)
                if value not in (None, ""):
                    return Decimal(str(value or 0))
            return Decimal("0")

        def _settlement_component_summary(settlement: FinancialSettlement, entry: FinancialEntry) -> Dict[str, Any]:
            try:
                serialized = FinancialService.serialize_settlement(settlement, entry=entry, include_components=True)
                if serialized:
                    return {
                        "component_summary": dict(serialized.get("settlement_component_summary") or {}),
                        "allocation_breakdown": dict(serialized.get("settlement_allocation_breakdown") or {}),
                    }
            except Exception:
                pass

            metadata = dict(getattr(settlement, "metadata_json", {}) or {})
            return {
                "component_summary": {
                    "principal": Decimal(getattr(settlement, "principal_amount", None) or 0),
                    "financial_correction": Decimal(getattr(settlement, "interest_amount", None) or 0)
                    + Decimal(getattr(settlement, "penalty_amount", None) or 0)
                    + Decimal(getattr(settlement, "fee_amount", None) or 0)
                    + Decimal(getattr(settlement, "other_adjustments_amount", None) or 0),
                    "discount": Decimal(getattr(settlement, "discount_amount", None) or 0),
                },
                "allocation_breakdown": dict(metadata.get("settlement_allocation_breakdown") or {}),
            }

        def _resolve_entry_principal_amount(
            entry: FinancialEntry,
            *,
            schedule: Optional[FinancialSchedule] = None,
        ) -> Decimal:
            metadata = dict(getattr(entry, "metadata_json", {}) or {})
            schedule_amount = metadata.get("schedule_template_amount")
            if schedule_amount not in (None, ""):
                return Decimal(str(schedule_amount or 0))
            if schedule is not None:
                return Decimal(str(getattr(schedule, "template_amount", None) or 0))
            return Decimal(str(getattr(entry, "original_amount", None) or 0))

        def _date_in_bucket(target_date: Optional[date], bucket: str) -> bool:
            if not target_date:
                return False
            if consolidated_by_period:
                return bool(period_start and period_end and period_start <= target_date <= period_end)

            if bucket == "competencia":
                if competence_start and target_date < competence_start:
                    return False
                if competence_end and target_date > competence_end:
                    return False
                return True

            if bucket == "vencimento":
                if due_start and target_date < due_start:
                    return False
                if due_end and target_date > due_end:
                    return False
                return True

            if settlement_start and target_date < settlement_start:
                return False
            if settlement_end and target_date > settlement_end:
                return False
            return True

        def _apply_settlement_breakdown(
            *,
            settlement: FinancialSettlement,
            entry: FinancialEntry,
            fallback_chart_account_id: Optional[int],
            fallback_cost_center_id: Optional[int],
        ) -> Dict[str, bool]:
            summary_payload = _settlement_component_summary(settlement, entry)
            component_summary = summary_payload["component_summary"]
            allocation_breakdown = summary_payload["allocation_breakdown"]
            movement_multiplier = Decimal("1") if entry.movement_nature == "credit" else Decimal("-1")
            flags = {"competencia": False, "vencimento": False, "liquidacao": False}
            settlement_date = getattr(settlement, "settlement_date", None)

            component_specs = (
                ("principal", Decimal("1"), Decimal(getattr(settlement, "principal_amount", None) or 0)),
                ("financial_correction", Decimal("1"), Decimal("0")),
                ("discount", Decimal("-1"), Decimal("0")),
            )
            for component_key, component_multiplier, fallback_amount in component_specs:
                breakdown = dict(allocation_breakdown.get(component_key) or {})
                items = list(breakdown.get("items") or [])
                if items:
                    for item in items:
                        item_payload = dict(item or {})
                        amount = abs(_decimal_from_item(item_payload, "settled_allocated_amount", "allocated_amount", "amount"))
                        if amount == Decimal("0"):
                            continue
                        signed_amount = amount * movement_multiplier * component_multiplier
                        chart_account_id = item_payload.get("chart_account_id") or fallback_chart_account_id
                        cost_center_id = item_payload.get("cost_center_id") or fallback_cost_center_id

                        def _register(bucket: str) -> None:
                            slot = _account_slot(chart_account_id)
                            slot[bucket] += signed_amount
                            if cost_center_id:
                                slot["centros"].add(center_names.get(cost_center_id, str(cost_center_id)))
                            for project_id in FinancialReportService._entry_project_ids(entry):
                                if not filters.project_ids or project_id in filters.project_ids:
                                    slot["projetos"].add(project_names.get(project_id, str(project_id)))
                            _add_type(slot, entry.entry_type)
                            flags[bucket] = True

                        if component_key == "principal":
                            if _date_in_bucket(settlement_date, "liquidacao"):
                                _register("liquidacao")
                            continue

                        competence_date = None
                        due_date_value = None
                        raw_competence_date = item_payload.get("competence_date")
                        raw_due_date = item_payload.get("due_date")
                        if raw_competence_date:
                            try:
                                competence_date = date.fromisoformat(str(raw_competence_date))
                            except ValueError:
                                competence_date = settlement_date
                        else:
                            competence_date = settlement_date
                        if raw_due_date:
                            try:
                                due_date_value = date.fromisoformat(str(raw_due_date))
                            except ValueError:
                                due_date_value = settlement_date
                        else:
                            due_date_value = settlement_date

                        if _date_in_bucket(competence_date, "competencia"):
                            _register("competencia")
                        if _date_in_bucket(due_date_value, "vencimento"):
                            _register("vencimento")
                        if _date_in_bucket(settlement_date, "liquidacao"):
                            _register("liquidacao")
                    continue

                amount = abs(Decimal(str(component_summary.get(component_key) or fallback_amount or 0)))
                if amount == Decimal("0"):
                    continue
                signed_amount = amount * movement_multiplier * component_multiplier
                slot = _account_slot(fallback_chart_account_id)
                if component_key == "principal":
                    if _date_in_bucket(settlement_date, "liquidacao"):
                        slot["liquidacao"] += signed_amount
                        flags["liquidacao"] = True
                else:
                    if _date_in_bucket(settlement_date, "competencia"):
                        slot["competencia"] += signed_amount
                        flags["competencia"] = True
                    if _date_in_bucket(settlement_date, "vencimento"):
                        slot["vencimento"] += signed_amount
                        flags["vencimento"] = True
                    if _date_in_bucket(settlement_date, "liquidacao"):
                        slot["liquidacao"] += signed_amount
                        flags["liquidacao"] = True
                if fallback_cost_center_id:
                    slot["centros"].add(center_names.get(fallback_cost_center_id, str(fallback_cost_center_id)))
                for project_id in FinancialReportService._entry_project_ids(entry):
                    if not filters.project_ids or project_id in filters.project_ids:
                        slot["projetos"].add(project_names.get(project_id, str(project_id)))
                _add_type(slot, entry.entry_type)

            return flags

        if consolidated_by_period:
            schedule_query = FinancialSchedule.query.filter(
                FinancialSchedule.company_id == company_id,
                FinancialSchedule.deleted_at.is_(None),
            )
            if chart_account_ids:
                schedule_query = schedule_query.filter(FinancialSchedule.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                schedule_query = schedule_query.filter(FinancialSchedule.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                schedule_query = schedule_query.filter(FinancialSchedule.entry_type.in_(allowed_entry_types))

            schedules = schedule_query.order_by(FinancialSchedule.chart_account_id.asc(), FinancialSchedule.competence_date.asc(), FinancialSchedule.id.asc()).all()
            if filters.project_ids:
                schedules = [item for item in schedules if FinancialReportService._schedule_matches_projects(item, filters.project_ids)]

            schedule_ids = [item.id for item in schedules]
            schedule_refs = {f"financial_schedule:{item.id}": item.id for item in schedules}
            linked_entries: List[FinancialEntry] = []
            if schedule_ids:
                linked_entries = FinancialEntry.query.filter(
                    FinancialEntry.company_id == company_id,
                    FinancialEntry.deleted_at.is_(None),
                    or_(
                        FinancialEntry.financial_schedule_id.in_(schedule_ids),
                        FinancialEntry.external_reference.in_(list(schedule_refs.keys())),
                    ),
                ).all()

            entries_by_schedule: Dict[int, List[FinancialEntry]] = {item.id: [] for item in schedules}
            linked_entry_ids: List[int] = []
            for entry in linked_entries:
                schedule_id = getattr(entry, "financial_schedule_id", None) or schedule_refs.get(entry.external_reference)
                if schedule_id in entries_by_schedule:
                    entries_by_schedule[schedule_id].append(entry)
                    linked_entry_ids.append(entry.id)

            settlement_totals_by_entry: Dict[int, Decimal] = {}
            settlement_items_by_entry: Dict[int, List[FinancialSettlement]] = {}
            if linked_entry_ids:
                for settlement in FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(linked_entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                ).all():
                    settlement_principal = Decimal(settlement.principal_amount or 0)
                    settlement_amount = Decimal(settlement.net_amount or 0)
                    settlement_totals_by_entry.setdefault(settlement.financial_entry_id, Decimal("0"))
                    settlement_totals_by_entry[settlement.financial_entry_id] += settlement_principal
                    if settlement_amount != Decimal("0"):
                        settlement_items_by_entry.setdefault(settlement.financial_entry_id, []).append(settlement)

            for schedule in schedules:
                schedule_entries = entries_by_schedule.get(schedule.id, [])
                competence_date = schedule.competence_date or schedule.start_date
                due_date = schedule.next_due_date or schedule.first_due_date or schedule.start_date
                title_amount = Decimal(str(schedule.template_amount or 0))
                if title_amount <= Decimal("0") and schedule_entries:
                    title_amount = sum(
                        (_resolve_entry_principal_amount(entry, schedule=schedule) for entry in schedule_entries),
                        Decimal("0"),
                    )
                settled_total = sum((settlement_totals_by_entry.get(entry.id, Decimal("0")) for entry in schedule_entries), Decimal("0"))
                derived_settlement_state = "open"
                if title_amount > Decimal("0") and settled_total >= title_amount:
                    derived_settlement_state = "settled"
                elif settled_total > Decimal("0"):
                    derived_settlement_state = "partial"
                operational_state = build_title_operational_state_metadata(
                    schedule_status=schedule.status,
                    settlement_state=derived_settlement_state,
                    entry_type=schedule.entry_type,
                    metadata_json=schedule.metadata_json,
                )
                passes_status, is_open = _passes_financial_status(title_amount, settled_total, operational_state["code"])
                if not passes_status:
                    continue

                in_competence = bool(competence_date and period_start <= competence_date <= period_end)
                in_due = bool(due_date and period_start <= due_date <= period_end)
                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for entry in schedule_entries:
                    for settlement in settlement_items_by_entry.get(entry.id, []):
                        current_flags = _apply_settlement_breakdown(
                            settlement=settlement,
                            entry=entry,
                            fallback_chart_account_id=schedule.chart_account_id,
                            fallback_cost_center_id=schedule.cost_center_id,
                        )
                        settlement_flags = {
                            key: settlement_flags[key] or current_flags[key]
                            for key in settlement_flags
                        }
                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue

                signed_title = Decimal(str(FinancialService.get_signed_amount(title_amount, schedule.movement_nature)))
                slot = _account_slot(schedule.chart_account_id)
                if schedule.entry_type == "forecast" and in_competence:
                    slot["orcamento"] += signed_title
                if schedule.entry_type == "forecast" and not filters.include_budget_vs_actual:
                    if schedule.cost_center_id:
                        slot["centros"].add(center_names.get(schedule.cost_center_id, str(schedule.cost_center_id)))
                    for project_id in FinancialReportService._schedule_project_ids(schedule):
                        if not filters.project_ids or project_id in filters.project_ids:
                            slot["projetos"].add(project_names.get(project_id, str(project_id)))
                    _add_type(slot, schedule.entry_type)
                    continue
                if in_competence:
                    slot["competencia"] += signed_title
                if in_due:
                    slot["vencimento"] += signed_title
                if is_open:
                    slot["aberto"] += signed_title
                else:
                    slot["baixado"] += signed_title
                if schedule.cost_center_id:
                    slot["centros"].add(center_names.get(schedule.cost_center_id, str(schedule.cost_center_id)))
                for project_id in FinancialReportService._schedule_project_ids(schedule):
                    if not filters.project_ids or project_id in filters.project_ids:
                        slot["projetos"].add(project_names.get(project_id, str(project_id)))
                _add_type(slot, schedule.entry_type)

            manual_query = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
            )
            if chart_account_ids:
                manual_query = manual_query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                manual_query = manual_query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                manual_query = manual_query.filter(FinancialEntry.entry_type.in_(allowed_entry_types))
            manual_entries = [entry for entry in manual_query.order_by(FinancialEntry.chart_account_id.asc(), FinancialEntry.competence_date.asc(), FinancialEntry.id.asc()).all() if not getattr(entry, "financial_schedule_id", None) and not str(entry.external_reference or "").startswith("financial_schedule:")]
            if filters.project_ids:
                manual_entries = [entry for entry in manual_entries if FinancialReportService._entry_matches_projects(entry, filters.project_ids)]

            manual_entry_ids = [entry.id for entry in manual_entries]
            manual_settlement_totals: Dict[int, Decimal] = {}
            manual_settlement_items: Dict[int, List[FinancialSettlement]] = {}
            if manual_entry_ids:
                for settlement in FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(manual_entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                ).all():
                    settlement_principal = Decimal(settlement.principal_amount or 0)
                    settlement_amount = Decimal(settlement.net_amount or 0)
                    manual_settlement_totals.setdefault(settlement.financial_entry_id, Decimal("0"))
                    manual_settlement_totals[settlement.financial_entry_id] += settlement_principal
                    if settlement_amount != Decimal("0"):
                        manual_settlement_items.setdefault(settlement.financial_entry_id, []).append(settlement)

            for entry in manual_entries:
                original_amount = Decimal(entry.original_amount or 0)
                total_settlement_amount = manual_settlement_totals.get(entry.id, Decimal("0"))
                passes_status, is_open = _passes_financial_status(original_amount, total_settlement_amount, entry.status)
                if not passes_status:
                    continue
                in_competence = bool(entry.competence_date and period_start <= entry.competence_date <= period_end)
                in_due = bool(entry.due_date and period_start <= entry.due_date <= period_end)
                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for settlement in manual_settlement_items.get(entry.id, []):
                    current_flags = _apply_settlement_breakdown(
                        settlement=settlement,
                        entry=entry,
                        fallback_chart_account_id=entry.chart_account_id,
                        fallback_cost_center_id=entry.cost_center_id,
                    )
                    settlement_flags = {
                        key: settlement_flags[key] or current_flags[key]
                        for key in settlement_flags
                    }
                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue
                signed_original = original_amount if entry.movement_nature == "credit" else -original_amount
                slot = _account_slot(entry.chart_account_id)
                if entry.entry_type == "forecast" and in_competence:
                    slot["orcamento"] += signed_original
                if entry.entry_type == "forecast" and not filters.include_budget_vs_actual:
                    if entry.cost_center_id:
                        slot["centros"].add(center_names.get(entry.cost_center_id, str(entry.cost_center_id)))
                    for project_id in FinancialReportService._entry_project_ids(entry):
                        if not filters.project_ids or project_id in filters.project_ids:
                            slot["projetos"].add(project_names.get(project_id, str(project_id)))
                    _add_type(slot, entry.entry_type)
                    continue
                if in_competence:
                    slot["competencia"] += signed_original
                if in_due:
                    slot["vencimento"] += signed_original
                if is_open:
                    slot["aberto"] += signed_original
                else:
                    slot["baixado"] += signed_original
                if entry.cost_center_id:
                    slot["centros"].add(center_names.get(entry.cost_center_id, str(entry.cost_center_id)))
                for project_id in FinancialReportService._entry_project_ids(entry):
                    if not filters.project_ids or project_id in filters.project_ids:
                        slot["projetos"].add(project_names.get(project_id, str(project_id)))
                _add_type(slot, entry.entry_type)
        else:
            query = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
            )
            if chart_account_ids:
                query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                query = query.filter(FinancialEntry.entry_type.in_(allowed_entry_types))

            entries = query.order_by(FinancialEntry.chart_account_id.asc(), FinancialEntry.competence_date.asc(), FinancialEntry.id.asc()).all()
            if filters.project_ids:
                entries = [entry for entry in entries if FinancialReportService._entry_matches_projects(entry, filters.project_ids)]

            entry_ids = [entry.id for entry in entries]
            settlement_totals_by_entry: Dict[int, Decimal] = {}
            settlement_items_by_entry: Dict[int, List[FinancialSettlement]] = {}
            if entry_ids:
                settlement_query = FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                )
                settlements = settlement_query.all()
                for settlement in settlements:
                    settlement_totals_by_entry.setdefault(settlement.financial_entry_id, Decimal("0"))
                    settlement_totals_by_entry[settlement.financial_entry_id] += Decimal(settlement.principal_amount or 0)
                    settlement_items_by_entry.setdefault(settlement.financial_entry_id, []).append(settlement)

            schedule_by_entry_id: Dict[int, FinancialSchedule] = {}
            schedule_ids = {
                int(getattr(entry, "financial_schedule_id", 0) or 0)
                for entry in entries
                if getattr(entry, "financial_schedule_id", None)
            }
            if schedule_ids:
                for schedule in FinancialSchedule.query.filter(
                    FinancialSchedule.company_id == company_id,
                    FinancialSchedule.id.in_(list(schedule_ids)),
                    FinancialSchedule.deleted_at.is_(None),
                ).all():
                    schedule_by_entry_id[int(getattr(schedule, "id", 0) or 0)] = schedule

            for entry in entries:
                total_settlement_amount = settlement_totals_by_entry.get(entry.id, Decimal("0"))
                linked_schedule = schedule_by_entry_id.get(int(getattr(entry, "financial_schedule_id", 0) or 0))
                original_amount = _resolve_entry_principal_amount(entry, schedule=linked_schedule)
                passes_status, is_open = _passes_financial_status(original_amount, total_settlement_amount, entry.status)
                if not passes_status:
                    continue

                in_competence = True
                if competence_start and ((not entry.competence_date) or entry.competence_date < competence_start):
                    in_competence = False
                if competence_end and ((not entry.competence_date) or entry.competence_date > competence_end):
                    in_competence = False

                in_due = True
                if due_start and ((not entry.due_date) or entry.due_date < due_start):
                    in_due = False
                if due_end and ((not entry.due_date) or entry.due_date > due_end):
                    in_due = False

                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for settlement in settlement_items_by_entry.get(entry.id, []):
                    current_flags = _apply_settlement_breakdown(
                        settlement=settlement,
                        entry=entry,
                        fallback_chart_account_id=(linked_schedule.chart_account_id if linked_schedule is not None else entry.chart_account_id),
                        fallback_cost_center_id=(linked_schedule.cost_center_id if linked_schedule is not None else entry.cost_center_id),
                    )
                    settlement_flags = {
                        key: settlement_flags[key] or current_flags[key]
                        for key in settlement_flags
                    }

                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue

                signed_original = original_amount if entry.movement_nature == "credit" else -original_amount
                slot = _account_slot(entry.chart_account_id)
                if entry.entry_type == "forecast" and in_competence:
                    slot["orcamento"] += signed_original
                if entry.entry_type == "forecast" and not filters.include_budget_vs_actual:
                    if entry.cost_center_id:
                        slot["centros"].add(center_names.get(entry.cost_center_id, str(entry.cost_center_id)))
                    for project_id in FinancialReportService._entry_project_ids(entry):
                        if not filters.project_ids or project_id in filters.project_ids:
                            slot["projetos"].add(project_names.get(project_id, str(project_id)))
                    _add_type(slot, entry.entry_type)
                    continue
                if in_competence:
                    slot["competencia"] += signed_original
                if in_due:
                    slot["vencimento"] += signed_original
                if is_open:
                    slot["aberto"] += signed_original
                else:
                    slot["baixado"] += signed_original
                if entry.cost_center_id:
                    slot["centros"].add(center_names.get(entry.cost_center_id, str(entry.cost_center_id)))
                for project_id in FinancialReportService._entry_project_ids(entry):
                    if not filters.project_ids or project_id in filters.project_ids:
                        slot["projetos"].add(project_names.get(project_id, str(project_id)))
                _add_type(slot, entry.entry_type)

        hierarchy_nodes: Dict[int, Dict[str, Any]] = {}
        for account in chart_accounts:
            data = grouped.get(account.id, {})
            hierarchy_nodes[account.id] = {
                "id": account.id,
                "parent_id": account.parent_id,
                "codigo": getattr(account, "code", None) or f"CTA-{account.id}",
                "descricao": account.name,
                "orcamento": Decimal(data.get("orcamento", Decimal("0"))),
                "competencia": Decimal(data.get("competencia", Decimal("0"))),
                "vencimento": Decimal(data.get("vencimento", Decimal("0"))),
                "liquidacao": Decimal(data.get("liquidacao", Decimal("0"))),
                "aberto": Decimal(data.get("aberto", Decimal("0"))),
                "baixado": Decimal(data.get("baixado", Decimal("0"))),
                "centros": set(data.get("centros", set())),
                "projetos": set(data.get("projetos", set())),
                "tipos": set(data.get("tipos", set())),
                "accepts_posting": bool(getattr(account, "accepts_posting", False)),
                "children": list(chart_children_map.get(account.id, [])),
            }

        for account_id, data in grouped.items():
            if account_id in hierarchy_nodes:
                continue
            hierarchy_nodes[account_id] = {
                "id": account_id,
                "parent_id": data.get("parent_id"),
                "codigo": data.get("codigo") or f"CTA-{account_id}",
                "descricao": data.get("descricao") or "Sem conta contábil",
                "orcamento": Decimal(data.get("orcamento", Decimal("0"))),
                "competencia": Decimal(data.get("competencia", Decimal("0"))),
                "vencimento": Decimal(data.get("vencimento", Decimal("0"))),
                "liquidacao": Decimal(data.get("liquidacao", Decimal("0"))),
                "aberto": Decimal(data.get("aberto", Decimal("0"))),
                "baixado": Decimal(data.get("baixado", Decimal("0"))),
                "centros": set(data.get("centros", set())),
                "projetos": set(data.get("projetos", set())),
                "tipos": set(data.get("tipos", set())),
                "accepts_posting": bool(data.get("accepts_posting", True)),
                "children": [],
            }

        def _aggregate_node(account_id: int) -> Dict[str, Any]:
            node = hierarchy_nodes[account_id]
            for child_id in node["children"]:
                if child_id not in hierarchy_nodes:
                    continue
                child = _aggregate_node(child_id)
                node["orcamento"] += child["orcamento"]
                node["competencia"] += child["competencia"]
                node["vencimento"] += child["vencimento"]
                node["liquidacao"] += child["liquidacao"]
                node["aberto"] += child["aberto"]
                node["baixado"] += child["baixado"]
                node["centros"].update(child["centros"])
                node["projetos"].update(child["projetos"])
                node["tipos"].update(child["tipos"])
            return node

        root_ids = FinancialReportService._resolve_income_statement_root_ids(
            hierarchy_nodes,
            order_by=filters.order_by,
            reverse=filters.order_direction == "desc",
        )
        for root_id in root_ids:
            _aggregate_node(root_id)

        def _node_has_value(node: Dict[str, Any]) -> bool:
            return any(
                Decimal(node[key] or 0) != Decimal("0")
                for key in ("orcamento", "competencia", "vencimento", "liquidacao", "aberto", "baixado")
            )

        hierarchy_rows: List[Dict[str, Any]] = []

        def _append_node(account_id: int, level: int = 0, parent_path: str = "") -> None:
            node = hierarchy_nodes.get(account_id)
            if not node or not _node_has_value(node):
                return
            row_id = f"dre-{account_id}"
            parent_row_id = parent_path or None
            has_children = any(_node_has_value(hierarchy_nodes.get(child_id, {})) for child_id in node["children"])
            if level <= 0:
                row_type = "group"
            elif has_children and not node["accepts_posting"]:
                row_type = "subgroup"
            elif has_children:
                row_type = "account-group"
            else:
                row_type = "account"

            hierarchy_rows.append(
                {
                    "id": row_id,
                    "parent_id": parent_row_id,
                    "chart_account_id": account_id,
                    "codigo": node["codigo"],
                    "descricao": node["descricao"],
                    "account_label": f"{node['codigo']} - {node['descricao']}",
                    "level": level,
                    "row_type": row_type,
                    "is_leaf": not has_children,
                    "has_children": has_children,
                    "orcamento": FinancialReportService._serialize_money(node["orcamento"]),
                    "orcamento_label": FinancialReportService._format_currency(node["orcamento"]),
                    "competencia": FinancialReportService._serialize_money(node["competencia"]),
                    "competencia_label": FinancialReportService._format_currency(node["competencia"]),
                    "vencimento": FinancialReportService._serialize_money(node["vencimento"]),
                    "vencimento_label": FinancialReportService._format_currency(node["vencimento"]),
                    "liquidacao": FinancialReportService._serialize_money(node["liquidacao"]),
                    "liquidacao_label": FinancialReportService._format_currency(node["liquidacao"]),
                    "aberto": FinancialReportService._serialize_money(node["aberto"]),
                    "aberto_label": FinancialReportService._format_currency(node["aberto"]),
                    "baixado": FinancialReportService._serialize_money(node["baixado"]),
                    "baixado_label": FinancialReportService._format_currency(node["baixado"]),
                    "centros": ", ".join(sorted(node["centros"])) or "Todos",
                    "projetos": ", ".join(sorted(node["projetos"])) or "Todos",
                    "tipos": ", ".join(sorted(node["tipos"])) or "N/D",
                }
            )
            for child_id in node["children"]:
                _append_node(child_id, level + 1, row_id)

        for root_id in root_ids:
            _append_node(root_id)

        rows = []
        total_budget = Decimal("0")
        total_comp = Decimal("0")
        total_due = Decimal("0")
        total_set = Decimal("0")
        total_open = Decimal("0")
        total_settled = Decimal("0")
        for root_id in root_ids:
            node = hierarchy_nodes.get(root_id)
            if not node or not _node_has_value(node):
                continue
            total_budget += node["orcamento"]
            total_comp += node["competencia"]
            total_due += node["vencimento"]
            total_set += node["liquidacao"]
            total_open += node["aberto"]
            total_settled += node["baixado"]
        for item in hierarchy_rows:
            row_payload = {
                "codigo": item["codigo"],
                "descricao": item["descricao"],
                "account_label": item["account_label"],
                "level": item["level"],
                "row_type": item["row_type"],
                "has_children": item["has_children"],
                "orcamento": item["orcamento_label"],
                "competencia": item["competencia_label"],
                "vencimento": item["vencimento_label"],
                "liquidacao": item["liquidacao_label"],
                "centros": item["centros"],
                "projetos": item["projetos"],
                "tipos": item["tipos"],
            }
            if not consolidated_by_period:
                row_payload["aberto"] = item["aberto_label"]
                row_payload["baixado"] = item["baixado_label"]
            rows.append(row_payload)

        columns: List[Dict[str, str]] = []
        if filters.show_code:
            columns.append({"key": "codigo", "label": "Código"})
        if filters.show_description:
            columns.append({"key": "descricao", "label": "Descrição"})
        if filters.show_budget_column:
            columns.append({"key": "orcamento", "label": "Orçamento"})
        if filters.show_competence_column:
            columns.append({"key": "competencia", "label": "Competência"})
        if filters.show_due_column:
            columns.append({"key": "vencimento", "label": "Vencimento"})
        if filters.show_liquidation_column:
            columns.append({"key": "liquidacao", "label": "Liquidação"})
        if not consolidated_by_period:
            columns.extend(
                [
                    {"key": "aberto", "label": "Em aberto"},
                    {"key": "baixado", "label": "Baixado"},
                ]
            )
        columns.extend(
            [
                {"key": "centros", "label": "Centros de resultado"},
                {"key": "projetos", "label": "Projetos"},
                {"key": "tipos", "label": "Tipos"},
            ]
        )

        if consolidated_by_period:
            summary_cards: List[Dict[str, Any]] = []
            general_info: List[Dict[str, Any]] = []
        else:
            summary_cards = [
                FinancialReportService._report_card("Resultado competência", FinancialReportService._format_currency(total_comp)),
                FinancialReportService._report_card("Resultado vencimento", FinancialReportService._format_currency(total_due)),
                FinancialReportService._report_card("Resultado baixa", FinancialReportService._format_currency(total_set)),
                FinancialReportService._report_card("Linhas da DRE", len(hierarchy_rows)),
            ]
            due_window = f"{due_start.isoformat()} até {due_end.isoformat()}" if due_start and due_end else "Livre"
            settlement_window = f"{settlement_start.isoformat()} até {settlement_end.isoformat()}" if settlement_start and settlement_end else "Livre"
            general_info = [
                FinancialReportService._report_info("Competência", f"{competence_start.isoformat()} até {competence_end.isoformat()}"),
                FinancialReportService._report_info("Vencimento", due_window),
                FinancialReportService._report_info("Baixa", settlement_window),
                FinancialReportService._report_info("Ordenação", f"{filters.order_by} / {filters.order_direction}"),
                FinancialReportService._report_info("Orientação PDF", "Paisagem" if filters.orientation == "landscape" else "Retrato"),
                FinancialReportService._report_info("Contas consolidadas", len(hierarchy_rows)),
            ]

        return FinancialReportService._report_payload(
            definition,
            summary_cards=summary_cards,
            general_info=general_info,
            columns=columns,
            rows=rows,
            totals={
                "budget": FinancialReportService._serialize_money(total_budget),
                "competence": FinancialReportService._serialize_money(total_comp),
                "due": FinancialReportService._serialize_money(total_due),
                "liquidation": FinancialReportService._serialize_money(total_set),
                "open": FinancialReportService._serialize_money(total_open),
                "settled": FinancialReportService._serialize_money(total_settled),
                "budget_label": FinancialReportService._format_currency(total_budget),
                "competence_label": FinancialReportService._format_currency(total_comp),
                "due_label": FinancialReportService._format_currency(total_due),
                "liquidation_label": FinancialReportService._format_currency(total_set),
                "open_label": FinancialReportService._format_currency(total_open),
                "settled_label": FinancialReportService._format_currency(total_settled),
            },
            extra={
                "orientation": filters.orientation,
                "hierarchy_rows": hierarchy_rows,
                "collapsed_row_ids": filters.collapsed_row_ids,
                "visible_row_ids": filters.visible_row_ids,
                "show_status_columns": not consolidated_by_period,
                "show_budget_column": filters.show_budget_column,
                "show_competence_column": filters.show_competence_column,
                "show_due_column": filters.show_due_column,
                "show_liquidation_column": filters.show_liquidation_column,
            },
        )

    @staticmethod
    def _build_cash_flow(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        bank_names = FinancialReportService._name_map(FinancialBankAccount, company_id)
        chart_names = FinancialReportService._name_map(FinancialChartAccount, company_id)
        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        project_names = FinancialReportService._name_map(Project, company_id)
        process_names = FinancialReportService._name_map(Process, company_id)
        bank_account_ids = FinancialReportService._selected_ids(
            filters.bank_account_id,
            filters.bank_account_ids,
            preserve_empty_marker=True,
        )
        settlements = FinancialReportService._settlement_query(company_id, filters).order_by(FinancialSettlement.settlement_date.asc(), FinancialSettlement.id.asc()).all()
        flow_entries = {
            item.id: item
            for item in FinancialReportService._cash_flow_dimension_filtered_entries(company_id, filters)
        }
        all_entries = {
            item.id: item
            for item in FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
            ).all()
        }
        if bank_account_ids == [-1]:
            bank_accounts_label = "Nenhuma conta selecionada"
        elif bank_account_ids:
            bank_accounts_label = ", ".join(bank_names.get(item, str(item)) for item in bank_account_ids)
        else:
            bank_accounts_label = "Todas"
        periodicity_label = FinancialReportService._cash_flow_frequency_label(filters.frequency)

        bank_accounts_query = FinancialBankAccount.query.filter(
            FinancialBankAccount.company_id == company_id,
            FinancialBankAccount.deleted_at.is_(None),
            FinancialBankAccount.is_active.is_(True),
        )
        all_active_bank_account_count = bank_accounts_query.count()
        if bank_account_ids == [-1]:
            selected_bank_accounts: List[FinancialBankAccount] = []
        else:
            if bank_account_ids:
                bank_accounts_query = bank_accounts_query.filter(FinancialBankAccount.id.in_(bank_account_ids))
            selected_bank_accounts = bank_accounts_query.order_by(
                FinancialBankAccount.code.asc(),
                FinancialBankAccount.name.asc(),
            ).all()
        has_unselected_bank_accounts = (
            all_active_bank_account_count > 0
            if bank_account_ids == [-1]
            else bool(bank_account_ids) and len(selected_bank_accounts) < all_active_bank_account_count
        )

        overdraft_limit = (
            FinancialDashboardAnalytics.calculate_overdraft_limit(
                company_id,
                bank_account_ids=[] if bank_account_ids == [-1] else (bank_account_ids or None),
            )
            if filters.include_overdraft
            else Decimal("0")
        )

        history_query = FinancialSettlement.query.filter(
            FinancialSettlement.company_id == company_id,
            FinancialSettlement.deleted_at.is_(None),
            FinancialSettlement.settlement_status != "cancelled",
            FinancialSettlement.settlement_date < filters.period_start,
        )
        if bank_account_ids:
            history_query = history_query.filter(FinancialSettlement.bank_account_id.in_(bank_account_ids))
        initial_balance = FinancialDashboardAnalytics.calculate_current_balance(
            settlements=history_query.all(),
            entries_by_id=flow_entries,
            as_of_date=filters.period_start,
        )

        def _empty_cash_flow_slot() -> Dict[str, Decimal]:
            return {
                "realized_in": Decimal("0"),
                "realized_out": Decimal("0"),
                "projected_in": Decimal("0"),
                "projected_out": Decimal("0"),
            }

        daily: Dict[date, Dict[str, Decimal]] = {}
        realized_in = Decimal("0")
        realized_out = Decimal("0")
        for settlement in settlements:
            entry = flow_entries.get(settlement.financial_entry_id)
            if not entry:
                continue
            slot = daily.setdefault(settlement.settlement_date, _empty_cash_flow_slot())
            amount = Decimal(settlement.net_amount or 0)
            if entry.movement_nature == "credit":
                slot["realized_in"] += amount
                realized_in += amount
            else:
                slot["realized_out"] += amount
                realized_out += amount

        excluded_projection_ref_set = (
            FinancialReportService._cash_flow_selected_projection_refs(filters)
            if filters.enable_title_exclusions
            else set()
        )
        excluded_titles: List[Dict[str, Any]] = []
        excluded_open_total = Decimal("0")
        projected_in_total = Decimal("0")
        projected_out_total = Decimal("0")
        schedule_projection_cache: Dict[int, Dict[str, Decimal]] = {}
        selected_receivables: List[Dict[str, Any]] = []
        selected_payables: List[Dict[str, Any]] = []
        receivable_title_total = Decimal("0")
        receivable_correction_total = Decimal("0")
        receivable_open_total = Decimal("0")
        payable_title_total = Decimal("0")
        payable_correction_total = Decimal("0")
        payable_open_total = Decimal("0")
        if filters.include_projected:
            projected_entries = FinancialReportService._cash_flow_projected_entry_query(company_id, filters).all()
            if filters.project_ids:
                projected_entries = [
                    entry
                    for entry in projected_entries
                    if FinancialReportService._entry_matches_projects(entry, filters.project_ids)
                ]
            if filters.process_ids:
                projected_entries = [
                    entry
                    for entry in projected_entries
                    if FinancialReportService._entry_matches_processes(entry, filters.process_ids)
                ]
            projected_settlement_totals = FinancialReportService._entry_settlement_totals(
                company_id,
                entry_ids=[entry.id for entry in projected_entries],
            )
            counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
            schedule_ids = {FinancialReportService._entry_schedule_id(entry) for entry in projected_entries}
            schedule_ids = {item for item in schedule_ids if item}
            schedule_cache: Dict[int, FinancialSchedule] = {}
            if schedule_ids:
                for schedule in FinancialSchedule.query.filter(
                    FinancialSchedule.company_id == company_id,
                    FinancialSchedule.id.in_(list(schedule_ids)),
                    FinancialSchedule.deleted_at.is_(None),
                ).all():
                    schedule_cache[int(schedule.id)] = schedule
            for entry in projected_entries:
                if not entry.due_date:
                    continue
                title_amount = Decimal(str(getattr(entry, "original_amount", None) or 0))
                settled_amount = projected_settlement_totals.get(entry.id, Decimal("0"))
                outstanding = FinancialReportService._entry_projected_open_amount(
                    entry,
                    settled_amount=settled_amount,
                    projected_values_mode=filters.projected_values_mode,
                    schedule_cache=schedule_cache,
                    schedule_projection_cache=schedule_projection_cache,
                )
                if outstanding <= Decimal("0"):
                    continue
                serialized = FinancialReportService._serialize_cash_flow_excluded_title(
                    entry,
                    settled_amount=settled_amount,
                    counterparty_names=counterparty_names,
                    open_amount=outstanding,
                )
                entry_projection_ref = FinancialReportService._cash_flow_projection_ref("entry", entry.id)
                is_excluded = entry_projection_ref in excluded_projection_ref_set
                title_row = {
                    "id": entry.id,
                    "projection_ref": entry_projection_ref,
                    "entry_code": serialized["entry_code"],
                    "type_code": FinancialReportService._cash_flow_entry_type_code(entry),
                    "type_label": "Recebimento" if entry.movement_nature == "credit" else "Pagamento",
                    "title_amount": FinancialReportService._format_currency(title_amount),
                    "title_amount_value": FinancialReportService._serialize_money(title_amount),
                    "open_amount": FinancialReportService._format_currency(outstanding),
                    "open_amount_value": FinancialReportService._serialize_money(outstanding),
                    "projected_amount_mode": filters.projected_values_mode,
                    "counterparty": serialized["counterparty"],
                    "due_date": FinancialReportService._format_date_br(entry.due_date),
                    "competence_date": FinancialReportService._format_date_br(entry.competence_date),
                    "number_installment": serialized["number_installment"],
                    "history": serialized["history"],
                    "is_excluded": is_excluded,
                    "status_label": "Retirado" if is_excluded else "No fluxo",
                }
                if entry.movement_nature == "credit":
                    receivable_title_total += title_amount
                    receivable_open_total += outstanding
                    selected_receivables.append(title_row)
                else:
                    payable_title_total += title_amount
                    payable_open_total += outstanding
                    selected_payables.append(title_row)
                if is_excluded:
                    excluded_titles.append(
                        {
                            **serialized,
                            "type_code": FinancialReportService._cash_flow_entry_type_code(entry),
                            "competence_date_display": FinancialReportService._format_date_br(entry.competence_date),
                            "due_date_display": FinancialReportService._format_date_br(entry.due_date),
                            "status_label": "Retirado do fluxo",
                        }
                    )
                    excluded_open_total += outstanding
                    continue
                slot = daily.setdefault(entry.due_date, _empty_cash_flow_slot())
                if entry.movement_nature == "credit":
                    slot["projected_in"] += outstanding
                    projected_in_total += outstanding
                else:
                    slot["projected_out"] += outstanding
                    projected_out_total += outstanding

            projected_schedules = FinancialReportService._cash_flow_projected_schedule_query(company_id, filters).all()
            if filters.process_ids:
                projected_schedules = [item for item in projected_schedules if FinancialReportService._schedule_matches_processes(item, filters.process_ids)]
            if filters.project_ids:
                projected_schedules = [item for item in projected_schedules if FinancialReportService._schedule_matches_projects(item, filters.project_ids)]

            schedule_ids = [item.id for item in projected_schedules]
            entry_refs = {f"financial_schedule:{item.id}": item.id for item in projected_schedules}
            linked_entries_by_schedule: Dict[int, List[FinancialEntry]] = {item.id: [] for item in projected_schedules}
            if schedule_ids:
                linked_entries = FinancialEntry.query.filter(
                    FinancialEntry.company_id == company_id,
                    FinancialEntry.deleted_at.is_(None),
                    or_(
                        FinancialEntry.financial_schedule_id.in_(schedule_ids),
                        FinancialEntry.external_reference.in_(list(entry_refs.keys())),
                    ),
                ).all()
                for linked_entry in linked_entries:
                    schedule_id = getattr(linked_entry, "financial_schedule_id", None) or entry_refs.get(linked_entry.external_reference)
                    if schedule_id in linked_entries_by_schedule:
                        linked_entries_by_schedule[schedule_id].append(linked_entry)

            for schedule in projected_schedules:
                if linked_entries_by_schedule.get(schedule.id):
                    continue

                open_payload = FinancialReportService._cash_flow_schedule_open_amounts(schedule, filters)
                if open_payload is None:
                    continue
                title_amount, open_amount, due_date = open_payload

                metadata = dict(getattr(schedule, "metadata_json", None) or {})
                counterparty_label = (
                    counterparty_names.get(getattr(schedule, "counterparty_id", None))
                    or metadata.get("counterparty_name")
                    or "Não informado"
                )
                schedule_projection_ref = FinancialReportService._cash_flow_projection_ref("schedule", schedule.id)
                is_excluded = schedule_projection_ref in excluded_projection_ref_set
                title_row = {
                    "id": schedule.id,
                    "projection_ref": schedule_projection_ref,
                    "entry_code": getattr(schedule, "schedule_code", None) or str(schedule.id),
                    "type_code": FinancialReportService._cash_flow_entry_type_code(schedule),
                    "type_label": "Recebimento" if schedule.movement_nature == "credit" else "Pagamento",
                    "title_amount": FinancialReportService._format_currency(title_amount),
                    "title_amount_value": FinancialReportService._serialize_money(title_amount),
                    "open_amount": FinancialReportService._format_currency(open_amount),
                    "open_amount_value": FinancialReportService._serialize_money(open_amount),
                    "projected_amount_mode": filters.projected_values_mode,
                    "counterparty": counterparty_label,
                    "due_date": FinancialReportService._format_date_br(due_date),
                    "competence_date": FinancialReportService._format_date_br(schedule.competence_date),
                    "number_installment": FinancialReportService._schedule_installment_label(schedule),
                    "history": FinancialReportService._schedule_history_label(schedule),
                    "is_excluded": is_excluded,
                    "status_label": "Retirado" if is_excluded else "No fluxo",
                }
                if schedule.movement_nature == "credit":
                    receivable_title_total += title_amount
                    receivable_open_total += open_amount
                    selected_receivables.append(title_row)
                else:
                    payable_title_total += title_amount
                    payable_open_total += open_amount
                    selected_payables.append(title_row)

                if is_excluded:
                    excluded_titles.append(
                        {
                            **FinancialReportService._serialize_cash_flow_projected_schedule(
                                schedule,
                                title_amount=title_amount,
                                open_amount=open_amount,
                                counterparty_label=counterparty_label,
                                due_date=due_date,
                            ),
                            "status_label": "Retirado do fluxo",
                        }
                    )
                    excluded_open_total += open_amount
                    continue

                slot = daily.setdefault(due_date, _empty_cash_flow_slot())
                if schedule.movement_nature == "credit":
                    slot["projected_in"] += open_amount
                    projected_in_total += open_amount
                else:
                    slot["projected_out"] += open_amount
                    projected_out_total += open_amount

        bucket_mode = (filters.frequency or "daily").lower()
        bucket_specs = FinancialReportService._cash_flow_period_buckets(
            filters.period_start,
            filters.period_end,
            bucket_mode,
        )
        aggregated: Dict[str, Dict[str, Decimal]] = {
            bucket["key"]: _empty_cash_flow_slot()
            for bucket in bucket_specs
        }
        for day, amounts in sorted(daily.items(), key=lambda item: item[0]):
            if day < filters.period_start or day > filters.period_end:
                continue
            for bucket in bucket_specs:
                if bucket["start"] <= day <= bucket["end"]:
                    slot = aggregated[bucket["key"]]
                    slot["realized_in"] += amounts["realized_in"]
                    slot["realized_out"] += amounts["realized_out"]
                    slot["projected_in"] += amounts["projected_in"]
                    slot["projected_out"] += amounts["projected_out"]
                    break

        running = Decimal(initial_balance)
        rows: List[Dict[str, Any]] = []
        for bucket in bucket_specs:
            slot = aggregated[bucket["key"]]
            opening = running
            inflow_amount = slot["realized_in"] + slot["projected_in"]
            outflow_amount = slot["realized_out"] + slot["projected_out"]
            closing = opening + inflow_amount - outflow_amount
            available_total = closing + overdraft_limit
            rows.append(
                {
                    "periodo": bucket["label"],
                    "data_inicial": FinancialReportService._format_date_br(bucket["start"]),
                    "data_final": FinancialReportService._format_date_br(bucket["end"]),
                    "saldo_inicial": FinancialReportService._format_currency(opening),
                    "entrada": FinancialReportService._format_currency(inflow_amount),
                    "saida": FinancialReportService._format_currency(outflow_amount),
                    "saldo_final": FinancialReportService._format_currency(closing),
                    "limite": FinancialReportService._format_currency(overdraft_limit),
                    "disponivel_total_final": FinancialReportService._format_currency(available_total),
                }
            )
            running = closing
        final_balance = running if rows else Decimal(initial_balance)
        final_with_limit = final_balance + overdraft_limit
        flow_in_total = realized_in + projected_in_total
        flow_out_total = realized_out + projected_out_total
        projected_titles_label = "Todos retirados"
        if filters.include_projected:
            projected_titles_label = (
                f"{len(excluded_titles)} título(s) retirado(s)"
                if filters.enable_title_exclusions
                else "Incluídos"
            )
        dimension_lines = [
            item
            for item in (
                FinancialReportService._cash_flow_header_dimension_line(
                    "Plano de Contas",
                    getattr(filters, "chart_account_ids", None)
                    or ([getattr(filters, "chart_account_id", None)] if getattr(filters, "chart_account_id", None) else []),
                    chart_names,
                ),
                FinancialReportService._cash_flow_header_dimension_line(
                    "Centro de Resultados",
                    getattr(filters, "cost_center_ids", None)
                    or ([getattr(filters, "cost_center_id", None)] if getattr(filters, "cost_center_id", None) else []),
                    center_names,
                ),
                FinancialReportService._cash_flow_header_dimension_line(
                    "Processos",
                    getattr(filters, "process_ids", None) or [],
                    process_names,
                ),
                FinancialReportService._cash_flow_header_dimension_line(
                    "Projetos",
                    getattr(filters, "project_ids", None) or [],
                    project_names,
                ),
            )
            if item
        ]
        if not dimension_lines:
            dimension_lines = [{"label": "Seleções", "value": "Nenhuma seleção específica"}]
        cash_flow_header_cards = [
            {
                "title": "Período e projeção",
                "lines": [
                    {
                        "label": "Período",
                        "value": f"{FinancialReportService._format_date_br(filters.period_start)} até {FinancialReportService._format_date_br(filters.period_end)}",
                    },
                    {
                        "label": "Correção",
                        "value": FinancialReportService._projected_values_mode_label(filters.projected_values_mode),
                    },
                ],
            },
            {
                "title": "Filtros dimensionais",
                "lines": dimension_lines,
            },
            {
                "title": "Controle do relatório",
                "lines": [
                    {"label": "Títulos retirados", "value": "Sim" if excluded_titles else "Não"},
                    {"label": "Contas bancárias não selecionadas", "value": "Sim" if has_unselected_bank_accounts else "Não"},
                ],
            },
        ]

        balance_reference_date = date.today()
        bank_balance_totals = {
            "limit": Decimal("0"),
            "balance": Decimal("0"),
            "available_total": Decimal("0"),
        }
        bank_account_summary_rows: List[Dict[str, Any]] = []
        bank_settlements: List[FinancialSettlement] = []
        if selected_bank_accounts:
            selected_bank_account_ids = [item.id for item in selected_bank_accounts]
            bank_settlements = FinancialSettlement.query.filter(
                FinancialSettlement.company_id == company_id,
                FinancialSettlement.deleted_at.is_(None),
                FinancialSettlement.settlement_status != "cancelled",
                FinancialSettlement.bank_account_id.in_(selected_bank_account_ids),
                FinancialSettlement.settlement_date <= balance_reference_date,
            ).order_by(
                FinancialSettlement.bank_account_id.asc(),
                FinancialSettlement.settlement_date.asc(),
                FinancialSettlement.id.asc(),
            ).all()
        settlements_by_bank: Dict[int, List[FinancialSettlement]] = {}
        for settlement in bank_settlements:
            if settlement.bank_account_id is None:
                continue
            settlements_by_bank.setdefault(settlement.bank_account_id, []).append(settlement)

        for account in selected_bank_accounts:
            account_limit = (
                FinancialReportService._cash_flow_bank_account_limit(account)
                if filters.include_overdraft
                else Decimal("0")
            )
            account_balance = FinancialDashboardAnalytics.calculate_current_balance(
                settlements=settlements_by_bank.get(account.id, []),
                entries_by_id=all_entries,
                as_of_date=balance_reference_date,
            )
            account_available_total = account_balance + account_limit
            bank_account_summary_rows.append(
                {
                    "id": account.id,
                    "description": account.name,
                    "limit": FinancialReportService._format_currency(account_limit),
                    "limit_value": FinancialReportService._serialize_money(account_limit),
                    "balance": FinancialReportService._format_currency(account_balance),
                    "balance_value": FinancialReportService._serialize_money(account_balance),
                    "available_total": FinancialReportService._format_currency(account_available_total),
                    "available_total_value": FinancialReportService._serialize_money(account_available_total),
                }
            )
            bank_balance_totals["limit"] += account_limit
            bank_balance_totals["balance"] += account_balance
            bank_balance_totals["available_total"] += account_available_total

        return FinancialReportService._report_payload(
            definition,
            summary_cards=[
                FinancialReportService._report_card(
                    "Saldo inicial",
                    FinancialReportService._format_currency(initial_balance),
                    "positive" if initial_balance > 0 else ("negative" if initial_balance < 0 else "neutral"),
                ),
                FinancialReportService._report_card(
                    "Entradas no fluxo",
                    FinancialReportService._format_currency(flow_in_total),
                    "positive" if flow_in_total > 0 else "neutral",
                ),
                FinancialReportService._report_card(
                    "Saídas no fluxo",
                    FinancialReportService._format_currency(flow_out_total),
                    "negative" if flow_out_total > 0 else "neutral",
                ),
                FinancialReportService._report_card(
                    "Saldo final do período",
                    FinancialReportService._format_currency(final_balance),
                    "positive" if final_balance > 0 else ("negative" if final_balance < 0 else "neutral"),
                ),
                FinancialReportService._report_card(
                    "Disponível c/ limite",
                    FinancialReportService._format_currency(final_with_limit),
                    "primary" if final_with_limit >= 0 else "negative",
                ),
                FinancialReportService._report_card(
                    "Títulos retirados",
                    FinancialReportService._format_currency(excluded_open_total),
                    "primary" if excluded_open_total > 0 else "neutral",
                ),
            ],
            general_info=[
                FinancialReportService._report_info(
                    "Janela analisada",
                    f"{FinancialReportService._format_date_br(filters.period_start)} até {FinancialReportService._format_date_br(filters.period_end)}",
                ),
                FinancialReportService._report_info("Contas correntes", bank_accounts_label),
                FinancialReportService._report_info("Periodicidade", periodicity_label),
                FinancialReportService._report_info("Valores projetados", FinancialReportService._projected_values_mode_label(filters.projected_values_mode)),
                FinancialReportService._report_info("Títulos financeiros em aberto", projected_titles_label),
                FinancialReportService._report_info("Títulos retirados manualmente", str(len(excluded_titles))),
                FinancialReportService._report_info(
                    f"Saldo em {FinancialReportService._format_date_br(balance_reference_date)}",
                    FinancialReportService._format_currency(bank_balance_totals["balance"]),
                ),
                FinancialReportService._report_info(
                    "Disponível total",
                    FinancialReportService._format_currency(bank_balance_totals["available_total"]),
                ),
                FinancialReportService._report_info(
                    "Limite total",
                    FinancialReportService._format_currency(bank_balance_totals["limit"]),
                ),
            ],
            columns=[
                {"key": "periodo", "label": "Período"},
                {"key": "data_inicial", "label": "Data Inicial"},
                {"key": "data_final", "label": "Data Final"},
                {"key": "saldo_inicial", "label": "Saldo Inicial"},
                {"key": "entrada", "label": "Entrada"},
                {"key": "saida", "label": "Saída"},
                {"key": "saldo_final", "label": "Saldo Final"},
                {"key": "limite", "label": "Limite"},
                {"key": "disponivel_total_final", "label": "Disp. Total Final"},
            ],
            rows=rows,
            totals={
                "opening_balance": FinancialReportService._serialize_money(initial_balance),
                "flow_in_total": FinancialReportService._serialize_money(flow_in_total),
                "flow_out_total": FinancialReportService._serialize_money(flow_out_total),
                "final_balance": FinancialReportService._serialize_money(final_balance),
                "overdraft_limit": FinancialReportService._serialize_money(overdraft_limit),
                "final_with_limit": FinancialReportService._serialize_money(final_with_limit),
                "excluded_projected_amount": FinancialReportService._serialize_money(excluded_open_total),
                "current_bank_balance": FinancialReportService._serialize_money(bank_balance_totals["balance"]),
                "current_available_total": FinancialReportService._serialize_money(bank_balance_totals["available_total"]),
            },
            extra={
                "periodicity_label": periodicity_label,
                "bank_balance_reference_label": FinancialReportService._format_date_br(balance_reference_date),
                "bank_account_summary_rows": bank_account_summary_rows,
                "bank_account_summary_totals": {
                    "limit": FinancialReportService._format_currency(bank_balance_totals["limit"]),
                    "limit_value": FinancialReportService._serialize_money(bank_balance_totals["limit"]),
                    "balance": FinancialReportService._format_currency(bank_balance_totals["balance"]),
                    "balance_value": FinancialReportService._serialize_money(bank_balance_totals["balance"]),
                    "available_total": FinancialReportService._format_currency(bank_balance_totals["available_total"]),
                    "available_total_value": FinancialReportService._serialize_money(bank_balance_totals["available_total"]),
                },
                "selected_receivables": selected_receivables,
                "selected_payables": selected_payables,
                "selected_receivables_totals": {
                    "count": len(selected_receivables),
                    "title_amount": FinancialReportService._format_currency(receivable_title_total),
                    "title_amount_value": FinancialReportService._serialize_money(receivable_title_total),
                    "open_amount": FinancialReportService._format_currency(receivable_open_total),
                    "open_amount_value": FinancialReportService._serialize_money(receivable_open_total),
                },
                "selected_payables_totals": {
                    "count": len(selected_payables),
                    "title_amount": FinancialReportService._format_currency(payable_title_total),
                    "title_amount_value": FinancialReportService._serialize_money(payable_title_total),
                    "open_amount": FinancialReportService._format_currency(payable_open_total),
                    "open_amount_value": FinancialReportService._serialize_money(payable_open_total),
                },
                "excluded_titles": excluded_titles,
                "cash_flow_header_cards": cash_flow_header_cards,
                "projected_values_mode": filters.projected_values_mode,
                "projected_values_mode_label": FinancialReportService._projected_values_mode_label(filters.projected_values_mode),
                "projected_amount_label": "Saldo do Principal Corrigido" if filters.projected_values_mode == "with_financial_correction" else "Saldo do Principal",
            },
        )

    @staticmethod
    def _build_ledger(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        chart_names = FinancialReportService._name_map(FinancialChartAccount, company_id)
        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        project_names = FinancialReportService._name_map(Project, company_id)
        process_names = FinancialReportService._name_map(Process, company_id)
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
        competence_start = filters.competence_start or filters.period_start
        competence_end = filters.competence_end or filters.period_end
        due_start = filters.due_start
        due_end = filters.due_end
        settlement_start = filters.settlement_start
        settlement_end = filters.settlement_end

        query = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
            FinancialEntry.competence_date >= competence_start,
            FinancialEntry.competence_date <= competence_end,
        )
        chart_account_ids = FinancialReportService._selected_ids(filters.chart_account_id, filters.chart_account_ids)
        if chart_account_ids:
            query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
        cost_center_ids = FinancialReportService._selected_ids(filters.cost_center_id, filters.cost_center_ids)
        if cost_center_ids:
            query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
        counterparty_ids = FinancialReportService._selected_ids(filters.counterparty_id, filters.counterparty_ids)
        if counterparty_ids:
            query = query.filter(FinancialEntry.counterparty_id.in_(counterparty_ids))
        if filters.movement_nature:
            query = query.filter(FinancialEntry.movement_nature == filters.movement_nature)
        entries = query.order_by(FinancialEntry.competence_date.asc(), FinancialEntry.id.asc()).all()
        if filters.project_ids:
            entries = [entry for entry in entries if FinancialReportService._entry_matches_projects(entry, filters.project_ids)]
        if filters.process_ids:
            entries = [entry for entry in entries if FinancialReportService._entry_matches_processes(entry, filters.process_ids)]
        entry_ids = [entry.id for entry in entries]
        settlements_by_entry: Dict[int, Decimal] = {}
        settlement_dates_by_entry: Dict[int, List[date]] = defaultdict(list)
        allocation_counts_by_entry: Dict[int, int] = defaultdict(int)
        if entry_ids:
            settlement_query = FinancialSettlement.query.filter(
                FinancialSettlement.company_id == company_id,
                FinancialSettlement.financial_entry_id.in_(entry_ids),
                FinancialSettlement.deleted_at.is_(None),
                FinancialSettlement.settlement_status != "cancelled",
            )
            if settlement_start:
                settlement_query = settlement_query.filter(FinancialSettlement.settlement_date >= settlement_start)
            if settlement_end:
                settlement_query = settlement_query.filter(FinancialSettlement.settlement_date <= settlement_end)
            for settlement in settlement_query.all():
                settlements_by_entry.setdefault(settlement.financial_entry_id, Decimal("0"))
                settlements_by_entry[settlement.financial_entry_id] += Decimal(settlement.net_amount or 0)
                if settlement.settlement_date:
                    settlement_dates_by_entry[settlement.financial_entry_id].append(settlement.settlement_date)

            for allocation in FinancialEntryAllocation.query.filter(
                FinancialEntryAllocation.company_id == company_id,
                FinancialEntryAllocation.financial_entry_id.in_(entry_ids),
                FinancialEntryAllocation.deleted_at.is_(None),
            ).all():
                allocation_counts_by_entry[allocation.financial_entry_id] += 1

        inflow_total = Decimal("0")
        outflow_total = Decimal("0")
        net_total = Decimal("0")
        rows = []
        grouping_label_map = {
            "code": "Plano de Contas",
            "description": "Centro de Resultado",
            "project": "Projeto / Processo",
            "counterparty": "Favorecido",
            "movement_nature": "Tipo",
        }
        movement_label_map = {
            "credit": "Entrada",
            "debit": "Saída",
        }

        def _project_process_label(entry: FinancialEntry) -> str:
            project_labels = [project_names.get(pid, str(pid)) for pid in FinancialReportService._entry_project_ids(entry)]
            process_labels = [process_names.get(pid, str(pid)) for pid in FinancialReportService._entry_process_ids(entry)]
            labels = project_labels + [label for label in process_labels if label not in project_labels]
            return " | ".join(labels) if labels else "Não informado"

        def _group_label_for_row(row: Dict[str, Any]) -> str:
            value_map = {
                "code": row["plano_contas"],
                "description": row["centro_resultado"],
                "project": row["projeto_processo"],
                "counterparty": row["favorecido"],
                "movement_nature": row["tipo"],
            }
            return value_map.get(filters.order_by, row["plano_contas"])

        for entry in entries:
            if due_start and ((not entry.due_date) or entry.due_date < due_start):
                continue
            if due_end and ((not entry.due_date) or entry.due_date > due_end):
                continue
            settled_amount = settlements_by_entry.get(entry.id, Decimal("0"))
            if entry.status == "settled":
                status_bucket = "Baixado"
            elif settled_amount > 0:
                status_bucket = "Baixado Parcial"
            else:
                status_bucket = "Aberto"
            if status_bucket == "Baixado" and not filters.include_settled:
                continue
            if status_bucket == "Baixado Parcial" and not filters.include_budget_vs_actual:
                continue
            if status_bucket == "Aberto" and not filters.include_open:
                continue
            amount = Decimal(entry.original_amount or 0)
            signed_amount = amount if entry.movement_nature == "credit" else -amount
            if signed_amount >= 0:
                inflow_total += signed_amount
            else:
                outflow_total += abs(signed_amount)
            net_total += signed_amount
            liquidation_dates = settlement_dates_by_entry.get(entry.id) or []
            liquidation_date = max(liquidation_dates).isoformat() if liquidation_dates else "-"
            project_process = _project_process_label(entry)
            rows.append(
                {
                    "agrupador": "",
                    "id": entry.id,
                    "historico": FinancialReportService._entry_history_label(entry),
                    "favorecido": counterparty_names.get(entry.counterparty_id, "Não informado"),
                    "tipo": movement_label_map.get(entry.movement_nature, "Não informado"),
                    "plano_contas": chart_names.get(entry.chart_account_id, "Sem conta contábil"),
                    "centro_resultado": center_names.get(entry.cost_center_id, "Não informado"),
                    "projeto_processo": project_process,
                    "competencia": entry.competence_date.isoformat() if entry.competence_date else "-",
                    "vencimento": entry.due_date.isoformat() if entry.due_date else "-",
                    "liquidacao": liquidation_date,
                    "valor": FinancialReportService._format_signed_currency(signed_amount, positive_sign=True),
                    "valor_value": FinancialReportService._serialize_money(signed_amount),
                    "totalizador": "",
                    "totalizador_value": 0.0,
                    "numero_qtd_rateio": f"{FinancialReportService._entry_number_installment_label(entry)} / {allocation_counts_by_entry.get(entry.id, 0)}",
                    "status": status_bucket,
                    "_group_label": "",
                    "_group_sort_value": "",
                }
            )
        reverse = filters.order_direction == "desc"
        for row in rows:
            group_label = _group_label_for_row(row)
            row["agrupador"] = group_label
            row["_group_label"] = group_label
            row["_group_sort_value"] = str(group_label).lower()
        rows.sort(
            key=lambda item: (
                item["_group_sort_value"],
                item["competencia"],
                item["id"],
            ),
            reverse=reverse,
        )

        grouped_rows: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        group_totals: Dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in rows:
            grouped_rows[row["_group_label"]].append(row)
            group_totals[row["_group_label"]] += Decimal(str(row["valor_value"] or 0))

        groups = []
        flat_rows = []
        for group_label in sorted(grouped_rows.keys(), key=lambda value: str(value).lower(), reverse=reverse):
            group_total = group_totals[group_label]
            items = grouped_rows[group_label]
            for row in items:
                row["totalizador"] = FinancialReportService._format_signed_currency(group_total, positive_sign=True)
                row["totalizador_value"] = FinancialReportService._serialize_money(group_total)
                flat_rows.append(row)
            groups.append(
                {
                    "label": group_label,
                    "item_count": len(items),
                    "total": FinancialReportService._format_signed_currency(group_total, positive_sign=True),
                    "total_value": FinancialReportService._serialize_money(group_total),
                    "rows": items,
                }
            )
        return {
            "title": definition["label"],
            "subtitle": definition["description"],
            "summary_cards": [
                {"label": "Total entradas", "value": FinancialReportService._format_currency(inflow_total)},
                {"label": "Total saídas", "value": FinancialReportService._format_currency(outflow_total)},
                {"label": "Saldo líquido", "value": FinancialReportService._format_signed_currency(net_total, positive_sign=True)},
                {"label": "Lançamentos", "value": len(flat_rows)},
            ],
            "general_info": [
                {"label": "Competência", "value": f"{FinancialReportService._format_date_br(competence_start)} até {FinancialReportService._format_date_br(competence_end)}"},
                {"label": "Vencimento", "value": f"{FinancialReportService._format_date_br(due_start)} até {FinancialReportService._format_date_br(due_end)}" if due_start and due_end else "Livre"},
                {"label": "Liquidação", "value": f"{FinancialReportService._format_date_br(settlement_start)} até {FinancialReportService._format_date_br(settlement_end)}" if settlement_start and settlement_end else "Livre"},
                {"label": "Agrupado por", "value": grouping_label_map.get(filters.order_by, filters.order_by)},
                {"label": "Orientação PDF", "value": "Paisagem" if filters.orientation == "landscape" else "Retrato"},
            ],
            "columns": [
                {"key": "agrupador", "label": "Agrupador"},
                {"key": "id", "label": "ID"},
                {"key": "historico", "label": "Histórico"},
                {"key": "favorecido", "label": "Favorecido"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "plano_contas", "label": "Plano de Contas"},
                {"key": "centro_resultado", "label": "Centro de Resultado"},
                {"key": "projeto_processo", "label": "Projeto/Processo"},
                {"key": "competencia", "label": "Competência"},
                {"key": "vencimento", "label": "Vencimento"},
                {"key": "liquidacao", "label": "Liquidação"},
                {"key": "valor", "label": "Valor"},
                {"key": "totalizador", "label": "Totalizador"},
                {"key": "numero_qtd_rateio", "label": "Número/Qtd Rateio"},
            ],
            "rows": flat_rows,
            "groups": groups,
            "grouped_by": filters.order_by,
            "grouped_by_label": grouping_label_map.get(filters.order_by, filters.order_by),
            "totals": {
                "inflow": FinancialReportService._serialize_money(inflow_total),
                "outflow": FinancialReportService._serialize_money(outflow_total),
                "net": FinancialReportService._serialize_money(net_total),
            },
            "orientation": filters.orientation,
        }

    @staticmethod
    def _build_working_capital(company_id: int, filters: FinancialManagementReportFiltersInput) -> Dict[str, Any]:
        definition = FinancialReportService.REPORT_DEFINITIONS[filters.report_type]
        reference_date = filters.reference_date or filters.period_end
        working_capital_accounts = FinancialReportService._get_working_capital_accounts(company_id)
        entries_by_id = {entry.id: entry for entry in FinancialEntry.query.filter(FinancialEntry.company_id == company_id, FinancialEntry.deleted_at.is_(None)).all()}
        settlement_query = FinancialSettlement.query.filter(
            FinancialSettlement.company_id == company_id,
            FinancialSettlement.deleted_at.is_(None),
            FinancialSettlement.settlement_status != "cancelled",
            FinancialSettlement.settlement_date <= reference_date,
        )
        bank_account_ids = FinancialReportService._selected_ids(filters.bank_account_id, filters.bank_account_ids)
        if bank_account_ids:
            settlement_query = settlement_query.filter(FinancialSettlement.bank_account_id.in_(bank_account_ids))
        bank_balance = FinancialDashboardAnalytics.calculate_current_balance(settlements=settlement_query.all(), entries_by_id=entries_by_id, as_of_date=reference_date)
        open_entries = FinancialEntry.query.filter(
            FinancialEntry.company_id == company_id,
            FinancialEntry.deleted_at.is_(None),
            FinancialEntry.status.in_(["draft", "pending_review", "scheduled", "posted", "partially_settled"]),
            FinancialEntry.due_date.isnot(None),
        )
        if bank_account_ids:
            open_entries = open_entries.filter(FinancialEntry.bank_account_id.in_(bank_account_ids))
        if filters.cost_center_id:
            open_entries = open_entries.filter(FinancialEntry.cost_center_id == filters.cost_center_id)
        global_chart_account_ids = set(FinancialReportService._selected_ids(filters.chart_account_id, filters.chart_account_ids))
        if global_chart_account_ids:
            open_entries = open_entries.filter(FinancialEntry.chart_account_id.in_(list(global_chart_account_ids)))
        open_entries = open_entries.order_by(FinancialEntry.due_date.asc(), FinancialEntry.id.asc()).all()
        chart_names = FinancialReportService._name_map(FinancialChartAccount, company_id)
        overdraft = FinancialDashboardAnalytics.calculate_overdraft_limit(company_id) if filters.include_overdraft else Decimal("0")

        def _entry_label(entry: FinancialEntry) -> str:
            return f"{chart_names.get(entry.chart_account_id, entry.description or 'Sem conta')} · {entry.description}"

        def _is_due(entry: FinancialEntry) -> bool:
            return bool(entry.due_date and entry.due_date >= reference_date)

        def _is_overdue(entry: FinancialEntry) -> bool:
            return bool(entry.due_date and entry.due_date < reference_date)

        def _classify_bucket(entry: FinancialEntry) -> str:
            label = f"{chart_names.get(entry.chart_account_id, '')} {entry.description or ''}".lower()
            if any(token in label for token in ["fornecedor", "compra", "mercadoria"]):
                return "supplier"
            if any(token in label for token in ["salário", "funcion", "folha", "colaborador"]):
                return "people"
            if any(token in label for token in ["imposto", "tribut", "taxa", "fiscal"]):
                return "tax"
            if any(token in label for token in ["emprést", "financi", "banco", "finame"]):
                return "financing"
            if any(token in label for token in ["invest", "aplica", "contrato"]):
                return "investment"
            return "generic"

        def _sum_entries(predicate) -> tuple[Decimal, list[str]]:
            total = Decimal("0")
            labels: list[str] = []
            for entry in open_entries:
                if predicate(entry):
                    total += Decimal(entry.original_amount or 0)
                    labels.append(_entry_label(entry))
            return total, labels

        receivables_due_180, receivables_due_180_labels = _sum_entries(lambda e: e.movement_nature == "credit" and _is_due(e) and (e.due_date - reference_date).days <= 180)
        receivables_overdue, receivables_overdue_labels = _sum_entries(lambda e: e.movement_nature == "credit" and _is_overdue(e))
        payable_supplier_due, payable_supplier_due_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "supplier" and _is_due(e))
        payable_supplier_overdue, payable_supplier_overdue_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "supplier" and _is_overdue(e))
        payable_people_due, payable_people_due_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "people" and _is_due(e))
        payable_people_overdue, payable_people_overdue_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "people" and _is_overdue(e))
        payable_tax_due, payable_tax_due_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "tax" and _is_due(e))
        payable_tax_overdue, payable_tax_overdue_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "tax" and _is_overdue(e))
        payable_financing_due, payable_financing_due_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "financing" and _is_due(e))
        payable_financing_overdue, payable_financing_overdue_labels = _sum_entries(lambda e: e.movement_nature == "debit" and _classify_bucket(e) == "financing" and _is_overdue(e))
        receivable_investment_due, receivable_investment_due_labels = _sum_entries(lambda e: e.movement_nature == "credit" and _classify_bucket(e) == "investment" and _is_due(e))
        receivable_investment_overdue, receivable_investment_overdue_labels = _sum_entries(lambda e: e.movement_nature == "credit" and _classify_bucket(e) == "investment" and _is_overdue(e))

        computed_accounts = {
            "bank_balance": (Decimal(bank_balance), []),
            "receivable_due_180": (receivables_due_180, receivables_due_180_labels),
            "receivable_overdue": (receivables_overdue, receivables_overdue_labels),
            "payable_supplier_due": (payable_supplier_due, payable_supplier_due_labels),
            "payable_supplier_overdue": (payable_supplier_overdue, payable_supplier_overdue_labels),
            "payable_people_due": (payable_people_due, payable_people_due_labels),
            "payable_people_overdue": (payable_people_overdue, payable_people_overdue_labels),
            "payable_tax_due": (payable_tax_due, payable_tax_due_labels),
            "payable_tax_overdue": (payable_tax_overdue, payable_tax_overdue_labels),
            "payable_financing_due": (payable_financing_due, payable_financing_due_labels),
            "payable_financing_overdue": (payable_financing_overdue, payable_financing_overdue_labels),
            "receivable_investment_due": (receivable_investment_due, receivable_investment_due_labels),
            "receivable_investment_overdue": (receivable_investment_overdue, receivable_investment_overdue_labels),
        }

        def _calculate_bank_balance_for_ids(selected_bank_ids: Sequence[int]) -> Decimal:
            ids = [int(item) for item in selected_bank_ids if item]
            if not ids:
                return Decimal("0")
            scoped_query = FinancialSettlement.query.filter(
                FinancialSettlement.company_id == company_id,
                FinancialSettlement.deleted_at.is_(None),
                FinancialSettlement.settlement_status != "cancelled",
                FinancialSettlement.settlement_date <= reference_date,
                FinancialSettlement.bank_account_id.in_(ids),
            )
            return Decimal(
                FinancialDashboardAnalytics.calculate_current_balance(
                    settlements=scoped_query.all(),
                    entries_by_id=entries_by_id,
                    as_of_date=reference_date,
                )
            )

        def _calculate_due_date_account(config: Dict[str, Any]) -> tuple[Decimal, list[str]]:
            metadata = dict(config.get("metadata_json") or {})
            chart_account_ids = {int(item) for item in metadata.get("chart_account_ids") or [] if item}
            if global_chart_account_ids:
                chart_account_ids = chart_account_ids.intersection(global_chart_account_ids) if chart_account_ids else set(global_chart_account_ids)
            if not chart_account_ids:
                return Decimal("0"), []

            patrimonial_type = str(metadata.get("patrimonial_type") or "asset").strip().lower()
            expected_nature = "debit" if patrimonial_type == "liability" else "credit"
            due_scope = str(metadata.get("due_scope") or "overdue").strip().lower()
            due_in_days = metadata.get("due_in_days")

            def _matches_scope(entry: FinancialEntry) -> bool:
                if entry.chart_account_id not in chart_account_ids:
                    return False
                if entry.movement_nature != expected_nature:
                    return False
                if not entry.due_date:
                    return False
                if due_scope == "all_future":
                    return entry.due_date >= reference_date
                if due_scope == "due_in_days":
                    if entry.due_date < reference_date:
                        return False
                    if due_in_days in (None, ""):
                        return True
                    return (entry.due_date - reference_date).days <= int(due_in_days)
                return entry.due_date < reference_date

            total = Decimal("0")
            labels: list[str] = []
            for entry in open_entries:
                if _matches_scope(entry):
                    total += Decimal(entry.original_amount or 0)
                    labels.append(_entry_label(entry))
            return total, labels

        selected_ids = set(filters.working_capital_accounts or [item["id"] for item in working_capital_accounts])
        detail_rows = []
        current_assets = Decimal("0")
        current_liabilities = Decimal("0")
        non_current_assets = Decimal("0")
        non_current_liabilities = Decimal("0")
        for config in working_capital_accounts:
            if config["id"] not in selected_ids:
                continue
            metadata = dict(config.get("metadata_json") or {})
            config_mode = str(config.get("config_mode") or config.get("rule") or "").strip().lower()
            if metadata:
                if config_mode == "bank_balances":
                    configured_bank_ids = [int(item) for item in metadata.get("bank_account_ids") or [] if item]
                    if bank_account_ids and configured_bank_ids:
                        effective_bank_ids = [item for item in configured_bank_ids if item in set(bank_account_ids)]
                    else:
                        effective_bank_ids = configured_bank_ids or list(bank_account_ids)
                    amount = _calculate_bank_balance_for_ids(effective_bank_ids)
                    labels = ["Contas bancárias selecionadas"] if effective_bank_ids else []
                elif config_mode == "manual_value":
                    amount = Decimal(filters.manual_values.get(config["id"], 0) or 0)
                    labels = ["Valor informado na emissão"] if amount else []
                else:
                    amount, labels = _calculate_due_date_account(config)
            else:
                amount, labels = computed_accounts.get(config["rule"], (Decimal("0"), []))
            account_type = str(config["type"] or "").strip()
            account_class = str(config["class_name"] or "").strip()
            is_current = account_class.lower() == "circulante"
            signal = Decimal("1") if account_type == "Ativo" else Decimal("-1")
            if account_type == "Ativo" and is_current:
                current_assets += amount
            elif account_type == "Passivo" and is_current:
                current_liabilities += amount
            elif account_type == "Ativo":
                non_current_assets += amount
            elif account_type == "Passivo":
                non_current_liabilities += amount
            detail_rows.append(
                {
                    "id": config["id"],
                    "descricao": config["description"],
                    "tipo": account_type,
                    "classe": account_class,
                    "categoria": config["category"],
                    "valor_data": config.get("value_label") or ("Saldo em conta" if config["rule"] == "bank_balance" else ("Vencidas" if "overdue" in config["rule"] else "Todas à vencer" if "payable" in config["rule"] or "investment" in config["rule"] else "À vencer em 180 dias.")),
                    "valor": FinancialReportService._serialize_money(amount * signal if account_type == "Passivo" else amount),
                    "valor_absoluto": FinancialReportService._serialize_money(amount),
                    "base_calculo": ", ".join(labels[:5]) if labels else ("Contas bancárias selecionadas" if config_mode == "bank_balances" or config.get("rule") == "bank_balance" else "Valor não informado para esta emissão" if config_mode == "manual_value" else "Sem títulos para a regra"),
                }
            )

        working_capital = current_assets - current_liabilities
        adjusted_liquidity = working_capital + overdraft
        liquidity_ratio = (current_assets / current_liabilities) if current_liabilities else Decimal("0")
        equity = (current_assets + non_current_assets) - (current_liabilities + non_current_liabilities)

        def _build_balance_groups(target_type: str, target_class: str) -> list[dict[str, Any]]:
            groups: dict[str, dict[str, Any]] = {}
            rows = [
                row for row in detail_rows
                if str(row.get("tipo") or "").strip().lower() == target_type.lower()
                and str(row.get("classe") or "").strip().lower() == target_class.lower()
            ]
            rows.sort(key=lambda item: str(item.get("descricao") or ""))
            for row in rows:
                description = str(row.get("descricao") or "").strip()
                match = re.match(r"^(?P<code>\d+(?:\.\d+)+)\s*-\s*(?P<label>.+)$", description)
                if match:
                    full_code = match.group("code")
                    item_label = match.group("label").strip()
                else:
                    full_code = ""
                    item_label = description
                code_parts = [part for part in full_code.split(".") if part]
                if len(code_parts) >= 2:
                    group_code = ".".join(code_parts[:2])
                    group_label = item_label.split(" - ")[0].strip()
                else:
                    group_code = full_code or description
                    group_label = item_label
                group_key = f"{group_code}|{group_label}"
                if group_key not in groups:
                    groups[group_key] = {
                        "code": group_code,
                        "label": group_label,
                        "total": Decimal("0"),
                        "items": [],
                    }
                amount = Decimal(str(row.get("valor_absoluto") or 0))
                groups[group_key]["total"] += amount
                groups[group_key]["items"].append(
                    {
                        "code": full_code or group_code,
                        "label": item_label,
                        "amount": FinancialReportService._format_currency(amount),
                    }
                )
            ordered_groups = []
            for group in sorted(groups.values(), key=lambda item: item["code"] or item["label"]):
                ordered_groups.append(
                    {
                        "code": group["code"],
                        "label": group["label"],
                        "amount": FinancialReportService._format_currency(group["total"]),
                        "items": group["items"],
                    }
                )
            return ordered_groups

        return {
            "title": definition["label"],
            "subtitle": definition["description"],
            "summary_cards": [
                {"label": "Ativo circulante", "value": FinancialReportService._format_currency(current_assets)},
                {"label": "Passivo circulante", "value": FinancialReportService._format_currency(current_liabilities)},
                {"label": "CCL", "value": FinancialReportService._format_currency(working_capital)},
                {"label": "Índice de liquidez", "value": f"{float(liquidity_ratio):.2f}" if current_liabilities else "N/A"},
            ],
            "general_info": [
                {"label": "Posição base", "value": reference_date.isoformat()},
                {"label": "Contas programadas", "value": len(detail_rows)},
                {"label": "Considera limites", "value": "Sim" if filters.include_overdraft else "Não"},
            ],
            "columns": [
                {"key": "id", "label": "ID"},
                {"key": "descricao", "label": "Descrição"},
                {"key": "tipo", "label": "Tipo"},
                {"key": "classe", "label": "Classe"},
                {"key": "categoria", "label": "Categoria"},
                {"key": "valor_data", "label": "Valor/Data"},
                {"key": "valor", "label": "Valor"},
                {"key": "base_calculo", "label": "Base de cálculo"},
            ],
            "rows": detail_rows,
            "totals": {
                "current_assets": FinancialReportService._serialize_money(current_assets),
                "current_liabilities": FinancialReportService._serialize_money(current_liabilities),
                "working_capital": FinancialReportService._serialize_money(working_capital),
                "adjusted_liquidity": FinancialReportService._serialize_money(adjusted_liquidity),
                "non_current_assets": FinancialReportService._serialize_money(non_current_assets),
                "non_current_liabilities": FinancialReportService._serialize_money(non_current_liabilities),
                "equity": FinancialReportService._serialize_money(equity),
            },
            "balance_sheet": {
                "asset": {
                    "title": "Ativo",
                    "current": {
                        "title": "Circulante",
                        "amount": FinancialReportService._format_currency(current_assets),
                        "groups": _build_balance_groups("Ativo", "Circulante"),
                    },
                    "non_current": {
                        "title": "Ativo Não Circulante",
                        "amount": FinancialReportService._format_currency(non_current_assets),
                    },
                },
                "liability": {
                    "title": "Passivo",
                    "current": {
                        "title": "Circulante",
                        "amount": FinancialReportService._format_currency(current_liabilities),
                        "groups": _build_balance_groups("Passivo", "Circulante"),
                    },
                    "non_current": {
                        "title": "Passivo Não Circulante",
                        "amount": FinancialReportService._format_currency(non_current_liabilities),
                    },
                    "equity": {
                        "title": "Patrimônio Líquido",
                        "amount": FinancialReportService._format_currency(equity),
                    },
                },
                "working_capital": {
                    "title": "Capital Circulante Líquido",
                    "amount": FinancialReportService._format_currency(working_capital),
                },
                "patrimonial_status": {
                    "title": "Situação Patrimonial",
                    "amount": FinancialReportService._format_currency(equity),
                },
            },
        }

    @staticmethod
    def _build_filter_labels(
        filters: FinancialManagementReportFiltersInput,
        company_id: int,
        raw_filters: Optional[Dict[str, Any]] = None,
    ) -> List[Dict[str, str]]:
        bank_names = FinancialReportService._name_map(FinancialBankAccount, company_id)
        chart_names = FinancialReportService._name_map(FinancialChartAccount, company_id)
        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)
        project_names = FinancialReportService._name_map(Project, company_id)
        process_names = FinancialReportService._name_map(Process, company_id)
        raw_filters = raw_filters or {}
        order_labels = {
            "code": "Plano de Contas",
            "description": "Centro de Resultado",
            "project": "Projeto / Processo",
            "movement_nature": "Tipo",
            "title_number": "Nº Título",
            "installment": "Parcela",
            "history": "Histórico",
            "counterparty": "Favorecido",
            "title_amount": "Valor Título",
            "balance_amount": "Valor Saldo",
            "competence_date": "Competência",
            "due_date": "Vencimento",
            "settlement_date": "Dt. da Última Liquid.",
        }
        default_start, default_end = FinancialReportService.default_period()
        values: List[Dict[str, str]] = []
        if filters.report_type not in {"schedule_report", "income_statement", "income_statement_2", "ledger"}:
            values.extend(
                [
                    {"label": "Período inicial", "value": filters.period_start.isoformat()},
                    {"label": "Período final", "value": filters.period_end.isoformat()},
                ]
            )
        if filters.report_type == "schedule_report":
            competence_explicit = (
                "competence_start" in raw_filters
                or "competence_end" in raw_filters
            )
            default_competence = (
                filters.competence_start == default_start
                and filters.competence_end == default_end
            )
            if filters.competence_start and filters.competence_end and (competence_explicit or not default_competence):
                values.append({"label": "Data competência", "value": f"{filters.competence_start.isoformat()} até {filters.competence_end.isoformat()}"})
        elif filters.report_type == "income_statement":
            if filters.period_start and filters.period_end:
                values.append({"label": "Período", "value": f"{filters.period_start.isoformat()} até {filters.period_end.isoformat()}"})
        elif filters.report_type == "income_statement_2":
            if (
                filters.competence_start and filters.competence_end
                and ("competence_start" in raw_filters or "competence_end" in raw_filters)
            ):
                values.append({"label": "Data competência", "value": f"{filters.competence_start.isoformat()} até {filters.competence_end.isoformat()}"})
        elif filters.competence_start and filters.competence_end:
            values.append({"label": "Data competência", "value": f"{filters.competence_start.isoformat()} até {filters.competence_end.isoformat()}"})
        if filters.due_start and filters.due_end and ("due_start" in raw_filters or "due_end" in raw_filters or filters.report_type not in {"schedule_report", "income_statement", "income_statement_2"}):
            values.append({"label": "Data vencimento", "value": f"{filters.due_start.isoformat()} até {filters.due_end.isoformat()}"})
        if filters.settlement_start and filters.settlement_end and ("settlement_start" in raw_filters or "settlement_end" in raw_filters or filters.report_type not in {"schedule_report", "income_statement", "income_statement_2"}):
            values.append({"label": "Data baixa", "value": f"{filters.settlement_start.isoformat()} até {filters.settlement_end.isoformat()}"})
        if filters.reference_date and filters.report_type != "schedule_report":
            values.append({"label": "Data de referência", "value": filters.reference_date.isoformat()})
        for current, label, names in [
            (filters.bank_account_id, "Conta bancária", bank_names),
            (filters.chart_account_id, "Conta contábil", chart_names),
            (filters.cost_center_id, "Centro de resultados", center_names),
        ]:
            if current:
                values.append({"label": label, "value": names.get(current, str(current))})
        if filters.bank_account_ids:
            if all(int(item) == -1 for item in filters.bank_account_ids):
                values.append({"label": "Contas correntes", "value": "Nenhuma conta selecionada"})
            else:
                positive_bank_ids = [int(item) for item in filters.bank_account_ids if int(item) > 0]
                values.append({"label": "Contas correntes", "value": ", ".join(bank_names.get(item, str(item)) for item in positive_bank_ids)})
        if filters.chart_account_ids:
            values.append({"label": "Planos de conta", "value": ", ".join(chart_names.get(item, str(item)) for item in filters.chart_account_ids)})
        if filters.cost_center_ids:
            values.append({"label": "Centros de resultado", "value": ", ".join(center_names.get(item, str(item)) for item in filters.cost_center_ids)})
        counterparty_ids = FinancialReportService._selected_ids(filters.counterparty_id, filters.counterparty_ids)
        if counterparty_ids:
            values.append({"label": "Favorecidos", "value": ", ".join(counterparty_names.get(item, str(item)) for item in counterparty_ids)})
        if filters.process_ids:
            values.append({"label": "Processos", "value": ", ".join(process_names.get(item, str(item)) for item in filters.process_ids)})
        if filters.project_ids:
            values.append({"label": "Projetos", "value": ", ".join(project_names.get(item, str(item)) for item in filters.project_ids)})
        if filters.working_capital_accounts:
            selected_map = {item["id"]: item for item in FinancialReportService._get_working_capital_accounts(company_id)}
            selected = [selected_map[item_id] for item_id in filters.working_capital_accounts if item_id in selected_map]
            values.append({"label": "Contas patrimoniais", "value": ", ".join(item["description"] for item in selected)})
        if filters.manual_values:
            manual_labels = []
            account_map = {item["id"]: item for item in FinancialReportService._get_working_capital_accounts(company_id)}
            for account_id, value in filters.manual_values.items():
                account = account_map.get(account_id)
                if not account:
                    continue
                manual_labels.append(f"{account['description']}: {FinancialReportService._format_currency(value)}")
            if manual_labels:
                values.append({"label": "Valores digitados", "value": " | ".join(manual_labels)})
        if filters.movement_nature:
            values.append({"label": "Movimento", "value": "Entrada" if filters.movement_nature == "credit" else "Saída"})
        if filters.schedule_status:
            values.append({"label": "Status do título", "value": filters.schedule_status})
        if filters.frequency:
            values.append({"label": "Frequência", "value": filters.frequency})
        if filters.report_type == "schedule_report":
            all_status_enabled = all([filters.include_settled, filters.include_partial, filters.include_open, filters.include_bordero])
            status_explicit = any(
                key in raw_filters
                for key in ("include_settled", "include_partial", "include_open", "include_bordero")
            )
            if status_explicit or not all_status_enabled:
                values.append({
                    "label": "Status",
                    "value": ", ".join(
                        [
                            label for enabled, label in [
                                (filters.include_settled, "Baixado"),
                                (filters.include_partial, "Baixado Parcial"),
                                (filters.include_open, "Aberto"),
                                (filters.include_bordero, "Borderô"),
                            ] if enabled
                        ]
                    ) or "Nenhum",
                })
            all_types_enabled = all([filters.include_payable, filters.include_receivable])
            type_explicit = any(
                key in raw_filters
                for key in ("include_payable", "include_receivable")
            )
            if type_explicit or not all_types_enabled:
                values.append({
                    "label": "Tipo",
                    "value": ", ".join(
                        [
                            label for enabled, label in [
                                (filters.include_payable, "Pagamento"),
                                (filters.include_receivable, "Recebimento"),
                            ] if enabled
                        ]
                    ) or "Nenhum",
                })
        elif filters.report_type in {"bank_statement", "bank_statement_dossier"}:
            bank_order_labels = {
                "settlement_date": "Data da baixa",
                "due_date": "Vencimento",
                "competence_date": "Competência",
                "code": "Liquidação",
                "title_number": "Lançamento",
                "description": "Descrição",
                "history": "Descrição",
                "counterparty": "Favorecido",
                "movement_nature": "Movimento",
                "title_amount": "Valor",
                "balance_amount": "Saldo",
            }
            values.append({"label": "Projetar abertos", "value": "Sim" if filters.include_projected else "Não"})
            values.append({"label": "Somente conciliados", "value": "Sim" if filters.include_reconciled_only else "Não"})
            values.append({"label": "Considerar limites", "value": "Sim" if filters.include_overdraft else "Não"})
            values.append({
                "label": "Status considerados",
                "value": ", ".join([label for enabled, label in [
                    (filters.include_settled, "Baixado"),
                    (filters.include_partial, "Baixado parcial"),
                    (filters.include_open, "Aberto"),
                ] if enabled]) or "Nenhum",
            })
            values.append({
                "label": "Tipos considerados",
                "value": ", ".join([label for enabled, label in [
                    (filters.include_receivable, "Recebimento"),
                    (filters.include_payable, "Pagamento"),
                ] if enabled]) or "Nenhum",
            })
            values.append({
                "label": "Exibir",
                "value": ", ".join([label for enabled, label in [
                    (filters.show_settlement_date, "Data"),
                    (filters.show_code, "Liquidação"),
                    (filters.show_title_number, "Lançamento"),
                    (filters.show_description, "Descrição"),
                    (filters.show_counterparty, "Favorecido"),
                    (filters.show_competence_date, "Competência"),
                    (filters.show_due_date, "Vencimento"),
                    (filters.show_title_amount, "Valor"),
                    (filters.show_balance_amount, "Saldo"),
                ] if enabled]) or "Nenhum",
            })
            values.append({"label": "Ordenar por", "value": bank_order_labels.get(filters.order_by, filters.order_by)})
            values.append({"label": "Direção", "value": "Crescente" if filters.order_direction == "asc" else "Decrescente"})
        elif filters.report_type in {"income_statement", "income_statement_2"}:
            status_explicit = any(key in raw_filters for key in ("include_settled", "include_open"))
            if status_explicit:
                values.append({
                    "label": "Status",
                    "value": ", ".join([label for enabled, label in [
                        (filters.include_settled, "Baixado"),
                        (filters.include_open, "Aberto"),
                    ] if enabled]) or "Nenhum",
                })
            type_explicit = any(key in raw_filters for key in ("include_receivable", "include_payable"))
            if type_explicit:
                values.append({
                    "label": "Tipo",
                    "value": ", ".join([label for enabled, label in [
                        (filters.include_receivable, "Recebimento"),
                        (filters.include_payable, "Pagamento"),
                    ] if enabled]) or "Nenhum",
                })
            column_explicit = any(key in raw_filters for key in ("show_competence_column", "show_due_column", "show_liquidation_column"))
            if column_explicit:
                values.append({
                    "label": "Colunas DRE",
                    "value": ", ".join([label for enabled, label in [
                        (filters.show_competence_column, "Competência"),
                        (filters.show_due_column, "Vencimento"),
                        (filters.show_liquidation_column, "Liquidação"),
                    ] if enabled]) or "Nenhuma",
                })
        elif filters.report_type == "ledger":
            values.append({
                "label": "Status considerados",
                "value": ", ".join([
                    label for enabled, label in [
                        (filters.include_settled, "Liquidado"),
                        (filters.include_budget_vs_actual, "Liquidado parcial"),
                        (filters.include_open, "Aberto"),
                    ] if enabled
                ]) or "Nenhum",
            })
            values.append({"label": "Agrupado por", "value": order_labels.get(filters.order_by, filters.order_by)})
            values.append({"label": "Orientação", "value": "Paisagem" if filters.orientation == "landscape" else "Retrato"})
        else:
            if filters.report_type == "cash_flow":
                if filters.include_projected:
                    values.append({"label": "Títulos financeiros em aberto", "value": "Incluídos"})
                    values.append({
                        "label": "Retirada manual de títulos",
                        "value": f"{len(FinancialReportService._cash_flow_selected_projection_refs(filters))} selecionado(s)" if filters.enable_title_exclusions else "Desativada",
                    })
                else:
                    values.append({"label": "Títulos financeiros em aberto", "value": "Todos retirados"})
            else:
                values.append({"label": "Projetar abertos", "value": "Sim" if filters.include_projected else "Não"})
            values.append({"label": "Somente conciliados", "value": "Sim" if filters.include_reconciled_only else "Não"})
            values.append({"label": "Considerar limites", "value": "Sim" if filters.include_overdraft else "Não"})
            values.append({"label": "Status considerados", "value": ", ".join([label for enabled, label in [(filters.include_settled, "Baixado"), (filters.include_open, "Aberto")] if enabled]) or "Nenhum"})
            values.append({"label": "Tipos considerados", "value": ", ".join([label for enabled, label in [(filters.include_receivable, "Recebimento"), (filters.include_payable, "Pagamento"), (filters.include_budget_vs_actual, "Orçado x Realizado")] if enabled]) or "Nenhum"})
            values.append({"label": "Exibir", "value": ", ".join([label for enabled, label in [(filters.show_code, "Código"), (filters.show_description, "Descrição")] if enabled]) or "Nenhum"})
            values.append({"label": "Ordenar por", "value": order_labels.get(filters.order_by, filters.order_by)})
            values.append({"label": "Direção", "value": filters.order_direction})
            values.append({"label": "Orientação", "value": "Paisagem" if filters.orientation == "landscape" else "Retrato"})
        return values

    @staticmethod
    def build_management_report(*, company_id: int, report_type: str, filters: Optional[Dict[str, Any]] = None, allowed_company_ids: Optional[Sequence[int]] = None) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error
        definition, error = FinancialReportService.get_report_definition_or_error(report_type)
        if error:
            return None, error
        normalized_filters, error = FinancialReportService._normalize_filters(definition["code"], filters)
        if error:
            return None, error
        builder_map = {
            "schedule_report": FinancialReportService._build_schedule_report,
            "bank_statement": FinancialReportService._build_bank_statement,
            "bank_statement_dossier": FinancialReportService._build_bank_statement,
            "income_statement": FinancialReportService._build_income_statement,
            "income_statement_2": FinancialReportService._build_income_statement_2,
            "cash_flow": FinancialReportService._build_cash_flow,
            "ledger": FinancialReportService._build_ledger,
            "working_capital": FinancialReportService._build_working_capital,
        }
        payload = builder_map[definition["code"]](company_id, normalized_filters)
        payload.update(
            {
                "report_type": definition["code"],
                "report_slug": definition["slug"],
                "generated_at": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "filters": FinancialReportService._build_filter_labels(normalized_filters, company_id, raw_filters=filters),
                "period_start": normalized_filters.period_start.isoformat(),
                "period_end": normalized_filters.period_end.isoformat(),
                "orientation": normalized_filters.orientation,
                "output_mode": normalized_filters.output_mode,
                "dossier_mode": normalized_filters.dossier_mode,
                "items": payload.get("rows", []),
            }
        )
        return payload, None

    @staticmethod
    def build_income_statement_drilldown(
        *,
        company_id: int,
        report_type: str,
        bucket: str,
        chart_account_id: Optional[int] = None,
        filters: Optional[Dict[str, Any]] = None,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        definition, error = FinancialReportService.get_report_definition_or_error(report_type)
        if error:
            return None, error
        if definition["code"] not in {"income_statement", "income_statement_2"}:
            return None, "Drill-down disponível apenas para a DRE."

        normalized_filters, error = FinancialReportService._normalize_filters(definition["code"], filters)
        if error:
            return None, error

        bucket_map = {
            "competence": {"key": "competencia", "label": "Competência", "source_label": "Títulos e ajustes"},
            "competencia": {"key": "competencia", "label": "Competência", "source_label": "Títulos e ajustes"},
            "due": {"key": "vencimento", "label": "Vencimento", "source_label": "Títulos e ajustes"},
            "vencimento": {"key": "vencimento", "label": "Vencimento", "source_label": "Títulos e ajustes"},
            "liquidation": {"key": "liquidacao", "label": "Baixa", "source_label": "Baixas"},
            "liquidacao": {"key": "liquidacao", "label": "Baixa", "source_label": "Baixas"},
            "settlement": {"key": "liquidacao", "label": "Baixa", "source_label": "Baixas"},
            "baixa": {"key": "liquidacao", "label": "Baixa", "source_label": "Baixas"},
            "open": {"key": "aberto", "label": "Em aberto", "source_label": "Títulos em aberto"},
            "aberto": {"key": "aberto", "label": "Em aberto", "source_label": "Títulos em aberto"},
            "settled": {"key": "baixado", "label": "Liquidado", "source_label": "Títulos liquidados"},
            "baixado": {"key": "baixado", "label": "Liquidado", "source_label": "Títulos liquidados"},
            "liquidado": {"key": "baixado", "label": "Liquidado", "source_label": "Títulos liquidados"},
        }
        bucket_meta = bucket_map.get(str(bucket or "").strip().lower())
        if not bucket_meta:
            return None, "Bucket inválido para drill-down da DRE."

        consolidated_by_period = definition["code"] == "income_statement"
        chart_accounts = FinancialChartAccount.query.filter(
            FinancialChartAccount.company_id == company_id,
            FinancialChartAccount.deleted_at.is_(None),
        ).order_by(FinancialChartAccount.code.asc(), FinancialChartAccount.name.asc()).all()
        chart_map = {
            item.id: {
                "id": item.id,
                "parent_id": item.parent_id,
                "code": getattr(item, "code", None) or f"CTA-{item.id}",
                "name": item.name,
                "accepts_posting": bool(getattr(item, "accepts_posting", False)),
            }
            for item in chart_accounts
        }
        chart_children_map: Dict[Optional[int], List[int]] = {}
        for item in chart_accounts:
            chart_children_map.setdefault(item.parent_id, []).append(item.id)
        for child_ids in chart_children_map.values():
            child_ids.sort(key=lambda account_id: ((chart_map.get(account_id) or {}).get("code") or "", (chart_map.get(account_id) or {}).get("name") or ""))

        if chart_account_id is not None and chart_account_id not in chart_map:
            return None, "Conta contábil não encontrada para o drill-down."

        center_names = FinancialReportService._name_map(FinancialCostCenter, company_id)
        counterparty_names = FinancialReportService._name_map(FinancialCounterparty, company_id)

        period_start = normalized_filters.period_start
        period_end = normalized_filters.period_end
        competence_start = normalized_filters.competence_start or normalized_filters.period_start
        competence_end = normalized_filters.competence_end or normalized_filters.period_end
        due_start = normalized_filters.due_start
        due_end = normalized_filters.due_end
        settlement_start = normalized_filters.settlement_start
        settlement_end = normalized_filters.settlement_end

        chart_account_ids = FinancialReportService._selected_ids(normalized_filters.chart_account_id, normalized_filters.chart_account_ids)
        cost_center_ids = FinancialReportService._selected_ids(normalized_filters.cost_center_id, normalized_filters.cost_center_ids)
        allowed_entry_types = []
        if normalized_filters.include_receivable:
            allowed_entry_types.append("receivable")
        if normalized_filters.include_payable:
            allowed_entry_types.append("payable")
        if normalized_filters.include_budget_vs_actual:
            allowed_entry_types.append("forecast")

        detail_items_by_bucket: Dict[str, Dict[int, List[Dict[str, Any]]]] = {
            "competencia": {},
            "vencimento": {},
            "liquidacao": {},
            "aberto": {},
            "baixado": {},
        }

        def _bucket_store(current_bucket: str, account_id: Optional[int]) -> List[Dict[str, Any]]:
            normalized_account_id = int(account_id or 0)
            return detail_items_by_bucket.setdefault(current_bucket, {}).setdefault(normalized_account_id, [])

        def _account_label(account_id: Optional[int]) -> str:
            account = chart_map.get(int(account_id or 0))
            if not account:
                return "Sem conta contábil"
            return f"{account['code']} - {account['name']}"

        def _date_iso(value: Optional[date]) -> Optional[str]:
            return value.isoformat() if hasattr(value, "isoformat") else None

        def _component_label(component_key: str) -> str:
            return {
                "principal": "Principal",
                "financial_correction": "Correção financeira",
                "discount": "Desconto",
            }.get(component_key, component_key)

        def _title_reference(schedule: Optional[FinancialSchedule], entry: Optional[FinancialEntry]) -> str:
            if schedule is not None:
                return getattr(schedule, "schedule_code", None) or getattr(schedule, "document_number_prefix", None) or f"Título {schedule.id}"
            if entry is not None:
                return getattr(entry, "entry_code", None) or getattr(entry, "document_number", None) or f"Lançamento {entry.id}"
            return "Título"

        def _title_description(schedule: Optional[FinancialSchedule], entry: Optional[FinancialEntry]) -> str:
            if schedule is not None:
                return getattr(schedule, "description", None) or getattr(schedule, "name", None) or "Sem histórico"
            if entry is not None:
                return getattr(entry, "description", None) or getattr(entry, "memo", None) or "Sem histórico"
            return "Sem histórico"

        def _title_counterparty(schedule: Optional[FinancialSchedule], entry: Optional[FinancialEntry]) -> str:
            schedule_counterparty_id = getattr(schedule, "counterparty_id", None) if schedule is not None else None
            entry_counterparty_id = getattr(entry, "counterparty_id", None) if entry is not None else None
            if schedule_counterparty_id:
                return counterparty_names.get(schedule_counterparty_id, "Não informado")
            if entry_counterparty_id:
                return counterparty_names.get(entry_counterparty_id, "Não informado")
            schedule_metadata = dict(getattr(schedule, "metadata_json", {}) or {}) if schedule is not None else {}
            entry_metadata = dict(getattr(entry, "metadata_json", {}) or {}) if entry is not None else {}
            return schedule_metadata.get("counterparty_name") or entry_metadata.get("counterparty_name") or "Não informado"

        def _relevant_date_iso(bucket_key: str, *, competence_date: Optional[date], due_date_value: Optional[date], settlement_date_value: Optional[date]) -> Optional[str]:
            if bucket_key == "competencia":
                return _date_iso(competence_date)
            if bucket_key == "vencimento":
                return _date_iso(due_date_value)
            if bucket_key == "liquidacao":
                return _date_iso(settlement_date_value)
            if bucket_key == "aberto":
                return _date_iso(due_date_value or competence_date)
            if bucket_key == "baixado":
                return _date_iso(settlement_date_value or due_date_value or competence_date)
            return None

        def _push_item(
            *,
            bucket_key: str,
            account_id: Optional[int],
            amount: Decimal,
            source_kind: str,
            source_code: str,
            component_key: str,
            description: str,
            counterparty: str,
            competence_date_value: Optional[date],
            due_date_value: Optional[date],
            settlement_date_value: Optional[date],
            entry: Optional[FinancialEntry] = None,
            schedule: Optional[FinancialSchedule] = None,
            settlement: Optional[FinancialSettlement] = None,
            cost_center_id: Optional[int] = None,
        ) -> None:
            item_amount = Decimal(amount or 0)
            if item_amount == Decimal("0"):
                return
            normalized_account_id = int(account_id or 0)
            cost_center_label = center_names.get(cost_center_id, str(cost_center_id)) if cost_center_id else "Todos"
            store = _bucket_store(bucket_key, normalized_account_id)
            store.append(
                {
                    "bucket": bucket_key,
                    "source_kind": source_kind,
                    "source_kind_label": "Baixa" if source_kind == "settlement" else "Título",
                    "source_code": source_code,
                    "component_key": component_key,
                    "component_label": _component_label(component_key),
                    "description": description,
                    "counterparty": counterparty or "Não informado",
                    "account_id": normalized_account_id,
                    "account_label": _account_label(normalized_account_id),
                    "cost_center_label": cost_center_label,
                    "competence_date": _date_iso(competence_date_value),
                    "due_date": _date_iso(due_date_value),
                    "settlement_date": _date_iso(settlement_date_value),
                    "relevant_date": _relevant_date_iso(
                        bucket_key,
                        competence_date=competence_date_value,
                        due_date_value=due_date_value,
                        settlement_date_value=settlement_date_value,
                    ),
                    "amount": FinancialReportService._serialize_money(item_amount),
                    "amount_label": FinancialReportService._format_currency(item_amount),
                    "schedule_id": getattr(schedule, "id", None),
                    "entry_id": getattr(entry, "id", None),
                    "settlement_id": getattr(settlement, "id", None),
                }
            )

        def _passes_financial_status(total_amount: Decimal, settled_amount: Decimal, explicit_status: Optional[str] = None) -> Tuple[bool, bool]:
            normalized_status = str(explicit_status or "").strip().lower()
            if normalized_status in {"draft", "cancelled"}:
                return False, False
            if normalized_status == "forecast":
                return bool(normalized_filters.include_budget_vs_actual), True
            is_settled = normalized_status in {"settled", "completed"} or (total_amount > Decimal("0") and settled_amount >= total_amount)
            is_partial = not is_settled and settled_amount > Decimal("0")
            is_open = not is_settled
            if is_settled:
                return bool(normalized_filters.include_settled), is_open
            if is_partial:
                return bool(normalized_filters.include_partial or normalized_filters.include_open), is_open
            return bool(normalized_filters.include_open), is_open

        def _settlement_component_summary(settlement: FinancialSettlement, entry: FinancialEntry) -> Dict[str, Any]:
            try:
                serialized = FinancialService.serialize_settlement(settlement, entry=entry, include_components=True)
                if serialized:
                    return {
                        "component_summary": dict(serialized.get("settlement_component_summary") or {}),
                        "allocation_breakdown": dict(serialized.get("settlement_allocation_breakdown") or {}),
                    }
            except Exception:
                pass

            metadata = dict(getattr(settlement, "metadata_json", {}) or {})
            return {
                "component_summary": {
                    "principal": Decimal(getattr(settlement, "principal_amount", None) or 0),
                    "financial_correction": Decimal(getattr(settlement, "interest_amount", None) or 0)
                    + Decimal(getattr(settlement, "penalty_amount", None) or 0)
                    + Decimal(getattr(settlement, "fee_amount", None) or 0)
                    + Decimal(getattr(settlement, "other_adjustments_amount", None) or 0),
                    "discount": Decimal(getattr(settlement, "discount_amount", None) or 0),
                },
                "allocation_breakdown": dict(metadata.get("settlement_allocation_breakdown") or {}),
            }

        def _resolve_entry_principal_amount(
            entry: FinancialEntry,
            *,
            schedule: Optional[FinancialSchedule] = None,
        ) -> Decimal:
            metadata = dict(getattr(entry, "metadata_json", {}) or {})
            schedule_amount = metadata.get("schedule_template_amount")
            if schedule_amount not in (None, ""):
                return Decimal(str(schedule_amount or 0))
            if schedule is not None:
                return Decimal(str(getattr(schedule, "template_amount", None) or 0))
            return Decimal(str(getattr(entry, "original_amount", None) or 0))

        def _date_in_bucket(target_date: Optional[date], bucket_key: str) -> bool:
            if not target_date:
                return False
            if consolidated_by_period:
                return bool(period_start and period_end and period_start <= target_date <= period_end)

            if bucket_key == "competencia":
                if competence_start and target_date < competence_start:
                    return False
                if competence_end and target_date > competence_end:
                    return False
                return True

            if bucket_key == "vencimento":
                if due_start and target_date < due_start:
                    return False
                if due_end and target_date > due_end:
                    return False
                return True

            if settlement_start and target_date < settlement_start:
                return False
            if settlement_end and target_date > settlement_end:
                return False
            return True

        def _push_settlement_breakdown(
            *,
            settlement: FinancialSettlement,
            entry: FinancialEntry,
            schedule: Optional[FinancialSchedule],
            fallback_chart_account_id: Optional[int],
            fallback_cost_center_id: Optional[int],
        ) -> Dict[str, bool]:
            summary_payload = _settlement_component_summary(settlement, entry)
            component_summary = summary_payload["component_summary"]
            allocation_breakdown = summary_payload["allocation_breakdown"]
            movement_multiplier = Decimal("1") if entry.movement_nature == "credit" else Decimal("-1")
            flags = {"competencia": False, "vencimento": False, "liquidacao": False}
            settlement_date_value = getattr(settlement, "settlement_date", None)
            title_reference = _title_reference(schedule, entry)
            title_description = _title_description(schedule, entry)
            title_counterparty = _title_counterparty(schedule, entry)

            component_specs = (
                ("principal", Decimal("1"), Decimal(getattr(settlement, "principal_amount", None) or 0)),
                ("financial_correction", Decimal("1"), Decimal("0")),
                ("discount", Decimal("-1"), Decimal("0")),
            )
            for component_key, component_multiplier, fallback_amount in component_specs:
                breakdown = dict(allocation_breakdown.get(component_key) or {})
                items = list(breakdown.get("items") or [])
                if items:
                    for item in items:
                        item_payload = dict(item or {})
                        amount = abs(Decimal(str(item_payload.get("settled_allocated_amount", item_payload.get("allocated_amount", item_payload.get("amount", 0))) or 0)))
                        if amount == Decimal("0"):
                            continue
                        signed_amount = amount * movement_multiplier * component_multiplier
                        allocated_chart_account_id = item_payload.get("chart_account_id") or fallback_chart_account_id
                        allocated_cost_center_id = item_payload.get("cost_center_id") or fallback_cost_center_id
                        raw_competence_date = item_payload.get("competence_date")
                        raw_due_date = item_payload.get("due_date")
                        competence_date_value = settlement_date_value
                        due_date_value = settlement_date_value
                        if raw_competence_date:
                            try:
                                competence_date_value = date.fromisoformat(str(raw_competence_date))
                            except ValueError:
                                competence_date_value = settlement_date_value
                        if raw_due_date:
                            try:
                                due_date_value = date.fromisoformat(str(raw_due_date))
                            except ValueError:
                                due_date_value = settlement_date_value

                        if component_key == "principal":
                            if _date_in_bucket(settlement_date_value, "liquidacao"):
                                _push_item(
                                    bucket_key="liquidacao",
                                    account_id=allocated_chart_account_id,
                                    amount=signed_amount,
                                    source_kind="settlement",
                                    source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                                    component_key=component_key,
                                    description=title_description,
                                    counterparty=title_counterparty,
                                    competence_date_value=competence_date_value,
                                    due_date_value=due_date_value,
                                    settlement_date_value=settlement_date_value,
                                    entry=entry,
                                    schedule=schedule,
                                    settlement=settlement,
                                    cost_center_id=allocated_cost_center_id,
                                )
                                flags["liquidacao"] = True
                            continue

                        if _date_in_bucket(competence_date_value, "competencia"):
                            _push_item(
                                bucket_key="competencia",
                                account_id=allocated_chart_account_id,
                                amount=signed_amount,
                                source_kind="settlement",
                                source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                                component_key=component_key,
                                description=title_description,
                                counterparty=title_counterparty,
                                competence_date_value=competence_date_value,
                                due_date_value=due_date_value,
                                settlement_date_value=settlement_date_value,
                                entry=entry,
                                schedule=schedule,
                                settlement=settlement,
                                cost_center_id=allocated_cost_center_id,
                            )
                            flags["competencia"] = True
                        if _date_in_bucket(due_date_value, "vencimento"):
                            _push_item(
                                bucket_key="vencimento",
                                account_id=allocated_chart_account_id,
                                amount=signed_amount,
                                source_kind="settlement",
                                source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                                component_key=component_key,
                                description=title_description,
                                counterparty=title_counterparty,
                                competence_date_value=competence_date_value,
                                due_date_value=due_date_value,
                                settlement_date_value=settlement_date_value,
                                entry=entry,
                                schedule=schedule,
                                settlement=settlement,
                                cost_center_id=allocated_cost_center_id,
                            )
                            flags["vencimento"] = True
                        if _date_in_bucket(settlement_date_value, "liquidacao"):
                            _push_item(
                                bucket_key="liquidacao",
                                account_id=allocated_chart_account_id,
                                amount=signed_amount,
                                source_kind="settlement",
                                source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                                component_key=component_key,
                                description=title_description,
                                counterparty=title_counterparty,
                                competence_date_value=competence_date_value,
                                due_date_value=due_date_value,
                                settlement_date_value=settlement_date_value,
                                entry=entry,
                                schedule=schedule,
                                settlement=settlement,
                                cost_center_id=allocated_cost_center_id,
                            )
                            flags["liquidacao"] = True
                    continue

                amount = abs(Decimal(str(component_summary.get(component_key) or fallback_amount or 0)))
                if amount == Decimal("0"):
                    continue
                signed_amount = amount * movement_multiplier * component_multiplier
                if component_key == "principal":
                    if _date_in_bucket(settlement_date_value, "liquidacao"):
                        _push_item(
                            bucket_key="liquidacao",
                            account_id=fallback_chart_account_id,
                            amount=signed_amount,
                            source_kind="settlement",
                            source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                            component_key=component_key,
                            description=title_description,
                            counterparty=title_counterparty,
                            competence_date_value=settlement_date_value,
                            due_date_value=settlement_date_value,
                            settlement_date_value=settlement_date_value,
                            entry=entry,
                            schedule=schedule,
                            settlement=settlement,
                            cost_center_id=fallback_cost_center_id,
                        )
                        flags["liquidacao"] = True
                else:
                    if _date_in_bucket(settlement_date_value, "competencia"):
                        _push_item(
                            bucket_key="competencia",
                            account_id=fallback_chart_account_id,
                            amount=signed_amount,
                            source_kind="settlement",
                            source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                            component_key=component_key,
                            description=title_description,
                            counterparty=title_counterparty,
                            competence_date_value=settlement_date_value,
                            due_date_value=settlement_date_value,
                            settlement_date_value=settlement_date_value,
                            entry=entry,
                            schedule=schedule,
                            settlement=settlement,
                            cost_center_id=fallback_cost_center_id,
                        )
                        flags["competencia"] = True
                    if _date_in_bucket(settlement_date_value, "vencimento"):
                        _push_item(
                            bucket_key="vencimento",
                            account_id=fallback_chart_account_id,
                            amount=signed_amount,
                            source_kind="settlement",
                            source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                            component_key=component_key,
                            description=title_description,
                            counterparty=title_counterparty,
                            competence_date_value=settlement_date_value,
                            due_date_value=settlement_date_value,
                            settlement_date_value=settlement_date_value,
                            entry=entry,
                            schedule=schedule,
                            settlement=settlement,
                            cost_center_id=fallback_cost_center_id,
                        )
                        flags["vencimento"] = True
                    if _date_in_bucket(settlement_date_value, "liquidacao"):
                        _push_item(
                            bucket_key="liquidacao",
                            account_id=fallback_chart_account_id,
                            amount=signed_amount,
                            source_kind="settlement",
                            source_code=getattr(settlement, "settlement_code", None) or f"Baixa {settlement.id}",
                            component_key=component_key,
                            description=title_description,
                            counterparty=title_counterparty,
                            competence_date_value=settlement_date_value,
                            due_date_value=settlement_date_value,
                            settlement_date_value=settlement_date_value,
                            entry=entry,
                            schedule=schedule,
                            settlement=settlement,
                            cost_center_id=fallback_cost_center_id,
                        )
                        flags["liquidacao"] = True

            return flags

        if consolidated_by_period:
            schedule_query = FinancialSchedule.query.filter(
                FinancialSchedule.company_id == company_id,
                FinancialSchedule.deleted_at.is_(None),
            )
            if chart_account_ids:
                schedule_query = schedule_query.filter(FinancialSchedule.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                schedule_query = schedule_query.filter(FinancialSchedule.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                schedule_query = schedule_query.filter(FinancialSchedule.entry_type.in_(allowed_entry_types))

            schedules = schedule_query.order_by(FinancialSchedule.chart_account_id.asc(), FinancialSchedule.competence_date.asc(), FinancialSchedule.id.asc()).all()
            if normalized_filters.project_ids:
                schedules = [item for item in schedules if FinancialReportService._schedule_matches_projects(item, normalized_filters.project_ids)]

            schedule_ids = [item.id for item in schedules]
            schedule_refs = {f"financial_schedule:{item.id}": item.id for item in schedules}
            linked_entries: List[FinancialEntry] = []
            if schedule_ids:
                linked_entries = FinancialEntry.query.filter(
                    FinancialEntry.company_id == company_id,
                    FinancialEntry.deleted_at.is_(None),
                    or_(
                        FinancialEntry.financial_schedule_id.in_(schedule_ids),
                        FinancialEntry.external_reference.in_(list(schedule_refs.keys())),
                    ),
                ).all()

            entries_by_schedule: Dict[int, List[FinancialEntry]] = {item.id: [] for item in schedules}
            linked_entry_ids: List[int] = []
            for entry in linked_entries:
                schedule_id = getattr(entry, "financial_schedule_id", None) or schedule_refs.get(entry.external_reference)
                if schedule_id in entries_by_schedule:
                    entries_by_schedule[schedule_id].append(entry)
                    linked_entry_ids.append(entry.id)

            settlement_totals_by_entry: Dict[int, Decimal] = {}
            settlement_items_by_entry: Dict[int, List[FinancialSettlement]] = {}
            if linked_entry_ids:
                for settlement in FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(linked_entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                ).all():
                    settlement_principal = Decimal(settlement.principal_amount or 0)
                    settlement_amount = Decimal(settlement.net_amount or 0)
                    settlement_totals_by_entry.setdefault(settlement.financial_entry_id, Decimal("0"))
                    settlement_totals_by_entry[settlement.financial_entry_id] += settlement_principal
                    if settlement_amount != Decimal("0"):
                        settlement_items_by_entry.setdefault(settlement.financial_entry_id, []).append(settlement)

            for schedule in schedules:
                schedule_entries = entries_by_schedule.get(schedule.id, [])
                competence_date_value = schedule.competence_date or schedule.start_date
                due_date_value = schedule.next_due_date or schedule.first_due_date or schedule.start_date
                title_amount = Decimal(str(schedule.template_amount or 0))
                if title_amount <= Decimal("0") and schedule_entries:
                    title_amount = sum((_resolve_entry_principal_amount(entry, schedule=schedule) for entry in schedule_entries), Decimal("0"))
                settled_total = sum((settlement_totals_by_entry.get(entry.id, Decimal("0")) for entry in schedule_entries), Decimal("0"))
                derived_settlement_state = "open"
                if title_amount > Decimal("0") and settled_total >= title_amount:
                    derived_settlement_state = "settled"
                elif settled_total > Decimal("0"):
                    derived_settlement_state = "partial"
                operational_state = build_title_operational_state_metadata(
                    schedule_status=schedule.status,
                    settlement_state=derived_settlement_state,
                    entry_type=schedule.entry_type,
                    metadata_json=schedule.metadata_json,
                )
                passes_status, is_open = _passes_financial_status(title_amount, settled_total, operational_state["code"])
                if not passes_status:
                    continue

                in_competence = bool(competence_date_value and period_start <= competence_date_value <= period_end)
                in_due = bool(due_date_value and period_start <= due_date_value <= period_end)
                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for entry in schedule_entries:
                    for settlement in settlement_items_by_entry.get(entry.id, []):
                        current_flags = _push_settlement_breakdown(
                            settlement=settlement,
                            entry=entry,
                            schedule=schedule,
                            fallback_chart_account_id=schedule.chart_account_id,
                            fallback_cost_center_id=schedule.cost_center_id,
                        )
                        settlement_flags = {key: settlement_flags[key] or current_flags[key] for key in settlement_flags}

                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue

                signed_title = Decimal(str(FinancialService.get_signed_amount(title_amount, schedule.movement_nature)))
                if in_competence:
                    _push_item(
                        bucket_key="competencia",
                        account_id=schedule.chart_account_id,
                        amount=signed_title,
                        source_kind="title",
                        source_code=_title_reference(schedule, None),
                        component_key="principal",
                        description=_title_description(schedule, None),
                        counterparty=_title_counterparty(schedule, None),
                        competence_date_value=competence_date_value,
                        due_date_value=due_date_value,
                        settlement_date_value=None,
                        schedule=schedule,
                        cost_center_id=schedule.cost_center_id,
                    )
                if in_due:
                    _push_item(
                        bucket_key="vencimento",
                        account_id=schedule.chart_account_id,
                        amount=signed_title,
                        source_kind="title",
                        source_code=_title_reference(schedule, None),
                        component_key="principal",
                        description=_title_description(schedule, None),
                        counterparty=_title_counterparty(schedule, None),
                        competence_date_value=competence_date_value,
                        due_date_value=due_date_value,
                        settlement_date_value=None,
                        schedule=schedule,
                        cost_center_id=schedule.cost_center_id,
                    )
                if is_open:
                    _push_item(
                        bucket_key="aberto",
                        account_id=schedule.chart_account_id,
                        amount=signed_title,
                        source_kind="title",
                        source_code=_title_reference(schedule, None),
                        component_key="principal",
                        description=_title_description(schedule, None),
                        counterparty=_title_counterparty(schedule, None),
                        competence_date_value=competence_date_value,
                        due_date_value=due_date_value,
                        settlement_date_value=None,
                        schedule=schedule,
                        cost_center_id=schedule.cost_center_id,
                    )
                else:
                    last_settlement_date = None
                    dated_settlements = [settlement for entry in schedule_entries for settlement in settlement_items_by_entry.get(entry.id, [])]
                    if dated_settlements:
                        dated_settlements.sort(key=lambda item: (getattr(item, "settlement_date", None) or date.max, getattr(item, "id", 0)))
                        last_settlement_date = getattr(dated_settlements[-1], "settlement_date", None)
                    _push_item(
                        bucket_key="baixado",
                        account_id=schedule.chart_account_id,
                        amount=signed_title,
                        source_kind="title",
                        source_code=_title_reference(schedule, None),
                        component_key="principal",
                        description=_title_description(schedule, None),
                        counterparty=_title_counterparty(schedule, None),
                        competence_date_value=competence_date_value,
                        due_date_value=due_date_value,
                        settlement_date_value=last_settlement_date,
                        schedule=schedule,
                        cost_center_id=schedule.cost_center_id,
                    )

            manual_query = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
            )
            if chart_account_ids:
                manual_query = manual_query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                manual_query = manual_query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                manual_query = manual_query.filter(FinancialEntry.entry_type.in_(allowed_entry_types))
            manual_entries = [
                entry
                for entry in manual_query.order_by(FinancialEntry.chart_account_id.asc(), FinancialEntry.competence_date.asc(), FinancialEntry.id.asc()).all()
                if not getattr(entry, "financial_schedule_id", None) and not str(entry.external_reference or "").startswith("financial_schedule:")
            ]
            if normalized_filters.project_ids:
                manual_entries = [entry for entry in manual_entries if FinancialReportService._entry_matches_projects(entry, normalized_filters.project_ids)]

            manual_entry_ids = [entry.id for entry in manual_entries]
            manual_settlement_totals: Dict[int, Decimal] = {}
            manual_settlement_items: Dict[int, List[FinancialSettlement]] = {}
            if manual_entry_ids:
                for settlement in FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(manual_entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                ).all():
                    settlement_principal = Decimal(settlement.principal_amount or 0)
                    settlement_amount = Decimal(settlement.net_amount or 0)
                    manual_settlement_totals.setdefault(settlement.financial_entry_id, Decimal("0"))
                    manual_settlement_totals[settlement.financial_entry_id] += settlement_principal
                    if settlement_amount != Decimal("0"):
                        manual_settlement_items.setdefault(settlement.financial_entry_id, []).append(settlement)

            for entry in manual_entries:
                original_amount = Decimal(entry.original_amount or 0)
                total_settlement_amount = manual_settlement_totals.get(entry.id, Decimal("0"))
                passes_status, is_open = _passes_financial_status(original_amount, total_settlement_amount, entry.status)
                if not passes_status:
                    continue
                in_competence = bool(entry.competence_date and period_start <= entry.competence_date <= period_end)
                in_due = bool(entry.due_date and period_start <= entry.due_date <= period_end)
                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for settlement in manual_settlement_items.get(entry.id, []):
                    current_flags = _push_settlement_breakdown(
                        settlement=settlement,
                        entry=entry,
                        schedule=None,
                        fallback_chart_account_id=entry.chart_account_id,
                        fallback_cost_center_id=entry.cost_center_id,
                    )
                    settlement_flags = {key: settlement_flags[key] or current_flags[key] for key in settlement_flags}
                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue

                signed_original = original_amount if entry.movement_nature == "credit" else -original_amount
                if in_competence:
                    _push_item(
                        bucket_key="competencia",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(None, entry),
                        component_key="principal",
                        description=_title_description(None, entry),
                        counterparty=_title_counterparty(None, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        cost_center_id=entry.cost_center_id,
                    )
                if in_due:
                    _push_item(
                        bucket_key="vencimento",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(None, entry),
                        component_key="principal",
                        description=_title_description(None, entry),
                        counterparty=_title_counterparty(None, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        cost_center_id=entry.cost_center_id,
                    )
                if is_open:
                    _push_item(
                        bucket_key="aberto",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(None, entry),
                        component_key="principal",
                        description=_title_description(None, entry),
                        counterparty=_title_counterparty(None, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        cost_center_id=entry.cost_center_id,
                    )
                else:
                    last_settlement_date = None
                    entry_settlements = list(manual_settlement_items.get(entry.id, []))
                    if entry_settlements:
                        entry_settlements.sort(key=lambda item: (getattr(item, "settlement_date", None) or date.max, getattr(item, "id", 0)))
                        last_settlement_date = getattr(entry_settlements[-1], "settlement_date", None)
                    _push_item(
                        bucket_key="baixado",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(None, entry),
                        component_key="principal",
                        description=_title_description(None, entry),
                        counterparty=_title_counterparty(None, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=last_settlement_date,
                        entry=entry,
                        cost_center_id=entry.cost_center_id,
                    )
        else:
            query = FinancialEntry.query.filter(
                FinancialEntry.company_id == company_id,
                FinancialEntry.deleted_at.is_(None),
            )
            if chart_account_ids:
                query = query.filter(FinancialEntry.chart_account_id.in_(chart_account_ids))
            if cost_center_ids:
                query = query.filter(FinancialEntry.cost_center_id.in_(cost_center_ids))
            if allowed_entry_types:
                query = query.filter(FinancialEntry.entry_type.in_(allowed_entry_types))

            entries = query.order_by(FinancialEntry.chart_account_id.asc(), FinancialEntry.competence_date.asc(), FinancialEntry.id.asc()).all()
            if normalized_filters.project_ids:
                entries = [entry for entry in entries if FinancialReportService._entry_matches_projects(entry, normalized_filters.project_ids)]

            entry_ids = [entry.id for entry in entries]
            settlement_totals_by_entry: Dict[int, Decimal] = {}
            settlement_items_by_entry: Dict[int, List[FinancialSettlement]] = {}
            if entry_ids:
                settlements = FinancialSettlement.query.filter(
                    FinancialSettlement.company_id == company_id,
                    FinancialSettlement.financial_entry_id.in_(entry_ids),
                    FinancialSettlement.deleted_at.is_(None),
                    FinancialSettlement.settlement_status != "cancelled",
                ).all()
                for settlement in settlements:
                    settlement_totals_by_entry.setdefault(settlement.financial_entry_id, Decimal("0"))
                    settlement_totals_by_entry[settlement.financial_entry_id] += Decimal(settlement.principal_amount or 0)
                    settlement_items_by_entry.setdefault(settlement.financial_entry_id, []).append(settlement)

            schedule_by_entry_id: Dict[int, FinancialSchedule] = {}
            schedule_ids = {
                int(getattr(entry, "financial_schedule_id", 0) or 0)
                for entry in entries
                if getattr(entry, "financial_schedule_id", None)
            }
            if schedule_ids:
                for schedule in FinancialSchedule.query.filter(
                    FinancialSchedule.company_id == company_id,
                    FinancialSchedule.id.in_(list(schedule_ids)),
                    FinancialSchedule.deleted_at.is_(None),
                ).all():
                    schedule_by_entry_id[int(getattr(schedule, "id", 0) or 0)] = schedule

            for entry in entries:
                linked_schedule = schedule_by_entry_id.get(int(getattr(entry, "financial_schedule_id", 0) or 0))
                total_settlement_amount = settlement_totals_by_entry.get(entry.id, Decimal("0"))
                original_amount = _resolve_entry_principal_amount(entry, schedule=linked_schedule)
                passes_status, is_open = _passes_financial_status(original_amount, total_settlement_amount, entry.status)
                if not passes_status:
                    continue

                in_competence = True
                if competence_start and ((not entry.competence_date) or entry.competence_date < competence_start):
                    in_competence = False
                if competence_end and ((not entry.competence_date) or entry.competence_date > competence_end):
                    in_competence = False

                in_due = True
                if due_start and ((not entry.due_date) or entry.due_date < due_start):
                    in_due = False
                if due_end and ((not entry.due_date) or entry.due_date > due_end):
                    in_due = False

                settlement_flags = {"competencia": False, "vencimento": False, "liquidacao": False}
                for settlement in settlement_items_by_entry.get(entry.id, []):
                    current_flags = _push_settlement_breakdown(
                        settlement=settlement,
                        entry=entry,
                        schedule=linked_schedule,
                        fallback_chart_account_id=(linked_schedule.chart_account_id if linked_schedule is not None else entry.chart_account_id),
                        fallback_cost_center_id=(linked_schedule.cost_center_id if linked_schedule is not None else entry.cost_center_id),
                    )
                    settlement_flags = {key: settlement_flags[key] or current_flags[key] for key in settlement_flags}

                in_liquidation = settlement_flags["liquidacao"]
                if not any([in_competence, in_due, settlement_flags["competencia"], settlement_flags["vencimento"], in_liquidation]):
                    continue

                signed_original = original_amount if entry.movement_nature == "credit" else -original_amount
                if in_competence:
                    _push_item(
                        bucket_key="competencia",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(linked_schedule, entry),
                        component_key="principal",
                        description=_title_description(linked_schedule, entry),
                        counterparty=_title_counterparty(linked_schedule, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        schedule=linked_schedule,
                        cost_center_id=entry.cost_center_id,
                    )
                if in_due:
                    _push_item(
                        bucket_key="vencimento",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(linked_schedule, entry),
                        component_key="principal",
                        description=_title_description(linked_schedule, entry),
                        counterparty=_title_counterparty(linked_schedule, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        schedule=linked_schedule,
                        cost_center_id=entry.cost_center_id,
                    )
                if is_open:
                    _push_item(
                        bucket_key="aberto",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(linked_schedule, entry),
                        component_key="principal",
                        description=_title_description(linked_schedule, entry),
                        counterparty=_title_counterparty(linked_schedule, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=None,
                        entry=entry,
                        schedule=linked_schedule,
                        cost_center_id=entry.cost_center_id,
                    )
                else:
                    last_settlement_date = None
                    entry_settlements = list(settlement_items_by_entry.get(entry.id, []))
                    if entry_settlements:
                        entry_settlements.sort(key=lambda item: (getattr(item, "settlement_date", None) or date.max, getattr(item, "id", 0)))
                        last_settlement_date = getattr(entry_settlements[-1], "settlement_date", None)
                    _push_item(
                        bucket_key="baixado",
                        account_id=entry.chart_account_id,
                        amount=signed_original,
                        source_kind="title",
                        source_code=_title_reference(linked_schedule, entry),
                        component_key="principal",
                        description=_title_description(linked_schedule, entry),
                        counterparty=_title_counterparty(linked_schedule, entry),
                        competence_date_value=entry.competence_date,
                        due_date_value=entry.due_date,
                        settlement_date_value=last_settlement_date,
                        entry=entry,
                        schedule=linked_schedule,
                        cost_center_id=entry.cost_center_id,
                    )

        def _descendant_ids(root_account_id: int) -> List[int]:
            ordered: List[int] = []
            queue = [root_account_id]
            seen = set()
            while queue:
                current = queue.pop(0)
                if current in seen:
                    continue
                seen.add(current)
                ordered.append(current)
                queue.extend(chart_children_map.get(current, []))
            return ordered

        requested_bucket = bucket_meta["key"]
        scoped_items: List[Dict[str, Any]] = []
        account_scope_ids = _descendant_ids(chart_account_id) if chart_account_id is not None else list((detail_items_by_bucket.get(requested_bucket) or {}).keys())
        for scoped_account_id in account_scope_ids:
            scoped_items.extend(list((detail_items_by_bucket.get(requested_bucket) or {}).get(scoped_account_id, [])))

        scoped_items.sort(
            key=lambda item: (
                item.get("relevant_date") or "9999-12-31",
                item.get("account_label") or "",
                item.get("source_code") or "",
                item.get("component_label") or "",
                item.get("description") or "",
            )
        )

        total_amount = sum((Decimal(str(item.get("amount") or 0)) for item in scoped_items), Decimal("0"))
        account_label = "Total consolidado"
        if chart_account_id is not None:
            account_label = _account_label(chart_account_id)

        return {
            "report_type": definition["code"],
            "report_slug": definition["slug"],
            "bucket": requested_bucket,
            "bucket_label": bucket_meta["label"],
            "source_label": bucket_meta["source_label"],
            "chart_account_id": chart_account_id,
            "account_label": account_label,
            "item_count": len(scoped_items),
            "total": FinancialReportService._serialize_money(total_amount),
            "total_label": FinancialReportService._format_currency(total_amount),
            "items": scoped_items,
        }, None

    @staticmethod
    def generate_report(*, company_id: int, report_type: str, period_start: str, period_end: str, allowed_company_ids: Optional[Sequence[int]] = None) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        return FinancialReportService.build_management_report(
            company_id=company_id,
            report_type=report_type,
            filters={"period_start": period_start, "period_end": period_end},
            allowed_company_ids=allowed_company_ids,
        )

    @staticmethod
    def export_xlsx(report_payload: Dict[str, Any]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        if report_payload.get("report_type") == "working_capital":
            return FinancialReportService._export_working_capital_xlsx(
                report_payload,
                workbook_factory=Workbook,
                alignment_cls=Alignment,
                border_cls=Border,
                font_cls=Font,
                fill_cls=PatternFill,
                side_cls=Side,
                get_column_letter_fn=get_column_letter,
            )

        if report_payload.get("report_type") == "cash_flow":
            return FinancialReportService._export_cash_flow_xlsx(
                report_payload,
                workbook_factory=Workbook,
                alignment_cls=Alignment,
                border_cls=Border,
                font_cls=Font,
                fill_cls=PatternFill,
                side_cls=Side,
                get_column_letter_fn=get_column_letter,
            )

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Filtros e resumo"
        summary_sheet["A1"] = report_payload.get("title", "Relatório")
        summary_sheet["A1"].font = Font(bold=True, size=14)
        summary_sheet["A3"] = "Gerado em"
        summary_sheet["B3"] = report_payload.get("generated_at")

        row_cursor = 5
        for section_title, items in [
            ("Filtros", report_payload.get("filters", [])),
            ("Informações gerais", report_payload.get("general_info", [])),
            ("Resumo executivo", report_payload.get("summary_cards", [])),
        ]:
            summary_sheet[f"A{row_cursor}"] = section_title
            summary_sheet[f"A{row_cursor}"].font = Font(bold=True)
            row_cursor += 1
            for item in items:
                summary_sheet[f"A{row_cursor}"] = item.get("label")
                summary_sheet[f"B{row_cursor}"] = item.get("value")
                row_cursor += 1
            row_cursor += 1

        data_sheet = workbook.create_sheet("Dados")
        columns = report_payload.get("columns", [])
        for index, column in enumerate(columns, start=1):
            cell = data_sheet.cell(row=1, column=index, value=column.get("label"))
            cell.font = Font(bold=True)
        for row_index, item in enumerate(report_payload.get("rows", []), start=2):
            for col_index, column in enumerate(columns, start=1):
                data_sheet.cell(row=row_index, column=col_index, value=item.get(column.get("key"), ""))

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _build_pdf_document_bytes(*, elements: List[Any], pagesize, report_payload: Dict[str, Any]) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=24, rightMargin=24, topMargin=26, bottomMargin=36)
        doc.build(
            elements,
            onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
        )
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _merge_pdf_bytes(parts: Sequence[bytes]) -> bytes:
        from pypdf import PdfReader, PdfWriter

        writer = PdfWriter()
        for part in parts:
            if not part:
                continue
            reader = PdfReader(io.BytesIO(part))
            for page in reader.pages:
                writer.add_page(page)
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _export_bank_statement_dossier_pdf(report_payload: Dict[str, Any]) -> bytes:
        styles = getSampleStyleSheet()

        portrait_doc = SimpleDocTemplate(io.BytesIO(), pagesize=A4, leftMargin=24, rightMargin=24, topMargin=26, bottomMargin=36)
        portrait_width = A4[0] - portrait_doc.leftMargin - portrait_doc.rightMargin

        statement_payload = {
            **report_payload,
            "report_type": "bank_statement",
            "orientation": "portrait",
            "title": report_payload.get("statement_title") or "Extrato Bancário",
            "subtitle": report_payload.get("statement_subtitle") or FinancialReportService.REPORT_DEFINITIONS["bank_statement"]["description"],
        }
        portrait_elements: List[Any] = []
        portrait_elements.extend(
            FinancialReportService._build_bank_statement_pdf_elements(
                report_payload=statement_payload,
                styles=styles,
                available_width=portrait_width,
            )
        )
        portrait_elements.append(PageBreak())
        income_statement_payload = {
            **(report_payload.get("dossier_income_statement") or {}),
            "company_name": report_payload.get("company_name") or (report_payload.get("dossier_income_statement") or {}).get("company_name"),
        }
        portrait_elements.extend(
            FinancialReportService._build_income_statement_liquidation_pdf_elements(
                report_payload=income_statement_payload,
                styles=styles,
                available_width=portrait_width,
            )
        )

        portrait_pdf = FinancialReportService._build_pdf_document_bytes(
            elements=portrait_elements,
            pagesize=A4,
            report_payload={**report_payload, "orientation": "portrait"},
        )

        landscape_size = landscape(A4)
        landscape_doc = SimpleDocTemplate(io.BytesIO(), pagesize=landscape_size, leftMargin=20, rightMargin=20, topMargin=22, bottomMargin=32)
        landscape_width = landscape_size[0] - landscape_doc.leftMargin - landscape_doc.rightMargin
        landscape_height = landscape_size[1] - landscape_doc.topMargin - landscape_doc.bottomMargin
        documents_pdf = FinancialReportService._build_pdf_document_bytes(
            elements=FinancialReportService._build_bank_statement_dossier_documents_pages(
                report_payload=report_payload,
                styles=styles,
                available_width=landscape_width,
                available_height=landscape_height,
            ),
            pagesize=landscape_size,
            report_payload={**report_payload, "orientation": "landscape"},
        )

        return FinancialReportService._merge_pdf_bytes([portrait_pdf, documents_pdf])

    @staticmethod
    def _pdf_text(value: Any) -> str:
        return (
            str(value if value is not None else "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    @staticmethod
    def _ledger_amount_color(value: Any) -> colors.Color:
        amount = FinancialReportService._parse_currency_label(value)
        if amount > 0:
            return colors.HexColor("#2563eb")
        if amount < 0:
            return colors.HexColor("#dc2626")
        return colors.HexColor("#334155")

    @staticmethod
    def _build_ledger_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "LedgerTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=17,
            leading=20,
            textColor=colors.white,
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "LedgerSubtitle",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#dbeafe"),
        )
        section_label_style = ParagraphStyle(
            "LedgerSectionLabel",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=6.2,
            leading=7.2,
            textColor=colors.HexColor("#2563eb"),
            uppercase=True,
        )
        card_value_style = ParagraphStyle(
            "LedgerCardValue",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.2,
            leading=10,
            textColor=colors.HexColor("#0f172a"),
        )
        group_title_style = ParagraphStyle(
            "LedgerGroupTitle",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=9.8,
            leading=12,
            textColor=colors.white,
        )
        group_meta_style = ParagraphStyle(
            "LedgerGroupMeta",
            parent=styles["BodyText"],
            fontSize=7.2,
            leading=9,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#dbeafe"),
        )
        cell_style = ParagraphStyle(
            "LedgerCell",
            parent=styles["BodyText"],
            fontSize=6.1,
            leading=7.4,
            textColor=colors.HexColor("#334155"),
        )
        small_cell_style = ParagraphStyle(
            "LedgerSmallCell",
            parent=cell_style,
            fontSize=5.6,
            leading=6.8,
        )
        money_style = ParagraphStyle(
            "LedgerMoney",
            parent=cell_style,
            fontName="Helvetica-Bold",
            alignment=TA_RIGHT,
        )
        header_style = ParagraphStyle(
            "LedgerHeaderCell",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=5.8,
            leading=7,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0f172a"),
        )

        def p(value: Any, style: ParagraphStyle = cell_style) -> Paragraph:
            return Paragraph(FinancialReportService._pdf_text(value), style)

        def section_cards(items: Sequence[Dict[str, Any]], *, columns: int = 4) -> Optional[Table]:
            normalized = list(items or [])
            if not normalized:
                return None
            rows: List[List[Any]] = []
            for start in range(0, len(normalized), columns):
                current = normalized[start:start + columns]
                row = []
                for item in current:
                    row.append([
                        p(str(item.get("label", "")).upper(), section_label_style),
                        p(item.get("value", ""), card_value_style),
                    ])
                while len(row) < columns:
                    row.append("")
                rows.append(row)
            table = Table(rows, colWidths=[available_width / columns] * columns, hAlign="LEFT")
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#bfdbfe")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#dbeafe")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            return table

        elements: List[Any] = []
        hero_data = [[
            [p("GESTÃO FINANCEIRA · RAZÃO ANALÍTICO", subtitle_style), p(report_payload.get("title", "Razão"), title_style)],
            p(
                f"{report_payload.get('company_name') or 'Empresa ativa'} · {report_payload.get('subtitle') or ''}",
                subtitle_style,
            ),
        ]]
        hero = Table(hero_data, colWidths=[available_width * 0.62, available_width * 0.38])
        hero.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0f172a")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 14),
            ("RIGHTPADDING", (0, 0), (-1, -1), 14),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))
        elements.extend([hero, Spacer(1, 8)])

        insight_table = section_cards([
            {"label": "Agrupado por", "value": report_payload.get("grouped_by_label", "-")},
            {"label": "Competência", "value": next((item.get("value") for item in report_payload.get("general_info", []) if item.get("label") == "Competência"), "-")},
            {"label": "Status", "value": next((item.get("value") for item in report_payload.get("filters", []) if item.get("label") == "Status considerados"), "-")},
        ], columns=3)
        if insight_table:
            elements.extend([insight_table, Spacer(1, 7)])
        stats_table = section_cards(report_payload.get("summary_cards", []), columns=4)
        if stats_table:
            elements.extend([stats_table, Spacer(1, 9)])

        table_headers = ["ID", "Histórico", "Favorecido", "Tipo", "Plano de Contas", "Centro", "Projeto/Processo", "Compet.", "Venc.", "Liq.", "Valor", "Nº/Rateio"]
        col_widths = [
            available_width * 0.045,
            available_width * 0.15,
            available_width * 0.125,
            available_width * 0.055,
            available_width * 0.13,
            available_width * 0.095,
            available_width * 0.12,
            available_width * 0.055,
            available_width * 0.055,
            available_width * 0.055,
            available_width * 0.07,
            available_width * 0.045,
        ]

        for group in report_payload.get("groups", []) or []:
            group_header = Table(
                [[
                    [p("AGRUPADOR", subtitle_style), p(group.get("label", "-"), group_title_style)],
                    Paragraph(
                        f"{FinancialReportService._pdf_text(group.get('item_count', 0))} lançamento(s)<br/><b>{FinancialReportService._pdf_text(group.get('total', ''))}</b>",
                        group_meta_style,
                    ),
                ]],
                colWidths=[available_width * 0.72, available_width * 0.28],
            )
            group_header.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#111827")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#111827")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]))

            data: List[List[Any]] = [[p(label, header_style) for label in table_headers]]
            row_styles: List[Tuple[Any, ...]] = []
            for idx, row in enumerate(group.get("rows", []) or [], start=1):
                amount = row.get("valor", "")
                amount_style = ParagraphStyle(
                    f"LedgerMoney{idx}_{abs(hash(str(amount))) % 99999}",
                    parent=money_style,
                    textColor=FinancialReportService._ledger_amount_color(amount),
                )
                type_style = ParagraphStyle(
                    f"LedgerType{idx}",
                    parent=small_cell_style,
                    fontName="Helvetica-Bold",
                    alignment=TA_CENTER,
                    textColor=colors.HexColor("#166534") if row.get("tipo") == "Entrada" else colors.HexColor("#991b1b"),
                )
                data.append([
                    p(row.get("id", ""), small_cell_style),
                    p(row.get("historico", ""), cell_style),
                    p(row.get("favorecido", ""), small_cell_style),
                    p(row.get("tipo", ""), type_style),
                    p(row.get("plano_contas", ""), small_cell_style),
                    p(row.get("centro_resultado", ""), small_cell_style),
                    p(row.get("projeto_processo", ""), small_cell_style),
                    p(row.get("competencia", ""), small_cell_style),
                    p(row.get("vencimento", ""), small_cell_style),
                    p(row.get("liquidacao", ""), small_cell_style),
                    p(amount, amount_style),
                    p(row.get("numero_qtd_rateio", ""), small_cell_style),
                ])
                if idx % 2 == 0:
                    row_styles.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#f8fbff")))

            ledger_table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
            ledger_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#dbeafe")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                *row_styles,
            ]))
            elements.extend([group_header, ledger_table, Spacer(1, 8)])

        if not report_payload.get("groups"):
            empty_style = ParagraphStyle("LedgerEmpty", parent=styles["BodyText"], alignment=TA_CENTER, textColor=colors.HexColor("#64748b"), fontSize=10)
            elements.append(Paragraph("Nenhum lançamento encontrado para os filtros informados.", empty_style))

        return elements

    @staticmethod
    def export_pdf(report_payload: Dict[str, Any]) -> bytes:
        if report_payload.get("report_type") == "bank_statement_dossier":
            return FinancialReportService._export_bank_statement_dossier_pdf(report_payload)

        buffer = io.BytesIO()
        pagesize = landscape(A4) if report_payload.get("orientation", "landscape") == "landscape" else A4
        doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=24, rightMargin=24, topMargin=26, bottomMargin=36)
        styles = getSampleStyleSheet()
        available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
        available_height = pagesize[1] - doc.topMargin - doc.bottomMargin

        if report_payload.get("report_type") == "schedule_report":
            elements = FinancialReportService._build_schedule_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        if report_payload.get("report_type") == "cash_flow":
            elements = FinancialReportService._build_cash_flow_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        if report_payload.get("report_type") == "bank_statement":
            elements = FinancialReportService._build_bank_statement_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        if report_payload.get("report_type") == "working_capital":
            elements = FinancialReportService._build_working_capital_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        if report_payload.get("report_type") in {"income_statement", "income_statement_2"}:
            pagesize = A4
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=22, rightMargin=22, topMargin=24, bottomMargin=36)
            available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
            elements = FinancialReportService._build_income_statement_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        if report_payload.get("report_type") == "ledger":
            pagesize = landscape(A4) if report_payload.get("orientation", "landscape") == "landscape" else A4
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=18, rightMargin=18, topMargin=20, bottomMargin=34)
            available_width = pagesize[0] - doc.leftMargin - doc.rightMargin
            elements = FinancialReportService._build_ledger_pdf_elements(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
            )
            doc.build(
                elements,
                onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
                onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            )
            buffer.seek(0)
            return buffer.getvalue()

        title_style = ParagraphStyle("FinancialReportTitle", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0f172a"), spaceAfter=10)
        subtitle_style = ParagraphStyle("FinancialReportSubtitle", parent=styles["BodyText"], fontSize=9, textColor=colors.HexColor("#475569"), spaceAfter=8)

        elements = [
            Paragraph(report_payload.get("title", "Relatório financeiro"), title_style),
            Paragraph(report_payload.get("subtitle", ""), subtitle_style),
            Paragraph(f"Gerado em: {report_payload.get('generated_at', '-')}", subtitle_style),
            Spacer(1, 8),
        ]

        summary_rows = [["Resumo", "Valor"]] + [[item.get("label", ""), str(item.get("value", ""))] for item in report_payload.get("summary_cards", [])]
        summary_table = Table(summary_rows, colWidths=[180, 220])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
                ]
            )
        )
        elements.extend([summary_table, Spacer(1, 10)])

        columns = [column.get("label", "") for column in report_payload.get("columns", [])]
        rows = [[str(item.get(column.get("key"), "")) for column in report_payload.get("columns", [])] for item in report_payload.get("rows", [])]
        if columns:
            data = [columns] + rows
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8),
                        ("FONTSIZE", (0, 1), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            elements.append(table)

        doc.build(
            elements,
            onFirstPage=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
            onLaterPages=lambda canvas, current_doc: FinancialReportService._draw_default_pdf_footer(canvas, current_doc, report_payload),
        )
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _parse_currency_label(value: Any) -> Decimal:
        raw_value = str(value or "0").strip()
        if not raw_value:
            return Decimal("0")
        negative = "-" in raw_value
        normalized = (
            raw_value.replace("R$ ", "")
            .replace("R$", "")
            .replace("- ", "")
            .replace("-", "")
            .replace(".", "")
            .replace(",", ".")
            .strip()
        )
        try:
            amount = Decimal(normalized or "0")
        except Exception:
            return Decimal("0")
        return -amount if negative else amount

    @staticmethod
    def _cash_flow_plain_amount_label(value: Any, amount_value: Any = None) -> str:
        amount = (
            Decimal(str(amount_value))
            if amount_value is not None and str(amount_value).strip() != ""
            else FinancialReportService._parse_currency_label(value)
        )
        absolute_label = str(FinancialReportService._format_currency(abs(amount))).replace("R$ ", "").replace("R$", "").strip()
        if amount < 0:
            return f"- {absolute_label}"
        return absolute_label

    @staticmethod
    def _cash_flow_amount_color_hex(value: Any, amount_value: Any = None) -> str:
        amount = (
            Decimal(str(amount_value))
            if amount_value is not None and str(amount_value).strip() != ""
            else FinancialReportService._parse_currency_label(value)
        )
        if amount > 0:
            return "#2563eb"
        if amount < 0:
            return "#dc2626"
        return "#334155"

    @staticmethod
    def _cash_flow_numeric_amount(value: Any, amount_value: Any = None) -> float:
        amount = (
            Decimal(str(amount_value))
            if amount_value is not None and str(amount_value).strip() != ""
            else FinancialReportService._parse_currency_label(value)
        )
        return FinancialReportService._serialize_money(amount)

    @staticmethod
    def _export_cash_flow_xlsx(
        report_payload: Dict[str, Any],
        *,
        workbook_factory,
        alignment_cls,
        border_cls,
        font_cls,
        fill_cls,
        side_cls,
        get_column_letter_fn,
    ) -> bytes:
        workbook = workbook_factory()
        hero_fill = fill_cls(fill_type="solid", fgColor="0F172A")
        accent_fill = fill_cls(fill_type="solid", fgColor="DBEAFE")
        header_fill = fill_cls(fill_type="solid", fgColor="1E3A8A")
        section_fill = fill_cls(fill_type="solid", fgColor="EFF6FF")
        white_font = font_cls(color="FFFFFF", bold=True)
        title_font = font_cls(color="FFFFFF", bold=True, size=18)
        subtitle_font = font_cls(color="CBD5E1", size=10)
        heading_font = font_cls(color="0F172A", bold=True, size=11)
        label_font = font_cls(color="475569", bold=True)
        positive_font = font_cls(color="2563EB", bold=True)
        negative_font = font_cls(color="DC2626", bold=True)
        neutral_font = font_cls(color="334155", bold=True)
        thin_border = border_cls(
            left=side_cls(style="thin", color="CBD5E1"),
            right=side_cls(style="thin", color="CBD5E1"),
            top=side_cls(style="thin", color="CBD5E1"),
            bottom=side_cls(style="thin", color="CBD5E1"),
        )
        currency_format = '#,##0.00;[Red]- #,##0.00'

        def _style_amount(cell, raw_value: Any, amount_value: Any = None):
            amount = FinancialReportService._cash_flow_numeric_amount(raw_value, amount_value)
            cell.value = amount
            cell.number_format = currency_format
            cell.alignment = alignment_cls(horizontal="right", vertical="center")
            if amount > 0:
                cell.font = positive_font
            elif amount < 0:
                cell.font = negative_font
            else:
                cell.font = neutral_font
            cell.border = thin_border

        def _write_section_title(sheet, row: int, title: str, end_col: int = 6):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
            cell = sheet.cell(row=row, column=1, value=title)
            cell.fill = section_fill
            cell.font = heading_font
            cell.alignment = alignment_cls(horizontal="left", vertical="center")
            cell.border = thin_border
            for column in range(2, end_col + 1):
                sheet.cell(row=row, column=column).border = thin_border

        summary_sheet = workbook.active
        summary_sheet.title = "Resumo Executivo"
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet.freeze_panes = "A8"
        summary_sheet.merge_cells("A1:F2")
        summary_sheet["A1"] = report_payload.get("title", "Fluxo de Caixa")
        summary_sheet["A1"].fill = hero_fill
        summary_sheet["A1"].font = title_font
        summary_sheet["A1"].alignment = alignment_cls(horizontal="left", vertical="center")
        summary_sheet.merge_cells("A3:F3")
        summary_sheet["A3"] = str(report_payload.get("company_name") or "Versus Gestão Corporativa")
        summary_sheet["A3"].fill = hero_fill
        summary_sheet["A3"].font = subtitle_font
        summary_sheet["A3"].alignment = alignment_cls(horizontal="left", vertical="center")
        summary_sheet.merge_cells("A4:F4")
        summary_sheet["A4"] = f"Emitido em {FinancialReportService._pdf_generated_at_label(report_payload)}"
        summary_sheet["A4"].fill = hero_fill
        summary_sheet["A4"].font = subtitle_font
        summary_sheet["A4"].alignment = alignment_cls(horizontal="left", vertical="center")

        _write_section_title(summary_sheet, 6, "Filtros aplicados", 6)
        row_cursor = 7
        for item in report_payload.get("filters", []) or [{"label": "Resumo", "value": "Sem filtros adicionais."}]:
            summary_sheet.cell(row=row_cursor, column=1, value=item.get("label")).font = label_font
            summary_sheet.cell(row=row_cursor, column=2, value=item.get("value"))
            summary_sheet.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=6)
            for column in range(1, 7):
                current_cell = summary_sheet.cell(row=row_cursor, column=column)
                current_cell.border = thin_border
                current_cell.alignment = alignment_cls(vertical="center", wrap_text=True)
            row_cursor += 1

        row_cursor += 1
        _write_section_title(summary_sheet, row_cursor, "Indicadores executivos", 6)
        row_cursor += 1
        for card in report_payload.get("summary_cards", []):
            summary_sheet.cell(row=row_cursor, column=1, value=card.get("label")).font = label_font
            amount_cell = summary_sheet.cell(row=row_cursor, column=2)
            amount_cell.value = card.get("value")
            amount_cell.alignment = alignment_cls(horizontal="left", vertical="center")
            tone = str(card.get("tone") or "neutral").lower()
            if tone == "positive":
                amount_cell.font = positive_font
            elif tone == "negative":
                amount_cell.font = negative_font
            else:
                amount_cell.font = neutral_font
            summary_sheet.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=3)
            for column in range(1, 7):
                summary_sheet.cell(row=row_cursor, column=column).border = thin_border
            row_cursor += 1

        flow_sheet = workbook.create_sheet("Fluxo Consolidado")
        flow_sheet.sheet_view.showGridLines = False
        flow_sheet.freeze_panes = "A9"
        flow_sheet.merge_cells("A1:I2")
        flow_sheet["A1"] = report_payload.get("title", "Fluxo de Caixa")
        flow_sheet["A1"].fill = hero_fill
        flow_sheet["A1"].font = title_font
        flow_sheet["A1"].alignment = alignment_cls(horizontal="left", vertical="center")
        flow_sheet.merge_cells("A3:I3")
        flow_sheet["A3"] = str(report_payload.get("subtitle") or "")
        flow_sheet["A3"].fill = hero_fill
        flow_sheet["A3"].font = subtitle_font
        flow_sheet["A3"].alignment = alignment_cls(horizontal="left", vertical="center")

        _write_section_title(flow_sheet, 5, "Contas correntes", 4)
        account_headers = ["Descrição", "Limite", f"Saldo em {report_payload.get('bank_balance_reference_label', '-')}", "Disp. Total"]
        for col_index, header in enumerate(account_headers, start=1):
            cell = flow_sheet.cell(row=6, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = alignment_cls(horizontal="center", vertical="center")
            cell.border = thin_border
        current_row = 7
        for item in report_payload.get("bank_account_summary_rows", []):
            flow_sheet.cell(row=current_row, column=1, value=item.get("description")).border = thin_border
            _style_amount(flow_sheet.cell(row=current_row, column=2), item.get("limit"), item.get("limit_value"))
            _style_amount(flow_sheet.cell(row=current_row, column=3), item.get("balance"), item.get("balance_value"))
            _style_amount(flow_sheet.cell(row=current_row, column=4), item.get("available_total"), item.get("available_total_value"))
            current_row += 1
        totals = report_payload.get("bank_account_summary_totals", {}) or {}
        total_row = current_row
        flow_sheet.cell(row=total_row, column=1, value="Total").font = label_font
        flow_sheet.cell(row=total_row, column=1).fill = accent_fill
        flow_sheet.cell(row=total_row, column=1).border = thin_border
        _style_amount(flow_sheet.cell(row=total_row, column=2), totals.get("limit"), totals.get("limit_value"))
        _style_amount(flow_sheet.cell(row=total_row, column=3), totals.get("balance"), totals.get("balance_value"))
        _style_amount(flow_sheet.cell(row=total_row, column=4), totals.get("available_total"), totals.get("available_total_value"))
        for column in range(2, 5):
            flow_sheet.cell(row=total_row, column=column).fill = accent_fill

        current_row += 2
        _write_section_title(flow_sheet, current_row, "Fluxo do período", 9)
        current_row += 1
        columns = report_payload.get("columns", [])
        amount_keys = {"saldo_inicial", "entrada", "saida", "saldo_final", "limite", "disponivel_total_final"}
        for col_index, column in enumerate(columns, start=1):
            cell = flow_sheet.cell(row=current_row, column=col_index, value=column.get("label"))
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = alignment_cls(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        current_row += 1
        for item in report_payload.get("rows", []):
            for col_index, column in enumerate(columns, start=1):
                key = column.get("key")
                cell = flow_sheet.cell(row=current_row, column=col_index)
                cell.border = thin_border
                if key in amount_keys:
                    _style_amount(cell, item.get(key))
                else:
                    cell.value = item.get(key)
                    cell.alignment = alignment_cls(horizontal="center" if key != "periodo" else "left", vertical="center")
            current_row += 1

        def _write_titles_sheet(sheet_name: str, rows: List[Dict[str, Any]], totals_payload: Dict[str, Any], title_text: str):
            sheet = workbook.create_sheet(sheet_name)
            sheet.sheet_view.showGridLines = False
            sheet.freeze_panes = "A5"
            sheet.merge_cells("A1:F2")
            sheet["A1"] = title_text
            sheet["A1"].fill = hero_fill
            sheet["A1"].font = title_font
            sheet["A1"].alignment = alignment_cls(horizontal="left", vertical="center")
            headers = ["ID", "Tipo", "Valor Título", report_payload.get("projected_amount_label", "Saldo Projetado"), "Favorecido", "Vencimento"]
            for col_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=4, column=col_index, value=header)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = alignment_cls(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            row_cursor = 5
            for item in rows:
                sheet.cell(row=row_cursor, column=1, value=item.get("id")).border = thin_border
                sheet.cell(row=row_cursor, column=2, value=item.get("type_code")).border = thin_border
                _style_amount(sheet.cell(row=row_cursor, column=3), item.get("title_amount"), item.get("title_amount_value"))
                _style_amount(sheet.cell(row=row_cursor, column=4), item.get("open_amount"), item.get("open_amount_value"))
                sheet.cell(row=row_cursor, column=5, value=item.get("counterparty")).border = thin_border
                sheet.cell(row=row_cursor, column=6, value=item.get("due_date")).border = thin_border
                row_cursor += 1
            sheet.cell(row=row_cursor, column=1, value="Totais").font = label_font
            sheet.cell(row=row_cursor, column=1).fill = accent_fill
            sheet.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=2)
            _style_amount(sheet.cell(row=row_cursor, column=3), totals_payload.get("title_amount"), totals_payload.get("title_amount_value"))
            _style_amount(sheet.cell(row=row_cursor, column=4), totals_payload.get("open_amount"), totals_payload.get("open_amount_value"))
            sheet.cell(row=row_cursor, column=5, value=f"{totals_payload.get('count', 0)} título(s)").font = label_font
            sheet.cell(row=row_cursor, column=5).fill = accent_fill
            sheet.cell(row=row_cursor, column=5).border = thin_border
            sheet.cell(row=row_cursor, column=6).fill = accent_fill
            sheet.cell(row=row_cursor, column=6).border = thin_border
            return sheet

        receivables_sheet = _write_titles_sheet(
            "Titulos Receber",
            report_payload.get("selected_receivables", []) or [],
            report_payload.get("selected_receivables_totals", {}) or {},
            "Contas a Receber Selecionadas",
        )
        payables_sheet = _write_titles_sheet(
            "Titulos Pagar",
            report_payload.get("selected_payables", []) or [],
            report_payload.get("selected_payables_totals", {}) or {},
            "Contas a Pagar Selecionadas",
        )

        for sheet in [summary_sheet, flow_sheet, receivables_sheet, payables_sheet]:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.border == border_cls():
                        cell.border = thin_border
            width_map = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    width_map[cell.column] = max(width_map.get(cell.column, 0), len(str(cell.value)))
            for col_index, width in width_map.items():
                sheet.column_dimensions[get_column_letter_fn(col_index)].width = min(max(width + 2, 12), 34)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _export_working_capital_xlsx(
        report_payload: Dict[str, Any],
        *,
        workbook_factory,
        alignment_cls,
        border_cls,
        font_cls,
        fill_cls,
        side_cls,
        get_column_letter_fn,
    ) -> bytes:
        workbook = workbook_factory()
        hero_fill = fill_cls(fill_type="solid", fgColor="0F172A")
        header_fill = fill_cls(fill_type="solid", fgColor="E5E7EB")
        main_fill = fill_cls(fill_type="solid", fgColor="F8FAFC")
        group_fill = fill_cls(fill_type="solid", fgColor="F3F4F6")
        highlight_fill = fill_cls(fill_type="solid", fgColor="F5F7FA")
        thin_border = border_cls(
            left=side_cls(style="thin", color="CBD5E1"),
            right=side_cls(style="thin", color="CBD5E1"),
            top=side_cls(style="thin", color="CBD5E1"),
            bottom=side_cls(style="thin", color="CBD5E1"),
        )
        white_font = font_cls(color="FFFFFF", bold=True)
        title_font = font_cls(color="FFFFFF", bold=True, size=18)
        subtitle_font = font_cls(color="CBD5E1", size=10)
        section_font = font_cls(color="111827", bold=True, size=11)
        row_font = font_cls(color="111827", bold=True)
        detail_font = font_cls(color="374151", size=10)
        currency_format = '#,##0.00;[Red]- #,##0.00'

        def _style_money(cell, value: Any):
            amount = Decimal(str(value or 0))
            cell.value = FinancialReportService._serialize_money(amount)
            cell.number_format = currency_format
            cell.alignment = alignment_cls(horizontal="right", vertical="center")
            cell.font = row_font if amount != 0 else detail_font
            cell.border = thin_border

        def _style_label(cell, value: Any, *, font=None, fill=None, indent=0):
            cell.value = value
            cell.alignment = alignment_cls(horizontal="left", vertical="center", indent=indent, wrap_text=True)
            cell.font = font or detail_font
            cell.border = thin_border
            if fill is not None:
                cell.fill = fill

        def _write_balance_side(sheet, start_col: int, header_title: str, section_payload: Dict[str, Any], total_label: str) -> int:
            col_label = start_col
            col_amount = start_col + 3
            for offset in range(4):
                sheet.cell(row=6, column=start_col + offset).fill = header_fill
                sheet.cell(row=6, column=start_col + offset).border = thin_border
            _style_label(sheet.cell(row=6, column=col_label), header_title.upper(), font=section_font, fill=header_fill)
            _style_money(sheet.cell(row=6, column=col_amount), FinancialReportService._parse_currency_label(section_payload["current"]["amount"]))

            for offset in range(4):
                sheet.cell(row=7, column=start_col + offset).fill = main_fill
                sheet.cell(row=7, column=start_col + offset).border = thin_border
            _style_label(sheet.cell(row=7, column=col_label), total_label, font=row_font, fill=main_fill)
            _style_money(sheet.cell(row=7, column=col_amount), FinancialReportService._parse_currency_label(section_payload["current"]["amount"]))

            current_row = 8
            for group in section_payload["current"].get("groups", []):
                for offset in range(4):
                    sheet.cell(row=current_row, column=start_col + offset).fill = group_fill
                    sheet.cell(row=current_row, column=start_col + offset).border = thin_border
                _style_label(
                    sheet.cell(row=current_row, column=col_label),
                    f"{group.get('code')} - {group.get('label')}",
                    font=row_font,
                    fill=group_fill,
                )
                _style_money(sheet.cell(row=current_row, column=col_amount), FinancialReportService._parse_currency_label(group.get("amount")))
                current_row += 1
                for item in group.get("items", []):
                    _style_label(
                        sheet.cell(row=current_row, column=col_label),
                        f"{item.get('code')} - {item.get('label')}",
                        font=detail_font,
                        indent=1,
                    )
                    _style_money(sheet.cell(row=current_row, column=col_amount), FinancialReportService._parse_currency_label(item.get("amount")))
                    current_row += 1
            return current_row

        summary_sheet = workbook.active
        summary_sheet.title = "Balanço CCL"
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet.merge_cells("A1:H2")
        summary_sheet["A1"] = report_payload.get("title", "Capital Circulante Líquido")
        summary_sheet["A1"].fill = hero_fill
        summary_sheet["A1"].font = title_font
        summary_sheet["A1"].alignment = alignment_cls(horizontal="left", vertical="center")
        summary_sheet.merge_cells("A3:H3")
        summary_sheet["A3"] = str(report_payload.get("company_name") or "Versus Gestão Corporativa")
        summary_sheet["A3"].fill = hero_fill
        summary_sheet["A3"].font = subtitle_font
        summary_sheet["A3"].alignment = alignment_cls(horizontal="left", vertical="center")
        summary_sheet.merge_cells("A4:H4")
        summary_sheet["A4"] = f"Emitido em {FinancialReportService._pdf_generated_at_label(report_payload)}"
        summary_sheet["A4"].fill = hero_fill
        summary_sheet["A4"].font = subtitle_font
        summary_sheet["A4"].alignment = alignment_cls(horizontal="left", vertical="center")

        balance_sheet = report_payload.get("balance_sheet") or {}
        asset_payload = balance_sheet.get("asset") or {"current": {"amount": "0,00", "groups": []}}
        liability_payload = balance_sheet.get("liability") or {"current": {"amount": "0,00", "groups": []}}

        left_end = _write_balance_side(summary_sheet, 1, "Ativo", asset_payload, "Circulante")
        right_end = _write_balance_side(summary_sheet, 5, "Passivo", liability_payload, "Circulante")
        current_row = max(left_end, right_end) + 1

        for column in range(1, 9):
            summary_sheet.cell(row=current_row, column=column).fill = highlight_fill
            summary_sheet.cell(row=current_row, column=column).border = thin_border
        _style_label(
            summary_sheet.cell(row=current_row, column=1),
            balance_sheet.get("working_capital", {}).get("title", "Capital Circulante Líquido"),
            font=section_font,
            fill=highlight_fill,
        )
        _style_money(
            summary_sheet.cell(row=current_row, column=8),
            FinancialReportService._parse_currency_label(balance_sheet.get("working_capital", {}).get("amount")),
        )
        current_row += 2

        for offset in range(4):
            summary_sheet.cell(row=current_row, column=1 + offset).fill = main_fill
            summary_sheet.cell(row=current_row, column=1 + offset).border = thin_border
            summary_sheet.cell(row=current_row, column=5 + offset).fill = main_fill
            summary_sheet.cell(row=current_row, column=5 + offset).border = thin_border
        _style_label(summary_sheet.cell(row=current_row, column=1), asset_payload.get("non_current", {}).get("title", "Ativo Não Circulante"), font=row_font, fill=main_fill)
        _style_money(summary_sheet.cell(row=current_row, column=4), FinancialReportService._parse_currency_label(asset_payload.get("non_current", {}).get("amount")))
        _style_label(summary_sheet.cell(row=current_row, column=5), liability_payload.get("non_current", {}).get("title", "Passivo Não Circulante"), font=row_font, fill=main_fill)
        _style_money(summary_sheet.cell(row=current_row, column=8), FinancialReportService._parse_currency_label(liability_payload.get("non_current", {}).get("amount")))
        current_row += 1

        for offset in range(4):
            summary_sheet.cell(row=current_row, column=5 + offset).fill = main_fill
            summary_sheet.cell(row=current_row, column=5 + offset).border = thin_border
        _style_label(summary_sheet.cell(row=current_row, column=5), liability_payload.get("equity", {}).get("title", "Patrimônio Líquido"), font=row_font, fill=main_fill)
        _style_money(summary_sheet.cell(row=current_row, column=8), FinancialReportService._parse_currency_label(liability_payload.get("equity", {}).get("amount")))
        current_row += 2

        for column in range(1, 9):
            summary_sheet.cell(row=current_row, column=column).fill = highlight_fill
            summary_sheet.cell(row=current_row, column=column).border = thin_border
        _style_label(
            summary_sheet.cell(row=current_row, column=1),
            balance_sheet.get("patrimonial_status", {}).get("title", "Situação Patrimonial"),
            font=section_font,
            fill=highlight_fill,
        )
        _style_money(
            summary_sheet.cell(row=current_row, column=8),
            FinancialReportService._parse_currency_label(balance_sheet.get("patrimonial_status", {}).get("amount")),
        )

        filters_sheet = workbook.create_sheet("Filtros e resumo")
        filters_sheet["A1"] = report_payload.get("title", "Relatório")
        filters_sheet["A1"].font = font_cls(bold=True, size=14)
        filters_sheet["A3"] = "Gerado em"
        filters_sheet["B3"] = report_payload.get("generated_at")
        row_cursor = 5
        for section_title, items in [
            ("Filtros", report_payload.get("filters", [])),
            ("Informações gerais", report_payload.get("general_info", [])),
            ("Resumo executivo", report_payload.get("summary_cards", [])),
        ]:
            filters_sheet[f"A{row_cursor}"] = section_title
            filters_sheet[f"A{row_cursor}"].font = font_cls(bold=True)
            row_cursor += 1
            for item in items:
                filters_sheet[f"A{row_cursor}"] = item.get("label")
                filters_sheet[f"B{row_cursor}"] = item.get("value")
                row_cursor += 1
            row_cursor += 1

        data_sheet = workbook.create_sheet("Base analítica")
        columns = report_payload.get("columns", [])
        for index, column in enumerate(columns, start=1):
            cell = data_sheet.cell(row=1, column=index, value=column.get("label"))
            cell.font = font_cls(bold=True)
        for row_index, item in enumerate(report_payload.get("rows", []), start=2):
            for col_index, column in enumerate(columns, start=1):
                data_sheet.cell(row=row_index, column=col_index, value=item.get(column.get("key"), ""))

        for sheet in workbook.worksheets:
            width_map = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    width_map[cell.column] = max(width_map.get(cell.column, 0), len(str(cell.value)))
            for col_index, width in width_map.items():
                sheet.column_dimensions[get_column_letter_fn(col_index)].width = min(max(width + 2, 14), 40)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _build_cash_flow_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "CashFlowPdfTitle",
            parent=styles["Heading1"],
            fontSize=20,
            leading=24,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
        subtitle_style = ParagraphStyle(
            "CashFlowPdfSubtitle",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#CBD5E1"),
            alignment=TA_LEFT,
        )
        section_title_style = ParagraphStyle(
            "CashFlowPdfSectionTitle",
            parent=styles["BodyText"],
            fontSize=10,
            leading=12,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=5,
        )
        table_header_style = ParagraphStyle(
            "CashFlowPdfTableHeader",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8,
            fontName="Helvetica-Bold",
            alignment=TA_CENTER,
            textColor=colors.white,
        )
        table_cell_style = ParagraphStyle(
            "CashFlowPdfTableCell",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#0F172A"),
        )
        table_center_style = ParagraphStyle("CashFlowPdfTableCenter", parent=table_cell_style, alignment=TA_CENTER)
        table_right_style = ParagraphStyle("CashFlowPdfTableRight", parent=table_cell_style, alignment=TA_LEFT)
        stat_label_style = ParagraphStyle(
            "CashFlowPdfStatLabel",
            parent=styles["BodyText"],
            fontSize=6,
            leading=7,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#475569"),
        )
        stat_value_style = ParagraphStyle(
            "CashFlowPdfStatValue",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        filter_style = ParagraphStyle(
            "CashFlowPdfFilterCell",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#0F172A"),
        )

        company_name = str(report_payload.get("company_name") or "Versus Gestão Corporativa")
        hero_left = [
            Paragraph(report_payload.get("title", "Fluxo de Caixa"), title_style),
            Paragraph(company_name, subtitle_style),
            Paragraph(str(report_payload.get("subtitle") or ""), subtitle_style),
        ]
        top_panel = FinancialReportService._build_cash_flow_pdf_header_accounts_panel(
            hero_left=hero_left,
            report_payload=report_payload,
            available_width=available_width,
            header_style=table_header_style,
            cell_style=table_cell_style,
        )

        elements: List[Any] = [top_panel, Spacer(1, 10)]
        elements.append(Paragraph("Contas correntes", section_title_style))
        elements.append(
            FinancialReportService._build_cash_flow_pdf_accounts_table(
                report_payload=report_payload,
                available_width=available_width,
                header_style=table_header_style,
                cell_style=table_cell_style,
            )
        )
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Fluxo do período", section_title_style))
        elements.append(
            FinancialReportService._build_cash_flow_pdf_flow_table(
                report_payload=report_payload,
                available_width=available_width,
                header_style=table_header_style,
                cell_style=table_cell_style,
                center_style=table_center_style,
            )
        )
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Contas a receber selecionadas", section_title_style))
        elements.append(
            FinancialReportService._build_cash_flow_pdf_titles_table(
                title_rows=report_payload.get("selected_receivables") or [],
                totals_payload=report_payload.get("selected_receivables_totals") or {},
                projected_label=f"{report_payload.get('projected_amount_label', 'Saldo Projetado')} a Receber",
                available_width=available_width,
                header_style=table_header_style,
                cell_style=table_cell_style,
                center_style=table_center_style,
                right_style=table_right_style,
            )
        )
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Contas a pagar selecionadas", section_title_style))
        elements.append(
            FinancialReportService._build_cash_flow_pdf_titles_table(
                title_rows=report_payload.get("selected_payables") or [],
                totals_payload=report_payload.get("selected_payables_totals") or {},
                projected_label=f"{report_payload.get('projected_amount_label', 'Saldo Projetado')} a Pagar",
                available_width=available_width,
                header_style=table_header_style,
                cell_style=table_cell_style,
                center_style=table_center_style,
                right_style=table_right_style,
            )
        )
        return elements

    @staticmethod
    def _build_working_capital_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "WorkingCapitalPdfTitle",
            parent=styles["Heading1"],
            fontSize=20,
            leading=24,
            textColor=colors.white,
            alignment=TA_LEFT,
        )
        subtitle_style = ParagraphStyle(
            "WorkingCapitalPdfSubtitle",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#CBD5E1"),
            alignment=TA_LEFT,
        )
        section_title_style = ParagraphStyle(
            "WorkingCapitalPdfSectionTitle",
            parent=styles["BodyText"],
            fontSize=10,
            leading=12,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=5,
        )
        table_header_style = ParagraphStyle(
            "WorkingCapitalPdfTableHeader",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            fontName="Helvetica-Bold",
            alignment=TA_LEFT,
            textColor=colors.HexColor("#0F172A"),
        )
        cell_style = ParagraphStyle(
            "WorkingCapitalPdfCell",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#111827"),
        )

        company_name = str(report_payload.get("company_name") or "Versus Gestão Corporativa")
        hero = Table(
            [[
                Paragraph(report_payload.get("title", "Capital Circulante Líquido"), title_style),
                Paragraph(company_name, subtitle_style),
                Paragraph(str(report_payload.get("subtitle") or ""), subtitle_style),
            ]],
            colWidths=[available_width],
            hAlign="LEFT",
        )
        hero.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 14),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                    ("TOPPADDING", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ]
            )
        )

        balance_sheet = report_payload.get("balance_sheet") or {}
        elements: List[Any] = [hero, Spacer(1, 10)]
        elements.append(
            Table(
                [[
                    FinancialReportService._build_working_capital_pdf_side_table(
                        title=(balance_sheet.get("asset") or {}).get("title", "Ativo"),
                        current_section=(balance_sheet.get("asset") or {}).get("current") or {},
                        table_header_style=table_header_style,
                        cell_style=cell_style,
                        available_width=(available_width / 2) - 6,
                    ),
                    FinancialReportService._build_working_capital_pdf_side_table(
                        title=(balance_sheet.get("liability") or {}).get("title", "Passivo"),
                        current_section=(balance_sheet.get("liability") or {}).get("current") or {},
                        table_header_style=table_header_style,
                        cell_style=cell_style,
                        available_width=(available_width / 2) - 6,
                    ),
                ]],
                colWidths=[available_width / 2, available_width / 2],
                hAlign="LEFT",
            )
        )
        elements.append(Spacer(1, 10))
        elements.append(
            Paragraph(
                f"<b>{(balance_sheet.get('working_capital') or {}).get('title', 'Capital Circulante Líquido')}:</b> "
                f"{(balance_sheet.get('working_capital') or {}).get('amount', '0,00')}",
                section_title_style,
            )
        )
        elements.append(
            Table(
                [[
                    Paragraph(
                        f"{(balance_sheet.get('asset') or {}).get('non_current', {}).get('title', 'Ativo Não Circulante')}: "
                        f"<b>{(balance_sheet.get('asset') or {}).get('non_current', {}).get('amount', '0,00')}</b>",
                        cell_style,
                    ),
                    Paragraph(
                        f"{(balance_sheet.get('liability') or {}).get('non_current', {}).get('title', 'Passivo Não Circulante')}: "
                        f"<b>{(balance_sheet.get('liability') or {}).get('non_current', {}).get('amount', '0,00')}</b><br/>"
                        f"{(balance_sheet.get('liability') or {}).get('equity', {}).get('title', 'Patrimônio Líquido')}: "
                        f"<b>{(balance_sheet.get('liability') or {}).get('equity', {}).get('amount', '0,00')}</b>",
                        cell_style,
                    ),
                ]],
                colWidths=[available_width / 2, available_width / 2],
                hAlign="LEFT",
            )
        )
        elements.append(Spacer(1, 8))
        elements.append(
            Paragraph(
                f"<b>{(balance_sheet.get('patrimonial_status') or {}).get('title', 'Situação Patrimonial')}:</b> "
                f"{(balance_sheet.get('patrimonial_status') or {}).get('amount', '0,00')}",
                section_title_style,
            )
        )
        return elements

    @staticmethod
    def _build_working_capital_pdf_side_table(*, title: str, current_section: Dict[str, Any], table_header_style, cell_style, available_width: float) -> Table:
        rows: List[List[Any]] = [
            [
                Paragraph(str(title).upper(), table_header_style),
                Paragraph(f"<para alignment='right'><b>{current_section.get('amount', '0,00')}</b></para>", cell_style),
            ],
            [
                Paragraph(f"<b>{current_section.get('title', 'Circulante')}</b>", cell_style),
                Paragraph(f"<para alignment='right'><b>{current_section.get('amount', '0,00')}</b></para>", cell_style),
            ],
        ]
        for group in current_section.get("groups", []):
            rows.append(
                [
                    Paragraph(f"<b>{group.get('code')} - {group.get('label')}</b>", cell_style),
                    Paragraph(f"<para alignment='right'><b>{group.get('amount')}</b></para>", cell_style),
                ]
            )
            for item in group.get("items", []):
                rows.append(
                    [
                        Paragraph(f"&nbsp;&nbsp;&nbsp;{item.get('code')} - {item.get('label')}", cell_style),
                        Paragraph(f"<para alignment='right'>{item.get('amount')}</para>", cell_style),
                    ]
                )
        table = Table(rows, colWidths=[available_width * 0.74, available_width * 0.26], hAlign="LEFT")
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F8FAFC")),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_index in range(2, len(rows)):
            fill = colors.HexColor("#F3F4F6") if not str(rows[row_index][0].text).startswith("&nbsp;") else colors.white
            styles.append(("BACKGROUND", (0, row_index), (-1, row_index), fill))
        table.setStyle(TableStyle(styles))
        return table

    @staticmethod
    def _build_cash_flow_pdf_header_accounts_panel(*, hero_left: List[Any], report_payload: Dict[str, Any], available_width: float, header_style, cell_style) -> Table:
        accounts_table = FinancialReportService._build_cash_flow_pdf_accounts_table(
            report_payload=report_payload,
            available_width=available_width * 0.72,
            header_style=header_style,
            cell_style=cell_style,
        )
        hero_card = Table([[hero_left]], colWidths=[available_width * 0.28], hAlign="LEFT")
        hero_card.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 14),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                    ("TOPPADDING", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        panel = Table(
            [[hero_card, accounts_table]],
            colWidths=[available_width * 0.28, available_width * 0.72],
            hAlign="LEFT",
        )
        panel.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        return panel

    @staticmethod
    def _build_cash_flow_pdf_accounts_table(*, report_payload: Dict[str, Any], available_width: float, header_style, cell_style) -> Table:
        rows = report_payload.get("bank_account_summary_rows") or []
        totals = report_payload.get("bank_account_summary_totals") or {}
        headers = [
            "Descrição",
            "Limite",
            f"Saldo em {report_payload.get('bank_balance_reference_label', '-')}",
            "Disp. Total",
        ]
        data: List[List[Any]] = [[Paragraph(value, header_style) for value in headers]]
        for item in rows:
            data.append(
                [
                    Paragraph(str(item.get("description") or "-"), cell_style),
                    FinancialReportService._cash_flow_pdf_amount_paragraph(item.get("limit"), item.get("limit_value"), cell_style),
                    FinancialReportService._cash_flow_pdf_amount_paragraph(item.get("balance"), item.get("balance_value"), cell_style),
                    FinancialReportService._cash_flow_pdf_amount_paragraph(item.get("available_total"), item.get("available_total_value"), cell_style),
                ]
            )
        data.append(
            [
                Paragraph("<b>Total</b>", cell_style),
                FinancialReportService._cash_flow_pdf_amount_paragraph(totals.get("limit"), totals.get("limit_value"), cell_style),
                FinancialReportService._cash_flow_pdf_amount_paragraph(totals.get("balance"), totals.get("balance_value"), cell_style),
                FinancialReportService._cash_flow_pdf_amount_paragraph(totals.get("available_total"), totals.get("available_total_value"), cell_style),
            ]
        )
        col_widths = [available_width * 0.34, available_width * 0.18, available_width * 0.24, available_width * 0.24]
        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#EFF6FF")),
        ]
        for row_index in range(1, len(data) - 1):
            styles.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#FFFFFF") if row_index % 2 else colors.HexColor("#F8FAFC")))
        table.setStyle(TableStyle(styles))
        return table

    @staticmethod
    def _build_cash_flow_pdf_flow_table(*, report_payload: Dict[str, Any], available_width: float, header_style, cell_style, center_style) -> Table:
        columns = report_payload.get("columns") or []
        amount_keys = {"saldo_inicial", "entrada", "saida", "saldo_final", "limite", "disponivel_total_final"}
        ratio_map = {
            "periodo": 0.95,
            "data_inicial": 0.95,
            "data_final": 0.95,
            "saldo_inicial": 1.05,
            "entrada": 0.9,
            "saida": 0.9,
            "saldo_final": 1.05,
            "limite": 0.9,
            "disponivel_total_final": 1.2,
        }
        total_ratio = sum(ratio_map.get(column.get("key"), 1) for column in columns) or 1
        col_widths = [available_width * (ratio_map.get(column.get("key"), 1) / total_ratio) for column in columns]
        data = [[Paragraph(str(column.get("label") or ""), header_style) for column in columns]]
        for item in report_payload.get("rows") or []:
            row_cells: List[Any] = []
            for column in columns:
                key = column.get("key")
                if key in amount_keys:
                    row_cells.append(FinancialReportService._cash_flow_pdf_amount_paragraph(item.get(key), None, cell_style))
                else:
                    style = cell_style if key == "periodo" else center_style
                    row_cells.append(Paragraph(str(item.get(key) or "-"), style))
            data.append(row_cells)
        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_index in range(1, len(data)):
            styles.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#FFFFFF") if row_index % 2 else colors.HexColor("#F8FAFC")))
        table.setStyle(TableStyle(styles))
        return table

    @staticmethod
    def _build_cash_flow_pdf_titles_table(
        *,
        title_rows: List[Dict[str, Any]],
        totals_payload: Dict[str, Any],
        projected_label: str,
        available_width: float,
        header_style,
        cell_style,
        center_style,
        right_style,
    ) -> Table:
        headers = ["ID", "Tipo", "Valor Título", projected_label, "Favorecido", "Vencimento"]
        data: List[List[Any]] = [[Paragraph(header, header_style) for header in headers]]
        rows = title_rows or []
        if rows:
            for item in rows:
                data.append(
                    [
                        Paragraph(str(item.get("id") or "-"), center_style),
                        Paragraph(str(item.get("type_code") or "-"), center_style),
                        FinancialReportService._cash_flow_pdf_amount_paragraph(item.get("title_amount"), item.get("title_amount_value"), cell_style),
                        FinancialReportService._cash_flow_pdf_amount_paragraph(item.get("open_amount"), item.get("open_amount_value"), cell_style),
                        Paragraph(str(item.get("counterparty") or "-"), right_style),
                        Paragraph(str(item.get("due_date") or "-"), center_style),
                    ]
                )
        else:
            data.append([Paragraph("Nenhum título encontrado para os filtros informados.", center_style), "", "", "", "", ""])
        data.append(
            [
                Paragraph("<b>Totais</b>", cell_style),
                "",
                FinancialReportService._cash_flow_pdf_amount_paragraph(totals_payload.get("title_amount"), totals_payload.get("title_amount_value"), cell_style),
                FinancialReportService._cash_flow_pdf_amount_paragraph(totals_payload.get("open_amount"), totals_payload.get("open_amount_value"), cell_style),
                Paragraph(f"<b>{totals_payload.get('count', 0)} título(s)</b>", center_style),
                "",
            ]
        )
        col_widths = [
            available_width * 0.07,
            available_width * 0.08,
            available_width * 0.17,
            available_width * 0.21,
            available_width * 0.31,
            available_width * 0.16,
        ]
        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#EFF6FF")),
        ]
        if not rows:
            styles.append(("SPAN", (0, 1), (-1, 1)))
        for row_index in range(1, len(data) - 1):
            styles.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#FFFFFF") if row_index % 2 else colors.HexColor("#F8FAFC")))
        table.setStyle(TableStyle(styles))
        return table

    @staticmethod
    def _cash_flow_pdf_amount_paragraph(value: Any, amount_value: Any, base_style) -> Paragraph:
        label = FinancialReportService._cash_flow_plain_amount_label(value, amount_value)
        color = FinancialReportService._cash_flow_amount_color_hex(value, amount_value)
        return Paragraph(
            f"<para alignment='right'><font color='{color}'><b>{label}</b></font></para>",
            base_style,
        )

    @staticmethod
    def _pdf_generated_at_label(report_payload: Dict[str, Any]) -> str:
        generated_at = str(report_payload.get("generated_at") or "-").strip()
        if generated_at == "-":
            return generated_at
        if " às " in generated_at:
            return generated_at
        if " " in generated_at:
            date_part, time_part = generated_at.split(" ", 1)
            return f"{date_part} às {time_part}"
        return generated_at

    @staticmethod
    def _draw_default_pdf_footer(canvas, doc, report_payload: Dict[str, Any]) -> None:
        canvas.saveState()
        width, _ = doc.pagesize
        footer_y = 18
        line_y = footer_y + 10
        canvas.setStrokeColor(colors.HexColor("#cbd5e1"))
        canvas.setLineWidth(0.6)
        canvas.line(doc.leftMargin, line_y, width - doc.rightMargin, line_y)
        canvas.setFillColor(colors.HexColor("#0f172a"))
        canvas.setFont("Helvetica", 8)
        canvas.drawString(doc.leftMargin, footer_y, "Versus Gestão Corporativa - Todos os direitos reservados.")
        canvas.drawRightString(
            width - doc.rightMargin,
            footer_y,
            f"Emitido em: {FinancialReportService._pdf_generated_at_label(report_payload)}",
        )
        canvas.restoreState()

    @staticmethod
    def _resolve_dossier_attachment_path(attachment: Dict[str, Any]) -> Optional[str]:
        raw_url = str(attachment.get("url") or attachment.get("document_url") or "").strip()
        if raw_url.startswith(("http://", "https://")):
            return None

        relative_path = raw_url
        if relative_path.startswith("/uploads/"):
            relative_path = relative_path[len("/uploads/"):]
        elif relative_path.startswith("uploads/"):
            relative_path = relative_path[len("uploads/"):]

        relative_path = relative_path.replace("\\", "/").lstrip("/")
        if not relative_path or ".." in relative_path.split("/"):
            return None

        upload_root = current_app.config.get("UPLOAD_FOLDER", "uploads") if has_app_context() else "uploads"
        root_abs = os.path.abspath(upload_root)
        candidate = os.path.abspath(os.path.join(root_abs, relative_path))
        try:
            if os.path.commonpath([root_abs, candidate]) != root_abs:
                return None
        except ValueError:
            return None
        return candidate if os.path.exists(candidate) else None

    @staticmethod
    def _scaled_reportlab_image(source: Any, *, max_width: float, max_height: float, pixel_size: Optional[Tuple[int, int]] = None) -> Optional[RLImage]:
        try:
            if pixel_size is None:
                from PIL import Image as PILImage

                with PILImage.open(source) as image:
                    pixel_size = image.size
            raw_width, raw_height = pixel_size
            if raw_width <= 0 or raw_height <= 0:
                return None
            scale = min(max_width / float(raw_width), max_height / float(raw_height), 1.0)
            flowable = RLImage(source, width=float(raw_width) * scale, height=float(raw_height) * scale)
            if hasattr(source, "getvalue"):
                setattr(flowable, "_gv_source_buffer", source)
            return flowable
        except Exception:
            return None

    @staticmethod
    def _render_pdf_first_page_as_image(path: str, *, max_width: float, max_height: float) -> Optional[RLImage]:
        try:
            import fitz
            from PIL import Image as PILImage

            pdf = fitz.open(path)
            try:
                if pdf.page_count < 1:
                    return None
                page = pdf.load_page(0)
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                image_bytes = pixmap.tobytes("png")
            finally:
                pdf.close()
            buffer = io.BytesIO(image_bytes)
            with PILImage.open(io.BytesIO(image_bytes)) as image:
                size = image.size
            return FinancialReportService._scaled_reportlab_image(buffer, max_width=max_width, max_height=max_height, pixel_size=size)
        except Exception:
            return None

    @staticmethod
    def _build_dossier_attachment_preview(document: Dict[str, Any], *, max_width: float, max_height: float, placeholder_style) -> Any:
        attachment = dict(document.get("attachment") or {})
        path = FinancialReportService._resolve_dossier_attachment_path(attachment)
        name = str(document.get("document_name") or attachment.get("name") or "Anexo")
        content_type = str(document.get("content_type") or attachment.get("content_type") or "").lower()
        extension = os.path.splitext(name.lower())[1]
        flowable = None
        if path:
            if content_type.startswith("image/") or extension in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}:
                flowable = FinancialReportService._scaled_reportlab_image(path, max_width=max_width, max_height=max_height)
            elif content_type == "application/pdf" or extension == ".pdf":
                flowable = FinancialReportService._render_pdf_first_page_as_image(path, max_width=max_width, max_height=max_height)
        if flowable is not None:
            return flowable
        return Paragraph(
            f"<b>{name}</b><br/>Prévia indisponível para este anexo. Consulte o arquivo original no sistema.",
            placeholder_style,
        )

    @staticmethod
    def _build_income_statement_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "IncomeStatementTitle",
            parent=styles["Heading1"],
            fontSize=15,
            leading=18,
            textColor=colors.white,
            spaceAfter=2,
        )
        subtitle_style = ParagraphStyle(
            "IncomeStatementSubtitle",
            parent=styles["BodyText"],
            fontSize=7.2,
            leading=8.5,
            textColor=colors.HexColor("#E2E8F0"),
        )
        section_title_style = ParagraphStyle(
            "IncomeStatementSectionTitle",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=5,
        )
        filter_style = ParagraphStyle(
            "IncomeStatementFilter",
            parent=styles["BodyText"],
            fontSize=7.1,
            leading=8.4,
            textColor=colors.HexColor("#0F172A"),
        )
        card_label_style = ParagraphStyle(
            "IncomeStatementCardLabel",
            parent=styles["BodyText"],
            fontSize=6.8,
            leading=8,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#64748B"),
        )
        card_value_style = ParagraphStyle(
            "IncomeStatementCardValue",
            parent=styles["BodyText"],
            fontSize=9.2,
            leading=11,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        header_style = ParagraphStyle(
            "IncomeStatementTableHeader",
            parent=styles["BodyText"],
            fontSize=7.4,
            leading=8.4,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        account_style = ParagraphStyle(
            "IncomeStatementAccountCell",
            parent=styles["BodyText"],
            fontSize=7.4,
            leading=8.7,
            textColor=colors.HexColor("#0F172A"),
        )
        account_group_style = ParagraphStyle(
            "IncomeStatementAccountGroupCell",
            parent=account_style,
            fontSize=8.4,
            leading=9.8,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        account_subgroup_style = ParagraphStyle(
            "IncomeStatementAccountSubgroupCell",
            parent=account_style,
            fontSize=8,
            leading=9.3,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1E293B"),
        )
        account_leaf_style = ParagraphStyle(
            "IncomeStatementAccountLeafCell",
            parent=account_style,
            fontSize=7.6,
            leading=8.9,
            textColor=colors.HexColor("#334155"),
        )
        amount_base_style = ParagraphStyle(
            "IncomeStatementAmountCell",
            parent=styles["BodyText"],
            fontSize=7.4,
            leading=8.6,
            alignment=TA_RIGHT,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        amount_positive_style = ParagraphStyle("IncomeStatementAmountPositive", parent=amount_base_style, textColor=colors.HexColor("#2563EB"))
        amount_negative_style = ParagraphStyle("IncomeStatementAmountNegative", parent=amount_base_style, textColor=colors.HexColor("#DC2626"))
        amount_neutral_style = ParagraphStyle("IncomeStatementAmountNeutral", parent=amount_base_style, textColor=colors.HexColor("#0F172A"))
        amount_group_base_style = ParagraphStyle(
            "IncomeStatementAmountGroupBase",
            parent=amount_base_style,
            fontSize=8.4,
            leading=9.8,
            fontName="Helvetica-Bold",
        )
        amount_subgroup_base_style = ParagraphStyle(
            "IncomeStatementAmountSubgroupBase",
            parent=amount_base_style,
            fontSize=8,
            leading=9.3,
            fontName="Helvetica-Bold",
        )
        amount_leaf_base_style = ParagraphStyle(
            "IncomeStatementAmountLeafBase",
            parent=amount_base_style,
            fontSize=7.6,
            leading=8.9,
            fontName="Helvetica-Bold",
        )
        amount_group_positive_style = ParagraphStyle("IncomeStatementAmountGroupPositive", parent=amount_group_base_style, textColor=colors.HexColor("#2563EB"))
        amount_group_negative_style = ParagraphStyle("IncomeStatementAmountGroupNegative", parent=amount_group_base_style, textColor=colors.HexColor("#DC2626"))
        amount_group_neutral_style = ParagraphStyle("IncomeStatementAmountGroupNeutral", parent=amount_group_base_style, textColor=colors.HexColor("#0F172A"))
        amount_subgroup_positive_style = ParagraphStyle("IncomeStatementAmountSubgroupPositive", parent=amount_subgroup_base_style, textColor=colors.HexColor("#2563EB"))
        amount_subgroup_negative_style = ParagraphStyle("IncomeStatementAmountSubgroupNegative", parent=amount_subgroup_base_style, textColor=colors.HexColor("#DC2626"))
        amount_subgroup_neutral_style = ParagraphStyle("IncomeStatementAmountSubgroupNeutral", parent=amount_subgroup_base_style, textColor=colors.HexColor("#0F172A"))
        amount_leaf_positive_style = ParagraphStyle("IncomeStatementAmountLeafPositive", parent=amount_leaf_base_style, textColor=colors.HexColor("#2563EB"))
        amount_leaf_negative_style = ParagraphStyle("IncomeStatementAmountLeafNegative", parent=amount_leaf_base_style, textColor=colors.HexColor("#DC2626"))
        amount_leaf_neutral_style = ParagraphStyle("IncomeStatementAmountLeafNeutral", parent=amount_leaf_base_style, textColor=colors.HexColor("#0F172A"))
        total_positive_style = ParagraphStyle("IncomeStatementTotalPositive", parent=amount_base_style, textColor=colors.HexColor("#93C5FD"))
        total_negative_style = ParagraphStyle("IncomeStatementTotalNegative", parent=amount_base_style, textColor=colors.HexColor("#FCA5A5"))
        total_neutral_style = ParagraphStyle("IncomeStatementTotalNeutral", parent=amount_base_style, textColor=colors.white)
        empty_style = ParagraphStyle(
            "IncomeStatementEmpty",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        )

        def _as_decimal(raw_value: Any, fallback_label: Any = None) -> Decimal:
            if isinstance(raw_value, Decimal):
                return raw_value
            if isinstance(raw_value, (int, float)):
                return Decimal(str(raw_value or 0))
            raw_label = str(fallback_label if fallback_label is not None else raw_value if raw_value is not None else "0").strip()
            is_negative = raw_label.startswith("-") or raw_label.startswith("−") or "R$-" in raw_label.replace(" ", "")
            normalized = (
                raw_label.replace("R$", "")
                .replace(" ", "")
                .replace("−", "-")
                .replace(".", "")
                .replace(",", ".")
            )
            try:
                amount = Decimal(normalized or "0")
            except Exception:
                amount = Decimal("0")
            if is_negative and amount > 0:
                return -amount
            return amount

        def _amount_style(amount: Decimal):
            if amount < 0:
                return amount_negative_style
            if amount > 0:
                return amount_positive_style
            return amount_neutral_style

        def _amount_style_for_row(amount: Decimal, row_type: str, level: int):
            if row_type == "group" or level == 0:
                if amount < 0:
                    return amount_group_negative_style
                if amount > 0:
                    return amount_group_positive_style
                return amount_group_neutral_style
            if row_type in {"subgroup", "account-group"}:
                if amount < 0:
                    return amount_subgroup_negative_style
                if amount > 0:
                    return amount_subgroup_positive_style
                return amount_subgroup_neutral_style
            if amount < 0:
                return amount_leaf_negative_style
            if amount > 0:
                return amount_leaf_positive_style
            return amount_leaf_neutral_style

        def _total_amount_style(amount: Decimal):
            if amount < 0:
                return total_negative_style
            if amount > 0:
                return total_positive_style
            return total_neutral_style

        def _account_style_for_row(row_type: str, level: int):
            if row_type == "group" or level == 0:
                return account_group_style
            if row_type in {"subgroup", "account-group"}:
                return account_subgroup_style
            return account_leaf_style

        def _compact_filter_summary() -> str:
            fragments: List[str] = []
            subtitle_text = str(report_payload.get("subtitle") or "").strip().lower()
            for item in list(report_payload.get("filters") or []):
                label = str(item.get("label") or "").strip()
                value = str(item.get("value") or "").strip()
                if not label or not value:
                    continue
                if "emiss" in label.lower() or "gerado" in label.lower():
                    continue
                if label.lower() == "período" and value.lower() in subtitle_text:
                    continue
                fragments.append(f"{label}: {value}")
            return " · ".join(fragments[:4])

        company_name = str(report_payload.get("company_name") or "Empresa ativa").strip()
        title = report_payload.get("title") or "Demonstração de Resultados"
        subtitle = report_payload.get("subtitle") or "DRE contábil hierárquica por conta contábil."
        header_meta = " · ".join(part for part in [company_name, str(subtitle), _compact_filter_summary()] if part)
        hero = Table(
            [[
                [
                    Paragraph("Gestão Financeira · DRE Contábil", subtitle_style),
                    Paragraph(str(title), title_style),
                    Paragraph(header_meta, subtitle_style),
                ]
            ]],
            colWidths=[available_width],
            hAlign="LEFT",
        )
        hero.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.9, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 16),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 16),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )

        elements: List[Any] = [hero, Spacer(1, 7)]

        rows = list(report_payload.get("hierarchy_rows") or report_payload.get("rows") or [])
        visible_row_ids = {
            str(row_id)
            for row_id in (report_payload.get("visible_row_ids") or [])
            if str(row_id or "").strip()
        }
        if visible_row_ids:
            rows = [row for row in rows if str(row.get("id") or "") in visible_row_ids]
        collapsed_row_ids = {
            str(row_id)
            for row_id in (report_payload.get("collapsed_row_ids") or [])
            if str(row_id or "").strip()
        }
        if collapsed_row_ids and not visible_row_ids:
            parent_by_row_id = {
                str(row.get("id") or ""): str(row.get("parent_id") or "")
                for row in rows
                if row.get("id")
            }

            def _is_hidden_by_collapsed_parent(row: Dict[str, Any]) -> bool:
                parent_id = str(row.get("parent_id") or "")
                visited: set[str] = set()
                while parent_id:
                    if parent_id in collapsed_row_ids:
                        return True
                    if parent_id in visited:
                        break
                    visited.add(parent_id)
                    parent_id = parent_by_row_id.get(parent_id, "")
                return False

            rows = [row for row in rows if not _is_hidden_by_collapsed_parent(row)]
        if not rows:
            elements.append(Paragraph("Nenhuma conta encontrada para os filtros informados.", empty_style))
            return elements

        value_columns: List[Tuple[str, str, str]] = []
        if report_payload.get("show_budget_column", True):
            value_columns.append(("orcamento", "orcamento_label", "Orçamento"))
        value_columns.extend(
            [
                ("competencia", "competencia_label", "Competência"),
                ("vencimento", "vencimento_label", "Vencimento"),
                ("liquidacao", "liquidacao_label", "Liquidação"),
            ]
        )

        table_data: List[List[Any]] = [[Paragraph("Conta contábil", header_style)] + [Paragraph(label, header_style) for _, _, label in value_columns]]
        for row in rows:
            level = min(max(int(row.get("level") or 0), 0), 6)
            row_type = str(row.get("row_type") or "")
            code = str(row.get("codigo") or row.get("code") or "").strip()
            description = str(row.get("descricao") or row.get("description") or row.get("account_label") or "-").strip()
            indent = "&nbsp;" * (level * 4)
            account_label = f"{indent}<b>{code}</b> - {description}" if code else f"{indent}{description}"
            line: List[Any] = [Paragraph(account_label, _account_style_for_row(row_type, level))]
            for value_key, label_key, _ in value_columns:
                raw_value = row.get(value_key)
                label_value = row.get(label_key) or row.get(value_key) or row.get("budget_label" if value_key == "orcamento" else "") or "R$ 0,00"
                amount = _as_decimal(raw_value, label_value)
                line.append(Paragraph(str(label_value), _amount_style_for_row(amount, row_type, level)))
            table_data.append(line)

        account_width = available_width * 0.44
        amount_width = (available_width - account_width) / len(value_columns)
        table = Table(
            table_data,
            colWidths=[account_width] + [amount_width for _ in value_columns],
            repeatRows=1,
            hAlign="LEFT",
        )
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6E3F8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for row_index, row in enumerate(rows, start=1):
            row_type = str(row.get("row_type") or "")
            if row_type == "group":
                background = colors.HexColor("#EAF2FF")
                table_styles.append(("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold"))
            elif row_type == "subgroup":
                background = colors.HexColor("#F4F8FF")
            elif row_index % 2 == 0:
                background = colors.HexColor("#F8FAFC")
            else:
                background = colors.white
            table_styles.append(("BACKGROUND", (0, row_index), (-1, row_index), background))
        table.setStyle(TableStyle(table_styles))

        elements.append(table)

        totals = dict(report_payload.get("totals") or {})
        total_cells = [Paragraph("Total consolidado", header_style)]
        for value_key, label_key, label in value_columns:
            total_value_key = {
                "orcamento": "budget",
                "competencia": "competence",
                "vencimento": "due",
                "liquidacao": "liquidation",
                "aberto": "open",
                "baixado": "settled",
            }.get(value_key, value_key)
            total_label_key = f"{total_value_key}_label"
            label_value = totals.get(total_label_key) or totals.get(label_key) or "-"
            amount = _as_decimal(totals.get(total_value_key), label_value)
            total_cells.append(Paragraph(str(label_value), _total_amount_style(amount)))
        total_table = Table([total_cells], colWidths=[account_width] + [amount_width for _ in value_columns], hAlign="LEFT")
        total_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.extend([Spacer(1, 7), total_table])
        return elements

    @staticmethod
    def _build_income_statement_liquidation_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "DossierIncomeStatementTitle",
            parent=styles["Heading1"],
            fontSize=17,
            leading=21,
            textColor=colors.white,
            spaceAfter=3,
        )
        subtitle_style = ParagraphStyle(
            "DossierIncomeStatementSubtitle",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#E2E8F0"),
        )
        section_title_style = ParagraphStyle(
            "DossierIncomeStatementSection",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=5,
        )
        filter_cell_style = ParagraphStyle(
            "DossierIncomeStatementFilter",
            parent=styles["BodyText"],
            fontSize=7.4,
            leading=8.6,
            textColor=colors.HexColor("#0F172A"),
        )
        header_style = ParagraphStyle(
            "DossierIncomeStatementHeader",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        cell_style = ParagraphStyle(
            "DossierIncomeStatementCell",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8.3,
            textColor=colors.HexColor("#0F172A"),
        )
        amount_style = ParagraphStyle(
            "DossierIncomeStatementAmount",
            parent=cell_style,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        amount_positive_style = ParagraphStyle(
            "DossierIncomeStatementAmountPositive",
            parent=amount_style,
            textColor=colors.HexColor("#2563EB"),
        )
        amount_negative_style = ParagraphStyle(
            "DossierIncomeStatementAmountNegative",
            parent=amount_style,
            textColor=colors.HexColor("#DC2626"),
        )
        amount_neutral_style = ParagraphStyle(
            "DossierIncomeStatementAmountNeutral",
            parent=amount_style,
            textColor=colors.HexColor("#0F172A"),
        )
        result_label_style = ParagraphStyle(
            "DossierIncomeStatementResultLabel",
            parent=styles["BodyText"],
            fontSize=10,
            leading=12,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        result_value_style = ParagraphStyle(
            "DossierIncomeStatementResultValue",
            parent=result_label_style,
            alignment=TA_RIGHT,
            fontSize=11,
            leading=13,
        )
        result_value_positive_style = ParagraphStyle(
            "DossierIncomeStatementResultValuePositive",
            parent=result_value_style,
            textColor=colors.HexColor("#2563EB"),
        )
        result_value_negative_style = ParagraphStyle(
            "DossierIncomeStatementResultValueNegative",
            parent=result_value_style,
            textColor=colors.HexColor("#DC2626"),
        )
        result_value_neutral_style = ParagraphStyle(
            "DossierIncomeStatementResultValueNeutral",
            parent=result_value_style,
            textColor=colors.white,
        )
        empty_style = ParagraphStyle(
            "DossierIncomeStatementEmpty",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        )

        company_name = str(report_payload.get("company_name") or "Empresa ativa").strip()
        hero_content = [
            Paragraph("Gestão Financeira · Demonstração de Resultados", subtitle_style),
            Paragraph(report_payload.get("title") or "Demonstração de Resultados 01", title_style),
            Paragraph(company_name, subtitle_style),
            Paragraph("Face de liquidação", subtitle_style),
            Paragraph(f"Emitido em {FinancialReportService._pdf_generated_at_label(report_payload)}", subtitle_style),
        ]
        hero_table = Table([[hero_content]], colWidths=[available_width], hAlign="LEFT")
        hero_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 14),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        elements: List[Any] = [hero_table, Spacer(1, 8)]
        elements.append(Paragraph("Filtros aplicados", section_title_style))
        elements.append(
            FinancialReportService._build_schedule_pdf_filter_cards(
                report_filters=report_payload.get("filters") or [],
                available_width=available_width,
                content_style=filter_cell_style,
            )
        )
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("Resultado por liquidação", section_title_style))

        source_rows = list(report_payload.get("hierarchy_rows") or report_payload.get("rows") or [])
        if not source_rows:
            elements.append(Paragraph("Nenhum resultado por liquidação encontrado para os filtros informados.", empty_style))
            return elements

        def _row_liquidation_amount(row: Dict[str, Any]) -> Decimal:
            raw_value = row.get("liquidacao")
            if raw_value is None:
                raw_value = row.get("liquidation")
            if isinstance(raw_value, (Decimal, int, float)):
                return Decimal(str(raw_value or 0))
            raw_label = str(
                row.get("liquidacao_label")
                or row.get("liquidation_label")
                or raw_value
                or "0"
            ).strip()
            is_negative = raw_label.startswith("-") or raw_label.startswith("−") or "R$-" in raw_label.replace(" ", "")
            normalized = (
                raw_label.replace("R$", "")
                .replace(" ", "")
                .replace("−", "-")
                .replace(".", "")
                .replace(",", ".")
            )
            try:
                amount = Decimal(normalized or "0")
            except Exception:
                amount = Decimal("0")
            if is_negative and amount > 0:
                amount = -amount
            return amount

        def _dre_pdf_amount_style(amount: Decimal):
            if amount < 0:
                return amount_negative_style
            if amount > 0:
                return amount_positive_style
            return amount_neutral_style

        def _dre_pdf_result_value_style(amount: Decimal):
            if amount < 0:
                return result_value_negative_style
            if amount > 0:
                return result_value_positive_style
            return result_value_neutral_style

        table_data: List[List[Any]] = [
            [Paragraph("Código", header_style), Paragraph("Descrição", header_style), Paragraph("Liquidação", header_style)]
        ]
        for row in source_rows:
            level = int(row.get("level") or 0)
            description = f"{'&nbsp;' * (level * 4)}{row.get('descricao') or row.get('description') or row.get('account_label') or '-'}"
            liquidation_value = row.get("liquidacao_label") or row.get("liquidation_label") or row.get("liquidacao") or row.get("liquidation") or "R$ 0,00"
            liquidation_amount = _row_liquidation_amount(row)
            table_data.append(
                [
                    Paragraph(str(row.get("codigo") or row.get("code") or ""), cell_style),
                    Paragraph(str(description), cell_style),
                    Paragraph(str(liquidation_value), _dre_pdf_amount_style(liquidation_amount)),
                ]
            )

        table = Table(
            table_data,
            colWidths=[available_width * 0.18, available_width * 0.58, available_width * 0.24],
            repeatRows=1,
            hAlign="LEFT",
        )
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for row_index in range(1, len(table_data)):
            background = colors.HexColor("#FFFFFF") if row_index % 2 else colors.HexColor("#F8FAFC")
            table_styles.append(("BACKGROUND", (0, row_index), (-1, row_index), background))
        table.setStyle(TableStyle(table_styles))
        elements.append(table)

        result_rows = [row for row in source_rows if int(row.get("level") or 0) == 0] or source_rows
        result_amount = sum((_row_liquidation_amount(row) for row in result_rows), Decimal("0"))
        result_table = Table(
            [[
                Paragraph("Resultado", result_label_style),
                Paragraph(FinancialReportService._format_currency(result_amount), _dre_pdf_result_value_style(result_amount)),
            ]],
            colWidths=[available_width * 0.58, available_width * 0.42],
            hAlign="LEFT",
        )
        result_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(Spacer(1, 8))
        elements.append(result_table)
        return elements

    @staticmethod
    def _build_bank_statement_dossier_cover(*, report_payload: Dict[str, Any], styles, available_width: float, available_height: float) -> List[Any]:
        title_style = ParagraphStyle(
            "BankStatementDossierCoverTitle",
            parent=styles["Heading1"],
            fontSize=24,
            leading=28,
            alignment=TA_CENTER,
            textColor=colors.white,
            spaceAfter=8,
        )
        subtitle_style = ParagraphStyle(
            "BankStatementDossierCoverSubtitle",
            parent=styles["BodyText"],
            fontSize=10,
            leading=13,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#E2E8F0"),
        )
        meta_style = ParagraphStyle(
            "BankStatementDossierCoverMeta",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0F172A"),
        )
        card_style = ParagraphStyle(
            "BankStatementDossierCoverCard",
            parent=styles["BodyText"],
            fontSize=9,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0F172A"),
        )
        company_name = str(report_payload.get("company_name") or "Empresa ativa").strip()
        period = f"{report_payload.get('period_start', '-')} até {report_payload.get('period_end', '-')}"
        mode_label = "Completo com capa" if report_payload.get("dossier_mode") != "simple" else "Somente extrato e anexos"
        document_count = int(report_payload.get("dossier_document_count") or 0)

        hero = Table(
            [[
                [
                    Paragraph("Gestão Financeira", subtitle_style),
                    Paragraph("Dossiê do Extrato Bancário", title_style),
                    Paragraph(company_name, subtitle_style),
                    Paragraph(f"Período: {period} · Emitido em {FinancialReportService._pdf_generated_at_label(report_payload)}", subtitle_style),
                ]
            ]],
            colWidths=[available_width],
            hAlign="LEFT",
        )
        hero.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 20),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 20),
                    ("TOPPADDING", (0, 0), (-1, -1), 28),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 28),
                ]
            )
        )
        cards = Table(
            [[
                Paragraph(f"<b>Modo</b><br/>{mode_label}", card_style),
                Paragraph(f"<b>Movimentos</b><br/>{len(report_payload.get('rows') or [])}", card_style),
                Paragraph(f"<b>Comprovantes</b><br/>{document_count}", card_style),
            ]],
            colWidths=[available_width / 3, available_width / 3, available_width / 3],
            hAlign="LEFT",
        )
        cards.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ]
            )
        )
        note = Paragraph(
            "Este dossiê preserva o Extrato Bancário padrão e adiciona, em sequência, os documentos anexados aos títulos, lançamentos e baixas localizados no período.",
            meta_style,
        )
        return [Spacer(1, max(12, available_height * 0.08)), hero, Spacer(1, 18), cards, Spacer(1, 14), note]

    @staticmethod
    def _build_bank_statement_dossier_document_block(
        *,
        document: Dict[str, Any],
        available_width: float,
        block_height: float,
        preview_style,
        label_style,
        value_style,
    ) -> List[Any]:
        left_width = available_width * (2 / 3)
        right_width = available_width - left_width
        preview = FinancialReportService._build_dossier_attachment_preview(
            document,
            max_width=left_width - 18,
            max_height=block_height - 34,
            placeholder_style=preview_style,
        )
        left_cell = [
            Paragraph(
                f"<b>{document.get('source_label') or 'Anexo'}</b> · {document.get('document_name') or 'Documento'}",
                label_style,
            ),
            Spacer(1, 3),
            preview,
        ]
        detail_rows = []
        for label, key in [
            ("Competência", "competence_date"),
            ("Vencimento", "due_date"),
            ("Liquidação", "settlement_date"),
            ("Plano de contas", "chart_account"),
            ("Centro de resultados", "cost_center"),
            ("Fornecedor", "counterparty"),
            ("Lançamento", "entry_code"),
            ("Baixa", "settlement_code"),
            ("Valor", "amount"),
            ("Origem", "source_label"),
        ]:
            detail_rows.append([Paragraph(f"<b>{label}</b>", label_style), Paragraph(str(document.get(key) or "-"), value_style)])
        details = Table(detail_rows, colWidths=[right_width * 0.38, right_width * 0.62], hAlign="LEFT")
        details.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        return [left_cell, details]

    @staticmethod
    def _build_bank_statement_dossier_documents_pages(*, report_payload: Dict[str, Any], styles, available_width: float, available_height: float) -> List[Any]:
        documents = list(report_payload.get("dossier_documents") or [])
        title_style = ParagraphStyle(
            "BankStatementDossierAttachmentTitle",
            parent=styles["BodyText"],
            fontSize=6.8,
            leading=7.6,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        value_style = ParagraphStyle(
            "BankStatementDossierAttachmentValue",
            parent=styles["BodyText"],
            fontSize=5.8,
            leading=6.5,
            textColor=colors.HexColor("#0F172A"),
        )
        placeholder_style = ParagraphStyle(
            "BankStatementDossierAttachmentPlaceholder",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8.5,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        )
        page_title_style = ParagraphStyle(
            "BankStatementDossierAttachmentPageTitle",
            parent=styles["Heading2"],
            fontSize=11,
            leading=13,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=6,
        )
        page_range_style = ParagraphStyle(
            "BankStatementDossierAttachmentPageRange",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_RIGHT,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#1E293B"),
        )
        if not documents:
            return [
                Paragraph("Comprovantes", page_title_style),
                Spacer(1, 8),
                Paragraph("Nenhum anexo localizado nos títulos, lançamentos ou baixas do extrato para os filtros informados.", placeholder_style),
            ]

        elements: List[Any] = []
        # Quatro comprovantes por folha em faixa horizontal: uma única linha
        # com quatro colunas. Esse formato aproveita melhor comprovantes
        # verticais, recibos e notas, evitando o espaço morto do grid 2x2.
        content_height = max(120, available_height - 42)
        cell_width = available_width / 4
        cell_height = content_height

        def _document_card(document: Dict[str, Any]) -> Table:
            card_height = max(120, cell_height - 8)
            header_height = 18
            meta_height = 56
            preview_height = max(80, card_height - header_height - meta_height - 12)
            preview = FinancialReportService._build_dossier_attachment_preview(
                document,
                max_width=cell_width - 16,
                max_height=preview_height - 10,
                placeholder_style=placeholder_style,
            )
            preview_box = Table(
                [[preview]],
                colWidths=[cell_width - 18],
                rowHeights=[preview_height],
                hAlign="LEFT",
            )
            preview_box.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            meta_lines = [
                f"<b>Liquidação:</b> {document.get('settlement_date') or '-'}",
                f"<b>Fornecedor:</b> {document.get('counterparty') or '-'}",
                f"<b>Plano:</b> {document.get('chart_account') or '-'}",
                f"<b>Valor:</b> {document.get('amount') or '-'}",
            ]
            card = Table(
                [
                    [
                        Paragraph(
                            f"{document.get('source_label') or 'Anexo'} · {document.get('document_name') or 'Documento'}",
                            title_style,
                        )
                    ],
                    [preview_box],
                    [Paragraph("<br/>".join(meta_lines), value_style)],
                ],
                colWidths=[cell_width - 8],
                rowHeights=[header_height, preview_height, meta_height],
                hAlign="LEFT",
            )
            card.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
                        ("LINEBELOW", (0, 0), (-1, 0), 0.25, colors.HexColor("#E2E8F0")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("VALIGN", (0, 1), (0, 1), "MIDDLE"),
                    ]
                )
            )
            return card

        for page_index, start in enumerate(range(0, len(documents), 4)):
            chunk = documents[start:start + 4]
            if page_index:
                elements.append(PageBreak())
            page_start = start + 1
            page_end = start + len(chunk)
            header = Table(
                [[
                    Paragraph(f"Lote {page_index + 1}", page_title_style),
                    Paragraph(f"{page_start}-{page_end} de {len(documents)}", page_range_style),
                ]],
                colWidths=[available_width * 0.5, available_width * 0.5],
                hAlign="LEFT",
            )
            header.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            elements.append(header)
            cards = [_document_card(document) for document in chunk]
            while len(cards) < 4:
                cards.append(Paragraph("", placeholder_style))
            rows = [cards[0:4]]
            page_table = Table(
                rows,
                colWidths=[cell_width, cell_width, cell_width, cell_width],
                rowHeights=[cell_height],
                hAlign="LEFT",
            )
            page_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 3),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            elements.append(page_table)
        return elements

    @staticmethod
    def _build_bank_statement_dossier_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float, available_height: float) -> List[Any]:
        elements: List[Any] = []
        if report_payload.get("dossier_mode") != "simple":
            elements.extend(
                FinancialReportService._build_bank_statement_dossier_cover(
                    report_payload=report_payload,
                    styles=styles,
                    available_width=available_width,
                    available_height=available_height,
                )
            )
            elements.append(PageBreak())

        statement_payload = {
            **report_payload,
            "report_type": "bank_statement",
            "title": report_payload.get("statement_title") or "Extrato Bancário",
            "subtitle": report_payload.get("statement_subtitle") or FinancialReportService.REPORT_DEFINITIONS["bank_statement"]["description"],
        }
        elements.extend(
            FinancialReportService._build_bank_statement_pdf_elements(
                report_payload=statement_payload,
                styles=styles,
                available_width=available_width,
            )
        )
        elements.append(PageBreak())
        elements.extend(
            FinancialReportService._build_bank_statement_dossier_documents_pages(
                report_payload=report_payload,
                styles=styles,
                available_width=available_width,
                available_height=available_height,
            )
        )
        return elements

    @staticmethod
    def _build_bank_statement_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        eyebrow_style = ParagraphStyle(
            "BankStatementPdfEyebrow",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#CBD5E1"),
        )
        title_style = ParagraphStyle(
            "BankStatementPdfTitle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=22,
            textColor=colors.white,
            spaceAfter=3,
        )
        subtitle_style = ParagraphStyle(
            "BankStatementPdfSubtitle",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#E2E8F0"),
        )
        section_title_style = ParagraphStyle(
            "BankStatementPdfSectionTitle",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=5,
        )
        filter_cell_style = ParagraphStyle(
            "BankStatementPdfFilterCell",
            parent=styles["BodyText"],
            fontSize=7.5,
            leading=9,
            textColor=colors.HexColor("#0F172A"),
        )
        stat_label_style = ParagraphStyle(
            "BankStatementPdfStatLabel",
            parent=styles["BodyText"],
            fontSize=5.5,
            leading=7,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#475569"),
        )
        stat_value_style = ParagraphStyle(
            "BankStatementPdfStatValue",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=10,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0F172A"),
        )
        table_header_style = ParagraphStyle(
            "BankStatementPdfTableHeader",
            parent=styles["BodyText"],
            fontSize=6.4,
            leading=7.2,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        table_cell_style = ParagraphStyle(
            "BankStatementPdfTableCell",
            parent=styles["BodyText"],
            fontSize=6.3,
            leading=7.2,
            textColor=colors.HexColor("#0F172A"),
        )
        table_cell_center_style = ParagraphStyle(
            "BankStatementPdfTableCellCenter",
            parent=table_cell_style,
            alignment=TA_CENTER,
        )
        table_cell_amount_style = ParagraphStyle(
            "BankStatementPdfTableCellAmount",
            parent=table_cell_style,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        empty_style = ParagraphStyle(
            "BankStatementPdfEmpty",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
        )

        company_name = str(report_payload.get("company_name") or "Versus Gestão Corporativa").strip()
        hero_content = [
            Paragraph("Gestão Financeira · Extrato Bancário", eyebrow_style),
            Paragraph(report_payload.get("title", "Extrato Bancário"), title_style),
        ]
        if company_name:
            hero_content.append(Paragraph(company_name, subtitle_style))
        if report_payload.get("subtitle"):
            hero_content.append(Paragraph(str(report_payload.get("subtitle")), subtitle_style))
        hero_content.append(Paragraph(f"Emitido em {FinancialReportService._pdf_generated_at_label(report_payload)}", subtitle_style))

        hero_table = Table([[hero_content]], colWidths=[available_width], hAlign="LEFT")
        hero_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#0F172A")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 14),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        elements: List[Any] = [hero_table, Spacer(1, 8)]
        elements.append(Paragraph("Filtros aplicados", section_title_style))
        elements.append(
            FinancialReportService._build_schedule_pdf_filter_cards(
                report_filters=report_payload.get("filters") or [],
                available_width=available_width,
                content_style=filter_cell_style,
            )
        )
        elements.append(Spacer(1, 6))
        elements.append(Paragraph("Resumo do período", section_title_style))
        elements.append(
            FinancialReportService._build_schedule_pdf_summary_cards(
                report_summary_cards=report_payload.get("summary_cards") or [],
                available_width=available_width,
                label_style=stat_label_style,
                value_style=stat_value_style,
            )
        )
        elements.append(Spacer(1, 6))
        elements.append(Paragraph("Movimentações", section_title_style))

        report_columns = report_payload.get("columns") or []
        report_rows = report_payload.get("rows") or []
        if report_columns:
            elements.append(
                FinancialReportService._build_bank_statement_pdf_data_table(
                    report_columns=report_columns,
                    report_rows=report_rows,
                    available_width=available_width,
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    cell_center_style=table_cell_center_style,
                    cell_amount_style=table_cell_amount_style,
                )
            )
        else:
            elements.append(Paragraph("Nenhum movimento liquidado encontrado para os filtros informados.", empty_style))

        return elements

    @staticmethod
    def _build_bank_statement_pdf_data_table(
        *,
        report_columns: List[Dict[str, Any]],
        report_rows: List[Dict[str, Any]],
        available_width: float,
        header_style,
        cell_style,
        cell_center_style,
        cell_amount_style,
    ) -> Table:
        width_ratio_map = {
            "data": 0.82,
            "codigo": 0.92,
            "conta_bancaria": 1.18,
            "lancamento": 1.45,
            "descricao": 2.75,
            "favorecido": 1.2,
            "valor": 0.74,
            "conciliacao": 0.86,
            "saldo": 0.72,
        }
        total_ratio = sum(width_ratio_map.get(column.get("key"), 1.0) for column in report_columns) or 1.0
        col_widths = [
            available_width * (width_ratio_map.get(column.get("key"), 1.0) / total_ratio)
            for column in report_columns
        ]

        header_row = [Paragraph(str(column.get("label") or ""), header_style) for column in report_columns]
        body_rows: List[List[Any]] = []

        def _amount_label_without_currency(item: Dict[str, Any], key: str, fallback: str, tone: str) -> str:
            label = str(item.get(f"{key}_label") or fallback or "0").replace("R$", "").replace("+", "").strip()
            if tone == "negative" and label and not label.startswith(("-", "−")):
                return f"- {label}"
            return label

        for item in report_rows:
            row_cells: List[Any] = []
            for column in report_columns:
                key = str(column.get("key") or "")
                raw_value = str(item.get(key, "") or "")
                style = cell_style
                if key in {"data", "codigo", "conciliacao"}:
                    style = cell_center_style
                if key in {"valor", "saldo"}:
                    tone = str(item.get(f"{key}_tone") or item.get("movimento_tone") or "neutral").lower()
                    color = {
                        "positive": "#2563EB",
                        "negative": "#DC2626",
                        "primary": "#2563EB",
                    }.get(tone, "#0F172A")
                    style = ParagraphStyle(
                        f"BankStatementPdfAmount_{key}_{tone}",
                        parent=cell_amount_style,
                        textColor=colors.HexColor(color),
                    )
                    raw_value = _amount_label_without_currency(item, key, raw_value, tone)
                elif key == "movimento":
                    tone = str(item.get("movimento_tone") or "neutral").lower()
                    color = {
                        "positive": "#15803D",
                        "negative": "#B91C1C",
                        "primary": "#1D4ED8",
                    }.get(tone, "#475569")
                    style = ParagraphStyle(
                        f"BankStatementPdfMovement_{tone}",
                        parent=cell_center_style,
                        fontName="Helvetica-Bold",
                        textColor=colors.HexColor(color),
                    )
                elif key == "conciliacao":
                    tone = str(item.get("conciliacao_tone") or "neutral").lower()
                    color = {
                        "positive": "#15803D",
                        "negative": "#B91C1C",
                        "primary": "#1D4ED8",
                    }.get(tone, "#475569")
                    style = ParagraphStyle(
                        f"BankStatementPdfReconciliation_{tone}",
                        parent=cell_center_style,
                        fontName="Helvetica-Bold",
                        textColor=colors.HexColor(color),
                    )
                    raw_value = str(item.get("conciliacao_label") or raw_value)
                row_cells.append(Paragraph(raw_value, style))
            body_rows.append(row_cells)

        data = [header_row] + body_rows
        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for row_index in range(1, len(data)):
            background = colors.HexColor("#FFFFFF") if row_index % 2 else colors.HexColor("#F8FAFC")
            table_styles.append(("BACKGROUND", (0, row_index), (-1, row_index), background))
        table.setStyle(TableStyle(table_styles))
        return table

    @staticmethod
    def _build_schedule_pdf_elements(*, report_payload: Dict[str, Any], styles, available_width: float) -> List[Any]:
        title_style = ParagraphStyle(
            "SchedulePdfTitle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=21,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=4,
        )
        company_style = ParagraphStyle(
            "SchedulePdfCompany",
            parent=styles["BodyText"],
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#334155"),
            spaceAfter=8,
        )
        section_title_style = ParagraphStyle(
            "SchedulePdfSectionTitle",
            parent=styles["BodyText"],
            fontSize=10,
            leading=12,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=6,
        )
        filter_cell_style = ParagraphStyle(
            "SchedulePdfFilterCell",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            textColor=colors.HexColor("#0f172a"),
        )
        stat_label_style = ParagraphStyle(
            "SchedulePdfStatLabel",
            parent=styles["BodyText"],
            fontSize=6,
            leading=7,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#334155"),
        )
        stat_value_style = ParagraphStyle(
            "SchedulePdfStatValue",
            parent=styles["BodyText"],
            fontSize=8.5,
            leading=10,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#0f172a"),
        )
        table_header_style = ParagraphStyle(
            "SchedulePdfTableHeader",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        table_cell_style = ParagraphStyle(
            "SchedulePdfTableCell",
            parent=styles["BodyText"],
            fontSize=7,
            leading=8,
            textColor=colors.HexColor("#0f172a"),
        )
        table_cell_center_style = ParagraphStyle(
            "SchedulePdfTableCellCenter",
            parent=table_cell_style,
            alignment=TA_CENTER,
        )
        table_cell_currency_style = ParagraphStyle(
            "SchedulePdfTableCellCurrency",
            parent=table_cell_style,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )

        elements: List[Any] = [
            Paragraph(report_payload.get("title", "Relatório financeiro"), title_style),
        ]
        company_name = str(report_payload.get("company_name") or "").strip()
        if company_name:
            elements.append(Paragraph(company_name, company_style))
        elements.append(Spacer(1, 6))

        report_filters = report_payload.get("filters") or []
        elements.append(Paragraph("FILTROS APLICADOS", section_title_style))
        elements.append(
            FinancialReportService._build_schedule_pdf_filter_cards(
                report_filters=report_filters,
                available_width=available_width,
                content_style=filter_cell_style,
            )
        )
        elements.append(Spacer(1, 8))

        report_summary_cards = report_payload.get("summary_cards") or []
        elements.append(
            FinancialReportService._build_schedule_pdf_summary_cards(
                report_summary_cards=report_summary_cards,
                available_width=available_width,
                label_style=stat_label_style,
                value_style=stat_value_style,
            )
        )
        elements.append(Spacer(1, 8))

        report_columns = report_payload.get("columns") or []
        report_rows = report_payload.get("rows") or []
        if report_columns:
            elements.append(
                FinancialReportService._build_schedule_pdf_data_table(
                    report_columns=report_columns,
                    report_rows=report_rows,
                    available_width=available_width,
                    header_style=table_header_style,
                    cell_style=table_cell_style,
                    cell_center_style=table_cell_center_style,
                    cell_currency_style=table_cell_currency_style,
                )
            )
        else:
            empty_style = ParagraphStyle(
                "SchedulePdfEmpty",
                parent=styles["BodyText"],
                fontSize=9,
                leading=11,
                alignment=TA_CENTER,
                textColor=colors.HexColor("#475569"),
            )
            elements.append(Paragraph("Nenhum dado encontrado para os filtros informados.", empty_style))

        return elements

    @staticmethod
    def _build_schedule_pdf_filter_cards(*, report_filters: List[Dict[str, Any]], available_width: float, content_style) -> Table:
        items = report_filters or [{"label": "Resumo", "value": "Sem filtros adicionais."}]
        cols = 3
        gutter = 6
        col_width = (available_width - (gutter * (cols - 1))) / cols
        cells: List[Any] = []
        for item in items:
            label = str(item.get("label") or "").strip()
            value = str(item.get("value") or "-").strip()
            cells.append(Paragraph(f"<b>{label}</b><br/>{value}", content_style))
        while len(cells) % cols != 0:
            cells.append(Paragraph("", content_style))
        matrix = [cells[index:index + cols] for index in range(0, len(cells), cols)]
        table = Table(matrix, colWidths=[col_width] * cols, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                    ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#e2e8f0")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return table

    @staticmethod
    def _build_schedule_pdf_summary_cards(*, report_summary_cards: List[Dict[str, Any]], available_width: float, label_style, value_style) -> Table:
        cards = report_summary_cards or []
        cols = min(max(len(cards), 1), 7)
        gutter = 4
        col_width = (available_width - (gutter * (cols - 1))) / cols
        rows: List[List[Any]] = []
        current_row: List[Any] = []
        tone_color_codes = {
            "positive": "#16a34a",
            "negative": "#dc2626",
            "primary": "#2563eb",
            "neutral": "#0f172a",
        }
        for card in cards:
            tone = str(card.get("tone") or "neutral").lower()
            value = Paragraph(
                f"<font color='{tone_color_codes.get(tone, '#0f172a')}'>{card.get('value', '-')}</font>",
                value_style,
            )
            label = Paragraph(str(card.get("label") or "-").upper(), label_style)
            current_row.append(Table([[value], [label]], colWidths=[col_width]))
            if len(current_row) == cols:
                rows.append(current_row)
                current_row = []
        if current_row:
            while len(current_row) < cols:
                current_row.append(Paragraph("", value_style))
            rows.append(current_row)

        table = Table(rows, colWidths=[col_width] * cols, hAlign="LEFT")
        base_styles = [
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                if isinstance(cell, Table):
                    cell.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ffffff")),
                                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
                                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                                ("TOPPADDING", (0, 0), (-1, -1), 4),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                            ]
                        )
                    )
        table.setStyle(TableStyle(base_styles))
        return table

    @staticmethod
    def _build_schedule_pdf_data_table(
        *,
        report_columns: List[Dict[str, Any]],
        report_rows: List[Dict[str, Any]],
        available_width: float,
        header_style,
        cell_style,
        cell_center_style,
        cell_currency_style,
    ) -> Table:
        width_ratio_map = {
            "title_number": 1.0,
            "counterparty": 1.7,
            "title_amount": 1.0,
            "balance_amount": 1.0,
            "competence_date": 1.0,
            "due_date": 1.0,
            "settlement_date": 1.15,
        }
        total_ratio = sum(width_ratio_map.get(column.get("key"), 1.0) for column in report_columns) or 1.0
        col_widths = [
            available_width * (width_ratio_map.get(column.get("key"), 1.0) / total_ratio)
            for column in report_columns
        ]

        header_row = [Paragraph(str(column.get("label") or ""), header_style) for column in report_columns]
        body_rows: List[List[Any]] = []
        for item in report_rows:
            row_cells: List[Any] = []
            for column in report_columns:
                key = column.get("key")
                raw_value = str(item.get(key, "") or "")
                if key in {"title_amount", "balance_amount"}:
                    style = cell_currency_style
                elif key in {"competence_date", "due_date", "settlement_date"}:
                    style = cell_center_style
                else:
                    style = cell_style
                row_cells.append(Paragraph(raw_value, style))
            body_rows.append(row_cells)

        data = [header_row] + body_rows
        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_index in range(1, len(data)):
            background = colors.HexColor("#ffffff") if row_index % 2 else colors.HexColor("#f8fafc")
            table_styles.append(("BACKGROUND", (0, row_index), (-1, row_index), background))
        table.setStyle(TableStyle(table_styles))
        return table
