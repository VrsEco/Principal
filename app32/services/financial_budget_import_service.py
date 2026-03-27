from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Sequence, Tuple

from models import FinancialBudgetVersion, FinancialChartAccount, FinancialCostCenter
from services.financial_budget_service import FinancialBudgetService
from services.financial_service import FinancialService


class FinancialBudgetImportService:
    FIXED_COLUMNS = [
        ("line_code", "Código da Linha *"),
        ("line_name", "Nome da Linha *"),
        ("budget_view", "Visão Orçamentária *"),
        ("movement_nature", "Natureza *"),
        ("chart_account_code", "Código Conta Contábil"),
        ("cost_center_code", "Código Centro de Custo"),
        ("notes", "Observações"),
    ]

    @staticmethod
    def import_matrix_file(
        *,
        company_id: int,
        version_id: int,
        file_name: str,
        file_bytes: bytes,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        version = FinancialBudgetVersion.query.filter(
            FinancialBudgetVersion.id == version_id,
            FinancialBudgetVersion.company_id == company_id,
            FinancialBudgetVersion.deleted_at.is_(None),
        ).first()
        if not version:
            return None, "Versão orçamentária não encontrada no escopo da empresa."

        rows, error = FinancialBudgetImportService._parse_rows(file_name=file_name, file_bytes=file_bytes)
        if error:
            return None, error
        if not rows:
            return None, "A planilha do orçamento está vazia."

        payload_lines, error = FinancialBudgetImportService._normalize_rows(
            company_id=company_id,
            version=version,
            rows=rows,
        )
        if error:
            return None, error

        result, error = FinancialBudgetService.upsert_matrix(
            payload={
                "company_id": company_id,
                "version_id": version_id,
                "lines": payload_lines,
            },
            allowed_company_ids=allowed_company_ids,
        )
        if error:
            return None, error

        result["import_summary"] = {
            "file_name": file_name,
            "rows_received": len(rows),
            "lines_imported": len(payload_lines),
        }
        return result, None

    @staticmethod
    def build_template(
        *,
        company_id: int,
        version_id: int,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[bytes], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        version = FinancialBudgetVersion.query.filter(
            FinancialBudgetVersion.id == version_id,
            FinancialBudgetVersion.company_id == company_id,
            FinancialBudgetVersion.deleted_at.is_(None),
        ).first()
        if not version:
            return None, "Versão orçamentária não encontrada no escopo da empresa."

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            return None, f"openpyxl não está disponível para geração do modelo XLSX: {exc}"

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Orçamento"
        header_fill = PatternFill("solid", fgColor="0F766E")

        month_columns = list(FinancialBudgetImportService._iter_months(version.period_start, version.period_end))
        headers = [label for _key, label in FinancialBudgetImportService.FIXED_COLUMNS] + [month.strftime("%Y-%m") for month in month_columns]

        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[FinancialBudgetImportService._column_letter(idx)].width = 18

        sample_values = [
            "DESP-ADM-001",
            "Despesas Administrativas",
            "competence",
            "debit",
            "",
            "",
            "Linha de exemplo",
        ] + [0 for _ in month_columns]

        for idx, value in enumerate(sample_values, start=1):
            ws.cell(row=2, column=idx, value=value)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), None

    @staticmethod
    def _parse_rows(*, file_name: str, file_bytes: bytes) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        extension = (file_name or "").lower().rsplit(".", 1)[-1]
        if extension != "xlsx":
            return None, "A importação do orçamento matricial aceita apenas arquivos XLSX."
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            return None, f"openpyxl não está disponível para leitura XLSX: {exc}"

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], None

        headers = [str(item).strip() if item is not None else f"column_{idx+1}" for idx, item in enumerate(rows[0])]
        payload: List[Dict[str, Any]] = []
        for raw in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                item[header] = raw[idx] if idx < len(raw) else None
            if any(value not in (None, "") for value in item.values()):
                payload.append(item)
        return payload, None

    @staticmethod
    def _normalize_rows(
        *,
        company_id: int,
        version: FinancialBudgetVersion,
        rows: List[Dict[str, Any]],
    ) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        month_index = {month.strftime("%Y-%m"): month for month in FinancialBudgetImportService._iter_months(version.period_start, version.period_end)}
        payload_lines: List[Dict[str, Any]] = []

        for idx, row in enumerate(rows, start=2):
            normalized = {str(key).strip().lower(): value for key, value in row.items()}
            line_code = FinancialBudgetImportService._clean_text(
                normalized.get("código da linha *") or normalized.get("codigo da linha *") or normalized.get("line_code")
            )
            line_name = FinancialBudgetImportService._clean_text(
                normalized.get("nome da linha *") or normalized.get("line_name")
            )
            budget_view = FinancialBudgetImportService._clean_text(
                normalized.get("visão orçamentária *") or normalized.get("visao orcamentária *") or normalized.get("visao orcamentaria *") or normalized.get("budget_view")
            ) or "competence"
            movement_nature = FinancialBudgetImportService._clean_text(
                normalized.get("natureza *") or normalized.get("movement_nature")
            ) or "debit"

            if not line_code or not line_name:
                return None, f"Linha {idx}: código e nome da linha são obrigatórios."

            chart_account_code = FinancialBudgetImportService._clean_text(
                normalized.get("código conta contábil") or normalized.get("codigo conta contabil") or normalized.get("chart_account_code")
            )
            cost_center_code = FinancialBudgetImportService._clean_text(
                normalized.get("código centro de custo") or normalized.get("codigo centro de custo") or normalized.get("cost_center_code")
            )
            chart_account_id = FinancialBudgetImportService._resolve_chart_account(company_id, chart_account_code)
            cost_center_id = FinancialBudgetImportService._resolve_cost_center(company_id, cost_center_code)
            if chart_account_code and not chart_account_id:
                return None, f"Linha {idx}: conta contábil com código {chart_account_code} não encontrada."
            if cost_center_code and not cost_center_id:
                return None, f"Linha {idx}: centro de custo com código {cost_center_code} não encontrado."

            amounts = []
            for key, month in month_index.items():
                raw_value = normalized.get(key.lower()) if key.lower() in normalized else normalized.get(key)
                if raw_value in (None, ""):
                    raw_value = 0
                amount = FinancialBudgetImportService._parse_decimal(raw_value)
                if amount is None:
                    return None, f"Linha {idx}: valor inválido para o mês {key}."
                amounts.append({"period_month": month.isoformat(), "budget_amount": float(amount)})

            payload_lines.append(
                {
                    "line_code": line_code,
                    "line_name": line_name,
                    "budget_view": budget_view,
                    "movement_nature": movement_nature,
                    "planned_amount": float(sum((Decimal(str(item["budget_amount"])) for item in amounts), Decimal("0"))),
                    "chart_account_id": chart_account_id,
                    "cost_center_id": cost_center_id,
                    "notes": FinancialBudgetImportService._clean_text(normalized.get("observações") or normalized.get("observacoes") or normalized.get("notes")),
                    "amounts": amounts,
                }
            )
        return payload_lines, None

    @staticmethod
    def _resolve_chart_account(company_id: int, code: Optional[str]) -> Optional[int]:
        if not code:
            return None
        item = FinancialChartAccount.query.filter(
            FinancialChartAccount.company_id == company_id,
            FinancialChartAccount.deleted_at.is_(None),
            FinancialChartAccount.code == code,
        ).first()
        return item.id if item else None

    @staticmethod
    def _resolve_cost_center(company_id: int, code: Optional[str]) -> Optional[int]:
        if not code:
            return None
        item = FinancialCostCenter.query.filter(
            FinancialCostCenter.company_id == company_id,
            FinancialCostCenter.deleted_at.is_(None),
            FinancialCostCenter.code == code,
        ).first()
        return item.id if item else None

    @staticmethod
    def _iter_months(start_date: date, end_date: date):
        cursor = start_date.replace(day=1)
        limit = end_date.replace(day=1)
        while cursor <= limit:
            yield cursor
            year = cursor.year + (1 if cursor.month == 12 else 0)
            month = 1 if cursor.month == 12 else cursor.month + 1
            cursor = cursor.replace(year=year, month=month, day=1)

    @staticmethod
    def _parse_decimal(value: Any) -> Optional[Decimal]:
        if isinstance(value, Decimal):
            return value
        if value in (None, ""):
            return Decimal("0")
        text = str(value).strip().replace("R$", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _clean_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _column_letter(index: int) -> str:
        result = ""
        while index > 0:
            index, rem = divmod(index - 1, 26)
            result = chr(65 + rem) + result
        return result
