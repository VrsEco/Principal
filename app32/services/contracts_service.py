from __future__ import annotations

import re
import csv
import io
import mimetypes
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
import calendar
from pathlib import Path
from typing import Optional
from uuid import uuid4
import xml.etree.ElementTree as ET

from flask import current_app
from sqlalchemy import or_
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from models import Company, Employee, Project, db
from models.automation import AutomationRegistry, AutomationRule
from models.contracts import (
    Contract,
    ContractBillingItem,
    ContractCatalogItem,
    ContractClause,
    ContractDocument,
    ContractEvent,
    ContractFinancialTerm,
    ContractFiscalTerm,
    ContractItem,
    ContractNativeBilling,
    ContractNativeBillingItem,
    ContractNote,
    ContractParty,
    ContractingLegalEntity,
    ContractRetention,
    ContractTrigger,
)
from models.financial import (
    FinancialAssetAccount,
    FinancialBankAccount,
    FinancialCorrectionIndex,
    FinancialCounterparty,
    FinancialChartAccount,
    FinancialCostCenter,
    FinancialPaymentMethod,
    FinancialSatelliteExecution,
    FinancialSchedule,
    FinancialScheduleLink,
)
from services.contracts_catalog_service import ContractsCatalogService
from services.contract_fiscal_invoice_spreadsheet import (
    NFSE_XLSX_MIMETYPE,
    build_nfse_integration_workbook,
)


class ContractService:
    ACTIVE_STATUSES = {"active", "signed", "implanting"}
    INACTIVE_STATUSES = {"inactive", "closed", "draft", "suspended"}
    TAB_REGISTRY = (
        {"key": "cliente", "label": "Cliente", "scope": "core", "description": "Favorecido cliente vinculado ao contrato."},
        {"key": "itens", "label": "Itens do Contrato", "scope": "core", "description": "Escopo, serviços e itens negociados."},
        {"key": "clausulas", "label": "Cláusulas", "scope": "core", "description": "Redação, cláusulas e observações estruturadas do contrato."},
        {"key": "faturamento", "label": "Faturamento", "scope": "core", "description": "Itens e regras de faturamento."},
        {"key": "periodicidade", "label": "Agenda Nativa", "scope": "core", "description": "Datas-base, competência, faturamento e renovação nativos do contrato."},
        {"key": "automacoes", "label": "Automações", "scope": "core", "description": "Automações nativas do contrato com visão simplificada."},
        {"key": "financeiro", "label": "Financeiro", "scope": "core", "description": "Títulos, satélites e regras automáticas do contrato."},
        {"key": "fiscal", "label": "Fiscal", "scope": "core", "description": "Perfil fiscal e retenções."},
        {"key": "observacoes", "label": "Observações", "scope": "core", "description": "Contexto operacional e observações livres."},
        {"key": "historico", "label": "Histórico", "scope": "core", "description": "Linha do tempo do contrato, observações e eventos relevantes."},
        {"key": "gerar_pdf", "label": "Gerar / Editar Contrato", "scope": "capability", "description": "Upload/controle da versão em PDF do contrato."},
        {"key": "contrato_assinado", "label": "Contrato Assinado", "scope": "capability", "description": "Upload da via assinada escaneada."},
        {"key": "documentos", "label": "Documentos / Anexos", "scope": "capability", "description": "Artefatos gerais vinculados ao contrato."},
    )
    OPERATIONAL_PROFILE_FULL = "full_contract"
    OPERATIONAL_PROFILE_BILLING_FISCAL = "billing_fiscal"
    OPERATIONAL_PROFILE_CONFIG = {
        OPERATIONAL_PROFILE_FULL: {
            "label": "Contrato completo",
            "description": "Usa todas as faces operacionais do contrato dentro do Gestão Versus.",
            "visible_tabs": [item["key"] for item in TAB_REGISTRY],
        },
        OPERATIONAL_PROFILE_BILLING_FISCAL: {
            "label": "Faturamento / Fiscal / Financeiro",
            "description": "Contrato controlado parcialmente em sistema externo, mantendo aqui serviços, agenda, fiscal, faturamento e financeiro.",
            "visible_tabs": [
                "cliente",
                "itens",
                "faturamento",
                "periodicidade",
                "financeiro",
                "fiscal",
                "observacoes",
                "historico",
            ],
        },
    }
    ITEM_RETENTION_OPTIONS = (
        ("iss", "ISS"),
        ("irrf", "IRRF"),
        ("csrf", "CSRF"),
        ("inss", "INSS"),
        ("other", "Outras Retenções"),
    )
    ITEM_RETENTION_DEDUCTION_MODES = (
        ("percent", "%"),
        ("amount", "Valor"),
    )
    ITEM_RETENTION_VALUE_MODES = (
        ("percent", "%"),
        ("amount", "Valor"),
    )

    @staticmethod
    def _normalize_text(value: object) -> str:
        return str(value or "").strip()

    @staticmethod
    def _normalize_bool(value: object) -> bool:
        if isinstance(value, bool):
            return value
        return str(value or "").strip().lower() in {"1", "true", "on", "yes", "sim"}

    @staticmethod
    def _normalize_decimal(value: object, *, default: str = "0") -> Decimal:
        raw = str(value if value not in (None, "") else default).strip()
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        try:
            return Decimal(raw)
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    @staticmethod
    def _normalize_int(value: object) -> Optional[int]:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return int(text)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _normalize_operational_profile(value: object) -> str:
        profile = ContractService._normalize_text(value).lower()
        if profile in ContractService.OPERATIONAL_PROFILE_CONFIG:
            return profile
        return ContractService.OPERATIONAL_PROFILE_FULL

    @staticmethod
    def _normalize_date(value: object) -> Optional[date]:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            return None

    @staticmethod
    def _normalize_metadata_dict(value: object) -> dict:
        if isinstance(value, dict):
            return dict(value)
        return {}

    @staticmethod
    def _normalize_metadata_dict_list(value: object) -> list[dict]:
        if not isinstance(value, list):
            return []
        return [dict(item) for item in value if isinstance(item, dict)]

    @staticmethod
    def _serialize_date(value: Optional[date]) -> Optional[str]:
        return value.isoformat() if value else None

    @staticmethod
    def infer_document_type(value: object) -> Optional[str]:
        digits = re.sub(r"\D", "", str(value or ""))
        if len(digits) == 11:
            return "cpf"
        if len(digits) == 14:
            return "cnpj"
        return None

    @staticmethod
    def calculate_total_price(quantity: object, unit_price: object) -> Decimal:
        qty = ContractService._normalize_decimal(quantity)
        price = ContractService._normalize_decimal(unit_price)
        return (qty * price).quantize(Decimal("0.01"))

    @staticmethod
    def _resolve_company_code(company_id: int) -> str:
        company = ContractService.get_company(company_id)
        raw_code = ContractService._normalize_text(getattr(company, "client_code", ""))
        raw_name = ContractService._normalize_text(getattr(company, "name", ""))

        sanitized = re.sub(r"[^A-Z0-9]", "", raw_code.upper())
        if not sanitized and raw_name:
            tokens = [token[0] for token in re.findall(r"[A-Za-z0-9]+", raw_name.upper()) if token]
            sanitized = "".join(tokens)
        if not sanitized:
            sanitized = str(company_id or "XX")

        sanitized = (sanitized[:2] if len(sanitized) >= 2 else sanitized.ljust(2, "X")).upper()
        return sanitized

    @staticmethod
    def _next_structured_code(model, company_id: int, marker: str) -> str:
        company_code = ContractService._resolve_company_code(company_id)
        normalized_marker = re.sub(r"[^A-Z0-9]", "", str(marker or "").upper())[:1] or "X"
        code_prefix = f"{company_code}.{normalized_marker}."
        last_number = 0
        rows = model.query.with_entities(model.code).filter(model.company_id == company_id).all()
        for (code,) in rows:
            normalized_code = str(code or "").strip().upper()
            if not normalized_code.startswith(code_prefix):
                continue
            match = re.search(r"(\d+)$", normalized_code)
            if match:
                last_number = max(last_number, int(match.group(1)))
        return f"{code_prefix}{last_number + 1:03d}"

    @staticmethod
    def get_company(company_id: int) -> Optional[Company]:
        return Company.query.get(company_id)

    @staticmethod
    def get_dashboard(company_id: int) -> dict:
        contracts_query = Contract.query.filter(Contract.company_id == company_id, Contract.deleted_at.is_(None))
        parties_query = ContractParty.query.filter(ContractParty.company_id == company_id, ContractParty.deleted_at.is_(None))
        return {
            "counts": {
                "contracts": contracts_query.count(),
                "drafts": contracts_query.filter(Contract.status == "draft").count(),
                "active": contracts_query.filter(Contract.status.in_(["active", "signed", "implanting"])).count(),
                "parties": parties_query.count(),
            },
            "latest_contracts": contracts_query.order_by(Contract.updated_at.desc()).limit(8).all(),
            "latest_parties": parties_query.order_by(ContractParty.updated_at.desc()).limit(8).all(),
        }

    @staticmethod
    def list_parties(company_id: int):
        ContractService.sync_parties_from_counterparties(company_id)
        return (
            ContractParty.query.filter(ContractParty.company_id == company_id, ContractParty.deleted_at.is_(None))
            .order_by(ContractParty.name.asc())
            .all()
        )

    @staticmethod
    def list_customer_parties(company_id: int):
        ContractService.sync_parties_from_counterparties(company_id, only_customer=True)
        return (
            ContractParty.query.filter(
                ContractParty.company_id == company_id,
                ContractParty.deleted_at.is_(None),
                ContractParty.is_customer.is_(True),
            )
            .order_by(ContractParty.name.asc())
            .all()
        )

    @staticmethod
    def list_contracts(company_id: int):
        ContractService.sync_parties_from_counterparties(company_id)
        return (
            Contract.query.filter(Contract.company_id == company_id, Contract.deleted_at.is_(None))
            .order_by(Contract.updated_at.desc())
            .all()
        )

    @staticmethod
    def list_contracts_filtered(company_id: int, filters: Optional[dict] = None):
        ContractService.sync_parties_from_counterparties(company_id)
        query = Contract.query.filter(Contract.company_id == company_id, Contract.deleted_at.is_(None))
        filters = filters or {}

        status = ContractService._normalize_text(filters.get("status")).lower()
        if status:
            query = query.filter(Contract.status == status)

        party_id = ContractService._normalize_int(filters.get("party_id"))
        if party_id:
            query = query.filter(Contract.party_id == party_id)

        manager_employee_id = ContractService._normalize_int(filters.get("manager_employee_id"))
        if manager_employee_id:
            query = query.filter(Contract.manager_employee_id == manager_employee_id)

        search = ContractService._normalize_text(filters.get("search"))
        if search:
            ilike = f"%{search}%"
            query = query.join(ContractParty, ContractParty.id == Contract.party_id).filter(
                or_(
                    Contract.code.ilike(ilike),
                    Contract.title.ilike(ilike),
                    ContractParty.name.ilike(ilike),
                )
            )

        return query.order_by(Contract.updated_at.desc(), Contract.id.desc()).all()

    @staticmethod
    def get_contracts_kpis(company_id: int) -> dict:
        contracts = ContractService.list_contracts(company_id)
        today = date.today()
        return {
            "total": len(contracts),
            "active": sum(1 for item in contracts if ContractService.get_contract_status_group(item) == "active"),
            "renewing": sum(1 for item in contracts if item.renewal_date and item.renewal_date >= today),
            "pending_billing": sum(1 for item in contracts if not item.last_billing_at and ContractService.get_contract_status_group(item) == "active"),
        }

    @staticmethod
    def get_contract_next_action(contract: Contract) -> dict:
        today = date.today()
        native_schedule = ContractService.get_native_schedule_overview(contract)
        next_event = native_schedule.get("next_event")
        if next_event and next_event.get("date") and next_event["date"] >= today:
            label_map = {
                "billing": "Faturar próxima competência",
                "billing_start": "Iniciar faturamento",
                "renewal": "Renovar contrato",
                "adjustment": "Aplicar reajuste",
                "termination": "Planejar encerramento",
                "alert": "Acompanhar alerta contratual",
            }
            return {"label": label_map.get(next_event["event_type"], next_event["label"]), "tone": next_event.get("tone", "info")}
        if contract.termination_date and contract.termination_date <= today:
            return {"label": "Encerrado", "tone": "neutral"}
        if contract.renewal_date and contract.renewal_date <= today:
            return {"label": "Renovar contrato", "tone": "warning"}
        if contract.last_billing_at is None and ContractService.get_contract_status_group(contract) == "active":
            return {"label": "Faturar primeira competência", "tone": "info"}
        if contract.adjustment_date and contract.adjustment_date <= today:
            return {"label": "Aplicar reajuste", "tone": "warning"}
        return {"label": "Acompanhar execução", "tone": "success"}

    @staticmethod
    def get_native_trigger_type_options() -> list[tuple[str, str]]:
        return [
            ("billing", "Faturamento"),
            ("renewal", "Renovação"),
            ("adjustment", "Reajuste"),
            ("termination", "Encerramento"),
            ("alert", "Alerta geral"),
        ]

    @staticmethod
    def get_reference_date_type_options() -> list[tuple[str, str]]:
        return [
            ("signed_at", "Assinatura"),
            ("service_start_at", "Início dos serviços"),
            ("billing_start_at", "Início do faturamento"),
            ("renewal_date", "Data de renovação"),
            ("adjustment_date", "Data de reajuste"),
            ("termination_date", "Data de finalização"),
            ("manual", "Data manual"),
        ]

    @staticmethod
    def get_contract_type_options() -> list[tuple[str, str]]:
        return [
            ("prestacao", "Prestação de serviços"),
            ("licenciamento", "Licenciamento"),
            ("manutencao_suporte", "Manutenção / suporte"),
            ("fornecimento", "Fornecimento"),
            ("consultoria", "Consultoria"),
            ("outsourcing", "Outsourcing"),
            ("locacao", "Locação"),
            ("parceria", "Parceria"),
        ]

    @staticmethod
    def get_currency_options() -> list[tuple[str, str]]:
        return [
            ("BRL", "BRL · Real"),
            ("USD", "USD · Dólar"),
            ("EUR", "EUR · Euro"),
        ]

    @staticmethod
    def get_periodicity_options() -> list[tuple[str, str]]:
        return [
            ("monthly", "Mensal"),
            ("weekly", "Semanal"),
            ("quarterly", "Trimestral"),
            ("semiannual", "Semestral"),
            ("annual", "Anual"),
        ]

    @staticmethod
    def get_competence_rule_options() -> list[tuple[str, str]]:
        return [
            ("mes atual", "Mês atual"),
            ("mes anterior", "Mês anterior"),
            ("antecipado", "Antecipado"),
            ("sob demanda", "Sob demanda"),
        ]

    @staticmethod
    def get_renewal_rule_options() -> list[tuple[str, str]]:
        return [
            ("manual", "Manual"),
            ("auto", "Automática"),
            ("aditivo", "Por aditivo"),
        ]

    @staticmethod
    def get_due_rule_reference_options() -> list[tuple[str, str]]:
        return [
            ("issue_month", "Mês da emissão"),
            ("issue_month_plus_1", "1º mês após a emissão"),
            ("issue_month_plus_2", "2º mês após a emissão"),
        ]

    @staticmethod
    def build_due_rule(*, reference: object, day: object) -> Optional[str]:
        reference_value = ContractService._normalize_text(reference)
        day_value = ContractService._normalize_int(day)
        if reference_value not in {"issue_month", "issue_month_plus_1", "issue_month_plus_2"} or not day_value:
            return None
        day_value = max(1, min(day_value, 31))
        return f"{reference_value}:{day_value:02d}"

    @staticmethod
    def parse_due_rule(value: object) -> dict:
        raw_value = ContractService._normalize_text(value)
        if ":" in raw_value:
            reference, day_text = raw_value.split(":", 1)
            day_value = ContractService._normalize_int(day_text)
            if reference in {"issue_month", "issue_month_plus_1", "issue_month_plus_2"} and day_value:
                return {
                    "reference": reference,
                    "day": max(1, min(day_value, 31)),
                    "label": f"{dict(ContractService.get_due_rule_reference_options()).get(reference, reference)} · dia {max(1, min(day_value, 31))}",
                    "is_structured": True,
                }
        return {
            "reference": None,
            "day": None,
            "label": raw_value or "-",
            "is_structured": False,
        }

    @staticmethod
    def _add_months(base_date: date, months: int) -> date:
        month_index = (base_date.month - 1) + months
        year = base_date.year + (month_index // 12)
        month = (month_index % 12) + 1
        day = min(base_date.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)

    @staticmethod
    def _month_end(base_date: date) -> date:
        last_day = calendar.monthrange(base_date.year, base_date.month)[1]
        return date(base_date.year, base_date.month, last_day)

    @staticmethod
    def _periodicity_month_interval(periodicity: object) -> int:
        periodicity_key = ContractService._normalize_text(periodicity).lower()
        return {
            "quarterly": 3,
            "trimestral": 3,
            "semiannual": 6,
            "semestral": 6,
            "annual": 12,
            "anual": 12,
        }.get(periodicity_key, 1)

    @staticmethod
    def resolve_due_date(*, issue_date: date, due_rule: object) -> Optional[date]:
        parsed = ContractService.parse_due_rule(due_rule)
        if not parsed["is_structured"] or not parsed["reference"] or not parsed["day"]:
            return None
        month_offset_map = {
            "issue_month": 0,
            "issue_month_plus_1": 1,
            "issue_month_plus_2": 2,
        }
        target_base = ContractService._add_months(issue_date, month_offset_map[parsed["reference"]])
        last_day = calendar.monthrange(target_base.year, target_base.month)[1]
        return date(target_base.year, target_base.month, min(parsed["day"], last_day))

    @staticmethod
    def get_contract_automation_template_options() -> list[dict]:
        return [
            {
                "key": "generate_billing_monthly",
                "label": "Faturamento mensal",
                "description": "Gera competência nativa mensal do contrato.",
            },
            {
                "key": "renewal_alert_before_date",
                "label": "Alerta de renovação",
                "description": "Avisa antes da data de renovação do contrato.",
            },
            {
                "key": "adjustment_on_date",
                "label": "Reajuste automático",
                "description": "Dispara ação de reajuste na data-base do contrato.",
            },
        ]

    @staticmethod
    def sync_parties_from_counterparties(company_id: int, *, only_customer: bool = False) -> None:
        counterparties = FinancialCounterparty.query.filter(
            FinancialCounterparty.company_id == company_id,
            FinancialCounterparty.deleted_at.is_(None),
        ).all()
        changed = False
        for counterparty in counterparties:
            metadata = dict(counterparty.metadata_json or {})
            is_customer = bool(metadata.get("is_customer"))
            is_supplier = bool(metadata.get("is_supplier"))
            if only_customer and not is_customer:
                continue
            if not is_customer and not is_supplier:
                continue
            party = ContractParty.query.filter(
                ContractParty.company_id == company_id,
                ContractParty.financial_counterparty_id == counterparty.id,
                ContractParty.deleted_at.is_(None),
            ).first()
            payload = {
                "name": counterparty.name,
                "legal_name": counterparty.legal_name,
                "document_type": ContractService.infer_document_type(counterparty.document_number),
                "document_number": counterparty.document_number,
                "email": counterparty.email,
                "phone": counterparty.phone,
                "is_customer": is_customer,
                "is_supplier": is_supplier,
                "status": "active" if counterparty.is_active else "inactive",
                "notes": counterparty.notes,
                "financial_counterparty_id": counterparty.id,
            }
            if party is None:
                party = ContractParty(
                    company_id=company_id,
                    code=ContractService._next_structured_code(ContractParty, company_id, "F"),
                )
                ContractService.update_party(party=party, payload=payload, user_id=None, is_new=True)
                db.session.add(party)
                changed = True
                continue
            before = (
                party.name,
                party.legal_name,
                party.document_type,
                party.document_number,
                party.email,
                party.phone,
                bool(party.is_customer),
                bool(party.is_supplier),
                party.status,
                party.notes,
                party.financial_counterparty_id,
            )
            ContractService.update_party(party=party, payload=payload, user_id=None, is_new=True)
            after = (
                party.name,
                party.legal_name,
                party.document_type,
                party.document_number,
                party.email,
                party.phone,
                bool(party.is_customer),
                bool(party.is_supplier),
                party.status,
                party.notes,
                party.financial_counterparty_id,
            )
            if before != after:
                changed = True
        if changed:
            db.session.commit()

    @staticmethod
    def get_contract_start_date(contract: Contract) -> Optional[date]:
        return contract.service_start_at or contract.billing_start_at or contract.signed_at

    @staticmethod
    def get_contract_status_group(contract: Contract) -> str:
        status = ContractService._normalize_text(getattr(contract, "status", "")).lower()
        return "active" if status in ContractService.ACTIVE_STATUSES else "inactive"

    @staticmethod
    def get_contract_status_label(contract: Contract) -> str:
        status = ContractService._normalize_text(getattr(contract, "status", "")).lower()
        label_map = {
            "draft": "Rascunho",
            "active": "Ativo",
            "signed": "Ativo",
            "implanting": "Implantação",
            "inactive": "Inativo",
            "suspended": "Suspenso",
            "closed": "Encerrado",
        }
        return label_map.get(status, "Ativo" if ContractService.get_contract_status_group(contract) == "active" else "Inativo")

    @staticmethod
    def get_contract_workspace_summary(contract: Optional[Contract]) -> dict:
        if contract is None:
            return {}

        contract_items = contract.items.order_by(ContractItem.order_index.asc(), ContractItem.id.asc()).all()
        billing_items = contract.billing_items.order_by(ContractBillingItem.order_index.asc(), ContractBillingItem.id.asc()).all()
        total_contract_value = sum((item.total_price or Decimal("0")) for item in contract_items)
        total_retention_value = Decimal("0.00")
        for item in contract_items:
            retention_summary = dict((item.metadata_json or {}).get("retention_summary") or {})
            total_retention_value += ContractService._normalize_decimal(retention_summary.get("total_retention_amount"))
        total_billing_value = sum((item.amount or Decimal("0")) for item in billing_items)
        total_contract_value = total_contract_value.quantize(Decimal("0.01")) if contract_items else Decimal("0.00")
        total_retention_value = total_retention_value.quantize(Decimal("0.01")) if contract_items else Decimal("0.00")
        net_contract_value = (total_contract_value - total_retention_value).quantize(Decimal("0.01")) if contract_items else Decimal("0.00")

        return {
            "status_group": ContractService.get_contract_status_group(contract),
            "status_label": ContractService.get_contract_status_label(contract),
            "start_date": ContractService.get_contract_start_date(contract),
            "renewal_date": contract.renewal_date,
            "adjustment_date": contract.adjustment_date,
            "termination_date": contract.termination_date,
            "manager_employee_id": contract.manager_employee_id,
            "contract_item_count": len(contract_items),
            "billing_item_count": len(billing_items),
            "clause_count": contract.clauses.count(),
            "note_count": contract.notes_log.count(),
            "event_count": contract.events.count(),
            "total_contract_value": total_contract_value,
            "total_retention_value": total_retention_value,
            "net_contract_value": net_contract_value,
            "total_billing_value": total_billing_value.quantize(Decimal("0.01")) if billing_items else Decimal("0.00"),
            "updated_at": contract.updated_at,
            "created_at": contract.created_at,
            "operational_profile": ContractService.get_contract_operational_profile(contract),
            "operational_profile_label": ContractService.get_operational_profile_label(contract),
        }

    @staticmethod
    def build_contract_review_flags(contract: Contract) -> dict:
        financial_terms = ContractFinancialTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
        fiscal_terms = ContractFiscalTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
        fiscal_ok = bool(
            contract.contracting_legal_entity_id
            and fiscal_terms
            and fiscal_terms.integration_mode
        )
        return {
            "cliente": bool(contract.party_id),
            "itens": "OK" if contract.items.count() else "Pendente",
            "faturamento": "OK" if contract.billing_items.count() else "Pendente",
            "periodicidade": bool(contract.periodicity or contract.renewal_date or contract.adjustment_date or contract.triggers.count()),
            "fiscal": fiscal_ok,
            "financeiro": bool(financial_terms),
            "cobranca": bool(financial_terms),
            "pdf": "OK" if contract.documents.filter(ContractDocument.document_type == "pdf_gerado").count() else "Pendente",
            "assinado": "OK" if contract.documents.filter(
                or_(ContractDocument.document_type == "contrato_assinado", ContractDocument.is_signed_version.is_(True))
            ).count() else "Pendente",
        }

    @staticmethod
    def _event_date_payload(event_date: Optional[date], label: str, event_type: str, source: str) -> Optional[dict]:
        if not event_date:
            return None
        today = date.today()
        if event_date < today:
            tone = "warning"
        elif event_date == today:
            tone = "info"
        else:
            tone = "success"
        return {
            "event_type": event_type,
            "label": label,
            "date": event_date,
            "date_label": event_date.strftime("%d/%m/%Y"),
            "tone": tone,
            "source": source,
        }

    @staticmethod
    def _resolve_trigger_reference_date(contract: Contract, trigger: ContractTrigger) -> Optional[date]:
        reference_type = ContractService._normalize_text(trigger.reference_date_type)
        if reference_type == "manual":
            return trigger.reference_date_value
        return {
            "signed_at": contract.signed_at,
            "service_start_at": contract.service_start_at,
            "billing_start_at": contract.billing_start_at,
            "renewal_date": contract.renewal_date,
            "adjustment_date": contract.adjustment_date,
            "termination_date": contract.termination_date,
        }.get(reference_type) or trigger.reference_date_value

    @staticmethod
    def get_native_schedule_overview(contract: Contract) -> dict:
        triggers = contract.triggers.filter(ContractTrigger.is_active.is_(True)).order_by(ContractTrigger.reference_date_value.asc().nulls_last(), ContractTrigger.id.asc()).all()
        events = []
        base_events = [
            ContractService._event_date_payload(contract.billing_start_at, "Início da cobrança contratual", "billing_start", "contract"),
            ContractService._event_date_payload(contract.renewal_date, "Janela de renovação", "renewal", "contract"),
            ContractService._event_date_payload(contract.adjustment_date, "Janela de reajuste", "adjustment", "contract"),
            ContractService._event_date_payload(contract.termination_date, "Data de finalização", "termination", "contract"),
        ]
        events.extend(item for item in base_events if item)

        for trigger in triggers:
            reference_date = ContractService._resolve_trigger_reference_date(contract, trigger)
            if reference_date and trigger.offset_days:
                reference_date = reference_date + timedelta(days=int(trigger.offset_days))
            label = {
                "billing": "Gatilho nativo de faturamento",
                "renewal": "Gatilho nativo de renovação",
                "adjustment": "Gatilho nativo de reajuste",
                "termination": "Gatilho nativo de encerramento",
                "alert": "Alerta contratual",
            }.get(trigger.trigger_type, trigger.trigger_type)
            payload = ContractService._event_date_payload(reference_date, label, trigger.trigger_type, "trigger")
            if payload:
                payload["periodicity"] = trigger.periodicity
                payload["alert_before_days"] = trigger.alert_before_days
                payload["reference_date_type"] = trigger.reference_date_type
                payload["reference_date_type_label"] = dict(ContractService.get_reference_date_type_options()).get(trigger.reference_date_type, trigger.reference_date_type)
                events.append(payload)

        events.sort(key=lambda item: item["date"])
        today = date.today()
        next_event = next((item for item in events if item["date"] >= today), events[0] if events else None)
        return {
            "events": events,
            "next_event": next_event,
            "trigger_count": len(triggers),
        }

    @staticmethod
    def _build_contract_automation_next_execution(contract: Contract, template_key: str) -> Optional[datetime]:
        today = date.today()
        if template_key == "generate_billing_monthly":
            anchor = contract.billing_start_at or contract.service_start_at or today
            day = min(max(anchor.day, 1), 28)
            year = today.year
            month = today.month
            candidate = date(year, month, day)
            if candidate <= today:
                if month == 12:
                    candidate = date(year + 1, 1, day)
                else:
                    candidate = date(year, month + 1, day)
            return datetime.combine(candidate, datetime.min.time())
        if template_key == "renewal_alert_before_date" and contract.renewal_date:
            candidate = contract.renewal_date - timedelta(days=30)
            return datetime.combine(candidate, datetime.min.time())
        if template_key == "adjustment_on_date" and contract.adjustment_date:
            return datetime.combine(contract.adjustment_date, datetime.min.time())
        return None

    @staticmethod
    def list_contract_automations(contract: Contract):
        return (
            AutomationRegistry.query.filter(
                AutomationRegistry.company_id == contract.company_id,
                AutomationRegistry.entity_type == "contract",
                AutomationRegistry.entity_id == contract.id,
            )
            .order_by(AutomationRegistry.next_execution_at.asc().nullslast(), AutomationRegistry.name.asc())
            .all()
        )

    @staticmethod
    def create_contract_automation(*, contract: Contract, template_key: str, user_id: Optional[int]):
        template_key = ContractService._normalize_text(template_key)
        template_map = {item["key"]: item for item in ContractService.get_contract_automation_template_options()}
        if template_key not in template_map:
            raise ValueError("Modelo de automação inválido para o contrato.")

        existing = AutomationRegistry.query.filter(
            AutomationRegistry.company_id == contract.company_id,
            AutomationRegistry.entity_type == "contract",
            AutomationRegistry.entity_id == contract.id,
            AutomationRegistry.action_type == template_key,
            AutomationRegistry.is_active.is_(True),
        ).first()
        if existing:
            raise ValueError("Já existe uma automação ativa deste tipo para o contrato.")

        template = template_map[template_key]
        trigger_type = "date"
        action_type = template_key
        execution_mode = "automatic"
        requires_approval = False
        origin_type = "native"
        next_execution_at = ContractService._build_contract_automation_next_execution(contract, template_key)

        if template_key == "renewal_alert_before_date":
            action_type = "send_alert"
        elif template_key == "adjustment_on_date":
            action_type = "apply_adjustment"

        registry = AutomationRegistry(
            company_id=contract.company_id,
            name=template["label"],
            module_key="contracts",
            origin_type=origin_type,
            entity_type="contract",
            entity_id=contract.id,
            trigger_type=trigger_type,
            action_type=action_type,
            execution_mode=execution_mode,
            status="active",
            requires_approval=requires_approval,
            is_active=True,
            next_execution_at=next_execution_at,
            created_by_user_id=user_id,
            updated_by_user_id=user_id,
        )
        db.session.add(registry)
        db.session.flush()

        if template_key == "generate_billing_monthly":
            trigger_config = {"mode": "monthly", "reference": "billing_start_at"}
            action_config = {"service": "BillingService.generate_from_contract"}
        elif template_key == "renewal_alert_before_date":
            trigger_config = {"mode": "before_date", "reference": "renewal_date", "days_before": 30}
            action_config = {"channel": "in_app", "action": "renewal_alert"}
        else:
            trigger_config = {"mode": "on_date", "reference": "adjustment_date"}
            action_config = {"service": "ContractService.apply_adjustment"}

        db.session.add(
            AutomationRule(
                company_id=contract.company_id,
                automation_registry_id=registry.id,
                rule_code=template_key,
                trigger_config_json=trigger_config,
                action_config_json=action_config,
                policy_config_json={"entity_type": "contract", "entity_id": contract.id},
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
        )
        ContractService.record_event(
            contract=contract,
            event_type="contract.automation_created",
            description=f"Automação '{template['label']}' criada para o contrato.",
            payload={"automation_template_key": template_key, "next_execution_at": next_execution_at.isoformat() if next_execution_at else None},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return registry

    @staticmethod
    def update_contract_automation_status(*, contract: Contract, automation_id: int, activate: bool, user_id: Optional[int]):
        registry = AutomationRegistry.query.filter(
            AutomationRegistry.id == automation_id,
            AutomationRegistry.company_id == contract.company_id,
            AutomationRegistry.entity_type == "contract",
            AutomationRegistry.entity_id == contract.id,
        ).first()
        if not registry:
            raise ValueError("Automação do contrato não localizada.")
        registry.is_active = bool(activate)
        registry.status = "active" if activate else "paused"
        registry.updated_by_user_id = user_id
        ContractService.record_event(
            contract=contract,
            event_type="contract.automation_status_updated",
            description=f"Automação '{registry.name}' {'ativada' if activate else 'pausada'}.",
            payload={"automation_id": registry.id, "status": registry.status},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return registry

    @staticmethod
    def list_customer_contract_tree(company_id: int) -> list[dict]:
        ContractService.sync_parties_from_counterparties(company_id, only_customer=True)
        parties = ContractService.list_customer_parties(company_id)
        contracts = (
            Contract.query.filter(
                Contract.company_id == company_id,
                Contract.deleted_at.is_(None),
            )
            .order_by(ContractParty.name.asc(), Contract.title.asc())
            .join(ContractParty, ContractParty.id == Contract.party_id)
            .all()
        )
        contracts_by_party: dict[int, list[Contract]] = {}
        for contract in contracts:
            contracts_by_party.setdefault(contract.party_id, []).append(contract)

        tree = []
        for party in parties:
            party_contracts = contracts_by_party.get(party.id, [])
            tree.append(
                {
                    "party": party,
                    "contracts": party_contracts,
                    "contract_count": len(party_contracts),
                    "active_count": sum(1 for item in party_contracts if ContractService.get_contract_status_group(item) == "active"),
                }
            )
        return tree

    @staticmethod
    def build_contract_list_tree(company_id: int, filters: Optional[dict] = None) -> list[dict]:
        ContractService.sync_parties_from_counterparties(company_id, only_customer=True)
        parties = ContractService.list_customer_parties(company_id)
        contracts = ContractService.list_contracts_filtered(company_id, filters or {})
        parties_by_id = {party.id: party for party in parties}
        for contract in contracts:
            if contract.party and contract.party.id not in parties_by_id:
                parties_by_id[contract.party.id] = contract.party

        contracts_by_party: dict[int, list[Contract]] = {}
        for contract in contracts:
            contracts_by_party.setdefault(contract.party_id, []).append(contract)

        tree = []
        for party in sorted(parties_by_id.values(), key=lambda item: ((item.name or "").lower(), item.id)):
            party_contracts = contracts_by_party.get(party.id, [])
            tree.append(
                {
                    "party": party,
                    "contracts": party_contracts,
                    "contract_count": len(party_contracts),
                    "active_count": sum(
                        1 for item in party_contracts if ContractService.get_contract_status_group(item) == "active"
                    ),
                }
            )
        return tree

    @staticmethod
    def get_customer_portfolio_summary(company_id: int) -> dict:
        tree = ContractService.list_customer_contract_tree(company_id)
        total_customers = len(tree)
        customers_with_contracts = sum(1 for item in tree if item.get("contract_count"))
        total_contracts = sum(item.get("contract_count", 0) for item in tree)
        active_contracts = sum(item.get("active_count", 0) for item in tree)
        return {
            "total_customers": total_customers,
            "customers_with_contracts": customers_with_contracts,
            "customers_without_contracts": max(total_customers - customers_with_contracts, 0),
            "total_contracts": total_contracts,
            "active_contracts": active_contracts,
        }

    @staticmethod
    def list_contracts_billing_view(company_id: int, filters: Optional[dict] = None) -> list[dict]:
        normalized_filters = dict(filters or {})
        billing_state = ContractService._normalize_text(normalized_filters.get("billing_state") or "eligible").lower()
        contracts = ContractService.list_contracts_filtered(company_id, normalized_filters)
        rows: list[dict] = []
        for contract in contracts:
            native_billings = ContractService.list_native_billings(contract)
            last_native_billing = native_billings[0] if native_billings else None
            next_period = ContractService.build_contract_next_billing_period(contract)
            preview_payload = {
                "competence_start": next_period["competence_start"].isoformat(),
                "competence_end": next_period["competence_end"].isoformat(),
                "issue_date": next_period["issue_date"].isoformat(),
                "due_date": next_period["due_date"].isoformat() if next_period.get("due_date") else None,
            }
            preview = ContractService.preview_native_billing(contract, preview_payload)
            eligibility = ContractService.get_contract_billing_eligibility(contract, preview)
            if billing_state == "eligible" and not eligibility["eligible"]:
                continue
            if billing_state == "blocked" and eligibility["eligible"]:
                continue
            rows.append(
                {
                    "contract": contract,
                    "billing_item_count": preview["item_count"],
                    "native_billing_count": len(native_billings),
                    "last_native_billing": last_native_billing,
                    "next_period": next_period,
                    "preview": preview,
                    "eligibility": eligibility,
                }
            )
        return rows

    @staticmethod
    def get_last_native_billing(contract: Contract) -> Optional[ContractNativeBilling]:
        if not hasattr(contract, "native_billings"):
            return None
        return (
            contract.native_billings.filter(ContractNativeBilling.status != "cancelled")
            .order_by(ContractNativeBilling.competence_start.desc(), ContractNativeBilling.id.desc())
            .first()
        )

    @staticmethod
    def build_contract_next_billing_period(contract: Contract, issue_date: Optional[date] = None) -> dict:
        base_issue_date = issue_date or date.today()
        periodicity_key = ContractService._normalize_text(getattr(contract, "periodicity", None)).lower()
        last_billing = ContractService.get_last_native_billing(contract)

        if periodicity_key in {"weekly", "semanal"}:
            if last_billing and last_billing.competence_start:
                competence_start = last_billing.competence_start + timedelta(days=7)
            else:
                competence_start = getattr(contract, "billing_start_at", None) or getattr(contract, "service_start_at", None) or base_issue_date
            competence_end = competence_start + timedelta(days=6)
        else:
            month_interval = ContractService._periodicity_month_interval(periodicity_key)
            if last_billing and last_billing.competence_start:
                anchor = date(last_billing.competence_start.year, last_billing.competence_start.month, 1)
                competence_start = ContractService._add_months(anchor, month_interval)
            else:
                anchor = getattr(contract, "billing_start_at", None) or getattr(contract, "service_start_at", None) or base_issue_date
                competence_start = date(anchor.year, anchor.month, 1)
            competence_end = ContractService._month_end(
                ContractService._add_months(competence_start, max(month_interval - 1, 0))
            )
        if not hasattr(contract, "items") and hasattr(contract, "billing_items"):
            competence_end = competence_start

        due_date = ContractService.resolve_due_date(issue_date=base_issue_date, due_rule=getattr(contract, "due_rule", None)) or base_issue_date
        return {
            "competence_start": competence_start,
            "competence_end": competence_end,
            "issue_date": base_issue_date,
            "due_date": due_date,
            "periodicity": periodicity_key or "monthly",
        }

    @staticmethod
    def has_existing_native_billing_for_period(
        contract: Contract,
        competence_start: Optional[date],
        competence_end: Optional[date],
    ) -> bool:
        if not competence_start or not competence_end:
            return False
        return (
            ContractNativeBilling.query.filter(
                ContractNativeBilling.company_id == contract.company_id,
                ContractNativeBilling.contract_id == contract.id,
                ContractNativeBilling.competence_start == competence_start,
                ContractNativeBilling.competence_end == competence_end,
                ContractNativeBilling.status != "cancelled",
            ).first()
            is not None
        )

    @staticmethod
    def get_contract_billing_eligibility(contract: Contract, preview: Optional[dict] = None) -> dict:
        reasons: list[str] = []
        if ContractService.get_contract_status_group(contract) != "active":
            reasons.append("Contrato ainda não está ativo.")
        if not contract.party_id:
            reasons.append("Cliente não definido.")

        preview_payload = preview or ContractService.preview_native_billing(contract, {})
        if int(preview_payload.get("item_count") or 0) <= 0:
            reasons.append("Nenhum item contratual cadastrado.")
        if ContractService._normalize_decimal(preview_payload.get("gross_amount")) <= Decimal("0.00"):
            reasons.append("Valor bruto igual a zero.")

        competence_start = preview_payload.get("competence_start")
        competence_end = preview_payload.get("competence_end")
        if contract.billing_end_at and competence_start and competence_start > contract.billing_end_at:
            reasons.append("Próxima competência está após o fim do faturamento.")
        if ContractService.has_existing_native_billing_for_period(contract, competence_start, competence_end):
            reasons.append("Competência já faturada.")

        return {
            "eligible": not reasons,
            "reasons": reasons,
            "label": "Apto para faturar" if not reasons else "Bloqueado",
        }

    @staticmethod
    def list_financial_counterparties(company_id: int):
        return (
            FinancialCounterparty.query.filter(
                FinancialCounterparty.company_id == company_id,
                FinancialCounterparty.deleted_at.is_(None),
            )
            .order_by(FinancialCounterparty.name.asc())
            .all()
        )

    @staticmethod
    def list_financial_references(company_id: int) -> dict:
        return {
            "bank_accounts": FinancialBankAccount.query.filter(
                FinancialBankAccount.company_id == company_id,
                FinancialBankAccount.deleted_at.is_(None),
            ).order_by(FinancialBankAccount.name.asc()).all(),
            "asset_accounts": FinancialAssetAccount.query.filter(
                FinancialAssetAccount.company_id == company_id,
                FinancialAssetAccount.deleted_at.is_(None),
                FinancialAssetAccount.is_active.is_(True),
            ).order_by(FinancialAssetAccount.name.asc()).all(),
            "payment_methods": FinancialPaymentMethod.query.filter(
                FinancialPaymentMethod.company_id == company_id,
                FinancialPaymentMethod.deleted_at.is_(None),
            ).order_by(FinancialPaymentMethod.name.asc()).all(),
            "chart_accounts": FinancialChartAccount.query.filter(
                FinancialChartAccount.company_id == company_id,
                FinancialChartAccount.deleted_at.is_(None),
                FinancialChartAccount.is_active.is_(True),
                FinancialChartAccount.accepts_posting.is_(True),
            ).order_by(FinancialChartAccount.code.asc(), FinancialChartAccount.name.asc()).all(),
            "cost_centers": FinancialCostCenter.query.filter(
                FinancialCostCenter.company_id == company_id,
                FinancialCostCenter.deleted_at.is_(None),
                FinancialCostCenter.is_active.is_(True),
                FinancialCostCenter.accepts_posting.is_(True),
            ).order_by(FinancialCostCenter.name.asc()).all(),
            "correction_indexes": FinancialCorrectionIndex.query.filter(
                FinancialCorrectionIndex.company_id == company_id,
                FinancialCorrectionIndex.deleted_at.is_(None),
            ).order_by(FinancialCorrectionIndex.name.asc()).all(),
            "projects": Project.query.filter(
                Project.company_id == company_id,
                Project.is_deleted.is_(False),
                Project.deleted_at.is_(None),
            ).order_by(Project.name.asc(), Project.id.asc()).all(),
        }

    @staticmethod
    def get_retention_trigger_options():
        return (
            ("emissao", "Na emissão"),
            ("vencimento", "No vencimento"),
            ("baixa", "Na baixa"),
        )

    @staticmethod
    def get_item_retention_options():
        return ContractService.ITEM_RETENTION_OPTIONS

    @staticmethod
    def get_item_retention_deduction_mode_options():
        return ContractService.ITEM_RETENTION_DEDUCTION_MODES

    @staticmethod
    def get_item_retention_value_mode_options():
        return ContractService.ITEM_RETENTION_VALUE_MODES

    @staticmethod
    def _retention_label(kind: str) -> str:
        normalized = ContractService._normalize_text(kind).lower()
        for key, label in ContractService.ITEM_RETENTION_OPTIONS:
            if key == normalized:
                return label
        return normalized.upper() if normalized else "Retenção"

    @staticmethod
    def _normalize_city_name(value: object) -> str:
        normalized = ContractService._normalize_text(value)
        if not normalized:
            return ""
        return (
            unicodedata.normalize("NFKD", normalized)
            .encode("ascii", "ignore")
            .decode("ascii")
            .strip()
            .lower()
        )

    @staticmethod
    def _is_salvador_city(value: object) -> bool:
        normalized = ContractService._normalize_city_name(value)
        return normalized in {"salvador", "salvador/ba", "salvador - ba", "salvador ba"}

    @staticmethod
    def _should_export_iss_as_other(fiscal_data: Optional[dict], fallback_sources: Optional[list[dict]] = None) -> bool:
        fiscal_data = fiscal_data or {}
        fallback_sources = fallback_sources or []
        iss_city = ContractService._normalize_text(fiscal_data.get("iss_city"))
        service_city = ContractService._normalize_text(fiscal_data.get("service_city"))
        resolved_city = (
            iss_city
            or service_city
            or ContractService._metadata_value(fallback_sources, "iss_city", "cidade_iss", "city_iss")
            or ContractService._metadata_value(fallback_sources, "service_city", "cidade_servico", "city_service")
        )
        if not resolved_city:
            return False
        return not ContractService._is_salvador_city(resolved_city)

    @staticmethod
    def _collect_retention_observation_lines(*, native_billing: ContractNativeBilling, fiscal_data: Optional[dict], fallback_sources: Optional[list[dict]] = None) -> list[str]:
        fallback_sources = fallback_sources or []
        iss_as_other = ContractService._should_export_iss_as_other(fiscal_data, fallback_sources)
        lines: list[str] = []
        seen: set[str] = set()
        for billing_item in native_billing.items.order_by(ContractNativeBillingItem.id.asc()).all():
            item_metadata = dict(billing_item.metadata_json or {})
            for detail in list(item_metadata.get("retention_details") or []):
                retention = dict(detail or {})
                kind = ContractService._normalize_text(retention.get("kind")).lower()
                amount = ContractService._normalize_decimal(
                    retention.get("calculated_amount") or retention.get("retention_amount")
                ).quantize(Decimal("0.01"))
                if amount <= Decimal("0.00"):
                    continue
                include_flag = ContractService._normalize_bool(
                    retention.get("include_in_fiscal_description")
                    or retention.get("include_in_observations")
                )
                observation_text = ContractService._normalize_text(
                    retention.get("fiscal_observation_text")
                    or retention.get("observation_text")
                )
                if kind == "iss" and iss_as_other:
                    include_flag = True
                    if not observation_text:
                        observation_text = f"ISS Retido: {ContractService._decimal_to_br_text(amount)}"
                if not include_flag:
                    continue
                if not observation_text:
                    observation_text = (
                        f"{ContractService._retention_label(kind)}: {ContractService._decimal_to_br_text(amount)}"
                    )
                normalized_key = observation_text.strip().lower()
                if normalized_key and normalized_key not in seen:
                    seen.add(normalized_key)
                    lines.append(observation_text.strip())
        return lines

    @staticmethod
    def _build_project_domain_metadata(project: Optional[Project]) -> dict:
        if not project:
            return {
                "domain_type": None,
                "domain_source_kind": None,
                "domain_source_id": None,
                "domain_label": None,
                "domain_value": None,
            }
        return {
            "domain_type": "project",
            "domain_source_kind": "manual",
            "domain_source_id": project.id,
            "domain_label": f"{project.code} · {project.name}",
            "domain_value": f"manual:project:{project.id}",
        }

    @staticmethod
    def _calculate_retention_amount(
        *,
        gross_amount: Decimal,
        deduction_mode: str,
        deduction_value: Decimal,
        value_mode: str,
        value_amount: Decimal,
    ) -> tuple[Decimal, Decimal]:
        effective_base = gross_amount
        if deduction_value > Decimal("0.00"):
            if deduction_mode == "percent":
                effective_base = gross_amount - ((gross_amount * deduction_value) / Decimal("100"))
            else:
                effective_base = gross_amount - deduction_value
        effective_base = max(effective_base, Decimal("0.00")).quantize(Decimal("0.01"))
        if value_amount <= Decimal("0.00"):
            return effective_base, Decimal("0.00")
        if value_mode == "percent":
            retention_amount = (effective_base * value_amount / Decimal("100")).quantize(Decimal("0.01"))
        else:
            retention_amount = value_amount.quantize(Decimal("0.01"))
        return effective_base, max(retention_amount, Decimal("0.00")).quantize(Decimal("0.01"))

    @staticmethod
    def _resolve_chart_account(company_id: int, account_id: object, *, field_label: str) -> Optional[FinancialChartAccount]:
        normalized_id = ContractService._normalize_int(account_id)
        if not normalized_id:
            return None
        account = FinancialChartAccount.query.filter(
            FinancialChartAccount.id == normalized_id,
            FinancialChartAccount.company_id == company_id,
            FinancialChartAccount.deleted_at.is_(None),
            FinancialChartAccount.is_active.is_(True),
            FinancialChartAccount.accepts_posting.is_(True),
        ).first()
        if not account:
            raise ValueError(f"{field_label} inválido para a empresa ativa.")
        return account

    @staticmethod
    def _resolve_cost_center(company_id: int, cost_center_id: object, *, field_label: str) -> Optional[FinancialCostCenter]:
        normalized_id = ContractService._normalize_int(cost_center_id)
        if not normalized_id:
            return None
        cost_center = FinancialCostCenter.query.filter(
            FinancialCostCenter.id == normalized_id,
            FinancialCostCenter.company_id == company_id,
            FinancialCostCenter.deleted_at.is_(None),
            FinancialCostCenter.is_active.is_(True),
            FinancialCostCenter.accepts_posting.is_(True),
        ).first()
        if not cost_center:
            raise ValueError(f"{field_label} inválido para a empresa ativa.")
        return cost_center

    @staticmethod
    def _resolve_asset_account(company_id: int, asset_account_id: object, *, field_label: str) -> Optional[FinancialAssetAccount]:
        normalized_id = ContractService._normalize_int(asset_account_id)
        if not normalized_id:
            return None
        asset_account = FinancialAssetAccount.query.filter(
            FinancialAssetAccount.id == normalized_id,
            FinancialAssetAccount.company_id == company_id,
            FinancialAssetAccount.deleted_at.is_(None),
            FinancialAssetAccount.is_active.is_(True),
        ).first()
        if not asset_account:
            raise ValueError(f"{field_label} inválida para a empresa ativa.")
        return asset_account

    @staticmethod
    def _resolve_bank_account(company_id: int, bank_account_id: object, *, field_label: str) -> Optional[FinancialBankAccount]:
        normalized_id = ContractService._normalize_int(bank_account_id)
        if not normalized_id:
            return None
        bank_account = FinancialBankAccount.query.filter(
            FinancialBankAccount.id == normalized_id,
            FinancialBankAccount.company_id == company_id,
            FinancialBankAccount.deleted_at.is_(None),
        ).first()
        if not bank_account:
            raise ValueError(f"{field_label} inválida para a empresa ativa.")
        return bank_account

    @staticmethod
    def _resolve_project(company_id: int, project_id: object, *, field_label: str) -> Optional[Project]:
        normalized_id = ContractService._normalize_int(project_id)
        if not normalized_id:
            return None
        project = Project.query.filter(
            Project.id == normalized_id,
            Project.company_id == company_id,
            Project.is_deleted.is_(False),
            Project.deleted_at.is_(None),
        ).first()
        if not project:
            raise ValueError(f"{field_label} inválido para a empresa ativa.")
        return project

    @staticmethod
    def _contracting_legal_entity_metadata(entity: Optional[ContractingLegalEntity]) -> dict:
        return dict(getattr(entity, "metadata_json", None) or {})

    @staticmethod
    def _normalize_legal_entity_iss_rules(raw_rules: object) -> list[dict]:
        normalized: list[dict] = []
        for raw in list(raw_rules or []):
            if not isinstance(raw, dict):
                continue
            effective_from = ContractService._normalize_date(raw.get("effective_from"))
            effective_to = ContractService._normalize_date(raw.get("effective_to"))
            percent = ContractService._normalize_decimal(raw.get("percent")).quantize(Decimal("0.0001"))
            if percent <= Decimal("0.00") or not effective_from:
                continue
            normalized.append(
                {
                    "effective_from": effective_from,
                    "effective_to": effective_to,
                    "percent": percent,
                }
            )
        normalized.sort(key=lambda item: (item["effective_from"], item.get("effective_to") or date.max))
        return normalized

    @staticmethod
    def list_contracting_legal_entity_iss_rules(entity: Optional[ContractingLegalEntity]) -> list[dict]:
        rules = ContractService._normalize_legal_entity_iss_rules(
            ContractService._contracting_legal_entity_metadata(entity).get("iss_rate_rules")
        )
        return [
            {
                "effective_from": ContractService._serialize_date(rule["effective_from"]),
                "effective_to": ContractService._serialize_date(rule.get("effective_to")),
                "percent": ContractService._decimal_to_export_text(rule["percent"], places=4, strip_trailing=True),
            }
            for rule in rules
        ]

    @staticmethod
    def get_contracting_legal_entity_active_iss_rule(
        entity: Optional[ContractingLegalEntity],
        reference_date: Optional[date] = None,
    ) -> Optional[dict]:
        rules = ContractService._normalize_legal_entity_iss_rules(
            ContractService._contracting_legal_entity_metadata(entity).get("iss_rate_rules")
        )
        if not rules:
            return None
        target_date = reference_date or date.today()
        for rule in sorted(rules, key=lambda item: item["effective_from"], reverse=True):
            effective_to = rule.get("effective_to")
            if rule["effective_from"] <= target_date and (effective_to is None or target_date <= effective_to):
                return {
                    "effective_from": ContractService._serialize_date(rule["effective_from"]),
                    "effective_to": ContractService._serialize_date(effective_to),
                    "percent": ContractService._decimal_to_export_text(rule["percent"], places=4, strip_trailing=True),
                }
        return None

    @staticmethod
    def get_contracting_legal_entity_latest_iss_rule(entity: Optional[ContractingLegalEntity]) -> Optional[dict]:
        rules = ContractService.list_contracting_legal_entity_iss_rules(entity)
        return rules[-1] if rules else None

    @staticmethod
    def _resolve_catalog_item_iss_rate_percent(catalog_item: Optional[ContractCatalogItem]) -> Decimal:
        metadata = dict(getattr(catalog_item, "metadata_json", None) or {})
        for key in ("iss_rate_percent", "iss_aliquot_percent", "aliquota_iss_percent", "aliquota_iss"):
            rate = ContractService._normalize_decimal(metadata.get(key))
            if rate > Decimal("0.00"):
                return rate.quantize(Decimal("0.0001"))
        return Decimal("0.00")

    @staticmethod
    def _resolve_contract_issuer_iss_rate(
        contract: Optional[Contract],
        reference_date: Optional[date] = None,
    ) -> tuple[Decimal, Optional[dict]]:
        if contract is None:
            return Decimal("0.00"), None
        legal_entity = getattr(contract, "contracting_legal_entity", None)
        if legal_entity is None and getattr(contract, "contracting_legal_entity_id", None):
            legal_entity = ContractService.get_contracting_legal_entity(contract.company_id, contract.contracting_legal_entity_id)
        if legal_entity is None and getattr(contract, "id", None) and getattr(contract, "company_id", None):
            fiscal_term = ContractFiscalTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
            if fiscal_term and fiscal_term.contracting_legal_entity_id:
                legal_entity = ContractService.get_contracting_legal_entity(contract.company_id, fiscal_term.contracting_legal_entity_id)
        if legal_entity is None:
            return Decimal("0.00"), None
        active_rule = ContractService.get_contracting_legal_entity_active_iss_rule(legal_entity, reference_date)
        if not active_rule:
            return Decimal("0.00"), None
        return ContractService._normalize_decimal(active_rule.get("percent")).quantize(Decimal("0.0001")), active_rule

    @staticmethod
    def _resolve_effective_iss_rate_percent(
        *,
        contract: Optional[Contract],
        catalog_item: Optional[ContractCatalogItem],
        reference_date: Optional[date] = None,
        fallback_rate: object = None,
    ) -> tuple[Decimal, Optional[str], Optional[dict]]:
        issuer_rate, issuer_rule = ContractService._resolve_contract_issuer_iss_rate(contract, reference_date)
        if issuer_rate > Decimal("0.00"):
            return issuer_rate, "issuer", issuer_rule
        service_rate = ContractService._resolve_catalog_item_iss_rate_percent(catalog_item)
        if service_rate > Decimal("0.00"):
            return service_rate, "service", {
                "effective_from": None,
                "effective_to": None,
                "percent": ContractService._decimal_to_export_text(service_rate, places=4, strip_trailing=True),
            }
        normalized_fallback = ContractService._normalize_decimal(fallback_rate)
        if normalized_fallback > Decimal("0.00"):
            return normalized_fallback.quantize(Decimal("0.0001")), "item", None
        return Decimal("0.00"), None, None

    @staticmethod
    def list_contracting_legal_entities(company_id: int):
        return (
            ContractingLegalEntity.query.filter(
                ContractingLegalEntity.company_id == company_id,
                ContractingLegalEntity.is_active.is_(True),
            )
            .order_by(ContractingLegalEntity.legal_name.asc(), ContractingLegalEntity.id.asc())
            .all()
        )

    @staticmethod
    def get_contracting_legal_entity(company_id: int, legal_entity_id: int) -> Optional[ContractingLegalEntity]:
        return ContractingLegalEntity.query.filter(
            ContractingLegalEntity.id == legal_entity_id,
            ContractingLegalEntity.company_id == company_id,
            ContractingLegalEntity.is_active.is_(True),
        ).first()

    @staticmethod
    def get_party(company_id: int, party_id: int) -> Optional[ContractParty]:
        ContractService.sync_parties_from_counterparties(company_id)
        return ContractParty.query.filter(
            ContractParty.id == party_id,
            ContractParty.company_id == company_id,
            ContractParty.deleted_at.is_(None),
        ).first()

    @staticmethod
    def get_party_by_counterparty_id(company_id: int, counterparty_id: int) -> Optional[ContractParty]:
        ContractService.sync_parties_from_counterparties(company_id)
        return ContractParty.query.filter(
            ContractParty.company_id == company_id,
            ContractParty.financial_counterparty_id == counterparty_id,
            ContractParty.deleted_at.is_(None),
        ).first()

    @staticmethod
    def get_contract(company_id: int, contract_id: int) -> Optional[Contract]:
        return Contract.query.filter(
            Contract.id == contract_id,
            Contract.company_id == company_id,
            Contract.deleted_at.is_(None),
        ).first()

    @staticmethod
    def get_tab_registry() -> list[dict]:
        return [dict(item) for item in ContractService.TAB_REGISTRY]

    @staticmethod
    def get_contract_operational_profile(contract: Optional[Contract]) -> str:
        metadata = dict((contract.metadata_json or {}) if contract else {})
        return ContractService._normalize_operational_profile(metadata.get("operational_profile"))

    @staticmethod
    def get_operational_profile_label(contract: Optional[Contract]) -> str:
        profile_key = ContractService.get_contract_operational_profile(contract)
        profile = ContractService.OPERATIONAL_PROFILE_CONFIG.get(profile_key, {})
        return profile.get("label") or "Contrato completo"

    @staticmethod
    def get_operational_profile_options() -> list[dict]:
        return [
            {"key": key, "label": config["label"], "description": config["description"]}
            for key, config in ContractService.OPERATIONAL_PROFILE_CONFIG.items()
        ]

    @staticmethod
    def get_visible_tabs(contract: Optional[Contract]) -> list[dict]:
        profile_key = ContractService.get_contract_operational_profile(contract)
        allowed_keys = set(ContractService.OPERATIONAL_PROFILE_CONFIG.get(profile_key, {}).get("visible_tabs") or [])
        registry = ContractService.get_tab_registry()
        if not allowed_keys:
            return registry
        return [tab for tab in registry if tab["key"] in allowed_keys]

    @staticmethod
    def contract_requires_financial_integration(contract: Optional[Contract]) -> bool:
        if not contract:
            return False
        visible_keys = {tab["key"] for tab in ContractService.get_visible_tabs(contract)}
        return "financeiro" in visible_keys

    @staticmethod
    def native_billing_has_financial_integration(native_billing: Optional[ContractNativeBilling]) -> bool:
        metadata = dict((native_billing.metadata_json or {}) if native_billing else {})
        integration = dict(metadata.get("financial_integration") or {})
        return ContractService._normalize_int(integration.get("main_schedule_id")) is not None

    @staticmethod
    def native_billing_has_financial_anomaly(*, contract: Optional[Contract], native_billing: Optional[ContractNativeBilling]) -> bool:
        if not native_billing or native_billing.status == "cancelled":
            return False
        if not ContractService.contract_requires_financial_integration(contract):
            return False
        return not ContractService.native_billing_has_financial_integration(native_billing)

    @staticmethod
    def resolve_active_tab(contract: Optional[Contract], requested_tab: Optional[str]) -> str:
        visible_tabs = ContractService.get_visible_tabs(contract)
        visible_keys = [tab["key"] for tab in visible_tabs]
        normalized = ContractService._normalize_text(requested_tab).lower() or "cliente"
        if normalized in visible_keys:
            return normalized
        return visible_keys[0] if visible_keys else "cliente"

    @staticmethod
    def preview_next_contracting_legal_entity_code(company_id: int) -> str:
        return ContractService._next_contracting_legal_entity_code(company_id)

    @staticmethod
    def _next_contracting_legal_entity_code(company_id: int) -> str:
        company_code = ContractService._resolve_company_code(company_id)
        code_pattern = re.compile(rf"^{re.escape(company_code)}\.([A-Z0-9]+)\.(\d+)$")
        next_emitter_code = 1
        next_sequence = 1

        rows = ContractingLegalEntity.query.with_entities(ContractingLegalEntity.code).filter(
            ContractingLegalEntity.company_id == company_id
        ).all()
        for (code,) in rows:
            normalized_code = str(code or "").strip().upper()
            match = code_pattern.match(normalized_code)
            if not match:
                continue
            emitter_token, sequence_token = match.groups()
            if emitter_token.isdigit():
                next_emitter_code = max(next_emitter_code, int(emitter_token) + 1)
            next_sequence = max(next_sequence, int(sequence_token) + 1)

        return f"{company_code}.{next_emitter_code}.{next_sequence:03d}"

    @staticmethod
    def create_contracting_legal_entity(*, company_id: int, payload: dict):
        entity = ContractingLegalEntity(
            company_id=company_id,
            code=ContractService._next_contracting_legal_entity_code(company_id),
            is_active=True,
        )
        ContractService.update_contracting_legal_entity(entity=entity, payload=payload, is_new=True)
        db.session.add(entity)
        db.session.commit()
        return entity

    @staticmethod
    def update_contracting_legal_entity(*, entity: ContractingLegalEntity, payload: dict, is_new: bool = False):
        legal_name = ContractService._normalize_text(payload.get("legal_name"))
        cnpj = ContractService._normalize_text(payload.get("cnpj"))
        if legal_name:
            entity.legal_name = legal_name
        if cnpj:
            entity.cnpj = cnpj
        entity.trade_name = ContractService._normalize_text(payload.get("trade_name")) or None
        entity.municipal_registration = ContractService._normalize_text(payload.get("municipal_registration")) or None
        entity.state_registration = ContractService._normalize_text(payload.get("state_registration")) or None
        entity.tax_regime = ContractService._normalize_text(payload.get("tax_regime")) or None
        entity.cnae = ContractService._normalize_text(payload.get("cnae")) or None
        entity.service_city = ContractService._normalize_text(payload.get("service_city")) or None
        entity.city_code_ibge = ContractService._normalize_text(payload.get("city_code_ibge")) or None
        entity.uf = ContractService._normalize_text(payload.get("uf")) or None
        entity.zip_code = ContractService._normalize_text(payload.get("zip_code")) or None
        entity.address_line = ContractService._normalize_text(payload.get("address_line")) or None
        entity.address_number = ContractService._normalize_text(payload.get("address_number")) or None
        entity.district = ContractService._normalize_text(payload.get("district")) or None
        entity.complement = ContractService._normalize_text(payload.get("complement")) or None
        entity.email = ContractService._normalize_text(payload.get("email")) or None
        entity.phone = ContractService._normalize_text(payload.get("phone")) or None
        entity.nfs_provider = ContractService._normalize_text(payload.get("nfs_provider")) or None
        entity.integration_mode = ContractService._normalize_text(payload.get("integration_mode")) or "manual"
        entity.api_profile_id = ContractService._normalize_int(payload.get("api_profile_id"))
        entity.spreadsheet_profile_id = ContractService._normalize_int(payload.get("spreadsheet_profile_id"))
        metadata = ContractService._contracting_legal_entity_metadata(entity)
        raw_iss_rate = payload.get("iss_rate_percent")
        raw_effective_from = payload.get("iss_rate_effective_from")
        raw_effective_to = payload.get("iss_rate_effective_to")
        if raw_iss_rate is not None or raw_effective_from is not None or raw_effective_to is not None:
            iss_rate_percent = ContractService._normalize_decimal(raw_iss_rate).quantize(Decimal("0.0001"))
            effective_from = ContractService._normalize_date(raw_effective_from)
            effective_to = ContractService._normalize_date(raw_effective_to)
            if iss_rate_percent > Decimal("0.00") and not effective_from:
                raise ValueError("Informe a data de início da vigência do ISS da PJ emissora.")
            if effective_to and effective_from and effective_to < effective_from:
                raise ValueError("A data final da vigência do ISS não pode ser menor que a inicial.")
            rules = ContractService._normalize_legal_entity_iss_rules(metadata.get("iss_rate_rules"))
            if iss_rate_percent > Decimal("0.00") and effective_from:
                replaced = False
                for rule in rules:
                    if rule["effective_from"] == effective_from:
                        rule["effective_to"] = effective_to
                        rule["percent"] = iss_rate_percent
                        replaced = True
                        break
                if not replaced:
                    rules.append(
                        {
                            "effective_from": effective_from,
                            "effective_to": effective_to,
                            "percent": iss_rate_percent,
                        }
                    )
                rules.sort(key=lambda item: (item["effective_from"], item.get("effective_to") or date.max))
            metadata["iss_rate_rules"] = [
                {
                    "effective_from": ContractService._serialize_date(rule["effective_from"]),
                    "effective_to": ContractService._serialize_date(rule.get("effective_to")),
                    "percent": ContractService._decimal_to_export_text(rule["percent"], places=4, strip_trailing=True),
                }
                for rule in rules
            ]
        entity.metadata_json = metadata
        entity.is_active = ContractService._normalize_bool(payload.get("is_active")) if payload.get("is_active") is not None else True
        if not entity.legal_name:
            raise ValueError("Informe a razão social da PJ contratada.")
        if not entity.cnpj:
            raise ValueError("Informe o CNPJ da PJ contratada.")
        if not is_new:
            db.session.commit()
        return entity

    @staticmethod
    def create_party(*, company_id: int, payload: dict, user_id: Optional[int]):
        party = ContractParty(
            company_id=company_id,
            code=ContractService._next_structured_code(ContractParty, company_id, "F"),
            created_by_user_id=user_id,
        )
        ContractService.update_party(party=party, payload=payload, user_id=user_id, is_new=True)
        db.session.add(party)
        db.session.commit()
        return party

    @staticmethod
    def update_party(*, party: ContractParty, payload: dict, user_id: Optional[int], is_new: bool = False):
        name = ContractService._normalize_text(payload.get("name"))
        if name:
            party.name = name
        party.legal_name = ContractService._normalize_text(payload.get("legal_name")) or None
        party.document_type = ContractService._normalize_text(payload.get("document_type")) or None
        party.document_number = ContractService._normalize_text(payload.get("document_number")) or None
        party.email = ContractService._normalize_text(payload.get("email")) or None
        party.phone = ContractService._normalize_text(payload.get("phone")) or None
        party.is_customer = ContractService._normalize_bool(payload.get("is_customer"))
        party.is_supplier = ContractService._normalize_bool(payload.get("is_supplier"))
        party.status = ContractService._normalize_text(payload.get("status")) or "active"
        party.notes = ContractService._normalize_text(payload.get("notes")) or None
        party.financial_counterparty_id = ContractService._normalize_int(payload.get("financial_counterparty_id"))
        party.updated_by_user_id = user_id
        if not party.is_customer and not party.is_supplier:
            raise ValueError("Selecione ao menos uma classificação: Cliente, Fornecedor ou ambos.")
        if not is_new:
            db.session.commit()
        return party

    @staticmethod
    def create_contract(*, company_id: int, payload: dict, user_id: Optional[int]) -> Contract:
        title = ContractService._normalize_text(payload.get("title"))
        party_id = ContractService._normalize_int(payload.get("party_id"))
        if not title:
            raise ValueError("Informe o título do contrato.")
        if not party_id:
            raise ValueError("Selecione o cliente do contrato.")
        contract = Contract(
            company_id=company_id,
            code=ContractService._next_structured_code(Contract, company_id, "N"),
            title=title,
            party_id=party_id,
            status="draft",
            currency_code="BRL",
            created_by_user_id=user_id,
            version=1,
        )
        ContractService.update_contract_general(contract=contract, payload=payload, user_id=user_id, is_new=True)
        db.session.add(contract)
        db.session.flush()
        ContractService.record_event(
            contract=contract,
            event_type="contract.created",
            description="Contrato criado.",
            payload={"status": contract.status, "party_id": contract.party_id, "title": contract.title},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return contract

    @staticmethod
    def update_contract_general(*, contract: Contract, payload: dict, user_id: Optional[int], is_new: bool = False):
        metadata = dict(contract.metadata_json or {})
        title = ContractService._normalize_text(payload.get("title"))
        if title:
            contract.title = title
        if "party_id" in payload:
            party_id = ContractService._normalize_int(payload.get("party_id")) or contract.party_id
            if party_id:
                party = ContractService.get_party(contract.company_id, party_id)
                if not party or not party.is_customer:
                    raise ValueError("Cliente inválido para a empresa ativa.")
                contract.party_id = party_id
        if "status" in payload:
            normalized_status = ContractService._normalize_text(payload.get("status")).lower()
            if normalized_status in {"active", "inactive"}:
                contract.status = normalized_status
            elif normalized_status:
                contract.status = normalized_status
            else:
                contract.status = contract.status or "draft"
        if "contract_type" in payload:
            contract.contract_type = ContractService._normalize_text(payload.get("contract_type")) or None
        if "currency_code" in payload:
            contract.currency_code = ContractService._normalize_text(payload.get("currency_code")) or "BRL"
        if "manager_employee_id" in payload:
            manager_employee_id = ContractService._normalize_int(payload.get("manager_employee_id"))
            if manager_employee_id:
                manager = Employee.query.filter_by(id=manager_employee_id, company_id=contract.company_id, status="active").first()
                if not manager:
                    raise ValueError("Gestor inválido para a empresa ativa.")
            contract.manager_employee_id = manager_employee_id
        if "contracting_legal_entity_id" in payload:
            contract.contracting_legal_entity_id = ContractService._normalize_int(payload.get("contracting_legal_entity_id"))
        if "signed_at" in payload:
            contract.signed_at = ContractService._normalize_date(payload.get("signed_at"))
        if "service_start_at" in payload:
            contract.service_start_at = ContractService._normalize_date(payload.get("service_start_at"))
        if "service_end_at" in payload:
            contract.service_end_at = ContractService._normalize_date(payload.get("service_end_at"))
        if "billing_start_at" in payload:
            contract.billing_start_at = ContractService._normalize_date(payload.get("billing_start_at"))
        if "billing_end_at" in payload:
            contract.billing_end_at = ContractService._normalize_date(payload.get("billing_end_at"))
        if "last_billing_at" in payload:
            contract.last_billing_at = ContractService._normalize_date(payload.get("last_billing_at"))
        if "renewal_date" in payload:
            contract.renewal_date = ContractService._normalize_date(payload.get("renewal_date"))
        if "adjustment_date" in payload:
            contract.adjustment_date = ContractService._normalize_date(payload.get("adjustment_date"))
        if "termination_date" in payload:
            contract.termination_date = ContractService._normalize_date(payload.get("termination_date"))
        if "periodicity" in payload:
            contract.periodicity = ContractService._normalize_text(payload.get("periodicity")) or None
        if "competence_rule" in payload:
            contract.competence_rule = ContractService._normalize_text(payload.get("competence_rule")) or None
        if "due_rule" in payload:
            contract.due_rule = ContractService._normalize_text(payload.get("due_rule")) or None
        if "renewal_rule" in payload:
            contract.renewal_rule = ContractService._normalize_text(payload.get("renewal_rule")) or None
        if "end_reason" in payload:
            contract.end_reason = ContractService._normalize_text(payload.get("end_reason")) or None
        if "previous_contract_id" in payload:
            contract.previous_contract_id = ContractService._normalize_int(payload.get("previous_contract_id"))
        if "notes" in payload:
            contract.notes = ContractService._normalize_text(payload.get("notes")) or None
        if "operational_profile" in payload:
            metadata["operational_profile"] = ContractService._normalize_operational_profile(payload.get("operational_profile"))
        contract.metadata_json = metadata
        contract.status = contract.status or "draft"
        contract.currency_code = contract.currency_code or "BRL"
        contract.updated_by_user_id = user_id
        if not is_new:
            ContractService.record_event(
                contract=contract,
                event_type="contract.summary_updated",
                description="Resumo do contrato atualizado.",
                payload=ContractService._build_contract_snapshot(contract),
                user_id=user_id,
                auto_commit=False,
            )
            db.session.commit()
        return contract

    @staticmethod
    def update_contract_summary(*, contract: Contract, payload: dict, user_id: Optional[int]):
        return ContractService.update_contract_general(contract=contract, payload=payload, user_id=user_id)

    @staticmethod
    def suspend_contract(*, contract: Contract, user_id: Optional[int], reason: Optional[str] = None):
        previous_status = ContractService._normalize_text(contract.status).lower() or "draft"
        if previous_status == "suspended":
            return contract
        contract.status = "suspended"
        contract.updated_by_user_id = user_id
        metadata = dict(contract.metadata_json or {})
        lifecycle = dict(metadata.get("lifecycle") or {})
        lifecycle["suspended_at"] = datetime.utcnow().isoformat()
        lifecycle["suspended_by_user_id"] = user_id
        lifecycle["suspension_reason"] = ContractService._normalize_text(reason) or None
        metadata["lifecycle"] = lifecycle
        contract.metadata_json = metadata
        ContractService.record_event(
            contract=contract,
            event_type="contract.suspended",
            description="Contrato suspenso.",
            payload={
                "previous_status": previous_status,
                "current_status": contract.status,
                "reason": lifecycle.get("suspension_reason"),
            },
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return contract

    @staticmethod
    def close_contract(*, contract: Contract, user_id: Optional[int], reason: Optional[str] = None, termination_date: Optional[date] = None):
        previous_status = ContractService._normalize_text(contract.status).lower() or "draft"
        if previous_status == "closed":
            return contract
        normalized_reason = ContractService._normalize_text(reason) or "manual_close"
        normalized_termination_date = termination_date or contract.termination_date or date.today()
        contract.status = "closed"
        contract.end_reason = normalized_reason
        contract.termination_date = normalized_termination_date
        contract.updated_by_user_id = user_id
        metadata = dict(contract.metadata_json or {})
        lifecycle = dict(metadata.get("lifecycle") or {})
        lifecycle["closed_at"] = datetime.utcnow().isoformat()
        lifecycle["closed_by_user_id"] = user_id
        lifecycle["close_reason"] = normalized_reason
        metadata["lifecycle"] = lifecycle
        contract.metadata_json = metadata
        ContractService.record_event(
            contract=contract,
            event_type="contract.closed",
            description="Contrato encerrado.",
            payload={
                "previous_status": previous_status,
                "current_status": contract.status,
                "reason": normalized_reason,
                "termination_date": normalized_termination_date.isoformat() if normalized_termination_date else None,
            },
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return contract

    @staticmethod
    def delete_contract(*, contract: Contract, user_id: Optional[int], reason: Optional[str] = None):
        if contract.deleted_at:
            return contract
        contract.deleted_at = datetime.utcnow()
        contract.updated_by_user_id = user_id
        metadata = dict(contract.metadata_json or {})
        deletion = dict(metadata.get("deletion") or {})
        deletion["deleted_at"] = contract.deleted_at.isoformat()
        deletion["deleted_by_user_id"] = user_id
        deletion["reason"] = ContractService._normalize_text(reason) or "manual_delete"
        metadata["deletion"] = deletion
        contract.metadata_json = metadata
        ContractService.record_event(
            contract=contract,
            event_type="contract.deleted",
            description="Contrato excluído logicamente.",
            payload={
                "status": contract.status,
                "reason": deletion.get("reason"),
                "deleted_at": contract.deleted_at.isoformat(),
            },
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return contract

    @staticmethod
    def update_contract_customer(*, contract: Contract, payload: dict, user_id: Optional[int]):
        party_id = ContractService._normalize_int(payload.get("party_id"))
        if not party_id:
            raise ValueError("Selecione um favorecido cliente para o contrato.")
        return ContractService.update_contract_general(
            contract=contract,
            payload={
                "party_id": party_id,
                "operational_profile": payload.get("operational_profile"),
            },
            user_id=user_id,
        )

    @staticmethod
    def update_contract_schedule(*, contract: Contract, payload: dict, user_id: Optional[int]):
        schedule_payload = {
            "signed_at": payload.get("signed_at"),
            "service_start_at": payload.get("service_start_at"),
            "service_end_at": payload.get("service_end_at"),
            "billing_start_at": payload.get("billing_start_at"),
            "billing_end_at": payload.get("billing_end_at"),
            "last_billing_at": payload.get("last_billing_at"),
            "renewal_date": payload.get("renewal_date"),
            "adjustment_date": payload.get("adjustment_date"),
            "termination_date": payload.get("termination_date"),
            "periodicity": payload.get("periodicity"),
            "competence_rule": payload.get("competence_rule"),
            "due_rule": payload.get("due_rule"),
            "renewal_rule": payload.get("renewal_rule"),
            "end_reason": payload.get("end_reason"),
        }
        return ContractService.update_contract_general(contract=contract, payload=schedule_payload, user_id=user_id)

    @staticmethod
    def update_contract_notes(*, contract: Contract, payload: dict, user_id: Optional[int]):
        updated_contract = ContractService.update_contract_general(
            contract=contract,
            payload={"notes": payload.get("notes")},
            user_id=user_id,
        )
        note_text = ContractService._normalize_text(payload.get("notes"))
        if note_text:
            ContractService.add_contract_note(
                contract=contract,
                payload={"note_type": "general", "note_text": note_text},
                user_id=user_id,
                auto_commit=True,
            )
        return updated_contract

    @staticmethod
    def update_contract_validation(*, contract: Contract, payload: dict, user_id: Optional[int]):
        metadata = dict(contract.metadata_json or {})
        metadata["validation_status"] = ContractService._normalize_text(payload.get("validation_status")) or "pending"
        metadata["validation_notes"] = ContractService._normalize_text(payload.get("validation_notes")) or None
        metadata["last_validation_user_id"] = user_id
        metadata["last_validation_at"] = datetime.utcnow().isoformat()
        contract.metadata_json = metadata
        contract.updated_by_user_id = user_id
        ContractService.record_event(
            contract=contract,
            event_type="contract.validation_updated",
            description="Status de validação do contrato atualizado.",
            payload={"validation_status": metadata["validation_status"]},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return contract

    @staticmethod
    def _build_contract_snapshot(contract: Contract) -> dict:
        return {
            "status": contract.status,
            "party_id": contract.party_id,
            "title": contract.title,
            "manager_employee_id": contract.manager_employee_id,
            "renewal_date": contract.renewal_date.isoformat() if contract.renewal_date else None,
            "adjustment_date": contract.adjustment_date.isoformat() if contract.adjustment_date else None,
            "termination_date": contract.termination_date.isoformat() if contract.termination_date else None,
            "previous_contract_id": contract.previous_contract_id,
        }

    @staticmethod
    def record_event(
        *,
        contract: Contract,
        event_type: str,
        description: Optional[str] = None,
        payload: Optional[dict] = None,
        user_id: Optional[int] = None,
        auto_commit: bool = True,
    ) -> ContractEvent:
        event = ContractEvent(
            company_id=contract.company_id,
            contract_id=contract.id,
            event_type=event_type,
            description=description,
            event_payload=payload or {},
            created_by_user_id=user_id,
        )
        db.session.add(event)
        if auto_commit:
            db.session.commit()
        return event

    @staticmethod
    def list_contract_history(contract: Contract) -> dict:
        return {
            "clauses": contract.clauses.order_by(ContractClause.order_index.asc(), ContractClause.id.asc()).all(),
            "notes": contract.notes_log.order_by(ContractNote.created_at.desc(), ContractNote.id.desc()).all(),
            "events": contract.events.order_by(ContractEvent.created_at.desc(), ContractEvent.id.desc()).all(),
        }

    @staticmethod
    def upsert_contract_clause(*, contract: Contract, payload: dict, user_id: Optional[int], clause_id: Optional[int] = None):
        clause = None
        if clause_id:
            clause = ContractClause.query.filter(
                ContractClause.id == clause_id,
                ContractClause.contract_id == contract.id,
                ContractClause.company_id == contract.company_id,
            ).first()
        if clause is None:
            clause = ContractClause(
                company_id=contract.company_id,
                contract_id=contract.id,
                created_by_user_id=user_id,
            )
            db.session.add(clause)

        clause.clause_type = ContractService._normalize_text(payload.get("clause_type")) or "general"
        clause.title = ContractService._normalize_text(payload.get("title")) or None
        clause.content = ContractService._normalize_text(payload.get("content")) or "Cláusula sem conteúdo."
        clause.order_index = ContractService._normalize_int(payload.get("order_index")) or 0
        clause.updated_by_user_id = user_id
        ContractService.record_event(
            contract=contract,
            event_type="contract.clause_upserted",
            description=f"Cláusula {'atualizada' if clause_id else 'criada'}.",
            payload={"clause_type": clause.clause_type, "title": clause.title},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return clause

    @staticmethod
    def delete_contract_clause(*, contract: Contract, clause_id: int, user_id: Optional[int]):
        clause = ContractClause.query.filter(
            ContractClause.id == clause_id,
            ContractClause.contract_id == contract.id,
            ContractClause.company_id == contract.company_id,
        ).first()
        if not clause:
            raise ValueError("Cláusula não encontrada para este contrato.")
        db.session.delete(clause)
        ContractService.record_event(
            contract=contract,
            event_type="contract.clause_deleted",
            description="Cláusula removida do contrato.",
            payload={"clause_id": clause_id},
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()

    @staticmethod
    def add_contract_note(*, contract: Contract, payload: dict, user_id: Optional[int], auto_commit: bool = True):
        note_text = ContractService._normalize_text(payload.get("note_text"))
        if not note_text:
            raise ValueError("Informe um texto para registrar a observação do contrato.")
        note = ContractNote(
            company_id=contract.company_id,
            contract_id=contract.id,
            note_type=ContractService._normalize_text(payload.get("note_type")) or "general",
            note_text=note_text,
            created_by_user_id=user_id,
        )
        db.session.add(note)
        ContractService.record_event(
            contract=contract,
            event_type="contract.note_added",
            description="Observação adicionada ao contrato.",
            payload={"note_type": note.note_type},
            user_id=user_id,
            auto_commit=False,
        )
        if auto_commit:
            db.session.commit()
        return note

    @staticmethod
    def add_contract_item(*, contract: Contract, payload: dict):
        item_data = ContractService._build_contract_item_data(contract=contract, payload=payload)
        item = ContractItem(
            company_id=contract.company_id,
            contract_id=contract.id,
            contract_catalog_item_id=item_data["contract_catalog_item_id"],
            item_code=item_data["item_code"],
            item_type=item_data["item_type"],
            description=item_data["description"],
            quantity=item_data["quantity"],
            unit_code=item_data["unit_code"],
            unit_price=item_data["unit_price"],
            total_price=item_data["total_price"],
            order_index=item_data["order_index"],
            notes=item_data["notes"],
            metadata_json=item_data["metadata_json"],
        )
        db.session.add(item)
        db.session.commit()
        return item

    @staticmethod
    def _build_contract_item_data(*, contract: Contract, payload: dict) -> dict:
        catalog_item_id = ContractService._normalize_int(payload.get("contract_catalog_item_id"))
        catalog_item = None
        if catalog_item_id:
            catalog_item = ContractCatalogItem.query.filter(
                ContractCatalogItem.id == catalog_item_id,
                ContractCatalogItem.company_id == contract.company_id,
                ContractCatalogItem.deleted_at.is_(None),
            ).first()
            if not catalog_item:
                raise ValueError("Item mestre não encontrado para este contrato.")
            if not ContractsCatalogService._is_selectable_level(catalog_item):
                raise ValueError("Somente itens do catálogo podem ser utilizados no contrato.")

        if not catalog_item and not ContractService._normalize_text(payload.get("description")):
            raise ValueError("Selecione um produto/serviço do catálogo para adicionar ao contrato.")

        description = catalog_item.name if catalog_item else (ContractService._normalize_text(payload.get("description")) or "Item contratual")
        item_code = catalog_item.code if catalog_item else (ContractService._normalize_text(payload.get("item_code")) or None)
        item_type = catalog_item.item_kind if catalog_item else (ContractService._normalize_text(payload.get("item_type")) or None)
        unit_code = catalog_item.unit_code if catalog_item else (ContractService._normalize_text(payload.get("unit_code")) or None)
        quantity = ContractService._normalize_decimal(payload.get("quantity"), default="1")
        unit_price_raw = ContractService._normalize_text(payload.get("unit_price"))
        total_price_raw = ContractService._normalize_text(payload.get("total_price"))
        unit_price = ContractService._normalize_decimal(unit_price_raw)
        total_price = (
            ContractService._normalize_decimal(total_price_raw)
            if total_price_raw
            else ContractService.calculate_total_price(quantity, unit_price)
        )
        if total_price_raw and not unit_price_raw and quantity > 0:
            unit_price = (total_price / quantity).quantize(Decimal("0.01"))
        metadata = dict(payload.get("metadata_json") or {})
        chart_account = ContractService._resolve_chart_account(
            contract.company_id,
            payload.get("chart_account_id"),
            field_label="Plano de contas do item",
        )
        cost_center = ContractService._resolve_cost_center(
            contract.company_id,
            payload.get("cost_center_id"),
            field_label="Centro de resultados do item",
        )
        project = ContractService._resolve_project(
            contract.company_id,
            payload.get("project_id"),
            field_label="Projeto do item",
        )
        if catalog_item:
            metadata["contract_catalog_item_id"] = catalog_item.id
            metadata["catalog_snapshot"] = {
                "code": catalog_item.code,
                "name": catalog_item.name,
                "item_kind": catalog_item.item_kind,
                "unit_code": catalog_item.unit_code,
            }
        metadata["allocation"] = {
            "chart_account_id": chart_account.id if chart_account else None,
            "chart_account_code": chart_account.code if chart_account else None,
            "chart_account_name": chart_account.name if chart_account else None,
            "cost_center_id": cost_center.id if cost_center else None,
            "cost_center_code": cost_center.code if cost_center else None,
            "cost_center_name": cost_center.name if cost_center else None,
            "project_id": project.id if project else None,
            "project_code": project.code if project else None,
            "project_name": project.name if project else None,
        }
        project_domain = ContractService._build_project_domain_metadata(project)
        retention_details = []
        retention_flags = {}
        for retention_key, retention_label in ContractService.get_item_retention_options():
            enabled = ContractService._normalize_bool(payload.get(f"retention_{retention_key}_enabled"))
            retention_flags[retention_key] = enabled
            if not enabled:
                continue
            deduction_mode = ContractService._normalize_text(payload.get(f"retention_{retention_key}_deduction_mode")).lower() or "percent"
            if deduction_mode not in {"percent", "amount"}:
                raise ValueError(f"Abatimento da base inválido para retenção {retention_label}.")
            deduction_value = ContractService._normalize_decimal(payload.get(f"retention_{retention_key}_deduction_value"))
            value_mode = ContractService._normalize_text(payload.get(f"retention_{retention_key}_value_mode")).lower() or "percent"
            if value_mode not in {"percent", "amount"}:
                raise ValueError(f"Tipo do valor da retenção inválido para {retention_label}.")
            value_amount = ContractService._normalize_decimal(payload.get(f"retention_{retention_key}_value"))
            rate_source = "manual"
            rate_rule = None
            if retention_key == "iss" and value_mode == "percent":
                resolved_rate, rate_source, rate_rule = ContractService._resolve_effective_iss_rate_percent(
                    contract=contract,
                    catalog_item=catalog_item,
                    reference_date=contract.billing_start_at or contract.service_start_at or date.today(),
                    fallback_rate=value_amount,
                )
                if resolved_rate > Decimal("0.00"):
                    value_amount = resolved_rate
            compensation_bank_account = ContractService._resolve_bank_account(
                contract.company_id,
                payload.get(f"retention_{retention_key}_bank_account_id"),
                field_label=f"Conta bancária para compensação da retenção {retention_label}",
            )
            retention_chart_account = ContractService._resolve_chart_account(
                contract.company_id,
                payload.get(f"retention_{retention_key}_chart_account_id"),
                field_label=f"Plano de contas da retenção {retention_label}",
            )
            retention_trigger = ContractService._normalize_text(payload.get(f"retention_{retention_key}_trigger")).lower() or None
            if retention_trigger not in {"emissao", "vencimento", "baixa"}:
                raise ValueError(f"Informe um gatilho válido para a retenção {retention_label}.")
            fiscal_observation_text = ContractService._normalize_text(
                payload.get(f"retention_{retention_key}_fiscal_observation_text")
            ) or None
            include_in_fiscal_description = ContractService._normalize_bool(
                payload.get(f"retention_{retention_key}_include_in_fiscal_description")
            )
            if not compensation_bank_account:
                raise ValueError(f"Informe a conta bancária para compensação da retenção {retention_label}.")
            if not retention_chart_account:
                raise ValueError(f"Informe o plano de contas da retenção {retention_label}.")
            if value_amount <= Decimal("0.00"):
                raise ValueError(f"Informe o valor da retenção {retention_label}.")
            calculation_base, retention_amount = ContractService._calculate_retention_amount(
                gross_amount=total_price,
                deduction_mode=deduction_mode,
                deduction_value=deduction_value,
                value_mode=value_mode,
                value_amount=value_amount,
            )
            retention_details.append(
                {
                    "kind": retention_key,
                    "label": retention_label,
                    "base_deduction_mode": deduction_mode,
                    "base_deduction_value": float(deduction_value),
                    "calculation_base": float(calculation_base),
                    "retention_value_mode": value_mode,
                    "retention_value": float(value_amount),
                    "retention_amount": float(retention_amount),
                    "bank_account_id": compensation_bank_account.id,
                    "bank_account_code": compensation_bank_account.code,
                    "bank_account_name": compensation_bank_account.name,
                    "asset_account_id": compensation_bank_account.id,
                    "asset_account_code": compensation_bank_account.code,
                    "asset_account_name": compensation_bank_account.name,
                    "chart_account_id": retention_chart_account.id,
                    "chart_account_code": retention_chart_account.code,
                    "chart_account_name": retention_chart_account.name,
                    "trigger": retention_trigger,
                    "fiscal_observation_text": fiscal_observation_text,
                    "include_in_fiscal_description": include_in_fiscal_description,
                    "rate_source": rate_source,
                    "issuer_rate_rule": rate_rule,
                    "project_id": project.id if project else None,
                    "project_code": project.code if project else None,
                    "project_name": project.name if project else None,
                    **project_domain,
                }
            )
        metadata["retention_flags"] = retention_flags
        metadata["retention_details"] = retention_details
        metadata["retention_summary"] = {
            "total_retention_amount": float(sum(Decimal(str(item.get("retention_amount") or 0)) for item in retention_details)),
            "retention_count": len(retention_details),
        }

        return {
            "contract_catalog_item_id": catalog_item.id if catalog_item else None,
            "item_code": item_code,
            "item_type": item_type,
            "description": description,
            "quantity": quantity,
            "unit_code": unit_code,
            "unit_price": unit_price,
            "total_price": total_price,
            "order_index": ContractService._normalize_int(payload.get("order_index")) or 0,
            "notes": ContractService._normalize_text(payload.get("notes")) or None,
            "metadata_json": metadata,
        }

    @staticmethod
    def update_contract_item(*, contract: Contract, item_id: int, payload: dict):
        item = ContractItem.query.filter_by(id=item_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not item:
            raise ValueError("Item do contrato não encontrado para edição.")
        item_data = ContractService._build_contract_item_data(contract=contract, payload=payload)
        item.contract_catalog_item_id = item_data["contract_catalog_item_id"]
        item.item_code = item_data["item_code"]
        item.item_type = item_data["item_type"]
        item.description = item_data["description"]
        item.quantity = item_data["quantity"]
        item.unit_code = item_data["unit_code"]
        item.unit_price = item_data["unit_price"]
        item.total_price = item_data["total_price"]
        item.order_index = item_data["order_index"]
        item.notes = item_data["notes"]
        item.metadata_json = item_data["metadata_json"]
        db.session.commit()
        return item

    @staticmethod
    def delete_contract_item(*, contract: Contract, item_id: int) -> bool:
        item = ContractItem.query.filter_by(id=item_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not item:
            return False
        db.session.delete(item)
        db.session.commit()
        return True

    @staticmethod
    def get_contract_item(company_id: int, contract_id: int, item_id: int) -> Optional[ContractItem]:
        return ContractItem.query.filter_by(
            id=item_id,
            company_id=company_id,
            contract_id=contract_id,
        ).first()

    @staticmethod
    def build_contract_item_form_state(item: Optional[ContractItem]) -> dict:
        if not item:
            return {}
        metadata = dict(item.metadata_json or {})
        allocation = dict(metadata.get("allocation") or {})
        retention_flags = dict(metadata.get("retention_flags") or {})
        retention_details = list(metadata.get("retention_details") or [])
        retention_by_kind = {
            str(detail.get("kind") or "").lower(): detail
            for detail in retention_details
            if detail.get("kind")
        }
        return {
            "id": item.id,
            "contract_catalog_item_id": item.contract_catalog_item_id,
            "quantity": item.quantity,
            "unit_price": item.unit_price,
            "total_price": item.total_price,
            "chart_account_id": allocation.get("chart_account_id"),
            "cost_center_id": allocation.get("cost_center_id"),
            "project_id": allocation.get("project_id"),
            "retention_flags": retention_flags,
            "retention_by_kind": retention_by_kind,
        }

    @staticmethod
    def add_billing_item(*, contract: Contract, payload: dict):
        item = ContractBillingItem(
            company_id=contract.company_id,
            contract_id=contract.id,
            contract_item_id=ContractService._normalize_int(payload.get("contract_item_id")),
            billing_code=ContractService._normalize_text(payload.get("billing_code")) or None,
            description=ContractService._normalize_text(payload.get("description")) or "Item de faturamento",
            amount=ContractService._normalize_decimal(payload.get("amount")),
            billing_periodicity=ContractService._normalize_text(payload.get("billing_periodicity")) or None,
            competence_rule=ContractService._normalize_text(payload.get("competence_rule")) or None,
            due_rule=ContractService._normalize_text(payload.get("due_rule")) or None,
            trigger_type=ContractService._normalize_text(payload.get("trigger_type")) or None,
            trigger_reference_date=ContractService._normalize_text(payload.get("trigger_reference_date")) or None,
            is_recurring=ContractService._normalize_bool(payload.get("is_recurring")),
            order_index=ContractService._normalize_int(payload.get("order_index")) or 0,
        )
        db.session.add(item)
        db.session.commit()
        return item

    @staticmethod
    def build_native_billing_idempotency_key(*, contract: Contract, competence_start: date, competence_end: date) -> str:
        return f"contract:{contract.company_id}:{contract.id}:{competence_start.isoformat()}:{competence_end.isoformat()}"

    @staticmethod
    def _cancelled_native_billing_idempotency_key(idempotency_key: str, native_billing_id: int) -> str:
        return f"{str(idempotency_key or '')[:130]}:cancelled:{native_billing_id}"

    @staticmethod
    def _next_native_billing_code(company_id: int) -> str:
        company_code = ContractService._resolve_company_code(company_id)
        code_prefix = f"{company_code}.B."
        last_number = 0
        rows = (
            ContractNativeBilling.query.with_entities(ContractNativeBilling.billing_code)
            .filter(ContractNativeBilling.company_id == company_id)
            .all()
        )
        for (billing_code,) in rows:
            normalized_code = str(billing_code or "").strip().upper()
            if not normalized_code.startswith(code_prefix):
                continue
            match = re.search(r"(\d+)$", normalized_code)
            if match:
                last_number = max(last_number, int(match.group(1)))
        return f"{code_prefix}{last_number + 1:03d}"

    @staticmethod
    def list_native_billings(contract: Contract, include_cancelled: bool = False):
        query = contract.native_billings
        if not include_cancelled:
            query = query.filter(ContractNativeBilling.status != "cancelled")
        return query.order_by(ContractNativeBilling.competence_start.desc(), ContractNativeBilling.id.desc()).all()

    @staticmethod
    def _normalize_id_list(value: object) -> list[int]:
        raw_values: list[object]
        if value is None:
            raw_values = []
        elif isinstance(value, (list, tuple, set)):
            raw_values = list(value)
        else:
            raw_values = re.split(r"[,;\s]+", str(value or ""))
        normalized: list[int] = []
        for raw_value in raw_values:
            item_id = ContractService._normalize_int(raw_value)
            if item_id and item_id not in normalized:
                normalized.append(item_id)
        return normalized

    @staticmethod
    def _resolve_native_billing_contract_items(contract: Contract, payload: Optional[dict] = None) -> list[ContractItem]:
        payload = payload or {}
        selected_ids: list[int] = []
        for key in ("contract_item_ids", "contract_item_id", "item_ids", "selected_item_ids"):
            raw_value = None
            if hasattr(payload, "getlist"):
                listed = payload.getlist(key)
                raw_value = listed if listed else None
            if raw_value is None:
                raw_value = payload.get(key) if hasattr(payload, "get") else None
            selected_ids.extend(ContractService._normalize_id_list(raw_value))

        if hasattr(contract, "items"):
            query = contract.items.order_by(ContractItem.order_index.asc(), ContractItem.id.asc())
        elif hasattr(contract, "billing_items"):
            query = contract.billing_items.order_by()
        else:
            return []
        if selected_ids:
            if hasattr(query, "filter"):
                return query.filter(ContractItem.id.in_(selected_ids)).all()
            return [item for item in query.all() if getattr(item, "id", None) in selected_ids]
        return query.all()

    @staticmethod
    def preview_native_billing(contract: Contract, payload: dict) -> dict:
        next_period = ContractService.build_contract_next_billing_period(contract)
        competence_start = ContractService._normalize_date(payload.get("competence_start")) or next_period["competence_start"]
        competence_end = ContractService._normalize_date(payload.get("competence_end")) or next_period["competence_end"]
        issue_date = ContractService._normalize_date(payload.get("issue_date")) or next_period["issue_date"]
        due_date = (
            ContractService._normalize_date(payload.get("due_date"))
            or ContractService.resolve_due_date(issue_date=issue_date, due_rule=getattr(contract, "due_rule", None))
            or issue_date
        )
        items = ContractService._resolve_native_billing_contract_items(contract, payload)
        item_snapshots = [ContractService._build_native_billing_item_snapshot(item, reference_date=issue_date) for item in items]
        gross_amount = sum((ContractService._normalize_decimal(item.get("gross_amount")) for item in item_snapshots), Decimal("0.00"))
        retention_amount = sum((ContractService._normalize_decimal(item.get("retention_amount")) for item in item_snapshots), Decimal("0.00"))
        net_amount = gross_amount - retention_amount
        return {
            "competence_start": competence_start,
            "competence_end": competence_end,
            "issue_date": issue_date,
            "due_date": due_date,
            "item_count": len(item_snapshots),
            "gross_amount": gross_amount.quantize(Decimal("0.01")) if item_snapshots else Decimal("0.00"),
            "retention_amount": retention_amount.quantize(Decimal("0.01")) if item_snapshots else Decimal("0.00"),
            "net_amount": net_amount.quantize(Decimal("0.01")) if item_snapshots else Decimal("0.00"),
            "items": item_snapshots,
        }

    @staticmethod
    def build_billing_review_rows(
        company_id: int,
        contract_ids: list[int],
        overrides_by_contract: Optional[dict[int, dict]] = None,
    ) -> list[dict]:
        overrides_by_contract = overrides_by_contract or {}
        rows: list[dict] = []
        for contract_id in contract_ids:
            contract = ContractService.get_contract(company_id, contract_id)
            if not contract:
                continue
            payload = dict(overrides_by_contract.get(contract.id) or {})
            if "contract_item_ids" not in payload and "item_ids" not in payload:
                payload["contract_item_ids"] = [item.id for item in contract.items.order_by(ContractItem.order_index.asc(), ContractItem.id.asc()).all()]
            preview = ContractService.preview_native_billing(contract, payload)
            eligibility = ContractService.get_contract_billing_eligibility(contract, preview)
            rows.append(
                {
                    "contract": contract,
                    "preview": preview,
                    "eligibility": eligibility,
                    "selected_item_ids": [item.get("contract_item_id") for item in preview.get("items", [])],
                    "last_native_billing": ContractService.get_last_native_billing(contract),
                    "review_notes": ContractService._normalize_text(payload.get("review_notes")),
                }
            )
        return rows

    @staticmethod
    def confirm_native_billing_review(*, company_id: int, review_payloads: list[dict], user_id: Optional[int]) -> dict:
        created: list[ContractNativeBilling] = []
        errors: list[str] = []
        for payload in review_payloads:
            contract_id = ContractService._normalize_int(payload.get("contract_id"))
            if not contract_id:
                continue
            contract = ContractService.get_contract(company_id, contract_id)
            if not contract:
                errors.append(f"Contrato {contract_id} não localizado para a empresa ativa.")
                continue
            try:
                preview = ContractService.preview_native_billing(contract, payload)
                eligibility = ContractService.get_contract_billing_eligibility(contract, preview)
                if not eligibility["eligible"]:
                    errors.append(f"{contract.code}: {'; '.join(eligibility['reasons'])}")
                    continue
                native_billing = ContractService.generate_native_billing(
                    contract=contract,
                    payload=payload,
                    user_id=user_id,
                )
                created.append(native_billing)
            except Exception as exc:  # noqa: BLE001 - retorno consolidado para conferência operacional
                db.session.rollback()
                errors.append(f"{contract.code}: {exc}")
        return {"created": created, "errors": errors}

    @staticmethod
    def list_native_billings_done(company_id: int, filters: Optional[dict] = None) -> list[dict]:
        filters = dict(filters or {})
        query = (
            ContractNativeBilling.query.filter(ContractNativeBilling.company_id == company_id)
            .outerjoin(Contract, Contract.id == ContractNativeBilling.contract_id)
            .outerjoin(ContractParty, ContractParty.id == ContractNativeBilling.party_id)
        )
        status = ContractService._normalize_text(filters.get("status"))
        if status:
            query = query.filter(ContractNativeBilling.status == status)
        party_id = ContractService._normalize_int(filters.get("party_id"))
        if party_id:
            query = query.filter(ContractNativeBilling.party_id == party_id)
        competence_from = ContractService._normalize_date(filters.get("competence_from"))
        if competence_from:
            query = query.filter(ContractNativeBilling.competence_start >= competence_from)
        competence_to = ContractService._normalize_date(filters.get("competence_to"))
        if competence_to:
            query = query.filter(ContractNativeBilling.competence_start <= competence_to)
        search = ContractService._normalize_text(filters.get("search"))
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                or_(
                    ContractNativeBilling.billing_code.ilike(pattern),
                    Contract.code.ilike(pattern),
                    Contract.title.ilike(pattern),
                    ContractParty.name.ilike(pattern),
                )
            )

        billings = query.order_by(ContractNativeBilling.competence_start.desc(), ContractNativeBilling.id.desc()).all()
        rows: list[dict] = []
        for billing in billings:
            metadata = dict(billing.metadata_json or {})
            financial_integration = dict(metadata.get("financial_integration") or {})
            gross_amount = ContractService._normalize_decimal(billing.gross_amount)
            net_amount = ContractService._normalize_decimal(billing.net_amount)
            contract = billing.contract
            financial_required = ContractService.contract_requires_financial_integration(contract)
            financial_anomaly = ContractService.native_billing_has_financial_anomaly(
                contract=contract,
                native_billing=billing,
            )
            rows.append(
                {
                    "billing": billing,
                    "contract": contract,
                    "party": billing.party,
                    "retention_amount": (gross_amount - net_amount).quantize(Decimal("0.01")),
                    "financial_integration": financial_integration,
                    "item_count": billing.items.count(),
                    "financial_required": financial_required,
                    "financial_anomaly": financial_anomaly,
                }
            )
        return rows

    @staticmethod
    def cancel_native_billing(*, company_id: int, native_billing_id: int, user_id: Optional[int], reason: Optional[str] = None):
        native_billing = ContractNativeBilling.query.filter(
            ContractNativeBilling.id == native_billing_id,
            ContractNativeBilling.company_id == company_id,
        ).first()
        if not native_billing:
            raise ValueError("Faturamento não localizado para a empresa ativa.")
        if native_billing.status == "cancelled":
            return native_billing

        now = datetime.utcnow()
        contract = native_billing.contract
        metadata = dict(native_billing.metadata_json or {})
        financial_integration = dict(metadata.get("financial_integration") or {})
        schedule_ids = []
        main_schedule_id = ContractService._normalize_int(financial_integration.get("main_schedule_id"))
        if main_schedule_id:
            schedule_ids.append(main_schedule_id)
        schedule_ids.extend(ContractService._normalize_id_list(financial_integration.get("satellite_schedule_ids")))
        schedule_ids = list(dict.fromkeys(schedule_ids))

        if schedule_ids:
            schedules = FinancialSchedule.query.filter(
                FinancialSchedule.company_id == company_id,
                FinancialSchedule.id.in_(schedule_ids),
                FinancialSchedule.deleted_at.is_(None),
            ).all()
            for schedule in schedules:
                schedule.status = "cancelled"
                schedule.deleted_at = now
                schedule.metadata_json = {
                    **dict(schedule.metadata_json or {}),
                    "cancelled_from_contract_billing": True,
                    "cancelled_native_billing_id": native_billing.id,
                    "cancelled_at": now.isoformat(),
                    "cancelled_by_user_id": user_id,
                }

            links = FinancialScheduleLink.query.filter(
                FinancialScheduleLink.company_id == company_id,
                FinancialScheduleLink.deleted_at.is_(None),
                or_(
                    FinancialScheduleLink.parent_schedule_id.in_(schedule_ids),
                    FinancialScheduleLink.child_schedule_id.in_(schedule_ids),
                ),
            ).all()
            for link in links:
                link.deleted_at = now
                link.metadata_json = {
                    **dict(link.metadata_json or {}),
                    "cancelled_from_contract_billing": True,
                    "cancelled_native_billing_id": native_billing.id,
                    "cancelled_at": now.isoformat(),
                    "cancelled_by_user_id": user_id,
                }

            executions = FinancialSatelliteExecution.query.filter(
                FinancialSatelliteExecution.company_id == company_id,
                FinancialSatelliteExecution.reversed_at.is_(None),
                or_(
                    FinancialSatelliteExecution.parent_schedule_id.in_(schedule_ids),
                    FinancialSatelliteExecution.child_schedule_id.in_(schedule_ids),
                ),
            ).all()
            for execution in executions:
                execution.reversed_at = now
                execution.execution_status = "reversed"
                execution.metadata_json = {
                    **dict(execution.metadata_json or {}),
                    "reversed_from_contract_billing": True,
                    "cancelled_native_billing_id": native_billing.id,
                    "reversed_at": now.isoformat(),
                    "reversed_by_user_id": user_id,
                }

        financial_integration["cancelled_at"] = now.isoformat()
        financial_integration["cancelled_schedule_ids"] = schedule_ids
        metadata["financial_integration"] = financial_integration
        metadata["cancellation"] = {
            "cancelled_at": now.isoformat(),
            "cancelled_by_user_id": user_id,
            "reason": ContractService._normalize_text(reason) or None,
            "original_idempotency_key": native_billing.idempotency_key,
        }
        native_billing.status = "cancelled"
        native_billing.idempotency_key = ContractService._cancelled_native_billing_idempotency_key(
            native_billing.idempotency_key,
            native_billing.id,
        )
        native_billing.metadata_json = metadata

        if contract:
            latest = ContractNativeBilling.query.filter(
                ContractNativeBilling.company_id == company_id,
                ContractNativeBilling.contract_id == contract.id,
                ContractNativeBilling.id != native_billing.id,
                ContractNativeBilling.status != "cancelled",
            ).order_by(ContractNativeBilling.issue_date.desc(), ContractNativeBilling.id.desc()).first()
            contract.last_billing_at = latest.issue_date if latest else None
            contract.updated_by_user_id = user_id
            ContractService.record_event(
                contract=contract,
                event_type="contract.billing_cancelled",
                description="Faturamento nativo cancelado a partir da tela Faturamentos Feitos.",
                payload={
                    "native_billing_id": native_billing.id,
                    "billing_code": native_billing.billing_code,
                    "cancelled_schedule_ids": schedule_ids,
                    "reason": ContractService._normalize_text(reason) or None,
                },
                user_id=user_id,
                auto_commit=False,
            )
        db.session.commit()
        return native_billing

    @staticmethod
    def _build_native_billing_item_snapshot(contract_item: ContractItem, reference_date: Optional[date] = None) -> dict:
        item_metadata = ContractService._normalize_metadata_dict(getattr(contract_item, "metadata_json", None))
        allocation = ContractService._normalize_metadata_dict(item_metadata.get("allocation"))
        raw_retention_details = ContractService._normalize_metadata_dict_list(item_metadata.get("retention_details"))
        retention_details = []
        gross_amount = ContractService._normalize_decimal(
            getattr(contract_item, "total_price", None) if getattr(contract_item, "total_price", None) is not None else getattr(contract_item, "amount", 0)
        )
        contract = getattr(contract_item, "contract", None)
        catalog_item = getattr(contract_item, "contract_catalog_item", None)
        for detail in raw_retention_details:
            normalized = dict(detail)
            if (
                (ContractService._normalize_text(normalized.get("kind")) or "").lower() == "iss"
                and (ContractService._normalize_text(normalized.get("retention_value_mode")) or "percent").lower() == "percent"
            ):
                effective_rate, rate_source, rate_rule = ContractService._resolve_effective_iss_rate_percent(
                    contract=contract,
                    catalog_item=catalog_item,
                    reference_date=reference_date,
                    fallback_rate=normalized.get("retention_value"),
                )
                if effective_rate > Decimal("0.00"):
                    normalized["retention_value"] = float(effective_rate)
                    normalized["rate_source"] = rate_source
                    normalized["issuer_rate_rule"] = rate_rule
                    calculation_base, retention_amount = ContractService._calculate_retention_amount(
                        gross_amount=gross_amount,
                        deduction_mode=normalized.get("base_deduction_mode"),
                        deduction_value=normalized.get("base_deduction_value"),
                        value_mode="percent",
                        value_amount=effective_rate,
                    )
                    normalized["calculation_base"] = float(calculation_base)
                else:
                    retention_amount = ContractService._normalize_decimal(
                        normalized.get("calculated_amount") or normalized.get("retention_amount")
                    )
            else:
                retention_amount = ContractService._normalize_decimal(
                    normalized.get("calculated_amount") or normalized.get("retention_amount")
                )
            retention_amount = retention_amount.quantize(Decimal("0.01"))
            if retention_amount <= Decimal("0.00"):
                continue
            normalized["retention_amount"] = float(retention_amount)
            normalized["calculated_amount"] = float(retention_amount)
            normalized["contract_item_id"] = contract_item.id
            retention_details.append(normalized)
        retention_total = sum(
            (ContractService._normalize_decimal(detail.get("calculated_amount")) for detail in retention_details),
            Decimal("0.00"),
        )
        return {
            "contract_item_id": getattr(contract_item, "id", None),
            "item_code": getattr(contract_item, "item_code", None),
            "description": getattr(contract_item, "description", None) or "Item de faturamento",
            "quantity": float(getattr(contract_item, "quantity", 1) or 0),
            "unit_price": float(getattr(contract_item, "unit_price", gross_amount) or 0),
            "gross_amount": float(gross_amount.quantize(Decimal("0.01"))),
            "net_amount": float((gross_amount - retention_total).quantize(Decimal("0.01"))),
            "retention_amount": float(retention_total.quantize(Decimal("0.01"))),
            "allocation": allocation,
            "retention_details": retention_details,
        }

    @staticmethod
    def generate_native_billing(*, contract: Contract, payload: dict, user_id: Optional[int]):
        if not contract.party_id:
            raise ValueError("Defina o cliente do contrato antes de gerar o faturamento nativo.")

        preview = ContractService.preview_native_billing(contract, payload)
        competence_start = preview["competence_start"]
        competence_end = preview["competence_end"]
        issue_date = preview["issue_date"]
        due_date = preview["due_date"]
        contract_items = ContractService._resolve_native_billing_contract_items(contract, payload)

        if not contract_items:
            raise ValueError("Cadastre ou selecione ao menos um item contratual antes de gerar a competência.")

        fiscal_snapshot = ContractService.build_contract_fiscal_snapshot(contract, reference_date=issue_date)

        idempotency_key = ContractService.build_native_billing_idempotency_key(
            contract=contract,
            competence_start=competence_start,
            competence_end=competence_end,
        )
        existing = ContractNativeBilling.query.filter_by(
            company_id=contract.company_id,
            idempotency_key=idempotency_key,
        ).first()
        if existing:
            if existing.status != "cancelled":
                raise ValueError("Já existe faturamento nativo gerado para esta competência.")
            existing_metadata = dict(existing.metadata_json or {})
            existing_metadata.setdefault("cancellation", {})
            existing_metadata["cancellation"]["original_idempotency_key"] = idempotency_key
            existing.metadata_json = existing_metadata
            existing.idempotency_key = ContractService._cancelled_native_billing_idempotency_key(idempotency_key, existing.id)
            db.session.flush()

        review_notes = ContractService._normalize_text(payload.get("review_notes")) or None
        native_billing = ContractNativeBilling(
            company_id=contract.company_id,
            contract_id=contract.id,
            party_id=contract.party_id,
            billing_code=ContractService._next_native_billing_code(contract.company_id),
            status="generated",
            source_type="native_contract",
            competence_start=competence_start,
            competence_end=competence_end,
            issue_date=issue_date,
            due_date=due_date,
            gross_amount=preview["gross_amount"],
            net_amount=preview["net_amount"],
            idempotency_key=idempotency_key,
            generated_by_user_id=user_id,
            metadata_json={
                "contract_version": contract.version,
                "generated_from": "contract_native_module",
                "reviewed_from": payload.get("reviewed_from") or "contract_native_module",
                "review_notes": review_notes,
                "item_count": len(contract_items),
                "contract_item_ids": [item.id for item in contract_items],
                "retention_amount": float(preview["retention_amount"]),
                "fiscal_snapshot": fiscal_snapshot,
            },
        )
        db.session.add(native_billing)
        db.session.flush()

        retention_summary = {}
        snapshots_by_item_id = {
            snapshot["contract_item_id"]: snapshot
            for snapshot in preview.get("items", [])
        }
        for item in contract_items:
            snapshot = snapshots_by_item_id.get(item.id) or ContractService._build_native_billing_item_snapshot(item, reference_date=issue_date)
            for retention in snapshot["retention_details"]:
                key = retention.get("kind")
                retention_summary[key] = round(
                    float(retention_summary.get(key) or 0) + float(retention.get("calculated_amount") or retention.get("retention_amount") or 0),
                    2,
                )
            db.session.add(
                ContractNativeBillingItem(
                    company_id=contract.company_id,
                    contract_native_billing_id=native_billing.id,
                    contract_billing_item_id=None,
                    contract_item_id=item.id,
                    description=item.description,
                    amount=ContractService._normalize_decimal(snapshot.get("gross_amount")),
                    competence_rule=contract.competence_rule,
                    due_rule=contract.due_rule,
                    trigger_type="contract_item",
                    trigger_reference_date="contract_rule",
                    metadata_json={
                        "billing_periodicity": contract.periodicity,
                        "is_recurring": True,
                        "review_notes": review_notes,
                        **snapshot,
                    },
                )
            )

        native_billing.metadata_json = {
            **dict(native_billing.metadata_json or {}),
            "retention_summary": retention_summary,
        }

        contract.last_billing_at = issue_date
        contract.updated_by_user_id = user_id
        ContractService.record_event(
            contract=contract,
            event_type="contract.billing_generated",
            description="Faturamento nativo gerado a partir do contrato.",
            payload={
                "native_billing_id": native_billing.id,
                "billing_code": native_billing.billing_code,
                "competence_start": competence_start.isoformat(),
                "competence_end": competence_end.isoformat(),
                "gross_amount": float(native_billing.gross_amount or 0),
                "net_amount": float(native_billing.net_amount or 0),
                "retention_amount": float(preview["retention_amount"]),
                "contract_item_ids": [item.id for item in contract_items],
                "idempotency_key": idempotency_key,
            },
            user_id=user_id,
            auto_commit=False,
        )
        from services.contract_financial_service import ContractFinancialService

        ContractFinancialService.ensure_financial_titles_for_native_billing(
            contract=contract,
            native_billing=native_billing,
            user_id=user_id,
            auto_commit=False,
        )
        if ContractService.native_billing_has_financial_anomaly(contract=contract, native_billing=native_billing):
            raise ValueError(
                "Faturamento gerado sem integração financeira obrigatória. Operação cancelada para preservar a consistência."
            )
        db.session.commit()
        return native_billing

    @staticmethod
    def build_contract_fiscal_snapshot(contract: Contract, reference_date: Optional[date] = None) -> dict:
        fiscal_terms = ContractFiscalTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
        legal_entity = None
        if fiscal_terms and fiscal_terms.contracting_legal_entity_id:
            legal_entity = ContractService.get_contracting_legal_entity(contract.company_id, fiscal_terms.contracting_legal_entity_id)
        if legal_entity is None and contract.contracting_legal_entity_id:
            legal_entity = ContractService.get_contracting_legal_entity(contract.company_id, contract.contracting_legal_entity_id)
        issuer_iss_rate, issuer_iss_rule = ContractService._resolve_contract_issuer_iss_rate(contract, reference_date)
        return {
            "contracting_legal_entity_id": legal_entity.id if legal_entity else None,
            "issuer_legal_name": legal_entity.legal_name if legal_entity else None,
            "issuer_trade_name": legal_entity.trade_name if legal_entity else None,
            "issuer_cnpj": legal_entity.cnpj if legal_entity else None,
            "issuer_municipal_registration": legal_entity.municipal_registration if legal_entity else None,
            "issuer_tax_regime": legal_entity.tax_regime if legal_entity else None,
            "issuer_cnae": legal_entity.cnae if legal_entity else None,
            "issuer_city_code_ibge": legal_entity.city_code_ibge if legal_entity else None,
            "integration_mode": (fiscal_terms.integration_mode if fiscal_terms else None) or (legal_entity.integration_mode if legal_entity else None),
            "nfs_provider": (fiscal_terms.nfs_provider if fiscal_terms else None) or (legal_entity.nfs_provider if legal_entity else None),
            "default_rps_series": fiscal_terms.default_rps_series if fiscal_terms else None,
            "service_code": fiscal_terms.service_code if fiscal_terms else None,
            "service_list_item": fiscal_terms.service_list_item if fiscal_terms else None,
            "operation_nature": fiscal_terms.operation_nature if fiscal_terms else None,
            "service_city": fiscal_terms.service_city if fiscal_terms else (legal_entity.service_city if legal_entity else None),
            "iss_city": fiscal_terms.iss_city if fiscal_terms else None,
            "issuer_iss_rate": ContractService._decimal_to_export_text(issuer_iss_rate, places=4, strip_trailing=True) if issuer_iss_rate else None,
            "issuer_iss_rate_effective_from": (issuer_iss_rule or {}).get("effective_from"),
            "issuer_iss_rate_effective_to": (issuer_iss_rule or {}).get("effective_to"),
            "withholding_flags": {},
            "fiscal_notes": fiscal_terms.tax_observation if fiscal_terms else None,
        }

    @staticmethod
    def build_native_billing_fiscal_export_payload(native_billing: ContractNativeBilling) -> dict:
        contract = native_billing.contract
        snapshot = dict((native_billing.metadata_json or {}).get("fiscal_snapshot") or {})
        if not snapshot:
            snapshot = ContractService.build_contract_fiscal_snapshot(contract)
        return {
            "billing_code": native_billing.billing_code,
            "contract_code": contract.code if contract else None,
            "customer_name": contract.party.name if contract and contract.party else None,
            "customer_document": contract.party.document_number if contract and contract.party else None,
            "issuer_cnpj": snapshot.get("issuer_cnpj"),
            "issuer_legal_name": snapshot.get("issuer_legal_name"),
            "issuer_cnae": snapshot.get("issuer_cnae"),
            "integration_mode": snapshot.get("integration_mode"),
            "nfs_provider": snapshot.get("nfs_provider"),
            "default_rps_series": snapshot.get("default_rps_series"),
            "service_code": snapshot.get("service_code"),
            "service_list_item": snapshot.get("service_list_item"),
            "operation_nature": snapshot.get("operation_nature"),
            "service_city": snapshot.get("service_city"),
            "iss_city": snapshot.get("iss_city"),
            "issuer_iss_rate": snapshot.get("issuer_iss_rate"),
            "fiscal_notes": snapshot.get("fiscal_notes"),
            "issue_date": native_billing.issue_date.isoformat() if native_billing.issue_date else None,
            "gross_amount": float(native_billing.gross_amount or 0),
            "net_amount": float(native_billing.net_amount or 0),
            "withholding_flags": snapshot.get("withholding_flags") or {},
            "items": [
                {
                    "description": item.description,
                    "amount": float(item.amount or 0),
                }
                for item in native_billing.items.order_by(ContractNativeBillingItem.id.asc()).all()
            ],
        }

    @staticmethod
    def _get_fiscal_invoice_state(native_billing: ContractNativeBilling) -> dict:
        metadata = dict(native_billing.metadata_json or {})
        state = dict(metadata.get("fiscal_invoice") or {})
        fiscal_payload = ContractService.build_native_billing_fiscal_export_payload(native_billing)
        fiscal_data = {
            **{
                "customer_name": fiscal_payload.get("customer_name"),
                "customer_document": fiscal_payload.get("customer_document"),
                "issuer_cnpj": fiscal_payload.get("issuer_cnpj"),
                "issuer_legal_name": fiscal_payload.get("issuer_legal_name"),
                "integration_mode": fiscal_payload.get("integration_mode"),
                "nfs_provider": fiscal_payload.get("nfs_provider"),
                "rps_series": fiscal_payload.get("default_rps_series"),
                "service_code": fiscal_payload.get("service_code"),
                "service_list_item": fiscal_payload.get("service_list_item"),
                "issuer_cnae": fiscal_payload.get("issuer_cnae"),
                "issuer_iss_rate": fiscal_payload.get("issuer_iss_rate"),
                "service_city": fiscal_payload.get("service_city"),
                "iss_city": fiscal_payload.get("iss_city"),
                "fiscal_notes": fiscal_payload.get("fiscal_notes"),
                "invoice_number": None,
                "issued_at": None,
            },
            **dict(state.get("fiscal_data") or {}),
        }
        state.setdefault("status", "pending")
        state.setdefault("batch_code", None)
        state.setdefault("batch_assigned_at", None)
        state["fiscal_data"] = fiscal_data
        state.setdefault("attachments", [])
        state.setdefault("integration_exports", [])
        return state

    @staticmethod
    def _set_fiscal_invoice_state(native_billing: ContractNativeBilling, state: dict):
        metadata = dict(native_billing.metadata_json or {})
        metadata["fiscal_invoice"] = state
        native_billing.metadata_json = metadata

    @staticmethod
    def _list_fiscal_invoice_billings(company_id: int, filters: Optional[dict] = None) -> list[ContractNativeBilling]:
        filters = dict(filters or {})
        query = (
            ContractNativeBilling.query.filter(
                ContractNativeBilling.company_id == company_id,
                ContractNativeBilling.status != "cancelled",
            )
            .outerjoin(Contract, Contract.id == ContractNativeBilling.contract_id)
            .outerjoin(ContractParty, ContractParty.id == ContractNativeBilling.party_id)
        )
        party_id = ContractService._normalize_int(filters.get("party_id"))
        if party_id:
            query = query.filter(ContractNativeBilling.party_id == party_id)
        search = ContractService._normalize_text(filters.get("search"))
        if search:
            pattern = f"%{search}%"
            query = query.filter(
                or_(
                    ContractNativeBilling.billing_code.ilike(pattern),
                    Contract.code.ilike(pattern),
                    Contract.title.ilike(pattern),
                    ContractParty.name.ilike(pattern),
                    ContractParty.document_number.ilike(pattern),
                )
            )
        return query.order_by(ContractNativeBilling.competence_start.desc(), ContractNativeBilling.id.desc()).all()

    @staticmethod
    def list_fiscal_invoice_workspace(company_id: int, filters: Optional[dict] = None) -> dict:
        filters = dict(filters or {})
        fiscal_status = ContractService._normalize_text(filters.get("fiscal_status") or "active").lower()
        batch_code = ContractService._normalize_text(filters.get("batch_code"))
        billings = ContractService._list_fiscal_invoice_billings(company_id, filters)
        rows: list[dict] = []
        batch_counts: dict[str, int] = {}
        status_counts: dict[str, int] = {}
        for billing in billings:
            state = ContractService._get_fiscal_invoice_state(billing)
            status = ContractService._normalize_text(state.get("status") or "pending").lower()
            current_batch = ContractService._normalize_text(state.get("batch_code"))
            status_counts[status] = status_counts.get(status, 0) + 1
            if current_batch and status != "deleted":
                batch_counts[current_batch] = batch_counts.get(current_batch, 0) + 1

            if fiscal_status == "active" and status == "deleted":
                continue
            if fiscal_status and fiscal_status not in {"active", "all"} and status != fiscal_status:
                continue
            if batch_code == "__without_batch__" and current_batch:
                continue
            if batch_code and batch_code != "__without_batch__" and current_batch != batch_code:
                continue

            gross_amount = ContractService._normalize_decimal(billing.gross_amount)
            net_amount = ContractService._normalize_decimal(billing.net_amount)
            rows.append(
                {
                    "billing": billing,
                    "contract": billing.contract,
                    "party": billing.party,
                    "fiscal_invoice": state,
                    "fiscal_data": dict(state.get("fiscal_data") or {}),
                    "batch_code": current_batch,
                    "retention_amount": (gross_amount - net_amount).quantize(Decimal("0.01")),
                    "item_count": billing.items.count(),
                }
            )
        batches = [
            {"batch_code": code, "item_count": count, "status": "active"}
            for code, count in sorted(batch_counts.items())
        ]
        return {
            "rows": rows,
            "batches": batches,
            "kpis": {
                "total": len(rows),
                "pending": sum(1 for row in rows if row["fiscal_invoice"].get("status") == "pending"),
                "batched": sum(1 for row in rows if row.get("batch_code")),
                "emitted": sum(1 for row in rows if row["fiscal_invoice"].get("status") == "emitted"),
                "cancelled": sum(1 for row in rows if row["fiscal_invoice"].get("status") == "cancelled"),
            },
            "status_counts": status_counts,
        }

    @staticmethod
    def _next_fiscal_invoice_batch_code(company_id: int) -> str:
        last_number = 0
        for billing in ContractNativeBilling.query.filter(ContractNativeBilling.company_id == company_id).all():
            batch_code = ContractService._normalize_text(
                ContractService._get_fiscal_invoice_state(billing).get("batch_code")
            ) or ""
            match = re.search(r"NF\.(\d+)$", batch_code)
            if match:
                last_number = max(last_number, int(match.group(1)))
        return f"NF.{last_number + 1:03d}"

    @staticmethod
    def _get_fiscal_invoice_billings_by_ids(company_id: int, billing_ids: list[int]) -> list[ContractNativeBilling]:
        normalized_ids = [item for item in ContractService._normalize_id_list(billing_ids) if item]
        if not normalized_ids:
            return []
        return (
            ContractNativeBilling.query.filter(
                ContractNativeBilling.company_id == company_id,
                ContractNativeBilling.id.in_(normalized_ids),
                ContractNativeBilling.status != "cancelled",
            )
            .order_by(ContractNativeBilling.id.asc())
            .all()
        )

    @staticmethod
    def assign_fiscal_invoice_batch(*, company_id: int, billing_ids: list[int], batch_code: Optional[str], user_id: Optional[int]) -> dict:
        billings = ContractService._get_fiscal_invoice_billings_by_ids(company_id, billing_ids)
        resolved_batch_code = ContractService._normalize_text(batch_code) or ContractService._next_fiscal_invoice_batch_code(company_id)
        now = datetime.utcnow().isoformat()
        for billing in billings:
            state = ContractService._get_fiscal_invoice_state(billing)
            state["batch_code"] = resolved_batch_code
            state["batch_assigned_at"] = now
            state["batch_assigned_by_user_id"] = user_id
            if state.get("status") in {None, "", "deleted"}:
                state["status"] = "pending"
            ContractService._set_fiscal_invoice_state(billing, state)
        db.session.commit()
        return {"updated": len(billings), "batch_code": resolved_batch_code}

    @staticmethod
    def remove_fiscal_invoice_batch(*, company_id: int, billing_ids: list[int], user_id: Optional[int]) -> dict:
        billings = ContractService._get_fiscal_invoice_billings_by_ids(company_id, billing_ids)
        now = datetime.utcnow().isoformat()
        for billing in billings:
            state = ContractService._get_fiscal_invoice_state(billing)
            state["previous_batch_code"] = state.get("batch_code")
            state["batch_code"] = None
            state["batch_removed_at"] = now
            state["batch_removed_by_user_id"] = user_id
            ContractService._set_fiscal_invoice_state(billing, state)
        db.session.commit()
        return {"updated": len(billings)}

    @staticmethod
    def update_fiscal_invoice_data(*, company_id: int, billing_id: int, payload: dict, user_id: Optional[int]) -> ContractNativeBilling:
        billing = ContractNativeBilling.query.filter(
            ContractNativeBilling.company_id == company_id,
            ContractNativeBilling.id == billing_id,
            ContractNativeBilling.status != "cancelled",
        ).first()
        if not billing:
            raise ValueError("Registro fiscal não localizado para a empresa ativa.")
        state = ContractService._get_fiscal_invoice_state(billing)
        fiscal_data = dict(state.get("fiscal_data") or {})
        for key in (
            "customer_name",
            "customer_document",
            "issuer_cnpj",
            "issuer_legal_name",
            "integration_mode",
            "nfs_provider",
            "rps_series",
            "service_city",
            "iss_city",
            "fiscal_notes",
            "invoice_number",
            "issued_at",
        ):
            if key in payload:
                fiscal_data[key] = ContractService._normalize_text(payload.get(key)) or None
        state["fiscal_data"] = fiscal_data
        state["updated_at"] = datetime.utcnow().isoformat()
        state["updated_by_user_id"] = user_id
        ContractService._set_fiscal_invoice_state(billing, state)
        db.session.commit()
        return billing

    @staticmethod
    def update_fiscal_invoice_status(
        *,
        company_id: int,
        billing_ids: list[int],
        status: str,
        payload: Optional[dict] = None,
        user_id: Optional[int],
    ) -> dict:
        status = (ContractService._normalize_text(status) or "").lower()
        if status not in {"pending", "emitted", "cancelled", "deleted"}:
            raise ValueError("Status fiscal inválido.")
        payload = payload or {}
        billings = ContractService._get_fiscal_invoice_billings_by_ids(company_id, billing_ids)
        now = datetime.utcnow().isoformat()
        for billing in billings:
            state = ContractService._get_fiscal_invoice_state(billing)
            fiscal_data = dict(state.get("fiscal_data") or {})
            if status == "emitted":
                fiscal_data["invoice_number"] = ContractService._normalize_text(payload.get("invoice_number")) or fiscal_data.get("invoice_number")
                fiscal_data["issued_at"] = (
                    ContractService._normalize_text(payload.get("issued_at"))
                    or fiscal_data.get("issued_at")
                    or date.today().isoformat()
                )
                state["emitted_at"] = now
                state["emitted_by_user_id"] = user_id
            elif status == "cancelled":
                state["cancelled_at"] = now
                state["cancelled_by_user_id"] = user_id
                state["cancellation_reason"] = ContractService._normalize_text(payload.get("reason")) or None
            elif status == "deleted":
                state["deleted_at"] = now
                state["deleted_by_user_id"] = user_id
                state["batch_code"] = None
            state["status"] = status
            state["fiscal_data"] = fiscal_data
            state["updated_at"] = now
            state["updated_by_user_id"] = user_id
            ContractService._set_fiscal_invoice_state(billing, state)
        db.session.commit()
        return {"updated": len(billings), "status": status}

    @staticmethod
    def _digits_only(value: object) -> Optional[str]:
        digits = re.sub(r"\D", "", str(value or ""))
        return digits or None

    @staticmethod
    def _normalize_export_code(value: object) -> Optional[str]:
        text = ContractService._normalize_text(value)
        if not text:
            return None
        digits = ContractService._digits_only(text)
        return digits or text

    @staticmethod
    def _normalize_export_uf(value: object) -> Optional[str]:
        text = ContractService._normalize_text(value)
        return text.upper() if text else None

    @staticmethod
    def _decimal_to_br_text(value: object, places: int = 2, strip_trailing: bool = False) -> Optional[str]:
        amount = ContractService._normalize_decimal(value)
        if amount == Decimal("0.00"):
            return None
        quantizer = Decimal("1") if places <= 0 else Decimal("1").scaleb(-places)
        text = f"{amount.quantize(quantizer):f}"
        if strip_trailing and "." in text:
            text = text.rstrip("0").rstrip(".")
        return text.replace(".", ",")

    @staticmethod
    def _decimal_to_export_text(value: object, places: int = 2, strip_trailing: bool = False) -> Optional[str]:
        amount = ContractService._normalize_decimal(value)
        if amount == Decimal("0.00"):
            return None
        quantizer = Decimal("1") if places <= 0 else Decimal("1").scaleb(-places)
        text = f"{amount.quantize(quantizer):f}"
        if strip_trailing and "." in text:
            text = text.rstrip("0").rstrip(".")
        return text

    @staticmethod
    def _metadata_value(sources: list[dict], *keys: str) -> Optional[str]:
        normalized_keys = {str(key or "").strip().lower(): key for key in keys}
        for source in sources:
            if not isinstance(source, dict):
                continue
            for key in keys:
                value = source.get(key)
                if value not in (None, ""):
                    return str(value).strip()
            lower_source = {str(key or "").strip().lower(): value for key, value in source.items()}
            for normalized_key in normalized_keys:
                value = lower_source.get(normalized_key)
                if value not in (None, ""):
                    return str(value).strip()
        return None

    @staticmethod
    def _metadata_value_with_fallback(primary_sources: list[dict], secondary_sources: list[dict], *keys: str) -> Optional[str]:
        return ContractService._metadata_value(primary_sources, *keys) or ContractService._metadata_value(secondary_sources, *keys)

    @staticmethod
    def _metadata_sources(*payloads: object) -> list[dict]:
        sources: list[dict] = []
        for payload in payloads:
            if not isinstance(payload, dict):
                continue
            sources.append(payload)
            for nested_key in ("address", "endereco", "billing_address", "fiscal_address", "tomador_address"):
                nested = payload.get(nested_key)
                if isinstance(nested, dict):
                    sources.append(nested)
        return sources

    @staticmethod
    def _party_metadata_sources(company_id: int, party: Optional[ContractParty]) -> list[dict]:
        if not party:
            return []
        sources = ContractService._metadata_sources(party.metadata_json or {})
        counterparty_id = ContractService._normalize_int(getattr(party, "financial_counterparty_id", None))
        if counterparty_id:
            counterparty = FinancialCounterparty.query.filter(
                FinancialCounterparty.id == counterparty_id,
                FinancialCounterparty.company_id == company_id,
                FinancialCounterparty.deleted_at.is_(None),
            ).first()
            if counterparty:
                sources.extend(ContractService._metadata_sources(counterparty.metadata_json or {}))
        return sources

    @staticmethod
    def _billing_item_metadata_sources(native_billing: ContractNativeBilling) -> list[dict]:
        sources: list[dict] = []
        for billing_item in native_billing.items.order_by(ContractNativeBillingItem.id.asc()).all():
            sources.extend(ContractService._metadata_sources(billing_item.metadata_json or {}))
            contract_item = billing_item.contract_item
            if contract_item:
                sources.extend(ContractService._metadata_sources(contract_item.metadata_json or {}))
                catalog_item = contract_item.contract_catalog_item
                if catalog_item:
                    sources.extend(ContractService._metadata_sources(catalog_item.metadata_json or {}))
        return sources

    @staticmethod
    def _billing_item_descriptions(native_billing: ContractNativeBilling) -> str:
        descriptions = [
            ContractService._normalize_text(item.description)
            for item in native_billing.items.order_by(ContractNativeBillingItem.id.asc()).all()
            if ContractService._normalize_text(item.description)
        ]
        return "; ".join(descriptions)

    @staticmethod
    def _retention_totals_by_kind(native_billing: ContractNativeBilling) -> tuple[dict[str, Decimal], dict[str, Decimal]]:
        totals: dict[str, Decimal] = {}
        rates: dict[str, Decimal] = {}
        for billing_item in native_billing.items.order_by(ContractNativeBillingItem.id.asc()).all():
            metadata = dict(billing_item.metadata_json or {})
            for detail in list(metadata.get("retention_details") or []):
                kind = (ContractService._normalize_text((detail or {}).get("kind")) or "").lower()
                if not kind:
                    continue
                amount = ContractService._normalize_decimal(
                    (detail or {}).get("calculated_amount") or (detail or {}).get("retention_amount")
                )
                totals[kind] = totals.get(kind, Decimal("0.00")) + amount
                if (ContractService._normalize_text((detail or {}).get("retention_value_mode")) or "").lower() == "percent":
                    rates[kind] = ContractService._normalize_decimal((detail or {}).get("retention_value"))
        return totals, rates

    @staticmethod
    def _issuer_iss_rate_from_fiscal_sources(fiscal_data: Optional[dict], fiscal_sources: Optional[list[dict]] = None) -> Decimal:
        rate = ContractService._normalize_decimal((fiscal_data or {}).get("issuer_iss_rate"))
        if rate > Decimal("0.00"):
            return rate.quantize(Decimal("0.0001"))
        rate = ContractService._normalize_decimal(
            ContractService._metadata_value(
                fiscal_sources,
                "issuer_iss_rate",
                "aliquota_iss_emissora",
                "iss_rate_percent",
                "aliquota_iss_percent",
            )
        )
        if rate > Decimal("0.00"):
            return rate.quantize(Decimal("0.0001"))
        return Decimal("0.00")

    @staticmethod
    def _build_fiscal_invoice_nfse_row(*, company_id: int, native_billing: ContractNativeBilling) -> dict:
        state = ContractService._get_fiscal_invoice_state(native_billing)
        fiscal_data = dict(state.get("fiscal_data") or {})
        party = native_billing.party
        party_sources = ContractService._party_metadata_sources(company_id, party)
        item_sources = ContractService._billing_item_metadata_sources(native_billing)
        fiscal_sources = ContractService._metadata_sources(fiscal_data, dict((native_billing.metadata_json or {}).get("fiscal_snapshot") or {}))
        all_sources = item_sources + fiscal_sources
        retention_totals, retention_rates = ContractService._retention_totals_by_kind(native_billing)
        gross_amount = ContractService._normalize_decimal(native_billing.gross_amount)
        iss_amount = retention_totals.get("iss", Decimal("0.00"))
        other_amount = retention_totals.get("other", Decimal("0.00"))
        iss_rate = retention_rates.get("iss")
        issuer_iss_rate = ContractService._issuer_iss_rate_from_fiscal_sources(fiscal_data, fiscal_sources)
        if issuer_iss_rate > Decimal("0.00"):
            iss_rate = issuer_iss_rate
        export_iss_as_other = ContractService._should_export_iss_as_other(fiscal_data, all_sources)
        if export_iss_as_other:
            other_amount += iss_amount
        if not iss_rate and iss_amount > Decimal("0.00") and gross_amount > Decimal("0.00"):
            iss_rate = (iss_amount / gross_amount * Decimal("100")).quantize(Decimal("0.0001"))

        description = (
            ContractService._normalize_text(fiscal_data.get("fiscal_notes"))
            or ContractService._metadata_value(all_sources, "fiscal_notes", "descricao", "description")
            or ContractService._billing_item_descriptions(native_billing)
        )
        description_lines = [description or "Serviços prestados"]
        for line in ContractService._collect_retention_observation_lines(
            native_billing=native_billing,
            fiscal_data=fiscal_data,
            fallback_sources=all_sources,
        ):
            if line and line not in description_lines:
                description_lines.append(line)
        description = "\n".join(line for line in description_lines if line)

        row = {
            "CPF_CNPJ": ContractService._digits_only(
                fiscal_data.get("customer_document") or (party.document_number if party else None)
            ),
            "Nome": fiscal_data.get("customer_name") or (party.legal_name or party.name if party else None),
            "Email": (party.email if party else None) or ContractService._metadata_value(party_sources, "email", "Email"),
            "Valor": float(gross_amount.quantize(Decimal("0.01"))),
            "Codigo_Servico": ContractService._normalize_export_code(
                ContractService._metadata_value_with_fallback(
                    fiscal_sources,
                    item_sources,
                    "Codigo_Servico",
                    "codigo_servico",
                    "service_code",
                    "municipal_service_code",
                )
            ),
            "Endereco_Pais": ContractService._metadata_value(
                party_sources,
                "Endereco_Pais",
                "endereco_pais",
                "country_code",
                "country",
                "pais",
            ) or "BRA",
            "Endereco_Cep": ContractService._digits_only(
                ContractService._metadata_value(party_sources, "Endereco_Cep", "endereco_cep", "zip_code", "zipcode", "postal_code", "cep")
            ),
            "Endereco_Logradouro": ContractService._metadata_value(
                party_sources,
                "Endereco_Logradouro",
                "endereco_logradouro",
                "address_line",
                "street",
                "logradouro",
                "address",
            ),
            "Endereco_Numero": ContractService._metadata_value(
                party_sources,
                "Endereco_Numero",
                "endereco_numero",
                "address_number",
                "numero",
                "number",
            ) or "s/n",
            "Endereco_Complemento": ContractService._metadata_value(
                party_sources,
                "Endereco_Complemento",
                "endereco_complemento",
                "complement",
                "complemento",
            ),
            "Endereco_Bairro": ContractService._metadata_value(
                party_sources,
                "Endereco_Bairro",
                "endereco_bairro",
                "district",
                "bairro",
                "neighborhood",
            ),
            "Endereco_Cidade_Codigo": ContractService._normalize_export_code(
                ContractService._metadata_value(
                    party_sources,
                    "Endereco_Cidade_Codigo",
                    "endereco_cidade_codigo",
                    "city_code_ibge",
                    "ibge_city_code",
                    "codigo_ibge",
                )
            ),
            "Endereco_Cidade_Nome": ContractService._metadata_value(
                party_sources,
                "Endereco_Cidade_Nome",
                "endereco_cidade_nome",
                "city_name",
                "city",
                "cidade",
                "municipio",
            ),
            "Endereco_Estado": ContractService._normalize_export_uf(
                ContractService._metadata_value(
                    party_sources,
                    "Endereco_Estado",
                    "endereco_estado",
                    "uf",
                    "state",
                    "estado",
                )
            ),
            "Descricao": description,
            "IBSCBS_Indicador_Operacao": ContractService._normalize_export_code(
                ContractService._metadata_value(
                    all_sources,
                    "IBSCBS_Indicador_Operacao",
                    "ibscbs_indicador_operacao",
                    "cindop",
                    "c_ind_op",
                )
            ),
            "IBSCBS_Codigo_Classificacao": ContractService._normalize_export_code(
                ContractService._metadata_value(
                    all_sources,
                    "IBSCBS_Codigo_Classificacao",
                    "ibscbs_codigo_classificacao",
                    "cclasstrib",
                    "c_class_trib",
                )
            ),
            "IBSCBS_Tipo_Operacao": ContractService._normalize_export_code(
                ContractService._metadata_value_with_fallback(
                    fiscal_sources,
                    item_sources,
                    "IBSCBS_Tipo_Operacao",
                    "ibscbs_tipo_operacao",
                    "tipo_operacao",
                    "service_list_code",
                    "service_list_item",
                )
            ),
            "NBS": ContractService._digits_only(
                ContractService._metadata_value_with_fallback(item_sources, fiscal_sources, "NBS", "nbs")
            ),
            "CNAE": ContractService._normalize_export_code(
                ContractService._metadata_value(all_sources, "CNAE", "cnae", "issuer_cnae")
            ),
            "Aliquota_ISS": ContractService._decimal_to_export_text(iss_rate, places=4, strip_trailing=True) if iss_rate else None,
            "Valor_ISS": ContractService._decimal_to_br_text(iss_amount) if iss_amount > Decimal("0.00") else None,
            "Retencao_IR": ContractService._decimal_to_br_text(retention_totals.get("irrf")) if retention_totals.get("irrf") else None,
            "Retencao_INSS": ContractService._decimal_to_br_text(retention_totals.get("inss")) if retention_totals.get("inss") else None,
            "Retencao_CSLL": ContractService._decimal_to_br_text(retention_totals.get("csrf")) if retention_totals.get("csrf") else None,
            "Retencao_ISS": ContractService._decimal_to_br_text(iss_amount) if iss_amount > Decimal("0.00") and not export_iss_as_other else None,
            "Retencao_OUTROS": ContractService._decimal_to_br_text(other_amount) if other_amount > Decimal("0.00") else None,
        }
        return {key: value for key, value in row.items() if value not in (None, "")}

    @staticmethod
    def build_fiscal_invoice_integration_spreadsheet(*, company_id: int, billing_ids: list[int], user_id: Optional[int]) -> dict:
        billings = ContractService._get_fiscal_invoice_billings_by_ids(company_id, billing_ids)
        rows = [
            ContractService._build_fiscal_invoice_nfse_row(company_id=company_id, native_billing=billing)
            for billing in billings
        ]
        content = build_nfse_integration_workbook(rows)
        now = datetime.utcnow().isoformat()
        for billing in billings:
            state = ContractService._get_fiscal_invoice_state(billing)
            exports = list(state.get("integration_exports") or [])
            exports.append({"exported_at": now, "exported_by_user_id": user_id, "format": "xlsx_nfse_save_water"})
            state["integration_exports"] = exports
            state["last_integration_export_at"] = now
            ContractService._set_fiscal_invoice_state(billing, state)
        db.session.commit()
        filename = f"notas_fiscais_integracao_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return {"filename": filename, "content": content, "mimetype": NFSE_XLSX_MIMETYPE, "row_count": len(rows)}

    @staticmethod
    def _save_fiscal_invoice_upload(company_id: int, billing_id: Optional[int], file: FileStorage) -> dict:
        safe_name = secure_filename(file.filename or "arquivo") or "arquivo"
        suffix = Path(safe_name).suffix.lower()
        stored_name = f"{uuid4().hex}{suffix}"
        relative_dir = Path("contracts") / "fiscal_invoices" / str(company_id) / (str(billing_id) if billing_id else "imports")
        target_dir = Path(current_app.config["UPLOAD_FOLDER"]) / relative_dir
        target_dir.mkdir(parents=True, exist_ok=True)
        file.stream.seek(0)
        file.save(target_dir / stored_name)
        mime_type = file.mimetype or mimetypes.guess_type(safe_name)[0] or "application/octet-stream"
        relative_path = (relative_dir / stored_name).as_posix()
        return {
            "id": uuid4().hex,
            "file_name": safe_name,
            "stored_name": stored_name,
            "mime_type": mime_type,
            "extension": suffix,
            "uploaded_at": datetime.utcnow().isoformat(),
            "url": f"/uploads/{relative_path}",
            "path": relative_path,
        }

    @staticmethod
    def _extract_xml_invoice_data(raw_bytes: bytes) -> dict:
        try:
            root = ET.fromstring(raw_bytes)
        except ET.ParseError:
            return {}

        def local_name(tag: str) -> str:
            return str(tag or "").split("}")[-1]

        values: dict[str, list[str]] = {}
        for element in root.iter():
            text = (element.text or "").strip()
            if not text:
                continue
            values.setdefault(local_name(element.tag), []).append(text)

        def first(*names: str) -> Optional[str]:
            for name in names:
                items = values.get(name)
                if items:
                    return items[0]
            return None

        return {
            "invoice_number": first("nNF", "Numero", "NumeroNfse", "numero"),
            "issued_at": first("dhEmi", "dEmi", "DataEmissao", "dataEmissao"),
            "customer_document": first("CNPJ", "CPF", "CpfCnpj", "cnpj", "cpf"),
            "customer_name": first("xNome", "RazaoSocial", "NomeRazaoSocial"),
            "gross_amount": first("vNF", "ValorServicos", "ValorNfse", "valor"),
        }

    @staticmethod
    def _match_fiscal_invoice_upload(company_id: int, extracted: dict) -> Optional[ContractNativeBilling]:
        document_digits = re.sub(r"\D", "", str(extracted.get("customer_document") or ""))
        gross_amount = ContractService._normalize_decimal(extracted.get("gross_amount")) if extracted.get("gross_amount") else None
        if not document_digits and gross_amount is None:
            return None
        candidates = ContractNativeBilling.query.filter(
            ContractNativeBilling.company_id == company_id,
            ContractNativeBilling.status != "cancelled",
        ).order_by(ContractNativeBilling.id.desc()).limit(300).all()
        for billing in candidates:
            party_digits = re.sub(r"\D", "", str(billing.party.document_number if billing.party else ""))
            if document_digits and party_digits and document_digits[-8:] != party_digits[-8:]:
                continue
            if gross_amount is not None:
                diff = abs(ContractService._normalize_decimal(billing.gross_amount) - gross_amount)
                if diff > Decimal("0.05"):
                    continue
                return billing
        return None

    @staticmethod
    def _extract_tabular_invoice_rows(raw_bytes: bytes, filename: str) -> list[dict]:
        suffix = Path(filename or "").suffix.lower()

        def normalize_key(value) -> str:
            return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

        def normalize_row(row: dict) -> dict:
            row = {normalize_key(key): value for key, value in row.items()}

            def first(*names: str):
                for name in names:
                    value = row.get(normalize_key(name))
                    if value not in (None, ""):
                        return str(value).strip()
                return None

            return {
                "invoice_number": first("invoice_number", "numero_nf", "n_nf", "nf", "nota_fiscal", "numero_nota"),
                "issued_at": first("issued_at", "data_emissao", "emissao", "data_da_emissao"),
                "customer_document": first("customer_document", "documento_cliente", "cnpj", "cpf", "cpf_cnpj"),
                "customer_name": first("customer_name", "cliente", "razao_social", "tomador"),
                "gross_amount": first("gross_amount", "valor_bruto", "valor_servicos", "valor_nf", "valor"),
            }

        if suffix == ".csv":
            text = None
            for encoding in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = raw_bytes.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if not text:
                return []
            sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
                reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            except csv.Error:
                reader = csv.DictReader(io.StringIO(text), delimiter=";")
            return [normalize_row(row) for row in reader if any(row.values())]

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception:
                return []
            workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                return []
            parsed_rows = []
            for values in rows:
                row = dict(zip(headers, values))
                if any(value not in (None, "") for value in row.values()):
                    parsed_rows.append(normalize_row(row))
            return parsed_rows

        return []

    @staticmethod
    def _apply_fiscal_invoice_upload_to_billing(
        *,
        company_id: int,
        billing: ContractNativeBilling,
        file: FileStorage,
        extracted: dict,
        user_id: Optional[int],
    ) -> None:
        attachment = ContractService._save_fiscal_invoice_upload(company_id, billing.id, file)
        state = ContractService._get_fiscal_invoice_state(billing)
        attachments = list(state.get("attachments") or [])
        attachments.append(attachment)
        fiscal_data = dict(state.get("fiscal_data") or {})
        for key in ("invoice_number", "issued_at", "customer_document", "customer_name"):
            if extracted.get(key):
                fiscal_data[key] = extracted[key]
        if extracted.get("invoice_number") or extracted.get("issued_at"):
            state["status"] = "emitted"
            state["emitted_at"] = state.get("emitted_at") or datetime.utcnow().isoformat()
            state["emitted_by_user_id"] = state.get("emitted_by_user_id") or user_id
        state["attachments"] = attachments
        state["fiscal_data"] = fiscal_data
        state["updated_at"] = datetime.utcnow().isoformat()
        state["updated_by_user_id"] = user_id
        ContractService._set_fiscal_invoice_state(billing, state)

    @staticmethod
    def upload_fiscal_invoice_files(*, company_id: int, billing_ids: list[int], files: list[FileStorage], user_id: Optional[int]) -> dict:
        selected_billings = ContractService._get_fiscal_invoice_billings_by_ids(company_id, billing_ids)
        updated = 0
        unmatched = 0
        for file in files:
            if not file or not file.filename:
                continue
            file.stream.seek(0)
            raw_bytes = file.read()
            file.stream.seek(0)
            lower_filename = file.filename.lower()
            extracted_rows = []
            if lower_filename.endswith(".xml"):
                extracted_rows = [ContractService._extract_xml_invoice_data(raw_bytes)]
            elif lower_filename.endswith((".csv", ".xlsx")):
                extracted_rows = ContractService._extract_tabular_invoice_rows(raw_bytes, file.filename)
            extracted_rows = [row for row in extracted_rows if row]

            if selected_billings:
                extracted = extracted_rows[0] if len(selected_billings) == 1 and extracted_rows else {}
                for billing in selected_billings:
                    ContractService._apply_fiscal_invoice_upload_to_billing(
                        company_id=company_id,
                        billing=billing,
                        file=file,
                        extracted=extracted,
                        user_id=user_id,
                    )
                    updated += 1
                continue

            if not extracted_rows:
                ContractService._save_fiscal_invoice_upload(company_id, None, file)
                unmatched += 1
                continue

            for extracted in extracted_rows:
                matched = ContractService._match_fiscal_invoice_upload(company_id, extracted)
                if not matched:
                    unmatched += 1
                    continue
                file.stream.seek(0)
                ContractService._apply_fiscal_invoice_upload_to_billing(
                    company_id=company_id,
                    billing=matched,
                    file=file,
                    extracted=extracted,
                    user_id=user_id,
                )
                updated += 1
        db.session.commit()
        return {"updated": updated, "unmatched": unmatched}


    @staticmethod
    def delete_billing_item(*, contract: Contract, item_id: int) -> bool:
        item = ContractBillingItem.query.filter_by(id=item_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not item:
            return False
        db.session.delete(item)
        db.session.commit()
        return True

    @staticmethod
    def upsert_financial_terms(*, contract: Contract, payload: dict, user_id: Optional[int] = None):
        record = ContractFinancialTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
        if not record:
            record = ContractFinancialTerm(company_id=contract.company_id, contract_id=contract.id)
            db.session.add(record)
        record.default_bank_account_id = ContractService._normalize_int(payload.get("default_bank_account_id"))
        record.default_payment_method_id = ContractService._normalize_int(payload.get("default_payment_method_id"))
        record.default_chart_account_id = ContractService._normalize_int(payload.get("default_chart_account_id"))
        record.default_cost_center_id = ContractService._normalize_int(payload.get("default_cost_center_id"))
        record.correction_index_id = ContractService._normalize_int(payload.get("correction_index_id"))
        record.payment_term_type = ContractService._normalize_text(payload.get("payment_term_type")) or None
        record.payment_term_days = ContractService._normalize_int(payload.get("payment_term_days"))
        record.billing_method = ContractService._normalize_text(payload.get("billing_method")) or None
        record.pricing_model = ContractService._normalize_text(payload.get("pricing_model")) or None
        record.adjustment_rule = ContractService._normalize_text(payload.get("adjustment_rule")) or None
        record.notes = ContractService._normalize_text(payload.get("notes")) or None
        record.metadata_json = {
            **(record.metadata_json or {}),
            "updated_by_user_id": user_id,
            "financial_defaults": {
                "bank_account_id": record.default_bank_account_id,
                "payment_method_id": record.default_payment_method_id,
                "chart_account_id": record.default_chart_account_id,
                "cost_center_id": record.default_cost_center_id,
                "correction_index_id": record.correction_index_id,
            },
        }
        ContractService.record_event(
            contract=contract,
            event_type="contract.financial_terms_updated",
            description="Condições financeiras do contrato atualizadas.",
            payload={
                "default_bank_account_id": record.default_bank_account_id,
                "default_chart_account_id": record.default_chart_account_id,
                "default_cost_center_id": record.default_cost_center_id,
            },
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return record

    @staticmethod
    def upsert_fiscal_terms(*, contract: Contract, payload: dict, user_id: Optional[int] = None):
        record = ContractFiscalTerm.query.filter_by(contract_id=contract.id, company_id=contract.company_id).first()
        if not record:
            record = ContractFiscalTerm(company_id=contract.company_id, contract_id=contract.id)
            db.session.add(record)
        legal_entity_id = ContractService._normalize_int(payload.get("contracting_legal_entity_id")) or contract.contracting_legal_entity_id
        if not legal_entity_id:
            raise ValueError("Selecione a PJ contratada que também emitirá a nota fiscal.")

        legal_entity = ContractService.get_contracting_legal_entity(contract.company_id, legal_entity_id)
        if not legal_entity:
            raise ValueError("PJ contratada inválida para a empresa ativa.")

        integration_mode = ContractService._normalize_text(payload.get("integration_mode")) or legal_entity.integration_mode or "manual"
        nfs_provider = ContractService._normalize_text(payload.get("nfs_provider")) or legal_entity.nfs_provider or None
        service_city = ContractService._normalize_text(payload.get("service_city")) or legal_entity.service_city or None
        iss_city = ContractService._normalize_text(payload.get("iss_city")) or service_city

        contract.contracting_legal_entity_id = legal_entity.id
        contract.updated_by_user_id = user_id
        record.contracting_legal_entity_id = legal_entity.id
        record.fiscal_profile_code = ContractService._normalize_text(payload.get("fiscal_profile_code")) or None
        record.integration_mode = integration_mode
        record.nfs_provider = nfs_provider
        record.default_rps_series = ContractService._normalize_text(payload.get("default_rps_series")) or None
        record.service_code = None
        record.service_list_item = None
        record.operation_nature = None
        record.service_city = service_city
        record.iss_city = iss_city
        record.tax_nature = None
        record.api_profile_id = legal_entity.api_profile_id
        record.spreadsheet_profile_id = legal_entity.spreadsheet_profile_id
        record.withholding_flags = {}
        record.tax_observation = ContractService._normalize_text(payload.get("tax_observation")) or None
        record.notes = None
        record.metadata_json = {
            **(record.metadata_json or {}),
            "issuer_cnpj": legal_entity.cnpj,
            "issuer_legal_name": legal_entity.legal_name,
            "compliance_rule": "contracting_legal_entity_equals_nf_issuer",
        }
        ContractService.record_event(
            contract=contract,
            event_type="contract.fiscal_config_updated",
            description="Configuração fiscal do contrato atualizada com PJ contratada vinculada à emissão.",
            payload={
                "contracting_legal_entity_id": legal_entity.id,
                "integration_mode": integration_mode,
                "nfs_provider": record.nfs_provider,
                "service_city": record.service_city,
                "iss_city": record.iss_city,
            },
            user_id=user_id,
            auto_commit=False,
        )
        db.session.commit()
        return record

    @staticmethod
    def add_retention(*, contract: Contract, payload: dict):
        retention = ContractRetention(
            company_id=contract.company_id,
            contract_id=contract.id,
            retention_type=ContractService._normalize_text(payload.get("retention_type")) or "retencao",
            calculation_mode=ContractService._normalize_text(payload.get("calculation_mode")) or None,
            rate_percent=ContractService._normalize_decimal(payload.get("rate_percent")),
            fixed_amount=ContractService._normalize_decimal(payload.get("fixed_amount")),
            notes=ContractService._normalize_text(payload.get("notes")) or None,
        )
        db.session.add(retention)
        db.session.commit()
        return retention

    @staticmethod
    def delete_retention(*, contract: Contract, retention_id: int) -> bool:
        retention = ContractRetention.query.filter_by(id=retention_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not retention:
            return False
        db.session.delete(retention)
        db.session.commit()
        return True

    @staticmethod
    def add_trigger(*, contract: Contract, payload: dict):
        trigger = ContractTrigger(
            company_id=contract.company_id,
            contract_id=contract.id,
            trigger_type=ContractService._normalize_text(payload.get("trigger_type")) or "alert",
            reference_date_type=ContractService._normalize_text(payload.get("reference_date_type")) or None,
            reference_date_value=ContractService._normalize_date(payload.get("reference_date_value")),
            offset_days=ContractService._normalize_int(payload.get("offset_days")),
            periodicity=ContractService._normalize_text(payload.get("periodicity")) or None,
            alert_before_days=ContractService._normalize_int(payload.get("alert_before_days")),
            is_active=ContractService._normalize_bool(payload.get("is_active")) if payload.get("is_active") not in (None, "") else True,
        )
        db.session.add(trigger)
        db.session.commit()
        return trigger

    @staticmethod
    def delete_trigger(*, contract: Contract, trigger_id: int) -> bool:
        trigger = ContractTrigger.query.filter_by(id=trigger_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not trigger:
            return False
        db.session.delete(trigger)
        db.session.commit()
        return True

    @staticmethod
    def save_document(*, contract: Contract, document_type: str, document_version: str, is_signed_version: bool, file: Optional[FileStorage], uploaded_by_user_id: Optional[int]):
        if file is None or not file.filename:
            raise ValueError("Selecione um arquivo para anexar ao contrato.")
        safe_name = secure_filename(file.filename)
        if not safe_name:
            raise ValueError("Nome de arquivo inválido.")
        relative_dir = Path("contracts") / f"company_{contract.company_id}" / f"contract_{contract.id}"
        target_dir = Path(current_app.config["UPLOAD_FOLDER"]) / relative_dir
        target_dir.mkdir(parents=True, exist_ok=True)
        final_name = f"{uuid4().hex}_{safe_name}"
        target_path = target_dir / final_name
        file.save(target_path)
        record = ContractDocument(
            company_id=contract.company_id,
            contract_id=contract.id,
            document_type=ContractService._normalize_text(document_type) or "documento",
            file_name=safe_name,
            file_path=str((relative_dir / final_name).as_posix()),
            mime_type=file.mimetype,
            document_version=ContractService._normalize_text(document_version) or None,
            source="manual",
            is_signed_version=bool(is_signed_version),
            uploaded_by_user_id=uploaded_by_user_id,
        )
        db.session.add(record)
        db.session.commit()
        return record

    @staticmethod
    def delete_document(*, contract: Contract, document_id: int) -> bool:
        document = ContractDocument.query.filter_by(id=document_id, company_id=contract.company_id, contract_id=contract.id).first()
        if not document:
            return False
        file_path = Path(current_app.config["UPLOAD_FOLDER"]) / str(document.file_path)
        if file_path.exists():
            try:
                file_path.unlink()
            except OSError:
                pass
        db.session.delete(document)
        db.session.commit()
        return True
