from __future__ import annotations

import csv
import hashlib
import io
import logging
import os
import re
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Sequence, Tuple

from models import db
from models.financial import (
    FinancialBankAccount,
    FinancialChartAccount,
    FinancialClassificationSuggestion,
    FinancialCounterparty,
    FinancialCostCenter,
    FinancialEntry,
    FinancialImportBatch,
    FinancialImportRow,
    FinancialReconciliationMatch,
)
from schemas.financial import FinancialImportBatchInput, FinancialImportRowInput
from services.financial_catalog_service import FinancialCatalogService
from services.financial_direct_entry_service import FinancialDirectEntryService
from services.financial_schedule_service import FinancialScheduleService
from services.financial_service import FinancialService


logger = logging.getLogger(__name__)


class FinancialImportService:
    """Hub determinístico de staging para importações financeiras."""

    EXTRA_COLUMNS_KEY = "__extra_columns__"

    IMPORT_TEMPLATE_COLUMNS = [
        {"key": "tipo_registro", "label": "Tipo de Registro *", "required": True, "example": "agendamento"},
        {"key": "tipo_titulo", "label": "Tipo do Título *", "required": True, "example": "Pagar ou Receber"},
        {"key": "historico", "label": "Histórico *", "required": True, "example": "Mensalidade escritório março/2026"},
        {"key": "numero_documento", "label": "Número do Documento", "required": False, "example": "NF-10293"},
        {"key": "favorecido", "label": "Favorecido *", "required": True, "example": "Fornecedor XPTO"},
        {"key": "documento_favorecido", "label": "CPF/CNPJ Favorecido", "required": False, "example": "12345678000199"},
        {"key": "valor", "label": "Valor *", "required": True, "example": "1500,75"},
        {"key": "competencia", "label": "Competência *", "required": True, "example": "31/03/2026"},
        {"key": "vencimento", "label": "Vencimento", "required": False, "example": "10/04/2026"},
        {"key": "data_lancamento", "label": "Data do Lançamento", "required": False, "example": "31/03/2026"},
        {"key": "plano_conta", "label": "Plano de Conta", "required": False, "example": "3.01.001 ou 145"},
        {"key": "centro_resultado", "label": "Centro de Resultado", "required": False, "example": "2.03.001 ou 245"},
        {"key": "projeto_processo", "label": "Projeto / Processo", "required": False, "example": "PRJ-001 Implantação ERP"},
        {"key": "correcao_financeira", "label": "Correção Financeira", "required": False, "example": "Padrão"},
        {"key": "desconto", "label": "Desconto", "required": False, "example": "Desconto Comercial"},
        {"key": "conta_bancaria", "label": "Conta Bancária", "required": False, "example": "001 - Banco do Brasil"},
        {"key": "observacoes", "label": "Observações", "required": False, "example": "Registro importado da planilha do cliente"},
    ]

    @staticmethod
    def _normalize_digits(value: Any) -> str:
        return "".join(ch for ch in str(value or "") if ch.isdigit())

    @staticmethod
    def _normalize_text_token(value: Any) -> str:
        return str(value or "").strip()

    @staticmethod
    def _normalize_account_variants(value: Any) -> set[str]:
        digits = FinancialImportService._normalize_digits(value)
        if not digits:
            return set()
        stripped = digits.lstrip("0")
        return {digits, stripped or "0"}

    @staticmethod
    def _split_account_components(value: Any) -> Tuple[Optional[str], Optional[str]]:
        text = str(value or "").strip()
        if not text:
            return None, None
        if "-" in text:
            number_part, digit_part = text.split("-", 1)
            number_digits = FinancialImportService._normalize_digits(number_part)
            digit_digits = FinancialImportService._normalize_digits(digit_part)
            return (number_digits or None), (digit_digits or None)
        digits = FinancialImportService._normalize_digits(text)
        return (digits or None), None

    @staticmethod
    def _parse_decimal(value: Any) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        if isinstance(value, Decimal):
            return value

        text = str(value).strip().replace("R$", "").replace(" ", "")
        if not text:
            return None

        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")

        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _parse_date(value: Any) -> Optional[datetime.date]:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "isoformat") and hasattr(value, "year"):
            return value

        text = str(value).strip()
        if "T" in text:
            try:
                return datetime.fromisoformat(text).date()
            except ValueError:
                pass
        if " " in text:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _guess_movement_nature(amount: Optional[Decimal]) -> Optional[str]:
        if amount is None:
            return None
        return "credit" if amount >= 0 else "debit"

    @staticmethod
    def _normalize_choice(value: Any) -> str:
        text = str(value or "").strip().lower()
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return re.sub(r"\s+", " ", text)

    @staticmethod
    def _extract_code_token(value: Any) -> Optional[str]:
        text = str(value or "").strip()
        if not text:
            return None
        token = re.split(r"\s+-\s+|\s{2,}", text, maxsplit=1)[0].strip(" -")
        return token or text

    @staticmethod
    def _map_entry_type(value: Any) -> Optional[str]:
        normalized = FinancialImportService._normalize_choice(value)
        if normalized in {"pagar", "conta a pagar", "a pagar", "payable"}:
            return "payable"
        if normalized in {"receber", "conta a receber", "a receber", "receivable"}:
            return "receivable"
        return None

    @staticmethod
    def _map_record_type(value: Any) -> Optional[str]:
        normalized = FinancialImportService._normalize_choice(value)
        if normalized in {"agendamento", "agendar"}:
            return "agendamento"
        if normalized in {"lancamento", "lançamento", "baixa", "quitado", "pago", "recebido"}:
            return "lancamento"
        return None

    @staticmethod
    def _movement_nature_from_entry_type(entry_type: Optional[str]) -> Optional[str]:
        if entry_type == "receivable":
            return "credit"
        if entry_type == "payable":
            return "debit"
        return None

    @staticmethod
    def _resolve_origin_type(source_type: Any) -> str:
        normalized = FinancialImportService._normalize_choice(source_type)
        if normalized == "xlsx":
            return "xls"
        if normalized in {"ofx", "csv", "xls", "csc", "api", "mcp"}:
            return normalized
        return "manual"

    @staticmethod
    def _resolve_counterparty_id_by_code(company_id: int, code: Any) -> Optional[int]:
        code_text = FinancialImportService._extract_code_token(code)
        if not code_text:
            return None
        item = FinancialCounterparty.query.filter(
            FinancialCounterparty.company_id == company_id,
            FinancialCounterparty.code == code_text,
            FinancialCounterparty.deleted_at.is_(None),
        ).first()
        return item.id if item else None

    @staticmethod
    def _resolve_chart_account_id_by_code(company_id: int, code: Any) -> Optional[int]:
        code_text = FinancialImportService._extract_code_token(code)
        if not code_text:
            return None
        item = FinancialChartAccount.query.filter(
            FinancialChartAccount.company_id == company_id,
            FinancialChartAccount.code == code_text,
            FinancialChartAccount.deleted_at.is_(None),
            FinancialChartAccount.accepts_posting.is_(True),
        ).first()
        return item.id if item else None

    @staticmethod
    def _resolve_cost_center_id_by_code(company_id: int, code: Any) -> Optional[int]:
        code_text = FinancialImportService._extract_code_token(code)
        if not code_text:
            return None
        item = FinancialCostCenter.query.filter(
            FinancialCostCenter.company_id == company_id,
            FinancialCostCenter.code == code_text,
            FinancialCostCenter.deleted_at.is_(None),
            FinancialCostCenter.accepts_posting.is_(True),
        ).first()
        return item.id if item else None

    @staticmethod
    def _resolve_bank_account_id_by_code(company_id: int, code: Any) -> Optional[int]:
        code_text = FinancialImportService._extract_code_token(code)
        if not code_text:
            return None
        item = FinancialBankAccount.query.filter(
            FinancialBankAccount.company_id == company_id,
            FinancialBankAccount.code == code_text,
            FinancialBankAccount.deleted_at.is_(None),
        ).first()
        if item:
            return item.id
        normalized_numeric_code = code_text.lstrip("0")
        if normalized_numeric_code and normalized_numeric_code.isdigit():
            item = FinancialBankAccount.query.filter(
                FinancialBankAccount.company_id == company_id,
                FinancialBankAccount.deleted_at.is_(None),
                db.func.ltrim(FinancialBankAccount.code, "0") == normalized_numeric_code,
            ).first()
        return item.id if item else None

    @staticmethod
    def _sanitize_raw_payload(raw_payload: Dict[str, Any] | None) -> Dict[str, Any]:
        sanitized: Dict[str, Any] = {}
        for index, (key, value) in enumerate((raw_payload or {}).items(), start=1):
            if key is None:
                target_key = FinancialImportService.EXTRA_COLUMNS_KEY
            else:
                target_key = str(key).strip() or f"column_{index}"
            sanitized[target_key] = value
        return sanitized

    @staticmethod
    def _canonicalize_column_key(value: Any) -> str:
        text = str(value or "").strip().lower()
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return text.strip("_")

    @staticmethod
    def _normalize_row(row_number: int, raw_payload: Dict[str, Any]) -> FinancialImportRowInput:
        sanitized_raw_payload = FinancialImportService._sanitize_raw_payload(raw_payload)
        normalized: Dict[str, Any] = {}
        for key, value in sanitized_raw_payload.items():
            raw_key = str(key).strip().lower()
            canonical_key = FinancialImportService._canonicalize_column_key(key)
            if raw_key and raw_key not in normalized:
                normalized[raw_key] = value
            if canonical_key and canonical_key not in normalized:
                normalized[canonical_key] = value
        competence_date = (
            FinancialImportService._parse_date(normalized.get("competence_date"))
            or FinancialImportService._parse_date(normalized.get("competencia"))
            or FinancialImportService._parse_date(normalized.get("competencia_"))
        )
        amount = (
            FinancialImportService._parse_decimal(normalized.get("amount"))
            or FinancialImportService._parse_decimal(normalized.get("valor"))
            or FinancialImportService._parse_decimal(normalized.get("valor_"))
            or FinancialImportService._parse_decimal(normalized.get("valor_documento"))
            or FinancialImportService._parse_decimal(normalized.get("amount_br"))
        )
        description = (
            normalized.get("description")
            or normalized.get("descricao")
            or normalized.get("memo")
            or normalized.get("historico")
            or normalized.get("historico_")
            or normalized.get("payee")
        )
        occurred_on = (
            FinancialImportService._parse_date(normalized.get("occurred_on"))
            or FinancialImportService._parse_date(normalized.get("data"))
            or FinancialImportService._parse_date(normalized.get("date"))
            or FinancialImportService._parse_date(normalized.get("posted_at"))
            or FinancialImportService._parse_date(normalized.get("data_do_lancamento"))
        )
        due_date = (
            FinancialImportService._parse_date(normalized.get("due_date"))
            or FinancialImportService._parse_date(normalized.get("vencimento"))
        )
        settlement_date = (
            FinancialImportService._parse_date(normalized.get("settlement_date"))
            or FinancialImportService._parse_date(normalized.get("data_da_baixa"))
            or FinancialImportService._parse_date(normalized.get("data_baixa"))
        )
        document_number = (
            normalized.get("document_number")
            or normalized.get("documento")
            or normalized.get("doc")
            or normalized.get("checknum")
            or normalized.get("numero_do_documento")
        )
        bank_reference = (
            normalized.get("bank_reference")
            or normalized.get("fitid")
            or normalized.get("ref")
            or normalized.get("reference")
        )
        counterparty_name = (
            normalized.get("counterparty_name")
            or normalized.get("favorecido")
            or normalized.get("favorecido_")
            or normalized.get("payee")
            or normalized.get("name")
        )
        entry_type = FinancialImportService._map_entry_type(
            normalized.get("entry_type")
            or normalized.get("tipo_titulo")
            or normalized.get("tipo_do_titulo")
            or normalized.get("tipo_do_titulo_")
        )
        record_type = FinancialImportService._map_record_type(
            normalized.get("record_type")
            or normalized.get("tipo_registro")
            or normalized.get("tipo_de_registro")
            or normalized.get("tipo_de_registro_")
        )
        movement_nature = (
            FinancialImportService._movement_nature_from_entry_type(entry_type)
            or normalized.get("movement_nature")
            or FinancialImportService._guess_movement_nature(amount)
        )
        if amount is not None and amount < 0:
            amount = abs(amount)

        error_message = None
        processing_status = "staged"
        if not description:
            error_message = "Descrição não identificada na linha importada."
            processing_status = "rejected"
        elif amount is None:
            error_message = "Valor financeiro não identificado na linha importada."
            processing_status = "rejected"

        return FinancialImportRowInput(
            company_id=0,
            row_number=row_number,
            processing_status=processing_status,
            document_number=document_number,
            description=description,
            occurred_on=occurred_on,
            due_date=due_date,
            amount=amount,
            movement_nature=movement_nature,
            bank_reference=bank_reference,
            counterparty_name=counterparty_name,
            raw_payload=sanitized_raw_payload,
            normalized_payload={
                "description": description,
                "competence_date": competence_date.isoformat() if competence_date else None,
                "occurred_on": occurred_on.isoformat() if occurred_on else None,
                "due_date": due_date.isoformat() if due_date else None,
                "settlement_date": settlement_date.isoformat() if settlement_date else None,
                "amount": float(amount) if amount is not None else None,
                "entry_type": entry_type,
                "record_type": record_type,
                "movement_nature": movement_nature,
                "document_number": document_number,
                "bank_reference": bank_reference,
                "counterparty_name": counterparty_name,
                "counterparty_hint": counterparty_name,
                "counterparty_code": FinancialImportService._extract_code_token(
                    normalized.get("counterparty_code")
                    or normalized.get("codigo_favorecido")
                    or normalized.get("codigo_do_favorecido")
                    or normalized.get("favorecido")
                    or normalized.get("favorecido_")
                ),
                "chart_account_code": FinancialImportService._extract_code_token(
                    normalized.get("chart_account_code")
                    or normalized.get("plano_conta")
                    or normalized.get("plano_de_conta")
                ),
                "cost_center_code": FinancialImportService._extract_code_token(
                    normalized.get("cost_center_code")
                    or normalized.get("centro_resultado")
                    or normalized.get("centro_de_resultado")
                ),
                "bank_account_code": FinancialImportService._extract_code_token(
                    normalized.get("bank_account_code")
                    or normalized.get("conta_bancaria")
                ),
            },
            error_message=error_message,
        )

    @staticmethod
    def _enrich_row_catalogs(company_id: int, row: FinancialImportRow) -> None:
        normalized = dict(row.normalized_payload or {})
        resolved_counterparty_id = FinancialImportService._resolve_counterparty_id_by_code(
            company_id,
            normalized.get("counterparty_code"),
        ) if normalized.get("counterparty_code") else None
        if resolved_counterparty_id:
            normalized["counterparty_id"] = resolved_counterparty_id
        resolved_chart_account_id = FinancialImportService._resolve_chart_account_id_by_code(
            company_id,
            normalized.get("chart_account_code"),
        ) if normalized.get("chart_account_code") else None
        if resolved_chart_account_id:
            normalized["chart_account_id"] = resolved_chart_account_id
        resolved_cost_center_id = FinancialImportService._resolve_cost_center_id_by_code(
            company_id,
            normalized.get("cost_center_code"),
        ) if normalized.get("cost_center_code") else None
        if resolved_cost_center_id:
            normalized["cost_center_id"] = resolved_cost_center_id
        resolved_bank_account_id = FinancialImportService._resolve_bank_account_id_by_code(
            company_id,
            normalized.get("bank_account_code"),
        ) if normalized.get("bank_account_code") else None
        if resolved_bank_account_id:
            normalized["bank_account_id"] = resolved_bank_account_id
        enriched = FinancialCatalogService.enrich_reference_payload(
            company_id=company_id,
            payload=normalized,
            counterparty_text=row.counterparty_name,
            description_text=row.description,
            bank_reference=row.bank_reference,
        )
        row.normalized_payload = enriched
        if row.processing_status == "staged" and (
            enriched.get("counterparty_id")
            or enriched.get("chart_account_id")
            or enriched.get("cost_center_id")
            or enriched.get("bank_account_id")
        ):
            row.processing_status = "validated"

    @staticmethod
    def _parse_csv_bytes(file_bytes: bytes, delimiter: str = ",") -> List[Dict[str, Any]]:
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        return [FinancialImportService._sanitize_raw_payload(dict(row)) for row in reader]

    @staticmethod
    def _extract_ofx_account_metadata(file_bytes: bytes) -> Dict[str, Any]:
        text = file_bytes.decode("latin-1", errors="ignore")
        account_block_match = re.search(
            r"<(?:BANKACCTFROM|CCACCTFROM)>(.*?)</(?:BANKACCTFROM|CCACCTFROM)>",
            text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        account_block = account_block_match.group(1) if account_block_match else text
        raw_account = FinancialImportService._extract_ofx_tag(account_block, "ACCTID")
        account_number, account_digit = FinancialImportService._split_account_components(raw_account)
        return {
            "source_type": "ofx",
            "validation_available": bool(
                FinancialImportService._extract_ofx_tag(account_block, "BANKID")
                or FinancialImportService._extract_ofx_tag(account_block, "BRANCHID")
                or account_number
            ),
            "bank_code": FinancialImportService._normalize_digits(
                FinancialImportService._extract_ofx_tag(account_block, "BANKID")
            ) or None,
            "branch_number": FinancialImportService._normalize_digits(
                FinancialImportService._extract_ofx_tag(account_block, "BRANCHID")
            ) or None,
            "account_number": account_number,
            "account_digit": account_digit,
            "account_raw": raw_account,
            "account_type": FinancialImportService._normalize_text_token(
                FinancialImportService._extract_ofx_tag(account_block, "ACCTTYPE")
            ) or None,
            "validation_source": "ofx_header",
        }

    @staticmethod
    def _extract_tabular_account_metadata(rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
        key_aliases = {
            "bank_code": {"bank_code", "codigo_banco", "cod_banco", "banco_codigo", "bankid"},
            "branch_number": {"branch", "agency", "agencia", "ag", "branch_number", "branchid"},
            "account_number": {
                "account",
                "account_number",
                "numero_conta",
                "num_conta",
                "conta",
                "acctid",
                "conta_corrente",
            },
            "account_digit": {"account_digit", "digito_conta", "digito", "dv_conta"},
        }
        for raw_row in rows[:5]:
            normalized = {str(key or "").strip().lower(): value for key, value in (raw_row or {}).items()}
            if not normalized:
                continue
            bank_code = next((normalized[key] for key in key_aliases["bank_code"] if key in normalized), None)
            branch_number = next((normalized[key] for key in key_aliases["branch_number"] if key in normalized), None)
            account_value = next((normalized[key] for key in key_aliases["account_number"] if key in normalized), None)
            account_digit = next((normalized[key] for key in key_aliases["account_digit"] if key in normalized), None)
            account_number, derived_digit = FinancialImportService._split_account_components(account_value)
            resolved_digit = FinancialImportService._normalize_digits(account_digit) or derived_digit
            if bank_code or branch_number or account_number:
                return {
                    "validation_available": True,
                    "bank_code": FinancialImportService._normalize_digits(bank_code) or None,
                    "branch_number": FinancialImportService._normalize_digits(branch_number) or None,
                    "account_number": account_number,
                    "account_digit": resolved_digit or None,
                    "account_raw": FinancialImportService._normalize_text_token(account_value) or None,
                    "validation_source": "tabular_columns",
                }
        return {"validation_available": False, "validation_source": "tabular_columns"}

    @staticmethod
    def _extract_source_account_metadata(
        source_type: str,
        file_bytes: bytes,
        parsed_rows: Sequence[Dict[str, Any]],
    ) -> Dict[str, Any]:
        source = str(source_type or "").strip().lower()
        if source == "ofx":
            return FinancialImportService._extract_ofx_account_metadata(file_bytes)
        metadata = FinancialImportService._extract_tabular_account_metadata(parsed_rows)
        metadata["source_type"] = source
        return metadata

    @staticmethod
    def _validate_batch_bank_account(
        *,
        company_id: int,
        metadata_json: Dict[str, Any],
        source_account_metadata: Dict[str, Any],
    ) -> Optional[str]:
        bank_account_id = int((metadata_json or {}).get("bank_account_id") or 0)
        if not bank_account_id:
            return None
        if not (source_account_metadata or {}).get("validation_available"):
            return None

        account = FinancialBankAccount.query.filter(
            FinancialBankAccount.id == bank_account_id,
            FinancialBankAccount.company_id == company_id,
            FinancialBankAccount.deleted_at.is_(None),
        ).first()
        if not account:
            return "Conta bancária selecionada para o upload não foi encontrada no escopo da empresa."

        source_bank_code = FinancialImportService._normalize_digits(source_account_metadata.get("bank_code"))
        target_bank_code = FinancialImportService._normalize_digits(account.bank_code)
        if source_bank_code and not target_bank_code:
            return (
                "A conta bancária selecionada não possui código do banco cadastrado no sistema. "
                "Preencha o cadastro da conta antes de importar este arquivo."
            )
        if source_bank_code and target_bank_code and source_bank_code != target_bank_code:
            return (
                f"O arquivo pertence ao banco {source_bank_code}, mas a conta selecionada está cadastrada no banco "
                f"{target_bank_code}. Selecione a conta correta antes do upload."
            )

        source_account_variants = FinancialImportService._normalize_account_variants(
            source_account_metadata.get("account_number")
        )
        target_account_variants = FinancialImportService._normalize_account_variants(account.account_number)
        if source_account_variants and not target_account_variants:
            return (
                "A conta bancária selecionada não possui número da conta cadastrado no sistema. "
                "Preencha o cadastro da conta antes de importar este arquivo."
            )
        if source_account_variants and target_account_variants and source_account_variants.isdisjoint(target_account_variants):
            return (
                "O número da conta identificado no arquivo não corresponde à conta bancária selecionada. "
                "Selecione a conta correta antes do upload."
            )

        source_branch_variants = FinancialImportService._normalize_account_variants(
            source_account_metadata.get("branch_number")
        )
        target_branch_variants = FinancialImportService._normalize_account_variants(account.branch_number)
        if source_branch_variants and target_branch_variants and source_branch_variants.isdisjoint(target_branch_variants):
            return (
                "A agência identificada no arquivo não corresponde à conta bancária selecionada. "
                "Selecione a conta correta antes do upload."
            )

        source_digit = FinancialImportService._normalize_digits(source_account_metadata.get("account_digit"))
        target_digit = FinancialImportService._normalize_digits(account.account_digit)
        if source_digit and target_digit and source_digit != target_digit:
            return (
                "O dígito da conta identificado no arquivo não corresponde à conta bancária selecionada. "
                "Selecione a conta correta antes do upload."
            )

        return None

    @staticmethod
    def get_import_batch_deletion_status(batch: FinancialImportBatch) -> Dict[str, Any]:
        active_row_ids = [
            int(item.id)
            for item in FinancialImportRow.query.filter(
                FinancialImportRow.company_id == batch.company_id,
                FinancialImportRow.import_batch_id == batch.id,
                FinancialImportRow.deleted_at.is_(None),
            ).all()
        ]
        confirmed_matches = 0
        active_created_entries = 0
        if active_row_ids:
            confirmed_matches = FinancialReconciliationMatch.query.filter(
                FinancialReconciliationMatch.company_id == batch.company_id,
                FinancialReconciliationMatch.import_batch_id == batch.id,
                FinancialReconciliationMatch.import_row_id.in_(active_row_ids),
                FinancialReconciliationMatch.match_status == "confirmed",
                FinancialReconciliationMatch.deleted_at.is_(None),
            ).count()
            created_entry_ids = [
                int(item[0])
                for item in FinancialImportRow.query.with_entities(FinancialImportRow.created_entry_id).filter(
                    FinancialImportRow.company_id == batch.company_id,
                    FinancialImportRow.import_batch_id == batch.id,
                    FinancialImportRow.deleted_at.is_(None),
                    FinancialImportRow.created_entry_id.isnot(None),
                ).all()
                if item[0]
            ]
            if created_entry_ids:
                active_created_entries = FinancialEntry.query.filter(
                    FinancialEntry.company_id == batch.company_id,
                    FinancialEntry.id.in_(created_entry_ids),
                    FinancialEntry.deleted_at.is_(None),
                ).count()

        can_delete = confirmed_matches == 0 and active_created_entries == 0
        blocked_reasons: List[str] = []
        if confirmed_matches:
            blocked_reasons.append(f"{confirmed_matches} linha(s) já conciliada(s)")
        if active_created_entries:
            blocked_reasons.append(f"{active_created_entries} lançamento(s) criado(s) pela conciliação")
        return {
            "can_delete": can_delete,
            "confirmed_matches": confirmed_matches,
            "created_entries": active_created_entries,
            "blocked_reasons": blocked_reasons,
        }

    @staticmethod
    def serialize_import_batch(batch: FinancialImportBatch) -> Dict[str, Any]:
        payload = batch.to_dict()
        if getattr(batch, "id", None) and getattr(batch, "company_id", None):
            payload["deletion"] = FinancialImportService.get_import_batch_deletion_status(batch)
        return payload

    @staticmethod
    def _parse_xlsx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl não está disponível para leitura XLSX.") from exc

        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(item).strip() if item is not None else f"column_{idx+1}" for idx, item in enumerate(rows[0])]
        items: List[Dict[str, Any]] = []
        for values in rows[1:]:
            if values is None:
                continue
            item = {}
            for idx, header in enumerate(headers):
                item[header] = values[idx] if idx < len(values) else None
            if any(value not in (None, "") for value in item.values()):
                items.append(item)
        return items

    @staticmethod
    def _parse_xls_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError("xlrd não está disponível para leitura XLS.") from exc

        workbook = xlrd.open_workbook(file_contents=file_bytes)
        if workbook.nsheets <= 0:
            return []

        sheet = workbook.sheet_by_index(0)
        if sheet.nrows <= 0:
            return []

        headers = [
            str(sheet.cell_value(0, idx)).strip() if sheet.cell_value(0, idx) not in (None, "") else f"column_{idx+1}"
            for idx in range(sheet.ncols)
        ]
        items: List[Dict[str, Any]] = []
        for row_idx in range(1, sheet.nrows):
            item: Dict[str, Any] = {}
            has_value = False
            for col_idx, header in enumerate(headers):
                cell = sheet.cell(row_idx, col_idx)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                if value not in (None, ""):
                    has_value = True
                item[header] = value
            if has_value:
                items.append(item)
        return items

    @staticmethod
    def _extract_ofx_tag(block: str, tag: str) -> Optional[str]:
        match = re.search(rf"<{tag}>([^\r\n<]+)", block, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return None

    @staticmethod
    def _parse_ofx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
        text = file_bytes.decode("latin-1", errors="ignore")
        blocks = re.findall(r"<STMTTRN>(.*?)</STMTTRN>", text, flags=re.IGNORECASE | re.DOTALL)
        items: List[Dict[str, Any]] = []
        for block in blocks:
            dtposted = FinancialImportService._extract_ofx_tag(block, "DTPOSTED")
            items.append(
                {
                    "date": dtposted[:8] if dtposted else None,
                    "amount": FinancialImportService._extract_ofx_tag(block, "TRNAMT"),
                    "fitid": FinancialImportService._extract_ofx_tag(block, "FITID"),
                    "memo": FinancialImportService._extract_ofx_tag(block, "MEMO"),
                    "payee": FinancialImportService._extract_ofx_tag(block, "NAME"),
                    "document_number": FinancialImportService._extract_ofx_tag(block, "CHECKNUM"),
                }
            )
        return items

    @staticmethod
    def _parse_source_rows(source_type: str, file_bytes: bytes) -> List[Dict[str, Any]]:
        source = (source_type or "").lower()
        if source == "csv":
            return FinancialImportService._parse_csv_bytes(file_bytes, delimiter=",")
        if source == "csc":
            return FinancialImportService._parse_csv_bytes(file_bytes, delimiter=";")
        if source == "xlsx":
            return FinancialImportService._parse_xlsx_bytes(file_bytes)
        if source == "xls":
            return FinancialImportService._parse_xls_bytes(file_bytes)
        if source == "ofx":
            return FinancialImportService._parse_ofx_bytes(file_bytes)
        raise ValueError("Fonte de importação não suportada. Utilize CSV, XLS, XLSX ou OFX.")

    @staticmethod
    def build_import_template() -> Tuple[Optional[bytes], Optional[str]]:
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError as exc:
            return None, f"openpyxl não está disponível para geração do modelo XLSX: {exc}"

        workbook = Workbook()
        ws_instructions = workbook.active
        ws_instructions.title = "Instruções"
        ws_data = workbook.create_sheet("Importação")
        ws_instructions.sheet_view.showGridLines = False

        title_fill = PatternFill("solid", fgColor="0F766E")
        header_fill = PatternFill("solid", fgColor="E0F2FE")
        required_fill = PatternFill("solid", fgColor="DCFCE7")
        optional_fill = PatternFill("solid", fgColor="F8FAFC")
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        for column, width in {
            "A": 8,
            "B": 14,
            "C": 14,
            "D": 14,
            "E": 4,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 14,
        }.items():
            ws_instructions.column_dimensions[column].width = width

        ws_instructions.row_dimensions[1].height = 30
        ws_instructions.row_dimensions[2].height = 24
        ws_instructions.row_dimensions[4].height = 22
        ws_instructions.row_dimensions[5].height = 36
        ws_instructions.row_dimensions[7].height = 22

        ws_instructions.merge_cells("A1:D1")
        ws_instructions.merge_cells("A2:D2")
        ws_instructions.merge_cells("A5:I5")
        ws_instructions.merge_cells("F1:I2")

        ws_instructions["A1"] = "Versus Gestão Corporativa"
        ws_instructions["A1"].font = Font(size=16, bold=True, color="0F172A")
        ws_instructions["A1"].alignment = Alignment(vertical="center")
        ws_instructions["A2"] = "Modelo de importação financeira APP32"
        ws_instructions["A2"].font = Font(size=12, bold=True, color="0F766E")
        ws_instructions["A2"].alignment = Alignment(vertical="center")
        ws_instructions["F1"].fill = PatternFill("solid", fgColor="000000")
        ws_instructions["A4"] = "Objetivo"
        ws_instructions["A4"].font = Font(bold=True)
        ws_instructions["A5"] = (
            "Use esta planilha para preparar dados de agendamentos e lançamentos financeiros. "
            "Após preencher, faça o upload no Hub de Importação do APP32."
        )
        ws_instructions["A5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_instructions["A7"] = "Regras de preenchimento"
        ws_instructions["A7"].font = Font(bold=True)
        rules = [
            "Preencha obrigatoriamente as colunas marcadas com *.",
            "Datas podem ser digitadas como dd/mm/aaaa ou ddmmaaaa.",
            "Valores devem ser informados no padrão brasileiro, por exemplo: 1500,75.",
            "Tipo de Registro: use 'agendamento' para conta a pagar / receber. Use 'lancamento' para conta já paga / já recebida.",
            "Tipo do Título: use 'Pagar' para contas a pagar e 'Receber' para contas a receber.",
            "Plano de Conta, Centro de Resultado e Projeto / Processo: informe apenas itens analíticos. O sistema deve aceitar código completo ou código reduzido. Exemplos: Plano de Conta 3.01.001 ou 145; Centro de Resultado 2.03.001 ou 245.",
            "Em Projeto / Processo, utilize apenas itens habilitados no Financeiro.",
            "Não altere a ordem das colunas da aba Importação.",
        ]
        for offset, rule in enumerate(rules, start=8):
            ws_instructions[f"A{offset}"] = f"{offset - 7}."
            ws_instructions[f"A{offset}"].alignment = Alignment(horizontal="right", vertical="top")
            ws_instructions.merge_cells(f"B{offset}:I{offset}")
            ws_instructions[f"B{offset}"] = rule
            ws_instructions[f"B{offset}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws_instructions.row_dimensions[offset].height = 34 if len(rule) < 120 else 50

        ws_instructions["A18"] = "Legenda"
        ws_instructions["A18"].font = Font(bold=True)
        ws_instructions["A19"] = "Obrigatório"
        ws_instructions["A19"].fill = required_fill
        ws_instructions["A20"] = "Opcional"
        ws_instructions["A20"].fill = optional_fill
        ws_instructions["A22"] = "Exemplo de preenchimento"
        ws_instructions["A22"].font = Font(bold=True)

        example_header_row = 23
        example_value_row = 24

        image_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "img")
        logo_candidates = [
            os.path.join(image_dir, "logo-versus-preta.png"),
            os.path.join(image_dir, "logo-versus-clara.png"),
            os.path.join(image_dir, "versus-logo.png"),
            os.path.join(image_dir, "logo-versus.png"),
            os.path.join(image_dir, "logo-versus-slogan.png"),
        ]
        logo_path = next((path for path in logo_candidates if os.path.exists(path)), None)
        if logo_path:
            try:
                image = XLImage(logo_path)
                image.width = 220
                image.height = 39
                ws_instructions.add_image(image, "F1")
            except Exception:
                logger.exception("Falha ao inserir logo no modelo financeiro XLSX")

        for col_idx, column in enumerate(FinancialImportService.IMPORT_TEMPLATE_COLUMNS, start=1):
            instruction_header = ws_instructions.cell(row=example_header_row, column=col_idx, value=column["label"])
            instruction_header.font = Font(bold=True, color="FFFFFF")
            instruction_header.fill = title_fill
            instruction_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            instruction_header.border = border

            instruction_sample = ws_instructions.cell(row=example_value_row, column=col_idx, value=column["example"])
            instruction_sample.font = Font(italic=True, color="475569")
            instruction_sample.fill = header_fill
            instruction_sample.alignment = Alignment(wrap_text=True, vertical="top")
            instruction_sample.border = border

            cell = ws_data.cell(row=1, column=col_idx, value=column["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            ws_data.column_dimensions[chr(64 + col_idx if col_idx <= 26 else 65)].width = max(18, min(34, len(column["label"]) + 6))
            ws_instructions.column_dimensions[chr(64 + col_idx if col_idx <= 26 else 65)].width = max(
                ws_instructions.column_dimensions[chr(64 + col_idx if col_idx <= 26 else 65)].width or 0,
                max(18, min(34, len(column["label"]) + 6)),
            )

        ws_data.freeze_panes = "A2"
        ws_data.sheet_view.showGridLines = True

        tipo_registro_validation = DataValidation(
            type="list",
            formula1='"agendamento,lancamento"',
            allow_blank=False,
        )
        tipo_titulo_validation = DataValidation(
            type="list",
            formula1='"Pagar,Receber"',
            allow_blank=False,
        )
        ws_data.add_data_validation(tipo_registro_validation)
        ws_data.add_data_validation(tipo_titulo_validation)
        tipo_registro_validation.add("A2:A500")
        tipo_titulo_validation.add("B2:B500")

        for row in range(2, 501):
            for col_idx, column in enumerate(FinancialImportService.IMPORT_TEMPLATE_COLUMNS, start=1):
                cell = ws_data.cell(row=row, column=col_idx)
                cell.border = border
                if cell.value in (None, ""):
                    cell.fill = required_fill if column["required"] else optional_fill

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue(), None

    @staticmethod
    def create_import_batch(
        *,
        payload: Dict[str, Any],
        file_bytes: bytes,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        try:
            data = FinancialImportBatchInput(**payload)
        except Exception as exc:
            return None, f"Payload inválido para lote de importação: {str(exc)}"

        scope_error = FinancialService._ensure_company_scope(data.company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        try:
            parsed_rows = FinancialImportService._parse_source_rows(data.source_type, file_bytes)
        except Exception as exc:
            return None, f"Falha ao interpretar arquivo importado: {str(exc)}"

        source_account_metadata = FinancialImportService._extract_source_account_metadata(
            data.source_type,
            file_bytes,
            parsed_rows,
        )
        validation_error = FinancialImportService._validate_batch_bank_account(
            company_id=data.company_id,
            metadata_json=data.metadata_json or {},
            source_account_metadata=source_account_metadata,
        )
        if validation_error:
            return None, validation_error

        try:
            batch_payload = data.model_dump(exclude={"file_hash"})
            batch_metadata = dict(batch_payload.get("metadata_json") or {})
            batch_metadata["source_account"] = source_account_metadata
            batch_payload["metadata_json"] = batch_metadata
            batch = FinancialImportBatch(
                **batch_payload,
                file_hash=hashlib.sha256(file_bytes).hexdigest(),
                status="parsed",
            )
            db.session.add(batch)
            db.session.flush()

            staged_rows: List[FinancialImportRow] = []
            valid_rows = 0
            error_rows = 0

            selected_bank_account_id = int((data.metadata_json or {}).get("bank_account_id") or 0) or None

            for idx, raw_row in enumerate(parsed_rows, start=1):
                row_input = FinancialImportService._normalize_row(idx, raw_row)
                row_payload = row_input.model_dump()
                row_payload["company_id"] = data.company_id
                row_payload["import_batch_id"] = batch.id
                row = FinancialImportRow(**row_payload)
                db.session.add(row)
                staged_rows.append(row)
                FinancialImportService._enrich_row_catalogs(data.company_id, row)
                force_selected_bank_account = data.source_type == "ofx" and selected_bank_account_id
                fallback_bank_account_id = selected_bank_account_id
                if force_selected_bank_account:
                    row.normalized_payload = {
                        **(row.normalized_payload or {}),
                        "bank_account_id": int(fallback_bank_account_id),
                    }
                    if row.processing_status == "staged":
                        row.processing_status = "validated"
                elif fallback_bank_account_id and not (row.normalized_payload or {}).get("bank_account_id"):
                    row.normalized_payload = {
                        **(row.normalized_payload or {}),
                        "bank_account_id": int(fallback_bank_account_id),
                    }
                    if row.processing_status == "staged":
                        row.processing_status = "validated"
                if row.processing_status == "rejected":
                    error_rows += 1
                else:
                    valid_rows += 1

            batch.total_rows = len(staged_rows)
            batch.valid_rows = valid_rows
            batch.error_rows = error_rows
            batch.finished_at = datetime.utcnow()

            db.session.commit()
            return {
                "batch": FinancialImportService.serialize_import_batch(batch),
                "rows": [row.to_dict() for row in staged_rows],
            }, None
        except Exception as exc:
            db.session.rollback()
            logger.exception("Erro ao criar lote de importação financeira")
            return None, f"Erro ao criar lote de importação: {str(exc)}"

    @staticmethod
    def list_import_batches(
        *,
        company_id: int,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        batches = FinancialImportBatch.query.filter(
            FinancialImportBatch.company_id == company_id,
            FinancialImportBatch.deleted_at.is_(None),
        ).order_by(FinancialImportBatch.imported_at.desc(), FinancialImportBatch.id.desc()).all()
        return [FinancialImportService.serialize_import_batch(batch) for batch in batches], None

    @staticmethod
    def get_import_batch(
        *,
        batch_id: int,
        company_id: int,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        batch = FinancialImportBatch.query.filter(
            FinancialImportBatch.id == batch_id,
            FinancialImportBatch.company_id == company_id,
            FinancialImportBatch.deleted_at.is_(None),
        ).first()
        if not batch:
            return None, "Lote de importação não encontrado no escopo da empresa."

        rows = FinancialImportRow.query.filter(
            FinancialImportRow.company_id == company_id,
            FinancialImportRow.import_batch_id == batch.id,
            FinancialImportRow.deleted_at.is_(None),
        ).order_by(FinancialImportRow.row_number.asc()).all()
        matches = FinancialReconciliationMatch.query.filter(
            FinancialReconciliationMatch.company_id == company_id,
            FinancialReconciliationMatch.import_batch_id == batch.id,
            FinancialReconciliationMatch.deleted_at.is_(None),
        ).order_by(FinancialReconciliationMatch.id.asc()).all()
        suggestions = FinancialClassificationSuggestion.query.filter(
            FinancialClassificationSuggestion.company_id == company_id,
            FinancialClassificationSuggestion.import_batch_id == batch.id,
            FinancialClassificationSuggestion.deleted_at.is_(None),
        ).order_by(
            FinancialClassificationSuggestion.import_row_id.asc(),
            FinancialClassificationSuggestion.rank_position.asc(),
        ).all()
        return {
            "batch": FinancialImportService.serialize_import_batch(batch),
            "rows": [row.to_dict() for row in rows],
            "matches": [match.to_dict() for match in matches],
            "suggestions": [suggestion.to_dict() for suggestion in suggestions],
        }, None

    @staticmethod
    def delete_import_batch(
        *,
        batch_id: int,
        company_id: int,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        batch = FinancialImportBatch.query.filter(
            FinancialImportBatch.id == batch_id,
            FinancialImportBatch.company_id == company_id,
            FinancialImportBatch.deleted_at.is_(None),
        ).first()
        if not batch:
            return None, "Lote de importação não encontrado no escopo da empresa."

        deletion_status = FinancialImportService.get_import_batch_deletion_status(batch)
        if not deletion_status.get("can_delete"):
            reasons = ", ".join(deletion_status.get("blocked_reasons") or []) or "há conciliação vinculada"
            return None, f"Este upload não pode ser excluído porque {reasons}."

        try:
            now = datetime.utcnow()
            deleted_matches = FinancialReconciliationMatch.query.filter(
                FinancialReconciliationMatch.company_id == company_id,
                FinancialReconciliationMatch.import_batch_id == batch.id,
                FinancialReconciliationMatch.deleted_at.is_(None),
            ).update(
                {"deleted_at": now, "updated_at": now},
                synchronize_session=False,
            )
            deleted_suggestions = FinancialClassificationSuggestion.query.filter(
                FinancialClassificationSuggestion.company_id == company_id,
                FinancialClassificationSuggestion.import_batch_id == batch.id,
                FinancialClassificationSuggestion.deleted_at.is_(None),
            ).update(
                {"deleted_at": now, "updated_at": now},
                synchronize_session=False,
            )
            deleted_rows = FinancialImportRow.query.filter(
                FinancialImportRow.company_id == company_id,
                FinancialImportRow.import_batch_id == batch.id,
                FinancialImportRow.deleted_at.is_(None),
            ).update(
                {"deleted_at": now, "updated_at": now},
                synchronize_session=False,
            )
            batch.deleted_at = now
            batch.updated_at = now
            batch.status = "cancelled"
            db.session.commit()
            return {
                "batch_id": batch.id,
                "batch_code": batch.batch_code,
                "deleted_rows": int(deleted_rows or 0),
                "deleted_matches": int(deleted_matches or 0),
                "deleted_suggestions": int(deleted_suggestions or 0),
            }, None
        except Exception as exc:
            db.session.rollback()
            logger.exception("Erro ao excluir lote financeiro %s", batch_id)
            return None, f"Erro ao excluir lote de importação: {str(exc)}"

    @staticmethod
    def _build_entry_payload_from_row(batch: FinancialImportBatch, row: FinancialImportRow) -> Dict[str, Any]:
        normalized = row.normalized_payload or {}
        normalized = dict(normalized)
        resolved_counterparty_id = FinancialImportService._resolve_counterparty_id_by_code(
            batch.company_id,
            normalized.get("counterparty_code"),
        ) if normalized.get("counterparty_code") else None
        if resolved_counterparty_id:
            normalized["counterparty_id"] = resolved_counterparty_id
        resolved_chart_account_id = FinancialImportService._resolve_chart_account_id_by_code(
            batch.company_id,
            normalized.get("chart_account_code"),
        ) if normalized.get("chart_account_code") else None
        if resolved_chart_account_id:
            normalized["chart_account_id"] = resolved_chart_account_id
        resolved_cost_center_id = FinancialImportService._resolve_cost_center_id_by_code(
            batch.company_id,
            normalized.get("cost_center_code"),
        ) if normalized.get("cost_center_code") else None
        if resolved_cost_center_id:
            normalized["cost_center_id"] = resolved_cost_center_id
        resolved_bank_account_id = FinancialImportService._resolve_bank_account_id_by_code(
            batch.company_id,
            normalized.get("bank_account_code"),
        ) if normalized.get("bank_account_code") else None
        if resolved_bank_account_id:
            normalized["bank_account_id"] = resolved_bank_account_id
        occurred_on = row.occurred_on or FinancialImportService._parse_date(normalized.get("occurred_on"))
        due_date = row.due_date or FinancialImportService._parse_date(normalized.get("due_date"))
        competence_date = FinancialImportService._parse_date(normalized.get("competence_date"))

        payload = {
            "company_id": batch.company_id,
            "entry_code": f"{batch.batch_code}-{row.row_number}",
            "entry_type": normalized.get("entry_type") or "bank_movement",
            "movement_nature": normalized.get("movement_nature") or row.movement_nature or "debit",
            "origin_type": FinancialImportService._resolve_origin_type(batch.source_type),
            "status": "posted",
            "review_status": "pending_review",
            "description": row.description or normalized.get("description") or f"Importação {batch.batch_code} linha {row.row_number}",
            "document_number": row.document_number,
            "external_reference": row.bank_reference,
            "origin_reference": batch.batch_code,
            "competence_date": competence_date or occurred_on or due_date or batch.imported_at.date(),
            "due_date": due_date,
            "occurred_on": occurred_on,
            "original_amount": row.amount or Decimal("0"),
            "bank_account_id": normalized.get("bank_account_id"),
            "counterparty_id": normalized.get("counterparty_id"),
            "chart_account_id": normalized.get("chart_account_id"),
            "cost_center_id": normalized.get("cost_center_id"),
            "activity_id": normalized.get("activity_id"),
            "process_instance_id": normalized.get("process_instance_id"),
            "routine_id": normalized.get("routine_id"),
            "notes": f"Gerado a partir do lote {batch.batch_code}, linha {row.row_number}.",
            "metadata_json": {
                "import_batch_id": batch.id,
                "import_row_id": row.id,
                "source_type": batch.source_type,
                "counterparty_name": row.counterparty_name,
                "classification_rule_id": normalized.get("classification_rule_id"),
                "classification_rule_name": normalized.get("classification_rule_name"),
                "raw_payload": row.raw_payload or {},
            },
        }
        return FinancialCatalogService.enrich_reference_payload(
            company_id=batch.company_id,
            payload=payload,
            counterparty_text=row.counterparty_name,
            description_text=row.description,
            bank_reference=row.bank_reference,
        )

    @staticmethod
    def _build_schedule_payload_from_row(batch: FinancialImportBatch, row: FinancialImportRow) -> Dict[str, Any]:
        normalized = FinancialImportService._build_entry_payload_from_row(batch, row)
        competence_date = (
            FinancialImportService._parse_date(normalized.get("competence_date"))
            or FinancialImportService._parse_date(normalized.get("occurred_on"))
            or FinancialImportService._parse_date(normalized.get("due_date"))
            or batch.imported_at.date()
        )
        due_date = (
            FinancialImportService._parse_date(normalized.get("due_date"))
            or FinancialImportService._parse_date(normalized.get("occurred_on"))
            or competence_date
        )
        notes = f"Gerado a partir do lote {batch.batch_code}, linha {row.row_number}."
        return {
            "company_id": batch.company_id,
            "name": (row.description or normalized.get('description') or f'Importação {batch.batch_code}')[:120],
            "entry_type": normalized.get("entry_type") or "payable",
            "movement_nature": normalized.get("movement_nature") or "debit",
            "origin_type": FinancialImportService._resolve_origin_type(batch.source_type),
            "status": "active",
            "frequency": "one_time",
            "interval_value": 1,
            "start_date": competence_date,
            "competence_date": competence_date,
            "first_due_date": due_date,
            "next_due_date": due_date,
            "description": row.description or normalized.get("description") or f"Importação {batch.batch_code} linha {row.row_number}",
            "document_number_prefix": row.document_number,
            "template_amount": row.amount or Decimal("0"),
            "bank_account_id": normalized.get("bank_account_id"),
            "counterparty_id": normalized.get("counterparty_id"),
            "chart_account_id": normalized.get("chart_account_id"),
            "cost_center_id": normalized.get("cost_center_id"),
            "activity_id": normalized.get("activity_id"),
            "process_instance_id": normalized.get("process_instance_id"),
            "routine_id": normalized.get("routine_id"),
            "notes": notes,
            "metadata_json": {
                **(normalized.get("metadata_json") or {}),
                "import_batch_id": batch.id,
                "import_row_id": row.id,
                "source_type": batch.source_type,
                "counterparty_name": row.counterparty_name,
                "classification_rule_id": normalized.get("classification_rule_id"),
                "classification_rule_name": normalized.get("classification_rule_name"),
                "raw_payload": row.raw_payload or {},
            },
        }

    @staticmethod
    def _build_direct_entry_payload_from_row(batch: FinancialImportBatch, row: FinancialImportRow) -> Dict[str, Any]:
        normalized = FinancialImportService._build_entry_payload_from_row(batch, row)
        competence_date = (
            FinancialImportService._parse_date(normalized.get("competence_date"))
            or FinancialImportService._parse_date(normalized.get("due_date"))
            or FinancialImportService._parse_date(normalized.get("occurred_on"))
            or batch.imported_at.date()
        )
        due_date = FinancialImportService._parse_date(normalized.get("due_date"))
        settlement_date = (
            FinancialImportService._parse_date((row.normalized_payload or {}).get("settlement_date"))
            or FinancialImportService._parse_date(normalized.get("occurred_on"))
            or due_date
            or competence_date
        )
        return {
            "company_id": batch.company_id,
            "entry_type": normalized.get("entry_type") or "payable",
            "description": row.description or normalized.get("description") or f"Importação {batch.batch_code} linha {row.row_number}",
            "document_number": row.document_number,
            "counterparty_id": normalized.get("counterparty_id"),
            "bank_account_id": normalized.get("bank_account_id"),
            "competence_date": competence_date,
            "occurred_on": settlement_date,
            "due_date": due_date or settlement_date,
            "original_amount": row.amount or Decimal("0"),
            "chart_account_id": normalized.get("chart_account_id"),
            "cost_center_id": normalized.get("cost_center_id"),
            "notes": f"Gerado a partir do lote {batch.batch_code}, linha {row.row_number}.",
            "metadata_json": {
                "import_batch_id": batch.id,
                "import_row_id": row.id,
                "source_type": batch.source_type,
                "counterparty_name": row.counterparty_name,
                "classification_rule_id": normalized.get("classification_rule_id"),
                "classification_rule_name": normalized.get("classification_rule_name"),
                "raw_payload": row.raw_payload or {},
            },
        }

    @staticmethod
    def process_import_batch(
        *,
        batch_id: int,
        company_id: int,
        allowed_company_ids: Optional[Sequence[int]] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        scope_error = FinancialService._ensure_company_scope(company_id, allowed_company_ids)
        if scope_error:
            return None, scope_error

        batch = FinancialImportBatch.query.filter(
            FinancialImportBatch.id == batch_id,
            FinancialImportBatch.company_id == company_id,
            FinancialImportBatch.deleted_at.is_(None),
        ).first()
        if not batch:
            return None, "Lote de importação não encontrado no escopo da empresa."

        rows = FinancialImportRow.query.filter(
            FinancialImportRow.company_id == company_id,
            FinancialImportRow.import_batch_id == batch.id,
            FinancialImportRow.deleted_at.is_(None),
        ).order_by(FinancialImportRow.row_number.asc()).all()

        created_count = 0
        skipped_count = 0
        rejected_count = 0

        try:
            for row in rows:
                if row.processing_status == "rejected":
                    rejected_count += 1
                    continue
                if row.created_entry_id:
                    skipped_count += 1
                    continue

                existing = FinancialEntry.query.filter(
                    FinancialEntry.company_id == company_id,
                    FinancialEntry.entry_code == f"{batch.batch_code}-{row.row_number}",
                ).first()
                if existing:
                    row.created_entry_id = existing.id
                    row.processing_status = "imported"
                    skipped_count += 1
                    continue

                normalized = dict(row.normalized_payload or {})
                record_type = normalized.get("record_type") or "lancamento"
                if record_type == "agendamento":
                    payload = FinancialImportService._build_schedule_payload_from_row(batch, row)
                    schedule_result, error = FinancialScheduleService.create_schedule(
                        payload=payload,
                        allowed_company_ids=allowed_company_ids,
                    )
                    if error:
                        row.processing_status = "rejected"
                        row.error_message = error
                        rejected_count += 1
                        continue
                    normalized["created_schedule_id"] = schedule_result.get("id")
                    normalized["created_schedule_code"] = schedule_result.get("schedule_code")
                    row.normalized_payload = normalized
                else:
                    payload = FinancialImportService._build_direct_entry_payload_from_row(batch, row)
                    result, error = FinancialDirectEntryService.create_direct_entry(
                        payload=payload,
                        allowed_company_ids=allowed_company_ids,
                    )
                    if error:
                        row.processing_status = "rejected"
                        row.error_message = error
                        rejected_count += 1
                        continue
                    row.created_entry_id = (result or {}).get("entry", {}).get("id")
                    normalized["created_schedule_id"] = (result or {}).get("schedule", {}).get("id")
                    normalized["created_schedule_code"] = (result or {}).get("schedule", {}).get("schedule_code")
                    normalized["created_settlement_id"] = (result or {}).get("settlement", {}).get("id")
                    row.normalized_payload = normalized
                row.processing_status = "imported"
                row.error_message = None
                created_count += 1

            batch.status = "processed_with_errors" if rejected_count else "processed"
            batch.finished_at = datetime.utcnow()
            batch.error_rows = rejected_count
            batch.valid_rows = sum(1 for row in rows if row.processing_status in {"validated", "imported", "staged"})
            db.session.commit()
            return {
                "batch": batch.to_dict(),
                "summary": {
                    "created_count": created_count,
                    "skipped_count": skipped_count,
                    "rejected_count": rejected_count,
                    "total_rows": len(rows),
                },
            }, None
        except Exception as exc:
            db.session.rollback()
            logger.exception("Erro ao processar lote financeiro %s", batch_id)
            return None, f"Erro ao processar lote de importação: {str(exc)}"
