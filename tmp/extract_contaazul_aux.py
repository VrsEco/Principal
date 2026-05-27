import json
import re
import unicodedata
from collections import OrderedDict
from openpyxl import load_workbook

path = r'C:\Users\mff20\OneDrive\Versus\Versus Consultoria\Versus - Clientes\AL - Save Water\Transição ContaAzul - Versus\ContaAzul - Final.xlsx'
wb = load_workbook(path, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
headers = [str(h).strip() if h is not None else '' for h in next(rows)]

def val(row, name, occurrence=1):
    hits = [i for i,h in enumerate(headers) if h == name]
    if len(hits) >= occurrence:
        return row[hits[occurrence-1]]
    return None

def clean_text(s):
    s = '' if s is None else str(s).strip()
    return s

def normalize_key(s):
    s = clean_text(s)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = s.lower()
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def category_parts(text):
    text = clean_text(text)
    if ' - ' in text:
        code, name = text.split(' - ', 1)
        return code.strip(), name.strip()
    return '', text

counterparties = OrderedDict()
categories = OrderedDict()
for row in rows:
    if not row:
        continue
    name = clean_text(val(row, 'Nome do fornecedor/cliente'))
    doc = clean_text(val(row, 'Identificador do fornecedor/cliente'))
    if name:
        key = normalize_key(name) + '|' + doc
        if key not in counterparties:
            counterparties[key] = {'name': name, 'document_number': doc or None, 'count': 0}
        counterparties[key]['count'] += 1
    for occ in (1,2,3):
        cat = clean_text(val(row, f'Categoria {occ}', 1))
        if cat:
            code, namep = category_parts(cat)
            key = normalize_key(namep)
            if key not in categories:
                categories[key] = {'source_text': cat, 'source_code': code, 'source_name': namep, 'count': 0}
            categories[key]['count'] += 1

out = {
    'counterparties': list(counterparties.values()),
    'categories': list(categories.values()),
}
print(json.dumps(out, ensure_ascii=False, indent=2))
