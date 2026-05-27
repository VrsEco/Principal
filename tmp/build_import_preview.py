import csv
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r'C:\GestaoVersus\app32')
XLSX = Path(r'C:\Users\mff20\OneDrive\Versus\Versus Consultoria\Versus - Clientes\AL - Save Water\Transição ContaAzul - Versus\ContaAzul - Final.xlsx')
OUT = BASE / 'outputs'
OUT.mkdir(exist_ok=True)

BANK_MAP = {
    'itau - conta corrente': {'target_name': 'ITAU - STELLA BAPTISTA', 'target_id': 7, 'reason': 'user_mapping'},
    'save water comercio e servicos': {'target_name': 'ITAU - SAVE WATER', 'target_id': 8, 'reason': 'user_mapping'},
    'corporativo - vcto dia 05': {'target_name': 'Corporativo - Vcto dia 05', 'target_id': 13, 'reason': 'direct_name'},
    'corporativo vcto dia 15': {'target_name': 'Corporativo - Vcto dia 15', 'target_id': 14, 'reason': 'user_mapping'},
    'compensacoes e retencoes': {'target_name': 'COMPENSAÇÃO E RETENÇÃO', 'target_id': 12, 'reason': 'normalized_name'},
}
DEFAULT_PAYMENT_METHOD = {'target_name': 'GERAL', 'target_id': 4}
DEFAULT_COST_CENTER = {'target_name': 'Geral', 'target_code': '7.1', 'target_id': 23}
SETTLED_STATUSES = {'quitado', 'conciliado'}
OPEN_STATUSES = {'em aberto', 'atrasado'}
TRANSFER_HINTS = {'transferencia de entrada', 'transferencia de saida', 'rec - contas de compensacao'}


def read_json_any(path: Path):
    raw = path.read_bytes()
    for enc in ('utf-8', 'utf-8-sig', 'cp1252', 'latin-1'):
        try:
            return json.loads(raw.decode(enc))
        except Exception:
            continue
    raise RuntimeError(f'Cannot decode {path}')


def norm(value):
    value = '' if value is None else str(value).strip()
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii').lower()
    value = re.sub(r'\s+', ' ', value)
    return value.strip()


def parse_date(value):
    if value in (None, ''):
        return None
    if hasattr(value, 'date'):
        try:
            return value.date().isoformat()
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def money_abs(value):
    if value in (None, ''):
        return 0.0
    try:
        val = float(value)
    except Exception:
        text = str(value).replace('R$', '').replace('.', '').replace(',', '.').strip()
        val = float(text) if text else 0.0
    return round(abs(val), 2)


def money_signed(value):
    if value in (None, ''):
        return 0.0
    try:
        val = float(value)
    except Exception:
        text = str(value).replace('R$', '').replace('.', '').replace(',', '.').strip()
        val = float(text) if text else 0.0
    return round(val, 2)

category_map_raw = read_json_any(BASE / 'tmp' / 'aux_prepare_result.json')
category_map = {}
for item in category_map_raw['category_map']:
    src = item['source']['source_text']
    category_map[norm(src)] = {
        'target_id': item['target']['id'],
        'target_code': item['target']['code'],
        'target_name': item['target']['name'].strip(),
        'reason': item['reason'],
    }

wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
headers = [str(h).strip() if h is not None else '' for h in next(rows)]
header_hits = defaultdict(list)
for i, h in enumerate(headers):
    header_hits[h].append(i)

def get(row, header, occurrence=1):
    hits = header_hits.get(header, [])
    if len(hits) >= occurrence:
        idx = hits[occurrence - 1]
        return row[idx] if idx < len(row) else None
    return None

preview_rows = []
rateio_rows = []
summary = Counter()
by_status = Counter()
by_class = Counter()
by_origin = Counter()
issues = []

for excel_row, row in enumerate(rows, start=2):
    if not row or not any(v not in (None, '') for v in row):
        continue

    descricao = str(get(row, 'Descrição') or '').strip()
    favorecido = str(get(row, 'Nome do fornecedor/cliente') or '').strip()
    identificador = str(get(row, 'Identificador do fornecedor/cliente') or '').strip() or None
    tipo = str(get(row, 'Tipo') or '').strip()
    origem = str(get(row, 'Origem do lançamento') or '').strip()
    situacao = str(get(row, 'Situação') or '').strip()
    conta_bancaria_src = str(get(row, 'Conta bancária') or '').strip()
    forma_src = str(get(row, 'Forma de pgto/recbto') or '').strip()

    movement_date = parse_date(get(row, 'Data movimento'))
    competence_date = parse_date(get(row, 'Data de competência')) or movement_date
    due_date = parse_date(get(row, 'Data original de vencimento')) or parse_date(get(row, 'Data prevista')) or movement_date
    predicted_date = parse_date(get(row, 'Data prevista'))

    valor = money_signed(get(row, 'Valor (R$)'))
    valor_original = money_signed(get(row, 'Valor original (R$)'))
    juros = money_abs(get(row, 'Juros (R$)'))
    multa = money_abs(get(row, 'Multa (R$)'))
    desconto = money_abs(get(row, 'Desconto (R$)'))
    taxas = money_abs(get(row, 'Taxas (R$)'))

    categories = []
    total_rateio = 0.0
    for occ in (1, 2, 3):
        cat_text = get(row, f'Categoria {occ}', 1)
        cat_val = get(row, f'Valor na Categoria {occ}', 1)
        cat_text = str(cat_text).strip() if cat_text not in (None, '') else ''
        if not cat_text:
            continue
        alloc = money_abs(cat_val)
        total_rateio += alloc
        mapped = category_map.get(norm(cat_text))
        categories.append({
            'source_text': cat_text,
            'allocation_amount': alloc,
            'mapped_account_id': mapped['target_id'] if mapped else None,
            'mapped_account_code': mapped['target_code'] if mapped else None,
            'mapped_account_name': mapped['target_name'] if mapped else None,
            'map_reason': mapped['reason'] if mapped else None,
        })

    normalized_status = norm(situacao)
    normalized_origin = norm(origem)
    category_norms = {norm(c['source_text']) for c in categories}

    if normalized_origin == 'transferencia' or category_norms & TRANSFER_HINTS:
        movement_class = 'transfer'
        title_kind = 'transfer'
    elif norm(tipo) == 'receita':
        movement_class = 'receipt'
        title_kind = 'receivable'
    elif norm(tipo) == 'despesa':
        movement_class = 'payment'
        title_kind = 'payable'
    else:
        movement_class = 'unknown'
        title_kind = 'unknown'

    if movement_class == 'transfer' and 'compensacao' in ' '.join(sorted(category_norms)):
        movement_class = 'compensation'
        title_kind = 'transfer'

    gross_title_amount = money_abs(valor_original if valor_original != 0 else valor)
    if gross_title_amount == 0 and total_rateio > 0:
        gross_title_amount = round(total_rateio, 2)

    if normalized_status in SETTLED_STATUSES:
        has_settlement = True
        open_balance = 0.0
        settlement_date = movement_date
        settlement_bank = BANK_MAP.get(norm(conta_bancaria_src))
        settlement_principal_amount = gross_title_amount
        settlement_net_amount = money_abs(valor)
    elif normalized_status in OPEN_STATUSES:
        has_settlement = False
        open_balance = gross_title_amount
        settlement_date = None
        settlement_bank = None
        settlement_principal_amount = 0.0
        settlement_net_amount = 0.0
    else:
        has_settlement = movement_class in {'transfer', 'compensation'}
        open_balance = 0.0 if has_settlement else gross_title_amount
        settlement_date = movement_date if has_settlement else None
        settlement_bank = BANK_MAP.get(norm(conta_bancaria_src)) if has_settlement else None
        settlement_principal_amount = gross_title_amount if has_settlement else 0.0
        settlement_net_amount = money_abs(valor) if has_settlement else 0.0

    if movement_class in {'transfer', 'compensation'} and not settlement_bank and conta_bancaria_src:
        settlement_bank = BANK_MAP.get(norm(conta_bancaria_src))

    if not categories and movement_class not in {'transfer', 'compensation'}:
        issues.append({'row': excel_row, 'issue': 'no_category_rateio', 'descricao': descricao})
    if any(c['mapped_account_id'] is None for c in categories if movement_class not in {'transfer', 'compensation'}):
        issues.append({'row': excel_row, 'issue': 'unmapped_category', 'descricao': descricao})
    if has_settlement and not settlement_bank and conta_bancaria_src:
        issues.append({'row': excel_row, 'issue': 'unmapped_bank_account', 'conta_bancaria': conta_bancaria_src})

    row_result = {
        'excel_row': excel_row,
        'descricao': descricao,
        'favorecido': favorecido,
        'identificador': identificador,
        'tipo_origem': tipo,
        'origem_lancamento': origem,
        'situacao': situacao,
        'movement_class': movement_class,
        'title_kind': title_kind,
        'competence_date': competence_date,
        'due_date': due_date,
        'predicted_date': predicted_date,
        'movement_date': movement_date,
        'gross_title_amount': gross_title_amount,
        'rateio_total': round(total_rateio, 2),
        'has_settlement': has_settlement,
        'settlement_date': settlement_date,
        'settlement_bank_account_source': conta_bancaria_src,
        'settlement_bank_account_target_id': settlement_bank['target_id'] if settlement_bank else None,
        'settlement_bank_account_target_name': settlement_bank['target_name'] if settlement_bank else None,
        'payment_method_source': forma_src,
        'payment_method_target_id': DEFAULT_PAYMENT_METHOD['target_id'],
        'payment_method_target_name': DEFAULT_PAYMENT_METHOD['target_name'],
        'cost_center_target_id': DEFAULT_COST_CENTER['target_id'],
        'cost_center_target_code': DEFAULT_COST_CENTER['target_code'],
        'cost_center_target_name': DEFAULT_COST_CENTER['target_name'],
        'principal_settlement_amount': round(settlement_principal_amount, 2),
        'net_settlement_amount': round(settlement_net_amount, 2),
        'interest_amount': juros,
        'penalty_amount': multa,
        'discount_amount': desconto,
        'fee_amount': taxas,
        'open_balance_amount': round(open_balance, 2),
        'category_count': len(categories),
        'notes': '',
    }

    if movement_class in {'transfer', 'compensation'}:
        row_result['notes'] = 'Movimento fora do plano operacional de receita/despesa; tratar como transferência/compensação.'
    elif normalized_status in OPEN_STATUSES:
        row_result['notes'] = 'Título em aberto sem baixa.'
    elif normalized_status in SETTLED_STATUSES:
        row_result['notes'] = 'Título liquidado com baixa identificada.'

    preview_rows.append(row_result)

    for seq, cat in enumerate(categories, start=1):
        rateio_rows.append({
            'excel_row': excel_row,
            'allocation_seq': seq,
            'movement_class': movement_class,
            'source_category': cat['source_text'],
            'allocation_amount': cat['allocation_amount'],
            'target_account_id': cat['mapped_account_id'],
            'target_account_code': cat['mapped_account_code'],
            'target_account_name': cat['mapped_account_name'],
            'map_reason': cat['map_reason'],
        })

    summary['rows'] += 1
    by_status[situacao] += 1
    by_class[movement_class] += 1
    by_origin[origem] += 1

# outputs
preview_csv = OUT / 'save_water_financial_import_preview.csv'
rateio_csv = OUT / 'save_water_financial_import_rateio.csv'
summary_json = OUT / 'save_water_financial_import_summary.json'
issues_csv = OUT / 'save_water_financial_import_issues.csv'

with open(preview_csv, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=list(preview_rows[0].keys()))
    writer.writeheader(); writer.writerows(preview_rows)
with open(rateio_csv, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=list(rateio_rows[0].keys()))
    writer.writeheader(); writer.writerows(rateio_rows)
with open(issues_csv, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=['row', 'issue', 'descricao', 'conta_bancaria'])
    writer.writeheader(); writer.writerows(issues)
summary_payload = {
    'rows_total': summary['rows'],
    'by_class': dict(by_class),
    'by_status': dict(by_status),
    'by_origin': dict(by_origin),
    'with_settlement': sum(1 for r in preview_rows if r['has_settlement']),
    'without_settlement': sum(1 for r in preview_rows if not r['has_settlement']),
    'rateio_rows_total': len(rateio_rows),
    'rows_with_rateio': sum(1 for r in preview_rows if r['category_count'] > 0),
    'rows_without_rateio': sum(1 for r in preview_rows if r['category_count'] == 0),
    'issues_total': len(issues),
    'issue_breakdown': dict(Counter(i['issue'] for i in issues)),
    'files': {
        'preview_csv': str(preview_csv),
        'rateio_csv': str(rateio_csv),
        'issues_csv': str(issues_csv),
    },
    'sample_preview': preview_rows[:10],
}
summary_json.write_text(json.dumps(summary_payload, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(summary_payload, ensure_ascii=False, indent=2))
